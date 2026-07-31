import json
import os
from datetime import datetime, time, timedelta
from pathlib import Path
from decimal import Decimal, InvalidOperation
from io import StringIO

from django.contrib import messages
from django.contrib.auth.models import User
from decouple import config
from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.core.management import call_command
from django.db.models import Count, Q, Sum
from django.http import Http404
from django.http import JsonResponse
from django.http import HttpResponse, QueryDict
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_http_methods, require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import requests

from .forms import (
    AttachmentUploadForm,
    BusinessUnitCreateForm,
    ChannelCreateForm,
    DailyAdSpendForm,
    DailyChannelSaleForm,
    ExportRequestForm,
    GlobalFilterForm,
    MasterImportForm,
    MarketplaceDailyForm,
    OperationalGoalTaskUpdateForm,
    ProfileForm,
    WeeklyTaskFilterForm,
)
from .models import AdPlatform, Attachment, BusinessUnit, Channel, Country, DailyAdSpend, DailyChannelSale, InsightAchievement, OperationalGoalTask, Product, SalesTarget, UserProfile, UserTask, Website, WebsiteHealthCheck
from .services.analytics import attachments, build_dashboard_summary, build_filter_dict, build_unit_summary, weekly_tasks
from .services.excel_master import build_master_workbook, commit_master_import, create_export_job, preview_master_import
from .services.marketplace_inventory import marketplace_inventory_snapshot
from .services.sales_dashboard import build_ad_platform_performance, build_awn_international_snapshot, build_bali_product_detail, build_bali_snapshot, build_comfama_snapshot, build_copa_uva_country_comparison, build_ecuador_snapshot, build_marketplace_product_detail, build_sales_snapshot, build_uva_category_country_comparison, build_uva_category_snapshot, build_uva_geo_map_data, build_uva_meta_ads_preview, build_uva_product_detail, remove_colombia_vat
from .query_cache import memoize_per_request
from .services.website_monitor import latest_checks_by_website

MASTER_IMPORT_SESSION_KEY = "master_import_preview"
MARKETPLACE_GROUP = "Marketplace"
BALI_WHATSAPP_GROUP = "Bali WhatsApp"
KATERINE_USERNAME = "katerine"
EDITRAFFICKER_USERNAME = "editrafficker"
# Tareas y Metas estan apagadas desde antes de este trabajo. Estaba escrito a mano
# aqui, asi que encenderlo exigia editar el codigo y desplegar. Ahora es una
# variable de entorno con el mismo valor por defecto: apagado.
#
# No se borra el modulo: son 8 modelos, 8 migraciones y ~1.100 lineas de plantillas
# que el equipo construyo. Borrarlo es una decision de producto, no de limpieza.
FEATURE_TASKS_GOALS_ENABLED = config("FEATURE_TASKS_GOALS_ENABLED", default=False, cast=bool)


def _bonus_tier(fulfillment):
    value = float(fulfillment or 0)
    if value >= 100:
        return {"name": "Oro", "payout": "Bono completo", "multiplier": 1}
    if value >= 90:
        return {"name": "Plata", "payout": "75% del bono", "multiplier": 0.75}
    if value >= 80:
        return {"name": "Bronce", "payout": "Bono bronce", "multiplier": None}
    return {"name": "Sin bono", "payout": "No aplica bono", "multiplier": 0}


def _fallback_insight_cards(messages):
    cards = []
    for message in messages or []:
        normalized = str(message).lower()
        signal = "danger" if any(value in normalized for value in ("alerta", "riesgo", "debajo", "bloque")) else "warning"
        cards.append(
            {
                "message": message,
                "signal": signal,
                "title": "Requiere revision" if signal == "danger" else "Seguimiento",
            }
        )
    return cards


def _is_marketplace_only_user(user):
    return user.is_authenticated and not user.is_superuser and user.groups.filter(name=MARKETPLACE_GROUP).exists()


def _is_bali_whatsapp_only_user(user):
    return user.is_authenticated and not user.is_superuser and user.groups.filter(name=BALI_WHATSAPP_GROUP).exists()


def _is_katerine_limited_user(user):
    return user.is_authenticated and not user.is_superuser and user.username.lower() == KATERINE_USERNAME


def _is_editrafficker_user(user):
    return user.is_authenticated and user.username.lower() == EDITRAFFICKER_USERNAME


def _can_run_external_sync(user):
    return _is_editrafficker_user(user)


def _safe_user_profile(user):
    if not getattr(user, "is_authenticated", False):
        return None
    try:
        return user.profile
    except Exception:
        return None


@memoize_per_request
def _managed_user_ids(user):
    if not user.is_authenticated:
        return []
    direct_ids = list(UserProfile.objects.filter(manager=user).values_list("user_id", flat=True))
    if direct_ids:
        return direct_ids
    if user.is_superuser:
        return list(UserProfile.objects.exclude(user=user).values_list("user_id", flat=True))
    return []


def _has_own_goals(user):
    return user.is_authenticated and (
        SalesTarget.objects.filter(user=user, is_active=True).exists()
        or OperationalGoalTask.objects.filter(assigned_to=user).exists()
        or InsightAchievement.objects.filter(user=user).exists()
    )


def _can_view_goals_dashboard(user):
    return FEATURE_TASKS_GOALS_ENABLED and user.is_authenticated and (user.is_superuser or UserProfile.objects.filter(manager=user).exists() or _has_own_goals(user))


def _can_view_tasks_dashboard(user):
    return FEATURE_TASKS_GOALS_ENABLED and user.is_authenticated


def _sidebar_context(active, request=None):
    today = timezone.localdate()
    uva_default_query = QueryDict("", mutable=True)
    uva_default_query["date_start"] = today.replace(day=1).isoformat()
    uva_default_query["date_end"] = today.isoformat()
    uva_default_query["business_unit"] = "uva"
    current_user_profile = _safe_user_profile(request.user) if request else None
    sync_context = {
        "can_run_external_sync": _can_run_external_sync(request.user) if request else False,
        "external_sync_today": today.isoformat(),
        "external_sync_default_date_from": today.isoformat(),
        "external_sync_default_date_to": today.isoformat(),
    }
    if request and _is_katerine_limited_user(request.user):
        return {
            "sidebar_items": [
                {"label": "Marketplace", "url": reverse("reports:marketplace"), "key": "marketplace"},
                {"label": "Bali", "url": f"{reverse('reports:bali')}?tab=physical", "key": "bali"},
                {"label": "Mi perfil", "url": reverse("reports:settings"), "key": "settings"},
            ],
            "active_nav": active,
            "current_user_profile": current_user_profile,
            "tasks_goals_enabled": FEATURE_TASKS_GOALS_ENABLED,
            **sync_context,
        }
    if request and _is_marketplace_only_user(request.user):
        sidebar_items = [
            {"label": "Inicio", "url": reverse("reports:dashboard"), "key": "dashboard"},
            {"label": "Marketplace", "url": reverse("reports:marketplace"), "key": "marketplace"},
        ]
        if _is_editrafficker_user(request.user):
            sidebar_items.append({"label": "Webs", "url": reverse("reports:websites"), "key": "websites"})
        sidebar_items.append({"label": "Mi perfil", "url": reverse("reports:settings"), "key": "settings"})
        return {
            "sidebar_items": sidebar_items,
            "active_nav": active,
            "current_user_profile": current_user_profile,
            "tasks_goals_enabled": FEATURE_TASKS_GOALS_ENABLED,
            **sync_context,
        }
    if request and _is_bali_whatsapp_only_user(request.user):
        return {
            "sidebar_items": [
                {"label": "Bali", "url": reverse("reports:bali"), "key": "bali"},
                {"label": "Mi perfil", "url": reverse("reports:settings"), "key": "settings"},
            ],
            "active_nav": active,
            "current_user_profile": current_user_profile,
            "tasks_goals_enabled": FEATURE_TASKS_GOALS_ENABLED,
            **sync_context,
        }
    sidebar_items = [
            {"label": "Inicio", "url": reverse("reports:dashboard"), "key": "dashboard"},
            {"label": "Uva", "url": f"{reverse('reports:uva')}?{uva_default_query.urlencode()}", "key": "uva"},
            {"label": "Uva Comfama", "url": reverse("reports:uva_comfama"), "key": "uva_comfama", "parent": "uva"},
            {"label": "Awn Internacional", "url": reverse("reports:awn_internacional"), "key": "awn_internacional", "parent": "uva"},
            {"label": "Bali", "url": reverse("reports:bali"), "key": "bali"},
            {"label": "Marketplace", "url": reverse("reports:marketplace"), "key": "marketplace"},
            {"label": "DistriSex", "url": reverse("reports:distrisex_ecuador"), "key": "distrisex"},
            {"label": "DistriSex Ecuador", "url": reverse("reports:distrisex_ecuador"), "key": "distrisex_ecuador", "parent": "distrisex"},
            {"label": "Webs", "url": reverse("reports:websites"), "key": "websites"},
            {"label": "Mi perfil", "url": reverse("reports:settings"), "key": "settings"},
    ]
    if request and _can_view_tasks_dashboard(request.user):
        sidebar_items.insert(6, {"label": "Tareas", "url": reverse("reports:tasks"), "key": "tasks"})
    if request and _can_view_goals_dashboard(request.user):
        sidebar_items.insert(7, {"label": "Metas", "url": reverse("reports:goals"), "key": "goals"})
    grouped_items = []
    children_by_parent = {}
    for item in sidebar_items:
        parent_key = item.get("parent")
        if parent_key:
            children_by_parent.setdefault(parent_key, []).append(item)
        else:
            grouped_items.append(item)
    for item in grouped_items:
        children = children_by_parent.get(item["key"], [])
        if children:
            item["children"] = children
            item["open"] = active == item["key"] or any(child["key"] == active for child in children)
            item["active"] = item["open"]
    return {
        "sidebar_items": grouped_items,
        "active_nav": active,
        "current_user_profile": current_user_profile,
        "tasks_goals_enabled": FEATURE_TASKS_GOALS_ENABLED,
        **sync_context,
    }


def _global_filter_context(request, defaults=None, overrides=None):
    if request.GET:
        params = request.GET.copy()
    elif defaults:
        params = QueryDict("", mutable=True)
        for key, value in defaults.items():
            params[key] = value
    else:
        params = None
    if overrides:
        if params is None:
            params = QueryDict("", mutable=True)
        for key, value in overrides.items():
            params[key] = value
    form = GlobalFilterForm(params or None)
    if form.is_bound and form.is_valid():
        filters = build_filter_dict(form.cleaned_data)
    elif form.is_bound:
        messages.warning(request, "Algunos filtros no son validos. Se aplico un rango seguro por defecto.")
        filters = build_filter_dict(defaults or {})
    else:
        filters = build_filter_dict({})
    today = timezone.localdate()
    if filters.get("period_type") == "weekly" and not filters.get("date_start") and not filters.get("date_end"):
        filters["date_start"] = (today - timedelta(days=6)).isoformat()
        filters["date_end"] = today.isoformat()
    elif filters.get("period_type") == "monthly" and not filters.get("date_start") and not filters.get("date_end"):
        filters["date_start"] = today.replace(day=1).isoformat()
        filters["date_end"] = today.isoformat()
    return form, filters


def _detail_filter_context(request, overrides=None):
    filters = build_filter_dict(_dashboard_default_filters())
    for key in ("period_type", "date_start", "date_end", "time_granularity", "compare_mode", "business_unit", "country", "channel"):
        value = request.GET.get(key)
        if value:
            filters[key] = value
    if overrides:
        filters.update(overrides)
    return filters


def _dashboard_default_filters():
    today = timezone.localdate()
    return {
        "period_type": "custom",
        "date_start": today.replace(day=1).isoformat(),
        "date_end": today.isoformat(),
        "time_granularity": "daily",
        "compare_mode": "previous_period",
    }


def _effective_filter_query(filters, extra=None, exclude=None):
    query = QueryDict("", mutable=True)
    exclude = set(exclude or [])
    for key, value in (filters or {}).items():
        if key in exclude:
            continue
        if value in (None, "", []):
            continue
        query[key] = value.isoformat() if hasattr(value, "isoformat") else str(value)
    for key, value in (extra or {}).items():
        if key in exclude or value in (None, "", []):
            continue
        query[key] = value.isoformat() if hasattr(value, "isoformat") else str(value)
    return query


def _redirect_bali_whatsapp_user(request):
    if _is_bali_whatsapp_only_user(request.user):
        return redirect("reports:bali")
    return None


def _redirect_katerine_limited_user(request):
    if _is_katerine_limited_user(request.user):
        return redirect(f"{reverse('reports:bali')}?tab=physical")
    return None


def _parse_sync_start(raw_value, today):
    if not raw_value:
        return today
    try:
        return datetime.fromisoformat(str(raw_value)).date()
    except (TypeError, ValueError) as exc:
        raise ValueError("Fecha de inicio invalida.") from exc


def _parse_sync_end(raw_value, today):
    if not raw_value:
        return today
    try:
        return datetime.fromisoformat(str(raw_value)).date()
    except (TypeError, ValueError) as exc:
        raise ValueError("Fecha final invalida.") from exc


def _metadata_json(path):
    response = requests.get(
        f"http://metadata.google.internal/computeMetadata/v1/{path}",
        headers={"Metadata-Flavor": "Google"},
        timeout=5,
    )
    response.raise_for_status()
    try:
        return response.json()
    except ValueError:
        return response.text


def _cloud_run_api_get(resource_name):
    token_payload = _metadata_json("instance/service-accounts/default/token")
    access_token = token_payload["access_token"]
    url = resource_name if str(resource_name).startswith("https://") else f"https://run.googleapis.com/v2/{str(resource_name).lstrip('/')}"
    response = requests.get(
        url,
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=20,
    )
    response.raise_for_status()
    return response.json()


def _trigger_cloud_run_sync_job(date_from, date_to):
    project_id = (
        os.environ.get("GOOGLE_CLOUD_PROJECT")
        or os.environ.get("GCLOUD_PROJECT")
        or _metadata_json("project/project-id")
    )
    token_payload = _metadata_json("instance/service-accounts/default/token")
    access_token = token_payload["access_token"]
    region = getattr(settings, "AXIS_SYNC_RUN_REGION", "us-central1")
    job_name = getattr(settings, "AXIS_SYNC_RUN_JOB_NAME", "axis-temp-sync-daily")
    url = f"https://run.googleapis.com/v2/projects/{project_id}/locations/{region}/jobs/{job_name}:run"
    args = [
        "manage.py",
        "sync_axis_history_range",
        "--date-from",
        date_from.isoformat(),
        "--date-to",
        date_to.isoformat(),
        "--all",
        "--continue-on-error",
    ]
    response = requests.post(
        url,
        headers={"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"},
        json={
            "overrides": {
                "containerOverrides": [{"args": args}],
                "taskCount": 1,
                "timeout": "3600s",
            }
        },
        timeout=20,
    )
    response.raise_for_status()
    payload = response.json()
    return {
        "operation_name": payload.get("name", ""),
        "execution_name": (payload.get("metadata") or {}).get("name", ""),
    }


def _sync_execution_status_payload(execution_payload):
    conditions = execution_payload.get("conditions") or []

    def find_condition(condition_type):
        for condition in conditions:
            if condition.get("type") == condition_type:
                return condition
        return {}

    completed = find_condition("Completed")
    started = find_condition("Started")
    resources = find_condition("ResourcesAvailable")
    container_ready = find_condition("ContainerReady")
    state = completed.get("state") or completed.get("status") or ""
    message = completed.get("message") or started.get("message") or resources.get("message") or ""
    execution_name = execution_payload.get("name", "")
    execution_id = execution_name.split("/")[-1] if execution_name else ""

    if state in {"CONDITION_FAILED", "False", False}:
        return {
            "ok": False,
            "status": "error",
            "progress": 100,
            "message": message or "La sincronizacion fallo.",
            "execution_id": execution_id,
        }
    if state in {"CONDITION_SUCCEEDED", "True", True} or execution_payload.get("completionTime"):
        return {
            "ok": True,
            "status": "completed",
            "progress": 100,
            "message": message or "Datos actualizados correctamente.",
            "execution_id": execution_id,
        }

    progress = 12
    if container_ready.get("state") in {"CONDITION_SUCCEEDED", "True", True} or container_ready.get("status") in {"True", True}:
        progress = 28
    if resources.get("state") in {"CONDITION_SUCCEEDED", "True", True} or resources.get("status") in {"True", True}:
        progress = 42
    if started.get("state") in {"CONDITION_SUCCEEDED", "True", True} or started.get("status") in {"True", True}:
        progress = 68

    return {
        "ok": True,
        "status": "running" if progress >= 68 else "pending",
        "progress": progress,
        "message": message or "Sincronizacion en curso.",
        "execution_id": execution_id,
    }


def _is_json_request(request):
    return request.headers.get("X-Requested-With") == "XMLHttpRequest" or "application/json" in request.headers.get("Accept", "")


def _run_inline_external_sync(date_from, date_to):
    current = date_from
    results = []
    decoder = json.JSONDecoder()
    while current <= date_to:
        capture = StringIO()
        call_command("sync_axis_daily_data", "--date", current.isoformat(), "--continue-on-error", stdout=capture)
        payload = None
        raw_output = capture.getvalue()
        for index, char in enumerate(raw_output):
            if char != "{":
                continue
            try:
                candidate, _ = decoder.raw_decode(raw_output[index:])
            except ValueError:
                continue
            payload = candidate
        results.append(
            {
                "date": current.isoformat(),
                "errors": len(payload.get("errors", [])) if isinstance(payload, dict) else 0,
            }
        )
        current += timedelta(days=1)
    call_command("fetch_onedrive_google_ads", stdout=StringIO())
    call_command("sync_websites_health", stdout=StringIO())
    return results


def _safe_next_url(request):
    fallback = reverse("reports:dashboard")
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or fallback
    if url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return next_url
    return fallback


@require_POST
def sync_external_data_now(request):
    if not _can_run_external_sync(request.user):
        raise PermissionDenied("No tienes permisos para ejecutar esta sincronizacion.")

    today = timezone.localdate()
    try:
        date_from = _parse_sync_start(request.POST.get("date_from"), today)
        date_to = _parse_sync_end(request.POST.get("date_to"), today)
    except ValueError as exc:
        if _is_json_request(request):
            return JsonResponse({"ok": False, "error": str(exc)}, status=400)
        messages.error(request, str(exc))
        return redirect(_safe_next_url(request))

    if date_from > date_to:
        if _is_json_request(request):
            return JsonResponse({"ok": False, "error": "La fecha de inicio no puede estar despues de la fecha final."}, status=400)
        messages.error(request, "La fecha de inicio no puede estar despues de la fecha final.")
        return redirect(_safe_next_url(request))

    if date_to > today:
        if _is_json_request(request):
            return JsonResponse({"ok": False, "error": "La fecha final no puede estar despues de hoy."}, status=400)
        messages.error(request, "La fecha final no puede estar despues de hoy.")
        return redirect(_safe_next_url(request))

    total_days = (date_to - date_from).days + 1
    duration_note = " Puede tardar varios minutos si el rango incluye muchos dias."
    try:
        if getattr(settings, "DEBUG", False):
            results = _run_inline_external_sync(date_from, date_to)
            failed_days = sum(item["errors"] for item in results)
            if failed_days:
                messages.warning(
                    request,
                    f"Sincronizacion ejecutada del {date_from:%d/%m/%Y} al {date_to:%d/%m/%Y} con {failed_days} errores reportados.{duration_note}",
                )
            else:
                messages.success(request, f"Sincronizacion ejecutada del {date_from:%d/%m/%Y} al {date_to:%d/%m/%Y}.{duration_note}")
        else:
            trigger_payload = _trigger_cloud_run_sync_job(date_from, date_to)
            operation_name = trigger_payload.get("operation_name", "") if isinstance(trigger_payload, dict) else str(trigger_payload or "")
            execution_name = trigger_payload.get("execution_name", "") if isinstance(trigger_payload, dict) else ""
            suffix = f" ({operation_name.split('/')[-1]})" if operation_name else ""
            if _is_json_request(request):
                status_url = reverse("reports:sync_external_data_status")
                return JsonResponse(
                    {
                        "ok": True,
                        "status": "queued",
                        "progress": 8,
                        "message": f"Sincronizacion lanzada para {total_days} dia(s).",
                        "operation_name": operation_name,
                        "execution_name": execution_name,
                        "status_url": status_url,
                    }
                )
            messages.success(
                request,
                f"Sincronizacion lanzada para {total_days} dia(s), del {date_from:%d/%m/%Y} al {date_to:%d/%m/%Y}{suffix}.{duration_note}",
            )
    except Exception as exc:
        if _is_json_request(request):
            return JsonResponse({"ok": False, "error": f"No fue posible lanzar la sincronizacion: {exc}"}, status=500)
        messages.error(request, f"No fue posible lanzar la sincronizacion: {exc}")

    return redirect(_safe_next_url(request))


@require_http_methods(["GET"])
def sync_external_data_status(request):
    if not _can_run_external_sync(request.user):
        raise PermissionDenied("No tienes permisos para consultar esta sincronizacion.")

    execution_name = request.GET.get("execution", "").strip()
    operation_name = request.GET.get("operation", "").strip()
    if not execution_name and not operation_name:
        return JsonResponse({"ok": False, "error": "Falta execution u operation."}, status=400)

    try:
        if execution_name:
            execution_payload = _cloud_run_api_get(execution_name)
        else:
            operation_payload = _cloud_run_api_get(operation_name)
            execution_payload = operation_payload.get("response") or operation_payload.get("metadata") or {}
        return JsonResponse(_sync_execution_status_payload(execution_payload))
    except Exception as exc:
        return JsonResponse({"ok": False, "status": "error", "progress": 100, "message": str(exc)}, status=500)


def _web_sales_default_filters():
    today = timezone.localdate()
    return {
        "period_type": "custom",
        "date_start": today.replace(day=1).isoformat(),
        "date_end": today.isoformat(),
        "time_granularity": "daily",
        "compare_mode": "previous_period",
        "business_unit": "uva",
    }


def _ad_spend_default_filters():
    today = timezone.localdate()
    return {
        "period_type": "custom",
        "date_start": today.replace(day=1).isoformat(),
        "date_end": today.isoformat(),
        "time_granularity": "daily",
        "compare_mode": "previous_period",
        "business_unit": "uva",
        "country": "CO",
    }


def _merge_sales_snapshot(summary, snapshot):
    if not snapshot["row_count"] and not snapshot.get("spend_rows"):
        return summary

    operational_alerts = [
        alert
        for alert in summary.get("alerts", [])
        if "tareas" in (alert or "").lower()
    ]

    summary["kpis"].update(
        {
            "sales_total": snapshot["kpis"]["sales_total"],
            "sales_month": snapshot["kpis"]["sales_month"],
            "sales_web": snapshot["kpis"]["sales_web"],
            "sales_whatsapp": snapshot["kpis"]["sales_whatsapp"],
            "ad_spend": snapshot["kpis"]["ad_spend"],
            "roas": snapshot["kpis"]["roas"],
            "orders": snapshot["kpis"]["orders"],
            "web_order_count": snapshot["kpis"].get("web_order_count", 0),
            "whatsapp_order_count": snapshot["kpis"].get("whatsapp_order_count", 0),
            "units": snapshot["kpis"]["units"],
            "product_units_web": snapshot["kpis"].get("product_units_web", 0),
            "product_units_whatsapp": snapshot["kpis"].get("product_units_whatsapp", 0),
            "average_ticket": snapshot["kpis"].get("average_ticket", 0),
            "web_average_ticket": snapshot["kpis"].get("web_average_ticket", 0),
            "whatsapp_average_ticket": snapshot["kpis"].get("whatsapp_average_ticket", 0),
        }
    )
    summary["sales_by_unit"] = snapshot["sales_by_unit"]
    summary["sales_by_channel"] = snapshot["sales_by_channel"]
    summary["roas_by_unit"] = snapshot.get("roas_by_unit", summary.get("roas_by_unit", []))
    summary["insights"] = snapshot.get("insights", [])

    sales_alerts = []
    if snapshot.get("coverage", {}).get("is_partial"):
        sales_alerts.append(snapshot["coverage"]["message"])
    if snapshot["kpis"].get("ad_spend") and snapshot["kpis"].get("roas", 0) < 3:
        sales_alerts.append("El ROAS consolidado del filtro actual esta por debajo del umbral de referencia 3.0.")
    summary["alerts"] = [*sales_alerts, *operational_alerts][:4]
    return summary


def _target_owner_ids(user):
    if not user.is_authenticated:
        return []
    managed_ids = _managed_user_ids(user)
    return managed_ids or [user.id]


def _marketplace_target_rows(user, filters):
    owner_ids = _target_owner_ids(user)
    if not owner_ids:
        return []
    business_unit = BusinessUnit.objects.filter(slug="marketplace").first()
    if not business_unit:
        return []

    channels = list(Channel.objects.filter(is_active=True, business_unit=business_unit).order_by("display_order", "name"))
    sales = DailyChannelSale.objects.filter(business_unit=business_unit)
    if filters.get("country"):
        sales = sales.filter(country__code=filters["country"])
    if filters.get("date_start"):
        sales = sales.filter(sale_date__gte=filters["date_start"])
    if filters.get("date_end"):
        sales = sales.filter(sale_date__lte=filters["date_end"])
    sales_by_channel = {
        row["channel_id"]: row["total"] or 0
        for row in sales.values("channel_id").annotate(total=Sum("sales_amount"))
    }

    targets = SalesTarget.objects.filter(user_id__in=owner_ids, business_unit=business_unit, is_active=True)
    if filters.get("date_start"):
        targets = targets.filter(date_end__gte=filters["date_start"])
    if filters.get("date_end"):
        targets = targets.filter(date_start__lte=filters["date_end"])
    target_by_channel = {
        row["channel_id"]: row["total"] or 0
        for row in targets.values("channel_id").annotate(total=Sum("target_amount"))
    }
    general_target = target_by_channel.get(None, 0)
    selected_channel = filters.get("channel")

    rows = []
    for channel in channels:
        if selected_channel and channel.slug != selected_channel:
            continue
        target_amount = float(target_by_channel.get(channel.id, 0) or general_target or 0)
        sales_amount_with_vat = sales_by_channel.get(channel.id, 0)
        sales_amount = float(remove_colombia_vat(sales_amount_with_vat))
        if not target_amount and not sales_amount:
            continue
        fulfillment = round((sales_amount / target_amount) * 100, 1) if target_amount else 0
        rows.append(
            {
                "label": channel.name,
                "slug": channel.slug,
                "sales": sales_amount,
                "sales_with_vat": float(sales_amount_with_vat),
                "target": target_amount,
                "fulfillment": min(fulfillment, 100),
                "fulfillment_width": int(min(fulfillment, 100)),
                "fulfillment_label": fulfillment,
                "bonus_tier": _bonus_tier(fulfillment),
            }
        )
    return rows


def _marketplace_goal_summary(target_rows, snapshot):
    if not target_rows:
        return []
    total_sales = sum(row["sales"] for row in target_rows)
    total_target = sum(row["target"] for row in target_rows)
    remaining = max(total_target - total_sales, 0)
    best = max(target_rows, key=lambda row: row["fulfillment_label"]) if target_rows else None
    rows = []
    if total_target:
        rows.append({"label": "Meta total", "value": total_target, "kind": "money"})
        rows.append({"label": "Cumplimiento total", "value": round((total_sales / total_target) * 100, 1), "kind": "percent"})
        if remaining:
            rows.append({"label": "Faltante", "value": remaining, "kind": "money"})
    if best and best["fulfillment_label"]:
        rows.append({"label": "Canal mas avanzado", "value": f"{best['label']} - {best['fulfillment_label']:.1f}%", "kind": "text"})
    if snapshot["kpis"].get("orders"):
        rows.append({"label": "Ticket por pedido", "value": snapshot["kpis"].get("average_ticket", 0), "kind": "money"})
    return rows


def _goals_dashboard_rows(user, filters):
    user_ids = _managed_user_ids(user)
    if not user_ids and _has_own_goals(user):
        user_ids = [user.id]
    if not user_ids:
        return [], []
    targets = SalesTarget.objects.select_related("user", "business_unit", "channel").filter(user_id__in=user_ids, is_active=True)
    if filters.get("date_start"):
        targets = targets.filter(date_end__gte=filters["date_start"])
    if filters.get("date_end"):
        targets = targets.filter(date_start__lte=filters["date_end"])
    target_rows = []
    for target in targets:
        sales = DailyChannelSale.objects.filter(business_unit=target.business_unit)
        if target.channel_id:
            sales = sales.filter(channel=target.channel)
        if filters.get("date_start"):
            sales = sales.filter(sale_date__gte=filters["date_start"])
        else:
            sales = sales.filter(sale_date__gte=target.date_start)
        if filters.get("date_end"):
            sales = sales.filter(sale_date__lte=filters["date_end"])
        else:
            sales = sales.filter(sale_date__lte=target.date_end)
        raw_sales_amount = sales.aggregate(total=Sum("sales_amount"))["total"] or 0
        if target.business_unit.slug == "marketplace":
            raw_sales_amount = remove_colombia_vat(raw_sales_amount)
        sales_amount = float(raw_sales_amount)
        target_amount = float(target.target_amount or 0)
        fulfillment = round((sales_amount / target_amount) * 100, 1) if target_amount else 0
        if not target_amount and not sales_amount:
            continue
        target_rows.append(
            {
                "user": target.user.get_full_name() or target.user.username,
                "business_unit": target.business_unit.name,
                "channel": target.channel.name if target.channel_id else "General",
                "sales": sales_amount,
                "target": target_amount,
                "remaining": max(target_amount - sales_amount, 0),
                "fulfillment": min(fulfillment, 100),
                "fulfillment_width": int(min(fulfillment, 100)),
                "fulfillment_label": fulfillment,
                "bonus_tier": _bonus_tier(fulfillment),
            }
        )
    task_rows = list(
        OperationalGoalTask.objects.select_related("assigned_to", "sales_target", "sales_target__business_unit", "sales_target__channel")
        .filter(assigned_to_id__in=user_ids)
        .order_by("assigned_to__first_name", "status", "due_date", "-created_at")[:30]
    )
    return target_rows, task_rows


def _selected_achievement_month(filters):
    reference = filters.get("date_end") or timezone.localdate()
    if isinstance(reference, str):
        reference = datetime.strptime(reference, "%Y-%m-%d").date()
    month_start = reference.replace(day=1)
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1)
    else:
        next_month = month_start.replace(month=month_start.month + 1)
    return month_start, next_month - timedelta(days=1)


def _sync_monthly_achievements(user_ids, month_start, month_end):
    if not user_ids:
        return []
    targets = (
        SalesTarget.objects.select_related("user", "business_unit", "channel")
        .filter(user_id__in=user_ids, is_active=True, date_end__gte=month_start, date_start__lte=month_end)
    )
    achievement_ids = []
    previous_end = month_start - timedelta(days=1)
    previous_start = previous_end.replace(day=1)
    for target in targets:
        current_sales = DailyChannelSale.objects.filter(
            business_unit=target.business_unit,
            sale_date__gte=month_start,
            sale_date__lte=month_end,
        )
        previous_sales = DailyChannelSale.objects.filter(
            business_unit=target.business_unit,
            sale_date__gte=previous_start,
            sale_date__lte=previous_end,
        )
        if target.channel_id:
            current_sales = current_sales.filter(channel=target.channel)
            previous_sales = previous_sales.filter(channel=target.channel)
        current = current_sales.aggregate(sales=Sum("sales_amount"), spend=Sum("spend_amount"))
        previous = previous_sales.aggregate(sales=Sum("sales_amount"), spend=Sum("spend_amount"))
        current_amount = Decimal(current["sales"] or 0)
        previous_amount = Decimal(previous["sales"] or 0)
        current_spend = Decimal(current["spend"] or 0)
        previous_spend = Decimal(previous["spend"] or 0)
        specs = []
        target_amount = Decimal(target.target_amount or 0)
        if target_amount and current_amount >= target_amount:
            fulfillment = (current_amount / target_amount) * Decimal("100")
            specs.append(
                (
                    InsightAchievement.AchievementType.SALES_TARGET,
                    "Meta de ventas alcanzada",
                    f"Cumplimiento de {fulfillment:.1f}% con ventas por ${current_amount:,.0f} COP.",
                    current_amount,
                    fulfillment - Decimal("100"),
                )
            )
        if previous_amount and current_amount > previous_amount:
            growth = ((current_amount - previous_amount) / previous_amount) * Decimal("100")
            specs.append(
                (
                    InsightAchievement.AchievementType.SALES_GROWTH,
                    "Aumento de ventas",
                    f"Las ventas crecieron {growth:.1f}% frente al mes anterior.",
                    current_amount,
                    growth,
                )
            )
        if previous_spend and current_spend < previous_spend and current_amount >= previous_amount:
            saving = ((previous_spend - current_spend) / previous_spend) * Decimal("100")
            specs.append(
                (
                    InsightAchievement.AchievementType.SPEND_EFFICIENCY,
                    "Reduccion eficiente de inversion",
                    f"La inversion bajo {saving:.1f}% manteniendo o aumentando las ventas.",
                    previous_spend - current_spend,
                    saving,
                )
            )
        if current_spend and previous_spend and previous_amount:
            current_roas = current_amount / current_spend
            previous_roas = previous_amount / previous_spend
            if current_roas > previous_roas:
                growth = ((current_roas - previous_roas) / previous_roas) * Decimal("100")
                specs.append(
                    (
                        InsightAchievement.AchievementType.ROAS_GROWTH,
                        "Mejora de ROAS",
                        f"El ROAS mejoro {growth:.1f}% y alcanzo {current_roas:.2f}.",
                        current_roas,
                        growth,
                    )
                )
        for achievement_type, title, description, value, delta in specs:
            achievement, _ = InsightAchievement.objects.update_or_create(
                sales_target=target,
                month=month_start,
                achievement_type=achievement_type,
                defaults={
                    "user": target.user,
                    "business_unit": target.business_unit,
                    "channel": target.channel,
                    "title": title,
                    "description": description,
                    "metric_value": value,
                    "delta_percent": delta,
                },
            )
            achievement_ids.append(achievement.id)
    stale = InsightAchievement.objects.filter(user_id__in=user_ids, month=month_start)
    if achievement_ids:
        stale.exclude(id__in=achievement_ids).delete()
    else:
        stale.delete()
    return list(
        InsightAchievement.objects.select_related("user", "business_unit", "channel")
        .filter(user_id__in=user_ids, month=month_start)
        .order_by("user__first_name", "user__username", "-metric_value")
    )


def _operational_task_context(user):
    if not user.is_authenticated:
        return {
            "operational_assigned_tasks": [],
            "operational_completed_tasks": [],
            "operational_managed_tasks": [],
        }
    assigned = (
        OperationalGoalTask.objects.select_related("sales_target", "sales_target__business_unit", "sales_target__channel", "assigned_by")
        .filter(assigned_to=user)
        .order_by("status", "due_date", "-created_at")
    )
    managed_ids = _managed_user_ids(user)
    managed = OperationalGoalTask.objects.none()
    if managed_ids:
        managed = (
            OperationalGoalTask.objects.select_related("assigned_to", "sales_target", "sales_target__business_unit", "sales_target__channel")
            .filter(assigned_to_id__in=managed_ids)
            .order_by("assigned_to__first_name", "status", "due_date", "-created_at")[:20]
        )
    return {
        "operational_assigned_tasks": list(assigned.exclude(status=OperationalGoalTask.Status.COMPLETED)[:10]),
        "operational_completed_tasks": list(assigned.filter(status=OperationalGoalTask.Status.COMPLETED)[:5]),
        "operational_managed_tasks": list(managed),
    }


def _user_task_context(user):
    if not user.is_authenticated:
        return {
            "user_pending_tasks": [],
            "user_completed_tasks": [],
            "user_managed_tasks": [],
            "user_task_agenda_weeks": [],
            "user_task_summary": {"pending": 0, "completed": 0, "hours": 0},
        }
    assigned = (
        UserTask.objects.select_related("assigned_to", "created_by")
        .prefetch_related("attachments", "task_links")
        .filter(assigned_to=user)
        .order_by("status", "due_date", "-created_at")
    )
    managed_ids = _managed_user_ids(user)
    managed = UserTask.objects.none()
    if managed_ids:
        managed = (
            UserTask.objects.select_related("assigned_to", "created_by")
            .filter(assigned_to_id__in=managed_ids)
            .order_by("assigned_to__first_name", "status", "due_date", "-created_at")[:20]
        )
    pending = assigned.exclude(status__in=[UserTask.Status.COMPLETED, UserTask.Status.CANCELED])
    completed = assigned.filter(status=UserTask.Status.COMPLETED)
    all_assigned_tasks = list(assigned.order_by("due_date", "created_at", "title"))
    return {
        "user_pending_tasks": list(pending[:8]),
        "user_completed_tasks": list(completed[:5]),
        "user_managed_tasks": list(managed),
        "user_task_agenda_weeks": _build_user_task_agenda(all_assigned_tasks),
        "user_task_summary": {
            "pending": pending.count(),
            "completed": completed.count(),
            "hours": assigned.aggregate(total=Sum("registered_hours"))["total"] or 0,
        },
    }


def _build_user_task_agenda(tasks):
    weekday_names = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes"]
    month_names = {
        1: "Enero",
        2: "Febrero",
        3: "Marzo",
        4: "Abril",
        5: "Mayo",
        6: "Junio",
        7: "Julio",
        8: "Agosto",
        9: "Septiembre",
        10: "Octubre",
        11: "Noviembre",
        12: "Diciembre",
    }
    dated_tasks = [task for task in tasks if task.due_date]
    if dated_tasks:
        start = min(task.due_date for task in dated_tasks)
        end = max(task.due_date for task in dated_tasks)
    else:
        today = timezone.localdate()
        start = today
        end = today
    current = start - timedelta(days=start.weekday())
    final = end + timedelta(days=6 - end.weekday())
    tasks_by_day = {}
    for task in dated_tasks:
        tasks_by_day.setdefault(task.due_date, []).append(task)

    weeks = []
    while current <= final:
        days = []
        month_labels = []
        for offset in range(5):
            day = current + timedelta(days=offset)
            day_tasks = tasks_by_day.get(day, [])
            days.append(
                {
                    "date": day,
                    "label": f"{weekday_names[offset]} {day.day}",
                    "tasks": day_tasks,
                    "hours": sum((task.registered_hours or 0) for task in day_tasks),
                }
            )
            month_name = month_names[day.month]
            if month_name not in month_labels:
                month_labels.append(month_name)
        weeks.append({"month_label": " / ".join(month_labels), "days": days})
        current += timedelta(days=7)
    return weeks


def _parse_task_date(value):
    try:
        return datetime.strptime(value or "", "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_task_time(value):
    try:
        return datetime.strptime(value or "", "%H:%M").time()
    except ValueError:
        return None


def _can_access_user_task(user, task):
    if not user.is_authenticated:
        return False
    return task.assigned_to_id == user.id or task.created_by_id == user.id or task.assigned_to_id in _managed_user_ids(user)


def _task_calendar_payload(task):
    start_hour = (task.due_time.hour if task.due_time else 8)
    start_minute = (task.due_time.minute if task.due_time else 0)
    duration = float(task.registered_hours or 1)
    return {
        "id": task.id,
        "title": task.title,
        "description": task.description,
        "links": task.links,
        "date": task.due_date.isoformat() if task.due_date else "",
        "time": task.due_time.strftime("%H:%M") if task.due_time else "08:00",
        "hour": start_hour,
        "minute": start_minute,
        "duration": max(duration, 0.5),
        "status": task.status,
        "status_label": task.get_status_display(),
        "hours": duration,
        "assigned_to": task.assigned_to.get_full_name() or task.assigned_to.username if task.assigned_to else "",
    }


def _task_calendar_context(user, week_start=None):
    selected = _parse_task_date(week_start) or timezone.localdate()
    start = selected - timedelta(days=selected.weekday())
    end = start + timedelta(days=4)
    tasks = (
        UserTask.objects.select_related("assigned_to", "created_by")
        .filter(due_date__gte=start, due_date__lte=end)
        .filter(Q(assigned_to=user) | Q(created_by=user) | Q(assigned_to_id__in=_managed_user_ids(user)))
        .distinct()
        .order_by("due_date", "due_time", "title")
    )
    weekdays = ["LUN", "MAR", "MIE", "JUE", "VIE"]
    month_names = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    assignees = [user, *[profile.user for profile in UserProfile.objects.select_related("user").filter(manager=user).order_by("user__first_name", "user__username")]]
    return {
        "calendar_start": start,
        "calendar_end": end,
        "calendar_title": f"{month_names[start.month - 1]} {start.year}" if start.month == end.month else f"{month_names[start.month - 1]} / {month_names[end.month - 1]} {end.year}",
        "calendar_days": [{"date": start + timedelta(days=index), "weekday": weekdays[index]} for index in range(5)],
        "calendar_hours": list(range(8, 19)),
        "calendar_prev": (start - timedelta(days=7)).isoformat(),
        "calendar_next": (start + timedelta(days=7)).isoformat(),
        "calendar_today": (timezone.localdate() - timedelta(days=timezone.localdate().weekday())).isoformat(),
        "today": timezone.localdate(),
        "calendar_tasks_json": json.dumps([_task_calendar_payload(task) for task in tasks]),
        "calendar_assignees": assignees,
    }


def _task_agent_users(user):
    ids = _managed_user_ids(user)
    if not ids:
        return []
    return list(User.objects.filter(id__in=ids).select_related("profile").order_by("first_name", "last_name", "username"))


def _spanish_date_label(day):
    month_names = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre",
    }
    return f"{day.day} de {month_names[day.month]}"


def _agent_week_title(start, end):
    if start.month == end.month:
        return f"Semana del lunes {start.day} al viernes {_spanish_date_label(end)}"
    return f"Semana del lunes {_spanish_date_label(start)} al viernes {_spanish_date_label(end)}"


def _agent_weekly_tasks(agent, start_date=None, end_date=None):
    today = timezone.localdate()
    start = start_date or (today - timedelta(days=today.weekday()))
    start = start - timedelta(days=start.weekday())
    base_queryset = UserTask.objects.select_related("assigned_to", "created_by").filter(assigned_to=agent, due_date__isnull=False)
    if end_date:
        end = end_date
    else:
        last_task_date = base_queryset.filter(due_date__gte=start).order_by("-due_date").values_list("due_date", flat=True).first()
        end = max(last_task_date or start, start + timedelta(days=27))
    end = end + timedelta(days=4 - end.weekday()) if end.weekday() < 5 else end - timedelta(days=end.weekday() - 4)
    queryset = base_queryset.filter(due_date__gte=start, due_date__lte=end).order_by("due_date", "status", "title")
    tasks_by_day = {}
    for task in queryset:
        if task.due_date.weekday() < 5:
            tasks_by_day.setdefault(task.due_date, []).append(task)

    weekday_names = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes"]
    weeks = []
    cursor = start
    while cursor <= end:
        friday = cursor + timedelta(days=4)
        days = []
        for offset, name in enumerate(weekday_names):
            day = cursor + timedelta(days=offset)
            tasks = tasks_by_day.get(day, [])
            days.append({"date": day, "label": f"{name} {day.day}", "tasks": tasks})
        week_tasks = [task for day in days for task in day["tasks"]]
        total_hours = sum(float(task.registered_hours or 0) for task in week_tasks)
        weeks.append(
            {
                "start": cursor,
                "end": friday,
                "title": _agent_week_title(cursor, friday),
                "days": days,
                "tasks": week_tasks,
                "task_count": len(week_tasks),
                "total_hours": total_hours,
                "completed_count": sum(1 for task in week_tasks if task.status == UserTask.Status.COMPLETED),
            }
        )
        cursor += timedelta(days=7)
    return weeks


def _task_agents_context(user, request):
    agents = _task_agent_users(user)
    selected_agent = None
    selected_id = request.GET.get("agent")
    if selected_id and selected_id.isdigit():
        selected_agent = next((agent for agent in agents if agent.id == int(selected_id)), None)
    if not selected_agent and agents:
        selected_agent = agents[0]
    start = _parse_task_date(request.GET.get("desde"))
    end = _parse_task_date(request.GET.get("hasta"))
    weeks = _agent_weekly_tasks(selected_agent, start, end) if selected_agent else []
    export_params = QueryDict(mutable=True)
    if selected_agent:
        export_params["agent"] = selected_agent.id
    if start:
        export_params["desde"] = start.isoformat()
    if end:
        export_params["hasta"] = end.isoformat()
    return {
        "task_tab": request.GET.get("tab") or "calendar",
        "task_agents": agents,
        "selected_agent": selected_agent,
        "agent_task_weeks": weeks,
        "agent_filter_start": start.isoformat() if start else "",
        "agent_filter_end": end.isoformat() if end else "",
        "agent_export_query": export_params.urlencode(),
    }


def _executive_dashboard_context(request, active_key, title, subtitle, filter_overrides=None):
    filter_form, filters = _global_filter_context(request, defaults=_dashboard_default_filters(), overrides=filter_overrides)
    summary = build_dashboard_summary(filters)
    sales_snapshot = build_sales_snapshot(filters)
    summary = _merge_sales_snapshot(summary, sales_snapshot)
    context = {
        **_sidebar_context(active_key, request),
        **_operational_task_context(request.user),
        **_user_task_context(request.user),
        "page_title": title,
        "page_subtitle": subtitle,
        "filter_form": filter_form,
        "summary": summary,
        "snapshot": sales_snapshot,
        "sales_by_unit_json": json.dumps(summary["sales_by_unit"]),
        "sales_by_channel_json": json.dumps(summary["sales_by_channel"]),
        "combined_series_json": json.dumps(sales_snapshot.get("combined_series", [])),
        "ad_platform_performance_json": json.dumps(build_ad_platform_performance(filters, sales_snapshot)),
        "is_uva_all_countries": active_key == "uva" and not filters.get("country"),
        "show_non_dashboard_kpis": active_key != "dashboard" and not (active_key == "uva" and not filters.get("country")),
        "show_comfama_panel": active_key == "dashboard" or (active_key == "uva" and filters.get("country") == "CO"),
    }
    if active_key == "uva":
        context["product_category_snapshot"] = build_uva_category_snapshot(filters)
        context["uva_meta_ads_preview"] = build_uva_meta_ads_preview(filters, allow_live_fetch=False)
        context["category_profitability_json"] = json.dumps(context["product_category_snapshot"].get("profitability_json", []))
        if not filters.get("country"):
            context["uva_category_country_snapshot"] = build_uva_category_country_comparison(filters)
            context["uva_category_country_comparison_json"] = json.dumps(context["uva_category_country_snapshot"].get("chart_rows", []))
    if active_key in {"dashboard", "uva"}:
        colombia_filters = dict(filters)
        colombia_filters["business_unit"] = "uva"
        colombia_filters["country"] = "CO"
        context["home_colombia_snapshot"] = build_sales_snapshot(colombia_filters, include_comparison=False)
        ecuador_sales_filters = dict(filters)
        ecuador_sales_filters["business_unit"] = "uva"
        ecuador_sales_filters["country"] = "EC"
        context["home_ecuador_sales_snapshot"] = build_sales_snapshot(ecuador_sales_filters, include_comparison=False)
        mexico_filters = dict(filters)
        mexico_filters["business_unit"] = "uva"
        mexico_filters["country"] = "MX"
        context["home_mexico_snapshot"] = build_sales_snapshot(mexico_filters, include_comparison=False)
        country_comparison = build_copa_uva_country_comparison(filters)
        context["home_comfama_snapshot"] = build_comfama_snapshot(filters)
        context["home_ecuador_snapshot"] = build_ecuador_snapshot(filters)
        context["home_ecuador_categories_json"] = json.dumps(context["home_ecuador_snapshot"].get("categories", []))
        context["home_country_comparison"] = country_comparison
        context["uva_geo_map"] = build_uva_geo_map_data(filters, country_comparison)
        context["copa_uva_country_comparison_json"] = json.dumps(country_comparison)
    if active_key == "dashboard":
        context["filter_fields"] = [filter_form["date_start"], filter_form["date_end"]]
        bali_filters = dict(filters)
        bali_filters["business_unit"] = "bali"
        bali_filters["country"] = "CO"
        context["home_bali_snapshot"] = build_bali_snapshot(bali_filters)
        context["home_bali_channels_json"] = json.dumps(context["home_bali_snapshot"].get("channels", []))
        context["home_bali_daily_json"] = json.dumps(context["home_bali_snapshot"].get("daily_series", []))
        marketplace_filters = dict(filters)
        marketplace_filters["business_unit"] = "marketplace"
        context["home_marketplace_snapshot"] = build_sales_snapshot(marketplace_filters, include_comparison=False)
        context["home_marketplace_channels_json"] = json.dumps(context["home_marketplace_snapshot"].get("sales_by_channel", []))
        context["home_marketplace_daily_json"] = json.dumps(context["home_marketplace_snapshot"].get("combined_series", []))
    return context


def _attachment_file_type(uploaded_file):
    suffix = Path(uploaded_file.name).suffix.lower().replace(".", "")
    if suffix in {"png", "jpg", "jpeg", "gif", "webp"}:
        return Attachment.FileType.IMAGE
    if suffix in {"xlsx", "xls", "csv"}:
        return Attachment.FileType.EXCEL
    if suffix in {"ppt", "pptx"}:
        return Attachment.FileType.PRESENTATION
    if suffix == "pdf":
        return Attachment.FileType.PDF
    return Attachment.FileType.DOCUMENT


def _build_attachment_ref(index):
    return f"ATT-UP-{index:04d}"


@never_cache
def dashboard(request):
    if _is_katerine_limited_user(request.user):
        return redirect(f"{reverse('reports:bali')}?tab=physical")
    if _is_marketplace_only_user(request.user):
        return redirect("reports:marketplace")
    if _is_bali_whatsapp_only_user(request.user):
        return redirect("reports:bali")
    context = _executive_dashboard_context(
        request,
        active_key="dashboard",
        title="Inicio",
        subtitle="Vista ejecutiva del negocio y de la operacion semanal.",
    )
    return render(request, "reports/dashboard.html", context)


def _website_status_class(status):
    return {
        WebsiteHealthCheck.OverallStatus.HEALTHY: "green",
        WebsiteHealthCheck.OverallStatus.WARNING: "yellow",
        WebsiteHealthCheck.OverallStatus.CRITICAL: "red",
    }.get(status, "muted")


def _website_score_class(score):
    if score is None:
        return "muted"
    if score >= 80:
        return "green"
    if score >= 50:
        return "yellow"
    return "red"


def _website_headers_class(check):
    if not check or not check.security_headers_total:
        return "muted"
    if check.security_headers_score == 0:
        return "red"
    if check.security_headers_score >= check.security_headers_total:
        return "green"
    return "yellow"


def _website_products_summary(check):
    """Etiqueta y color del catalogo visible.

    Distingue "no se pudo leer" de "leimos y hay cero", que antes se veian igual.
    """
    if not check:
        return "Sin dato", "muted"
    estado = check.products_visible_status
    if estado == "ok":
        agotados = check.products_out_of_stock_count or 0
        if agotados:
            return f"{check.products_in_stock_count or 0} en stock, {agotados} agotados", "yellow"
        return f"{check.products_in_stock_count or 0} en stock", "green"
    if estado == "empty":
        return "La tienda no devolvio productos", "red"
    if estado == "blocked":
        return "La tienda no permitio leer el catalogo", "yellow"
    if estado == "not_configured":
        return "Sin lectura de catalogo para esta plataforma", "muted"
    return "Sin dato", "muted"


def _website_score(check):
    if not check:
        return 0
    score = 100
    if check.overall_status == WebsiteHealthCheck.OverallStatus.CRITICAL:
        score -= 45
    elif check.overall_status == WebsiteHealthCheck.OverallStatus.WARNING:
        score -= 20
    if check.response_time_ms and check.response_time_ms > 3000:
        score -= 10
    if check.ssl_valid is False:
        score -= 25
    elif check.ssl_days_remaining is not None and check.ssl_days_remaining < 30:
        score -= 10
    if check.security_headers_total:
        missing_headers = check.security_headers_total - check.security_headers_score
        score -= min(20, missing_headers * 4)
    if check.performance_score is not None and check.performance_score < 50:
        score -= 15
    elif check.performance_score is not None and check.performance_score < 75:
        score -= 7
    if check.accessibility_score is not None and check.accessibility_score < 70:
        score -= 10
    if check.best_practices_score is not None and check.best_practices_score < 70:
        score -= 8
    return max(0, min(100, score))


def _website_history_rows(checks):
    rows = []
    for check in checks:
        technical_score = _website_score(check)
        performance_score = check.performance_score if check.performance_score is not None else technical_score
        rows.append(
            {
                "label": timezone.localtime(check.checked_at).strftime("%d/%m"),
                "checked_at": timezone.localtime(check.checked_at).strftime("%d/%m/%Y %H:%M"),
                "technical_score": technical_score,
                "performance_score": performance_score,
                "response_time_ms": check.response_time_ms or 0,
            }
        )
    return rows


def _website_history_by_website(websites, date_start, date_end):
    if not websites:
        return {}
    queryset = WebsiteHealthCheck.objects.filter(website__in=websites).select_related("website").order_by("website_id", "checked_at")
    if date_start:
        queryset = queryset.filter(checked_at__date__gte=date_start)
    if date_end:
        queryset = queryset.filter(checked_at__date__lte=date_end)
    grouped = {}
    for check in queryset:
        grouped.setdefault(check.website_id, []).append(check)
    return {website_id: _website_history_rows(rows) for website_id, rows in grouped.items()}


def _parse_website_date(raw_value, fallback):
    value = str(raw_value or "").strip()
    if not value:
        return fallback
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return fallback


def _website_cards(websites, latest_checks, history_by_website=None):
    cards = []
    history_by_website = history_by_website or {}
    availability_labels = {
        WebsiteHealthCheck.AvailabilityStatus.ONLINE: "En linea",
        WebsiteHealthCheck.AvailabilityStatus.REDIRECT: "Redireccion",
        WebsiteHealthCheck.AvailabilityStatus.OFFLINE: "Fuera de linea",
        WebsiteHealthCheck.AvailabilityStatus.ERROR: "Error",
        WebsiteHealthCheck.AvailabilityStatus.UNKNOWN: "Sin dato",
    }
    for website in websites:
        check = latest_checks.get(website.id)
        status = check.overall_status if check else WebsiteHealthCheck.OverallStatus.UNKNOWN
        pagespeed_label = "No medido"
        if check and check.pagespeed_status == "ok":
            pagespeed_label = "Medido"
        elif check and check.pagespeed_status == "stale":
            pagespeed_label = "Ultima medicion disponible"
        elif check and check.pagespeed_status == "quota_exceeded":
            pagespeed_label = "Sin cuota"
        performance_score = check.performance_score if check and check.performance_score is not None else _website_score(check)
        accessibility_score = check.accessibility_score if check else None
        best_practices_score = check.best_practices_score if check else None
        seo_score = check.seo_score if check else None
        products_label, products_status_class = _website_products_summary(check)
        cards.append(
            {
                "website": website,
                "check": check,
                "products_label": products_label,
                "products_status_class": products_status_class,
                "score": _website_score(check),
                "performance_circle_score": performance_score,
                "performance_status_class": _website_score_class(performance_score),
                "headers_status_class": _website_headers_class(check),
                "accessibility_status_class": _website_score_class(accessibility_score),
                "best_practices_status_class": _website_score_class(best_practices_score),
                "seo_status_class": _website_score_class(seo_score),
                "status": status,
                "status_class": _website_status_class(status),
                "status_label": dict(WebsiteHealthCheck.OverallStatus.choices).get(status, "Sin dato"),
                "availability_label": availability_labels.get(check.availability_status, "Sin dato") if check else "Sin dato",
                "pagespeed_label": pagespeed_label,
                "history_rows": history_by_website.get(website.id, []),
                "history_json": json.dumps(history_by_website.get(website.id, [])),
            }
        )
    return cards


@require_http_methods(["GET"])
@never_cache
def websites_module(request):
    limited_redirect = _redirect_katerine_limited_user(request)
    if limited_redirect:
        return limited_redirect
    limited_redirect = _redirect_bali_whatsapp_user(request)
    if limited_redirect:
        return limited_redirect

    # La siembra vive en la migracion 0055 y en sync_websites_health. Hacerla
    # aqui escribia 4 filas en cada carga de la pagina.

    today = timezone.localdate()
    default_start = today - timedelta(days=30)
    history_date_start = _parse_website_date(request.GET.get("date_start"), default_start)
    history_date_end = _parse_website_date(request.GET.get("date_end"), today)
    if history_date_start > history_date_end:
        history_date_start, history_date_end = history_date_end, history_date_start

    latest_checks = latest_checks_by_website()
    websites = list(Website.objects.select_related("business_unit").order_by("display_order", "name", "country_label"))
    active_websites = [website for website in websites if website.stage == Website.Stage.ACTIVE]
    history_by_website = _website_history_by_website(active_websites, history_date_start, history_date_end)
    active_cards = _website_cards(active_websites, latest_checks, history_by_website=history_by_website)
    status_counts = {
        "healthy": sum(1 for card in active_cards if card["status"] == WebsiteHealthCheck.OverallStatus.HEALTHY),
        "warning": sum(1 for card in active_cards if card["status"] == WebsiteHealthCheck.OverallStatus.WARNING),
        "critical": sum(1 for card in active_cards if card["status"] == WebsiteHealthCheck.OverallStatus.CRITICAL),
        "unknown": sum(1 for card in active_cards if card["status"] == WebsiteHealthCheck.OverallStatus.UNKNOWN),
    }
    checked_cards = [card for card in active_cards if card["check"]]
    average_score = round(sum(card["score"] for card in checked_cards) / len(checked_cards)) if checked_cards else 0

    context = {
        **_sidebar_context("websites", request),
        "page_title": "Webs",
        "page_subtitle": "Estado tecnico, seguridad, usabilidad y velocidad por sitio.",
        "active_cards": active_cards,
        "status_counts": status_counts,
        "average_score": average_score,
        "active_count": len(active_cards),
        "last_checked": max((card["check"].checked_at for card in checked_cards), default=None),
        "history_date_start": history_date_start.isoformat(),
        "history_date_end": history_date_end.isoformat(),
    }
    return render(request, "reports/websites.html", context)


@require_http_methods(["GET", "POST"])
@never_cache
def web_sales_report(request):
    limited_redirect = _redirect_katerine_limited_user(request)
    if limited_redirect:
        return limited_redirect
    limited_redirect = _redirect_bali_whatsapp_user(request)
    if limited_redirect:
        return limited_redirect
    # Los catalogos se siembran con `manage.py ensure_axis_catalogs` y en los
    # comandos de importacion. Hacerlo aqui escribia en la base en cada carga.
    filter_form, filters = _global_filter_context(request, defaults=_web_sales_default_filters())

    initial_business_unit = BusinessUnit.objects.filter(slug=filters.get("business_unit") or "uva").first()
    initial_country = Country.objects.filter(code=filters.get("country") or "CO").first()
    initial_channel = Channel.objects.filter(slug=filters.get("channel") or "ecommerce-uva").first()
    initial_date = timezone.localdate() - timedelta(days=1)
    entry_form = DailyChannelSaleForm(
        prefix="daily_sale",
        initial={
            "business_unit": initial_business_unit,
            "country": initial_country,
            "channel": initial_channel,
            "sale_date": initial_date,
            "order_count": 0,
        },
    )

    if request.method == "POST":
        entry_form = DailyChannelSaleForm(request.POST, prefix="daily_sale")
        if entry_form.is_valid():
            data = entry_form.cleaned_data
            entry, created = DailyChannelSale.objects.update_or_create(
                business_unit=data["business_unit"],
                country=data["country"],
                channel=data["channel"],
                sale_date=data["sale_date"],
                defaults={
                    "sales_amount": data["sales_amount"],
                    "order_count": data["order_count"],
                    "notes": data["notes"],
                    "source_type": DailyChannelSale.SourceType.MANUAL,
                    "source_file": "",
                },
            )
            action_label = "registro" if created else "actualizo"
            messages.success(request, f"Se {action_label} la venta web del {entry.sale_date} correctamente.")
            return redirect("reports:web_sales")
        messages.error(request, "No fue posible registrar la venta. Revisa los campos resaltados.")

    sales_snapshot = build_sales_snapshot(filters, limit=200)
    context = {
        **_sidebar_context("web_sales", request),
        "page_title": "Ventas diarias",
        "page_subtitle": "Registro diario y seguimiento historico de ventas por marca, pais y canal directo.",
        "filter_form": filter_form,
        "entry_form": entry_form,
        "sales_rows": sales_snapshot["rows"],
        "sales_row_count": sales_snapshot["row_count"],
        "sales_by_day_json": json.dumps(sales_snapshot.get("sales_by_day", [])),
        "sales_by_channel_json": json.dumps(sales_snapshot.get("sales_by_channel", [])),
        "ad_platform_performance_json": json.dumps(build_ad_platform_performance(filters, sales_snapshot)),
        "snapshot": sales_snapshot,
    }
    return render(request, "reports/web_sales.html", context)


@require_http_methods(["GET", "POST"])
def ad_spend_report(request):
    limited_redirect = _redirect_katerine_limited_user(request)
    if limited_redirect:
        return limited_redirect
    limited_redirect = _redirect_bali_whatsapp_user(request)
    if limited_redirect:
        return limited_redirect
    # Los catalogos se siembran con `manage.py ensure_axis_catalogs` y en los
    # comandos de importacion. Hacerlo aqui escribia en la base en cada carga.
    platforms = {platform.slug: platform for platform in AdPlatform.objects.filter(is_active=True)}
    filter_form, filters = _global_filter_context(request, defaults=_ad_spend_default_filters())

    initial_business_unit = BusinessUnit.objects.filter(slug=filters.get("business_unit") or "uva").first()
    initial_country = Country.objects.filter(code=filters.get("country") or "CO").first()
    initial_platform = AdPlatform.objects.filter(slug=request.GET.get("ad_platform") or "meta-ads").first() or platforms.get("meta-ads")
    initial_date = timezone.localdate() - timedelta(days=1)
    entry_form = DailyAdSpendForm(
        prefix="daily_spend",
        initial={
            "business_unit": initial_business_unit,
            "country": initial_country,
            "ad_platform": initial_platform,
            "spend_date": initial_date,
        },
    )

    if request.method == "POST":
        entry_form = DailyAdSpendForm(request.POST, prefix="daily_spend")
        if entry_form.is_valid():
            data = entry_form.cleaned_data
            spend, created = DailyAdSpend.objects.update_or_create(
                business_unit=data["business_unit"],
                country=data["country"],
                ad_platform=data["ad_platform"],
                spend_date=data["spend_date"],
                defaults={
                    "spend_amount": data["spend_amount"],
                    "notes": data["notes"],
                    "source_type": DailyAdSpend.SourceType.MANUAL,
                    "source_file": "",
                },
            )
            action_label = "registro" if created else "actualizo"
            messages.success(request, f"Se {action_label} la inversion del {spend.spend_date} correctamente.")
            return redirect("reports:ad_spend")
        messages.error(request, "No fue posible registrar la inversion. Revisa los campos resaltados.")

    snapshot = build_sales_snapshot(filters, limit=200)
    spend_rows = DailyAdSpend.objects.select_related("business_unit", "country", "ad_platform").filter(
        business_unit__slug=filters.get("business_unit") or "uva",
        country__code=filters.get("country") or "CO",
    )
    if filters.get("date_start"):
        spend_rows = spend_rows.filter(spend_date__gte=filters["date_start"])
    if filters.get("date_end"):
        spend_rows = spend_rows.filter(spend_date__lte=filters["date_end"])
    context = {
        **_sidebar_context("ad_spend", request),
        "page_title": "Inversion diaria",
        "page_subtitle": "Registro diario de Meta Ads y Google Ads para calcular ventas, inversion y ROAS.",
        "filter_form": filter_form,
        "entry_form": entry_form,
        "spend_rows": spend_rows.order_by("-spend_date", "ad_platform__name"),
        "spend_row_count": spend_rows.count(),
        "snapshot": snapshot,
        "sales_by_day_json": json.dumps(snapshot.get("sales_by_day", [])),
        "spend_by_day_json": json.dumps(snapshot.get("spend_by_day", [])),
        "roas_by_day_json": json.dumps(snapshot.get("roas_by_day", [])),
        "ad_platform_performance_json": json.dumps(build_ad_platform_performance(filters, snapshot)),
    }
    return render(request, "reports/ad_spend.html", context)


def _unit_module(request, unit_slug, template_name, nav_key, title, subtitle, filter_overrides=None):
    filter_form, filters = _global_filter_context(request)
    if filter_overrides:
        filters.update(filter_overrides)
    summary = build_unit_summary(unit_slug, filters)
    summary["insight_cards"] = _fallback_insight_cards(summary.get("insights", []))
    context = {
        **_sidebar_context(nav_key, request),
        "page_title": title,
        "page_subtitle": subtitle,
        "filter_form": filter_form,
        "summary": summary,
        "countries": Country.objects.filter(is_active=True),
        "channels": Channel.objects.filter(is_active=True, business_unit__slug=unit_slug),
        "products": Product.objects.filter(is_active=True, business_unit__slug=unit_slug),
        "sales_by_unit_json": json.dumps(summary["sales_by_unit"]),
        "sales_by_channel_json": json.dumps(summary["sales_by_channel"]),
        "roas_by_unit_json": json.dumps(summary["roas_by_unit"]),
        "country_sales_json": json.dumps(summary["country_sales"]),
        "product_sales_json": json.dumps(summary["product_sales"]),
        "investment_by_product_json": json.dumps(summary.get("investment_by_product_rows", [])),
        "ad_spend_by_country_json": json.dumps(summary.get("ad_spend_by_country", [])),
        "ad_spend_by_channel_json": json.dumps(summary.get("ad_spend_by_channel", [])),
        "messages_by_channel_json": json.dumps(summary.get("messages_by_channel", [])),
        "cpa_by_product_json": json.dumps(summary.get("cpa_by_product", [])),
        "cpl_whatsapp_json": json.dumps(summary.get("cpl_whatsapp", [])),
        "comfama_sales_json": json.dumps(summary.get("comfama_sales_by_product", [])),
        "comfama_messages_json": json.dumps(summary.get("comfama_messages_by_product", [])),
        "comfama_investment_json": json.dumps(summary.get("comfama_investment_by_product", [])),
    }
    return render(request, template_name, context)


@never_cache
def uva_module(request):
    limited_redirect = _redirect_katerine_limited_user(request)
    if limited_redirect:
        return limited_redirect
    limited_redirect = _redirect_bali_whatsapp_user(request)
    if limited_redirect:
        return limited_redirect
    context = _executive_dashboard_context(
        request,
        active_key="uva",
        title="Uva",
        subtitle="Dashboard ejecutivo exclusivo de la marca Uva.",
        filter_overrides={"business_unit": "uva"},
    )
    return render(request, "reports/dashboard.html", context)


def uva_comfama_module(request):
    limited_redirect = _redirect_katerine_limited_user(request)
    if limited_redirect:
        return limited_redirect
    limited_redirect = _redirect_bali_whatsapp_user(request)
    if limited_redirect:
        return limited_redirect
    filter_form, filters = _global_filter_context(request, defaults=_dashboard_default_filters(), overrides={"business_unit": "uva", "country": "CO"})
    comfama_snapshot = build_comfama_snapshot(filters)
    comfama_meta_ads_preview = build_uva_meta_ads_preview(filters, comfama_scope="only", allow_live_fetch=False)
    context = {
        **_sidebar_context("uva_comfama", request),
        "page_title": "Uva Comfama",
        "page_subtitle": "Informe de ventas WhatsApp Comfama, mensajes, conversion, CPL y ROAS.",
        "filter_form": filter_form,
        "filters": filters,
        "comfama_snapshot": comfama_snapshot,
        "comfama_meta_ads_preview": comfama_meta_ads_preview,
        "comfama_categories_json": json.dumps(comfama_snapshot.get("categories", [])),
        "comfama_daily_json": json.dumps(comfama_snapshot.get("daily_series", [])),
    }
    return render(request, "reports/uva_comfama.html", context)


def awn_internacional_module(request):
    limited_redirect = _redirect_katerine_limited_user(request)
    if limited_redirect:
        return limited_redirect
    limited_redirect = _redirect_bali_whatsapp_user(request)
    if limited_redirect:
        return limited_redirect
    filter_form, filters = _global_filter_context(request, defaults=_dashboard_default_filters())
    filter_form.fields["country"].choices = [("", "Todos los paises"), ("EC", "Ecuador"), ("MX", "Mexico")]
    awn_snapshot = build_awn_international_snapshot(filters)
    context = {
        **_sidebar_context("awn_internacional", request),
        "page_title": "Awn Internacional",
        "page_subtitle": "Seguimiento de campanas de seguidores en Instagram para Ecuador y Mexico.",
        "filter_form": filter_form,
        "filters": filters,
        "awn_snapshot": awn_snapshot,
        "awn_countries_json": json.dumps(awn_snapshot.get("countries", [])),
        "awn_daily_json": json.dumps(awn_snapshot.get("daily_series", [])),
    }
    return render(request, "reports/awn_internacional.html", context)


def _conekta_local_credential_status():
    credential_path = Path(r"C:\Users\trafficker.digital\Documents\conekta-api.txt")
    if not credential_path.exists():
        return {
            "file_found": False,
            "message": "No encontré el archivo local de credenciales de Conekta.",
            "entries": [],
        }
    entries = []
    try:
        for line in credential_path.read_text(encoding="utf-8").splitlines():
            if ":" not in line:
                continue
            label, value = line.split(":", 1)
            clean_value = value.strip().strip('"').strip("'")
            if not clean_value:
                continue
            entries.append(
                {
                    "label": label.strip(),
                    "length": len(clean_value),
                    "looks_like_conekta_private_key": clean_value.startswith("key_"),
                }
            )
    except OSError:
        return {
            "file_found": True,
            "message": "Encontré el archivo, pero no pude leerlo desde Axis.",
            "entries": [],
        }
    return {
        "file_found": True,
        "message": "Archivo localizado. Las credenciales deben validarse contra Conekta antes de sincronizar datos.",
        "entries": entries,
    }


@never_cache
def distrisex_ecuador_module(request):
    conekta_resources = [
        {"name": "Órdenes", "endpoint": "GET /orders", "can_power": "Ventas, estado de pago, moneda, productos/line_items, cliente, método de pago, impuestos, descuentos y envíos."},
        {"name": "Cargos", "endpoint": "GET /charges", "can_power": "Detalle transaccional: monto cobrado, estado, payment_method, referencia, order_id y búsquedas por cliente/correo/referencia."},
        {"name": "Clientes", "endpoint": "GET /customers", "can_power": "Base de clientes, recurrencia, medios de pago asociados y posibles cohortes de recompra."},
        {"name": "Balance", "endpoint": "GET /balance", "can_power": "Saldo disponible/pendiente para conciliación financiera de DistriSex Ecuador."},
        {"name": "Transacciones", "endpoint": "GET /transactions", "can_power": "Movimientos contables del balance: cargos, fees, refunds, payouts y ajustes según disponibilidad de cuenta."},
        {"name": "Transfers / Payouts", "endpoint": "GET /transfers y GET /payout_orders", "can_power": "Retiros/liquidaciones: montos, fechas, estados y conciliación contra bancos."},
        {"name": "Links de pago", "endpoint": "GET /payment_links", "can_power": "Ventas por links, links activos/cancelados, expiración y seguimiento para canales como WhatsApp."},
        {"name": "Eventos y webhooks", "endpoint": "GET /events, GET /webhooks", "can_power": "Auditoría de cambios de estado y sincronización automática casi en tiempo real."},
        {"name": "Suscripciones / Planes", "endpoint": "GET /subscriptions, GET /plans", "can_power": "Cobros recurrentes, churn, pausas/cancelaciones y eventos de retry si DistriSex usa billing."},
    ]
    dashboard_opportunities = [
        "Ventas aprobadas vs pendientes/fallidas por día.",
        "Valor cobrado, fees estimados y neto conciliable.",
        "Métodos de pago: tarjeta, efectivo, transferencia u otros habilitados por la cuenta.",
        "Tasa de aprobación y motivos de rechazo por método.",
        "Órdenes por producto o line_items si las órdenes se crean con carrito detallado.",
        "Clientes nuevos vs recurrentes y frecuencia de recompra.",
        "Links de pago activos, expirados, cancelados y conversión por canal.",
        "Eventos/webhooks fallidos para monitorear problemas de sincronización.",
    ]
    context = {
        **_sidebar_context("distrisex_ecuador", request),
        "page_title": "DistriSex Ecuador",
        "page_subtitle": "Diagnóstico inicial de integración Conekta para DistriSex Ecuador.",
        "credential_status": _conekta_local_credential_status(),
        "conekta_resources": conekta_resources,
        "dashboard_opportunities": dashboard_opportunities,
        "auth_note": "La auditoría local contra Conekta devolvió 401 con los dos valores del archivo. Para traer datos reales necesitamos una llave privada válida de Conekta, normalmente generada en el Panel y usada como Authorization Bearer.",
    }
    return render(request, "reports/distrisex_ecuador.html", context)


@never_cache
def bali_module(request):
    # Los catalogos se siembran con `manage.py ensure_axis_catalogs` y en los
    # comandos de importacion. Hacerlo aqui escribia en la base en cada carga.
    tab = request.GET.get("tab", "resumen")
    filter_form, filters = _global_filter_context(request, defaults=_dashboard_default_filters(), overrides={"business_unit": "bali", "country": "CO"})
    snapshot = build_bali_snapshot(filters)
    tab_query = _effective_filter_query(filters, exclude={"tab", "channel"})
    context = {
        **_sidebar_context("bali", request),
        "page_title": "Bali",
        "page_subtitle": "Dashboard ejecutivo de Shopify, Google Ads, WhatsApp y Comunidad Webcam.",
        "filter_form": filter_form,
        "filter_fields": [filter_form["date_start"], filter_form["date_end"]],
        "filters": filters,
        "bali_tab": tab,
        "bali_querystring": tab_query.urlencode(),
        "bali_snapshot": snapshot,
        "bali_channels_json": json.dumps(snapshot.get("channels", [])),
        "bali_daily_json": json.dumps(snapshot.get("daily_series", [])),
        "bali_web_daily_json": json.dumps(snapshot.get("web_daily", [])),
        "bali_whatsapp_daily_json": json.dumps(snapshot.get("whatsapp_daily", [])),
        "bali_physical_daily_json": json.dumps(snapshot.get("physical_daily", [])),
        "bali_community_daily_json": json.dumps(snapshot.get("community", {}).get("daily_series", [])),
    }
    return render(request, "reports/unit_bali.html", context)


@require_POST
def uva_meta_ads_panel_api(request):
    """Trae de Meta el panel de anuncios y lo deja en cache.

    La pagina ya no espera esta llamada: /uva/ tardaba 16 s con la cache fria
    porque el render se quedaba bloqueado en varias peticiones HTTP a Meta. Ahora
    la pagina sale de inmediato con el panel en `pending` y el navegador pide este
    endpoint aparte, con un timeout holgado.
    """
    filters = {
        "country": request.POST.get("country") or "",
        "date_start": request.POST.get("date_start") or "",
        "date_end": request.POST.get("date_end") or "",
    }
    scope = "only" if request.POST.get("comfama_scope") == "only" else "exclude"
    timeout = getattr(settings, "META_ADS_PREVIEW_PANEL_TIMEOUT", 60)
    preview = build_uva_meta_ads_preview(filters, comfama_scope=scope, timeout=timeout)
    return JsonResponse(
        {
            "ok": not preview.get("pending"),
            "ad_count": len(preview.get("ads") or []),
            "message": preview.get("message") or "",
        }
    )


@require_http_methods(["GET"])
def product_detail_api(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        raise PermissionDenied

    unit = request.GET.get("unit", "").strip().lower()
    if unit == "bali":
        filters = _detail_filter_context(request, overrides={"business_unit": "bali", "country": "CO"})
        payload = build_bali_product_detail(filters, request.GET.get("product"))
    elif unit == "uva":
        filters = _detail_filter_context(request, overrides={"business_unit": "uva"})
        payload = build_uva_product_detail(filters, request.GET.get("category_id") or request.GET.get("product"))
    elif unit == "marketplace":
        filters = _detail_filter_context(request, overrides={"business_unit": "marketplace"})
        payload = build_marketplace_product_detail(filters, request.GET.get("marketplace"), request.GET.get("item_id") or request.GET.get("product"))
    else:
        return JsonResponse({"detail": "Unidad no soportada."}, status=400)

    if not payload:
        raise Http404("No hay detalle disponible para el producto seleccionado.")
    return JsonResponse(payload)


@never_cache
def marketplace_module(request):
    limited_redirect = _redirect_bali_whatsapp_user(request)
    if limited_redirect:
        return limited_redirect
    # Los catalogos se siembran con `manage.py ensure_axis_catalogs` y en los
    # comandos de importacion. Hacerlo aqui escribia en la base en cada carga.
    filter_form, filters = _global_filter_context(request, defaults=_dashboard_default_filters(), overrides={"business_unit": "marketplace"})
    selected_channel = request.GET.get("channel", "")
    if selected_channel:
        filters["channel"] = selected_channel
    sales_snapshot = build_sales_snapshot({**filters, "business_unit": "marketplace"})
    marketplace_channels = list(Channel.objects.filter(is_active=True, business_unit__slug="marketplace").order_by("display_order", "name"))
    summary_params = _effective_filter_query(filters, exclude={"channel"})
    marketplace_channel_tabs = []
    for channel in marketplace_channels:
        tab_params = _effective_filter_query(filters, extra={"channel": channel.slug})
        marketplace_channel_tabs.append(
            {
                "name": channel.name,
                "slug": channel.slug,
                "url": f"?{tab_params.urlencode()}",
                "active": selected_channel == channel.slug,
            }
        )
    target_rows = _marketplace_target_rows(request.user, filters) if FEATURE_TASKS_GOALS_ENABLED else []
    comparisons = sales_snapshot.get("comparison", {})
    visible_kpis = [
        {
            "label": "Ventas",
            "value": sales_snapshot["kpis"].get("sales_total", 0),
            "value_with_vat": sales_snapshot["kpis"].get("sales_total_with_vat", 0),
            "kind": "money",
            "comparison": comparisons.get("sales_total"),
        },
        {"label": "Inversion", "value": sales_snapshot["kpis"].get("ad_spend", 0), "kind": "money", "comparison": comparisons.get("ad_spend")},
        {"label": "Pedidos", "value": sales_snapshot["kpis"].get("orders", 0), "kind": "number"},
        {"label": "Unidades", "value": sales_snapshot["kpis"].get("units", 0), "kind": "number"},
        {"label": "ROAS", "value": sales_snapshot["kpis"].get("roas", 0), "kind": "decimal", "comparison": comparisons.get("roas")},
        {
            "label": "Ticket promedio",
            "value": sales_snapshot["kpis"].get("average_ticket", 0),
            "value_with_vat": sales_snapshot["kpis"].get("average_ticket_with_vat", 0),
            "kind": "money",
            "comparison": comparisons.get("average_ticket"),
        },
    ]
    visible_kpis = [item for item in visible_kpis if item["value"]]
    chart_series = [row for row in sales_snapshot.get("combined_series", []) if row.get("sales") or row.get("spend")]
    inventory_snapshot = {"totals": {"total": 0}, "rows": [], "marketplace": "", "label": ""}
    if selected_channel in {"mercado-libre", "falabella"}:
        inventory_marketplace = "falabella" if selected_channel == "falabella" else "mercadolibre"
        inventory_snapshot = marketplace_inventory_snapshot(marketplace=inventory_marketplace)

    context = {
        **_sidebar_context("marketplace", request),
        **_operational_task_context(request.user),
        "page_title": "Ventas Marketplace",
        "page_subtitle": "Detalle diario por canal: ventas, inversion, pedidos y unidades.",
        "filter_form": filter_form,
        "summary": {"kpis": sales_snapshot["kpis"]},
        "snapshot": sales_snapshot,
        "marketplace_kpis": visible_kpis,
        "marketplace_channels": marketplace_channels,
        "marketplace_channel_tabs": marketplace_channel_tabs,
        "marketplace_summary_url": f"?{summary_params.urlencode()}" if summary_params else reverse("reports:marketplace"),
        "selected_marketplace_channel": selected_channel,
        "marketplace_targets": target_rows,
        "marketplace_goal_summary": _marketplace_goal_summary(target_rows, sales_snapshot),
        "marketplace_targets_json": json.dumps(target_rows),
        "marketplace_channel_json": json.dumps(sales_snapshot.get("sales_by_channel", [])),
        "marketplace_series_json": json.dumps(chart_series),
        "marketplace_inventory": inventory_snapshot,
    }
    return render(request, "reports/marketplace_daily.html", context)


@require_http_methods(["GET"])
def goals_dashboard(request):
    if not FEATURE_TASKS_GOALS_ENABLED:
        raise Http404()
    limited_redirect = _redirect_katerine_limited_user(request)
    if limited_redirect:
        return limited_redirect
    if not _can_view_goals_dashboard(request.user):
        raise Http404()
    filter_form, filters = _global_filter_context(request, defaults=_dashboard_default_filters())
    target_rows, task_rows = _goals_dashboard_rows(request.user, filters)
    achievement_owner_ids = _managed_user_ids(request.user) or ([request.user.id] if _has_own_goals(request.user) else [])
    achievement_month_start, achievement_month_end = _selected_achievement_month(filters)
    achievement_rows = _sync_monthly_achievements(achievement_owner_ids, achievement_month_start, achievement_month_end)
    total_target = sum(row["target"] for row in target_rows)
    total_sales = sum(row["sales"] for row in target_rows)
    goals_kpis = [
        {"label": "Meta total", "value": total_target, "kind": "money"},
        {"label": "Ventas acumuladas", "value": total_sales, "kind": "money"},
        {"label": "Cumplimiento", "value": round((total_sales / total_target) * 100, 1) if total_target else 0, "kind": "percent"},
        {"label": "Faltante", "value": max(total_target - total_sales, 0), "kind": "money"},
        {"label": "Logros del mes", "value": len(achievement_rows), "kind": "number"},
    ]
    goals_kpis = [item for item in goals_kpis if item["value"]]
    context = {
        **_sidebar_context("goals", request),
        **_operational_task_context(request.user),
        **_user_task_context(request.user),
        "page_title": "Metas",
        "page_subtitle": "Cumplimiento de metas por usuario y seguimiento de metas operativas.",
        "filter_form": filter_form,
        "goal_rows": target_rows,
        "goal_task_rows": task_rows,
        "achievement_rows": achievement_rows,
        "achievement_month_label": achievement_month_start.strftime("%B %Y").capitalize(),
        "goals_kpis": goals_kpis,
        "goal_rows_json": json.dumps(target_rows),
    }
    return render(request, "reports/goals.html", context)


@require_http_methods(["GET"])
def tasks_dashboard(request):
    if not FEATURE_TASKS_GOALS_ENABLED:
        raise Http404()
    limited_redirect = _redirect_katerine_limited_user(request)
    if limited_redirect:
        return limited_redirect
    if not request.user.is_authenticated:
        raise Http404()
    context = {
        **_sidebar_context("tasks", request),
        **_user_task_context(request.user),
        **_task_calendar_context(request.user, request.GET.get("week")),
        **_task_agents_context(request.user, request),
        "page_title": "Tareas",
        "page_subtitle": "Calendario interactivo para ubicar pendientes por fecha y hora.",
    }
    return render(request, "reports/tasks.html", context)


@require_POST
def create_user_task_from_calendar(request):
    if not request.user.is_authenticated:
        raise Http404()
    title = (request.POST.get("title") or "").strip()
    due_date = _parse_task_date(request.POST.get("due_date"))
    due_time = _parse_task_time(request.POST.get("due_time")) or time(8, 0)
    if not title or not due_date:
        return JsonResponse({"ok": False, "error": "Titulo y fecha son obligatorios."}, status=400)
    try:
        registered_hours = Decimal(request.POST.get("registered_hours") or "1")
    except (InvalidOperation, TypeError):
        registered_hours = Decimal("1")
    if registered_hours <= 0:
        registered_hours = Decimal("1")

    assigned_to = request.user
    assigned_to_id = request.POST.get("assigned_to")
    visible_ids = {request.user.id, *_managed_user_ids(request.user)}
    if assigned_to_id and assigned_to_id.isdigit() and int(assigned_to_id) in visible_ids:
        assigned_to = User.objects.filter(id=int(assigned_to_id)).first() or request.user

    task = UserTask.objects.create(
        created_by=request.user,
        assigned_to=assigned_to,
        title=title,
        description=(request.POST.get("description") or "").strip(),
        links=(request.POST.get("links") or "").strip(),
        due_date=due_date,
        due_time=due_time,
        registered_hours=registered_hours,
        status=request.POST.get("status") or UserTask.Status.PENDING,
    )
    return JsonResponse({"ok": True, "task": _task_calendar_payload(task)})


@require_POST
def update_user_task_schedule(request, pk):
    task = get_object_or_404(UserTask.objects.select_related("assigned_to", "created_by"), pk=pk)
    if not _can_access_user_task(request.user, task):
        raise Http404()
    due_date = _parse_task_date(request.POST.get("due_date"))
    due_time = _parse_task_time(request.POST.get("due_time"))
    if not due_date or not due_time:
        return JsonResponse({"ok": False, "error": "Fecha y hora invalidas."}, status=400)
    task.due_date = due_date
    task.due_time = due_time
    task.save(update_fields=["due_date", "due_time", "updated_at"])
    return JsonResponse({"ok": True, "task": _task_calendar_payload(task)})


@require_POST
def update_user_task_status(request, pk):
    task = get_object_or_404(UserTask.objects.select_related("assigned_to", "created_by"), pk=pk)
    if not _can_access_user_task(request.user, task):
        raise Http404()
    status = request.POST.get("status")
    valid_statuses = {choice[0] for choice in UserTask.Status.choices}
    if status not in valid_statuses:
        return JsonResponse({"ok": False, "error": "Estado invalido."}, status=400)
    task.status = status
    task.completed_at = timezone.now() if status == UserTask.Status.COMPLETED else None
    task.save(update_fields=["status", "completed_at", "updated_at"])
    return JsonResponse({"ok": True, "task": _task_calendar_payload(task)})


@require_POST
def update_user_task_from_calendar(request, pk):
    task = get_object_or_404(UserTask.objects.select_related("assigned_to", "created_by"), pk=pk)
    if not _can_access_user_task(request.user, task):
        raise Http404()
    title = (request.POST.get("title") or "").strip()
    due_date = _parse_task_date(request.POST.get("due_date"))
    due_time = _parse_task_time(request.POST.get("due_time"))
    if not title or not due_date or not due_time:
        return JsonResponse({"ok": False, "error": "Titulo, fecha y hora son obligatorios."}, status=400)
    try:
        registered_hours = Decimal(request.POST.get("registered_hours") or "1")
    except (InvalidOperation, TypeError):
        registered_hours = Decimal("1")
    status = request.POST.get("status") or task.status
    valid_statuses = {choice[0] for choice in UserTask.Status.choices}
    if status not in valid_statuses:
        status = task.status
    task.title = title
    task.description = (request.POST.get("description") or "").strip()
    task.links = (request.POST.get("links") or "").strip()
    task.due_date = due_date
    task.due_time = due_time
    task.registered_hours = registered_hours if registered_hours > 0 else Decimal("1")
    task.status = status
    task.completed_at = timezone.now() if status == UserTask.Status.COMPLETED else None
    task.save(update_fields=["title", "description", "links", "due_date", "due_time", "registered_hours", "status", "completed_at", "updated_at"])
    return JsonResponse({"ok": True, "task": _task_calendar_payload(task)})


@require_http_methods(["GET"])
def export_agent_tasks(request):
    if not request.user.is_authenticated:
        raise Http404()
    agents = _task_agent_users(request.user)
    agent_id = request.GET.get("agent")
    if not agent_id or not agent_id.isdigit():
        raise Http404()
    agent = next((item for item in agents if item.id == int(agent_id)), None)
    if not agent:
        raise Http404()
    start = _parse_task_date(request.GET.get("desde"))
    end = _parse_task_date(request.GET.get("hasta"))
    weeks = _agent_weekly_tasks(agent, start, end)
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="42122B")
    done_fill = PatternFill("solid", fgColor="B7E1CD")
    pending_fill = PatternFill("solid", fgColor="C9DAF8")
    canceled_fill = PatternFill("solid", fgColor="D9E2F3")
    for week_index, week in enumerate(weeks, start=1):
        ws = wb.create_sheet(f"Semana {week_index}")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws.cell(row=1, column=1, value=week["title"]).font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        max_rows = max([len(day["tasks"]) for day in week["days"]] or [1])
        for column, day in enumerate(week["days"], start=1):
            cell = ws.cell(row=2, column=column, value=day["label"])
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[cell.column_letter].width = 34
            for row_offset in range(max_rows):
                task = day["tasks"][row_offset] if row_offset < len(day["tasks"]) else None
                body = ws.cell(row=3 + row_offset, column=column)
                body.alignment = Alignment(wrap_text=True, vertical="top")
                if task:
                    body.value = f"{task.title}\n{task.get_status_display()} | {float(task.registered_hours or 0):.1f} h"
                    if task.status == UserTask.Status.COMPLETED:
                        body.fill = done_fill
                    elif task.status == UserTask.Status.CANCELED:
                        body.fill = canceled_fill
                    else:
                        body.fill = pending_fill
                else:
                    body.value = ""
        for row in range(3, 3 + max_rows):
            ws.row_dimensions[row].height = 48
    if not weeks:
        ws = wb.create_sheet("Tareas")
        ws.append(["Sin tareas en el rango seleccionado."])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"tareas_{agent.username}_{timezone.localdate().isoformat()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@require_POST
def update_operational_task(request, pk):
    task = get_object_or_404(OperationalGoalTask, pk=pk, assigned_to=request.user)
    form = OperationalGoalTaskUpdateForm(request.POST, instance=task)
    if not form.is_valid():
        messages.error(request, "No fue posible guardar la actualizacion de la tarea.")
        return redirect(_safe_next_url(request))

    task = form.save(commit=False)
    action = request.POST.get("action")
    if action == "complete":
        task.status = OperationalGoalTask.Status.COMPLETED
        task.completed_at = timezone.now()
        message = "Tarea marcada como completada."
    elif task.status == OperationalGoalTask.Status.PENDING:
        task.status = OperationalGoalTask.Status.IN_PROGRESS
        message = "Avance guardado."
    else:
        message = "Avance guardado."
    task.save()
    messages.success(request, message)
    return redirect(_safe_next_url(request))


@require_http_methods(["GET"])
def operation_weekly(request):
    limited_redirect = _redirect_katerine_limited_user(request)
    if limited_redirect:
        return limited_redirect
    filter_form = WeeklyTaskFilterForm(request.GET or None)
    filters = build_filter_dict(request.GET)
    tasks = weekly_tasks(filters)
    context = {
        **_sidebar_context("operation", request),
        "page_title": "Operacion semanal",
        "page_subtitle": "Seguimiento operativo, cierre ejecutivo y tareas criticas.",
        "filter_form": filter_form,
        "tasks": tasks,
        "task_summary": {
            "total": len(tasks),
            "completed": sum(1 for task in tasks if task.status == "completed"),
            "in_progress": sum(1 for task in tasks if task.status == "in_progress"),
            "blocked": sum(1 for task in tasks if task.status == "blocked"),
            "critical": sum(1 for task in tasks if task.priority == "critical"),
            "by_area": dict(__import__("collections").Counter(task.area for task in tasks)),
            "by_unit": dict(__import__("collections").Counter(task.business_unit.name if task.business_unit else "Sin unidad" for task in tasks)),
            "by_status": dict(__import__("collections").Counter(task.status for task in tasks)),
            "by_impact": dict(__import__("collections").Counter(task.impact for task in tasks)),
        },
    }
    return render(request, "reports/operation_weekly.html", context)


@require_http_methods(["GET", "POST"])
def files_module(request):
    limited_redirect = _redirect_katerine_limited_user(request)
    if limited_redirect:
        return limited_redirect
    filter_form = WeeklyTaskFilterForm(request.GET or None)
    filters = build_filter_dict(request.GET)
    upload_form = AttachmentUploadForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and upload_form.is_valid():
        files = upload_form.cleaned_data["files"]
        business_unit = upload_form.cleaned_data.get("business_unit")
        country = upload_form.cleaned_data.get("country")
        channel = upload_form.cleaned_data.get("channel")
        task = upload_form.cleaned_data.get("task")
        period_label = upload_form.cleaned_data.get("period_label") or (task.week_label if task else "")
        description = upload_form.cleaned_data.get("description", "")
        tags = upload_form.cleaned_data.get("tags", "")
        start_count = Attachment.objects.count() + 1
        for index, uploaded in enumerate(files, start=start_count):
            Attachment.objects.create(
                attachment_ref=_build_attachment_ref(index),
                business_unit=business_unit,
                country=country,
                channel=channel,
                period_label=period_label,
                task=task,
                file_name=uploaded.name,
                file_type=_attachment_file_type(uploaded),
                uploaded_file=uploaded,
                description=description,
                tags=tags,
                comment=description,
            )
        messages.success(request, f"Se cargaron {len(files)} archivo(s) correctamente.")
        return redirect("reports:files")

    attachment_rows = attachments(filters)
    context = {
        **_sidebar_context("files", request),
        "page_title": "Archivos",
        "page_subtitle": "Soportes, reportes y evidencias relacionadas con metricas y tareas.",
        "filter_form": filter_form,
        "attachments": attachment_rows,
        "upload_form": upload_form,
    }
    return render(request, "reports/files.html", context)


@require_http_methods(["GET", "POST"])
def excel_center(request):
    limited_redirect = _redirect_katerine_limited_user(request)
    if limited_redirect:
        return limited_redirect
    import_form = MasterImportForm()
    export_form = ExportRequestForm(request.GET or None)
    preview_payload = request.session.get(MASTER_IMPORT_SESSION_KEY)

    if request.method == "POST":
        step = request.POST.get("step", "preview")
        if step == "confirm":
            preview_payload = request.session.get(MASTER_IMPORT_SESSION_KEY)
            if not preview_payload:
                messages.error(request, "La sesion de importacion expiro.")
                return redirect("reports:excel_center")
            job = commit_master_import(preview_payload)
            request.session.pop(MASTER_IMPORT_SESSION_KEY, None)
            if job.status == job.Status.FAILED:
                messages.error(request, "La importacion tiene errores criticos. Corrige el archivo antes de guardar.")
            else:
                messages.success(request, "Importacion completada correctamente.")
            return redirect("reports:excel_center")

        import_form = MasterImportForm(request.POST, request.FILES)
        if import_form.is_valid():
            uploaded = import_form.cleaned_data["excel_file"]
            preview_payload = preview_master_import(uploaded, uploaded.name)
            request.session[MASTER_IMPORT_SESSION_KEY] = preview_payload
            if preview_payload["critical_errors"]:
                messages.error(request, "Se encontraron errores criticos en el archivo maestro.")
            else:
                messages.success(request, "Preview generado. Puedes confirmar la importacion.")

    context = {
        **_sidebar_context("excel", request),
        "page_title": "Importar / Exportar Excel",
        "page_subtitle": "Centro maestro para importacion, exportacion y plantilla reutilizable.",
        "import_form": import_form,
        "export_form": export_form,
        "preview": preview_payload,
    }
    return render(request, "reports/excel_center.html", context)


@require_http_methods(["GET"])
def export_master_template(request):
    content = build_master_workbook(template_only=True, include_data=False)
    response = HttpResponse(content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="master_template.xlsx"'
    return response


@require_http_methods(["GET"])
def export_master_data(request):
    filters = build_filter_dict(request.GET)
    scope = request.GET.get("export_scope") or "master"
    create_export_job(scope, filters)
    content = build_master_workbook(filters=filters, include_data=True, template_only=False)
    response = HttpResponse(content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="master_export.xlsx"'
    return response


@require_http_methods(["GET", "POST"])
def settings_view(request):
    user = request.user if request.user.is_authenticated else None
    if not user:
        messages.error(request, "Debes iniciar sesion para ver tu perfil.")
        return redirect("/admin/login/?next=/settings/")

    profile, _ = UserProfile.objects.get_or_create(user=user)
    profile_form = ProfileForm(request.POST or None, request.FILES or None, instance=profile, user=user)
    if request.method == "POST":
        if profile_form.is_valid():
            profile_form.save()
            messages.success(request, "Tu perfil se actualizo correctamente.")
            return redirect("reports:settings")
        messages.error(request, "No fue posible actualizar tu perfil. Revisa los campos resaltados.")

    context = {
        **_sidebar_context("settings", request),
        "page_title": "Mi perfil",
        "page_subtitle": "Consulta tu informacion personal y actualiza tus datos de contacto.",
        "profile_form": profile_form,
        "profile_obj": profile,
        "user_obj": user,
    }
    return render(request, "reports/profile.html", context)




