import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from reports.models import Task


@dataclass
class ParsedRow:
    row_number: int
    data: dict
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self):
        return not self.errors


HEADER_ALIASES = {
    "area": {"area"},
    "task_name": {"tarea actividad", "tarea", "actividad", "task", "task name"},
    "responsible": {"responsable", "responsible"},
    "status": {"estado", "status"},
    "priority": {"prioridad", "priority"},
    "observations": {"observaciones", "observacion", "notas", "notes"},
    "category": {"categoria", "category"},
}

SPANISH_MONTHS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}

AREA_ALIASES = {
    "ecommerce": Task.AREA_ECOMMERCE,
    "pauta": Task.AREA_PAUTA,
    "marketplace": Task.AREA_MARKETPLACE,
    "chat web bali": Task.AREA_CHAT_WEB_BALI,
}


def normalize_text(value):
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = "".join(char for char in text if unicodedata.category(char) not in {"So", "Cs"})
    return re.sub(r"\s+", " ", text).strip()


def normalize_header(value):
    return normalize_text(value).replace("/", " ").casefold()


def choice_map(choices):
    return {normalize_text(label).casefold(): value for value, label in choices}


AREA_MAP = choice_map(Task.AREA_CHOICES)
STATUS_MAP = choice_map(Task.STATUS_CHOICES)
PRIORITY_MAP = choice_map(Task.PRIORITY_CHOICES)


def _load_workbook(source):
    if isinstance(source, bytes):
        return load_workbook(filename=BytesIO(source), read_only=True, data_only=True)
    if hasattr(source, "read"):
        return load_workbook(filename=BytesIO(source.read()), read_only=True, data_only=True)
    raise TypeError("El origen del archivo no es compatible.")


def _cell_text(cell):
    return normalize_text(cell.value)


def _find_header_row(worksheet):
    for row in worksheet.iter_rows(min_row=1, max_row=min(30, worksheet.max_row)):
        columns = {}
        for index, cell in enumerate(row):
            header = normalize_header(cell.value)
            for target, aliases in HEADER_ALIASES.items():
                if header in aliases:
                    columns[target] = index
        if {"task_name", "status", "priority"}.issubset(columns):
            return row[0].row, columns
    return None, {}


def _extract_metadata_value(worksheet, label):
    for row in worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row)):
        for cell in row:
            text = _cell_text(cell)
            if f"{label}:" in text:
                return text.split(f"{label}:", 1)[1].strip()
    return ""


def _area_from_sheet_name(sheet_name):
    cleaned = normalize_text(sheet_name).casefold()
    for key, area in AREA_ALIASES.items():
        if key in cleaned:
            return area
    return ""


def _value_from_row(row, columns, key):
    index = columns.get(key)
    if index is None or index >= len(row):
        return ""
    return row[index]


def _discover_parse_targets(workbook):
    targets = []
    for worksheet in workbook.worksheets:
        header_row, columns = _find_header_row(worksheet)
        if header_row:
            targets.append((worksheet, header_row, columns))

    summary_targets = [target for target in targets if "area" in target[2]]
    return summary_targets or targets


def normalize_choice(value, valid_map):
    return valid_map.get(normalize_text(value).casefold())


def _parse_week_range(week_label):
    cleaned = normalize_text(week_label)
    match = re.search(
        r"(?P<start>\d{1,2})\s*(?:-|\u2013|\u2014)\s*(?P<end>\d{1,2})\s+(?P<month>[^\W\d_]+)\s+(?P<year>\d{4})",
        cleaned,
    )
    if not match:
        return None, None
    month = SPANISH_MONTHS.get(normalize_text(match.group("month")).casefold())
    if not month:
        return None, None
    year = int(match.group("year"))
    return date(year, month, int(match.group("start"))), date(year, month, int(match.group("end")))


def extract_report_metadata(source):
    workbook = _load_workbook(source)
    worksheet = workbook.worksheets[0]
    week_label = _extract_metadata_value(worksheet, "Semana")
    responsible = _extract_metadata_value(worksheet, "Responsable")
    date_start, date_end = _parse_week_range(week_label)
    return {
        "week_label": week_label,
        "date_start": date_start,
        "date_end": date_end,
        "responsible": responsible,
    }


def parse_excel(source):
    workbook = _load_workbook(source)
    rows = []
    targets = _discover_parse_targets(workbook)
    if not targets:
        return [ParsedRow(1, {}, ["No se encontraron encabezados validos en el Excel."])]

    for worksheet, header_row, columns in targets:
        inferred_area = _area_from_sheet_name(worksheet.title)
        default_responsible = _extract_metadata_value(worksheet, "Responsable")
        for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            raw = {
                "area": _value_from_row(row, columns, "area") or inferred_area,
                "task_name": _value_from_row(row, columns, "task_name"),
                "responsible": _value_from_row(row, columns, "responsible") or default_responsible,
                "status": _value_from_row(row, columns, "status"),
                "priority": _value_from_row(row, columns, "priority"),
                "observations": _value_from_row(row, columns, "observations"),
                "category": _value_from_row(row, columns, "category"),
            }

            if normalize_header(raw["task_name"]) in HEADER_ALIASES["task_name"]:
                continue
            if not any(raw.values()):
                continue

            observations = normalize_text(raw["observations"])
            category = normalize_text(raw["category"])
            if category and category != observations:
                observations = f"{observations} | Categoria: {category}" if observations else f"Categoria: {category}"

            data = {
                "area": normalize_choice(raw["area"], AREA_MAP),
                "task_name": normalize_text(raw["task_name"]),
                "responsible": normalize_text(raw["responsible"]),
                "status": normalize_choice(raw["status"], STATUS_MAP),
                "priority": normalize_choice(raw["priority"], PRIORITY_MAP),
                "observations": observations,
            }

            errors = []
            if not data["area"]:
                errors.append("Area invalida.")
            if not data["task_name"]:
                errors.append("La tarea es obligatoria.")
            if not data["status"]:
                errors.append("Estado invalido.")
            if not data["priority"]:
                errors.append("Prioridad invalida.")
            rows.append(ParsedRow(row_number=row_number, data=data, errors=errors))

    return rows


def import_valid_rows(parsed_rows, report):
    tasks = [Task(report=report, **row.data) for row in parsed_rows if row.is_valid]
    return Task.objects.bulk_create(tasks)

