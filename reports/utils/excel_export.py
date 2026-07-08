from io import BytesIO

from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

AREA_COLORS = {
    "Ecommerce": "1F77B4",
    "Pauta": "00897B",
    "Marketplace": "7B1FA2",
    "Chat Web Bali": "2E7D32",
}

STATUS_COLORS = {
    "Completado": "C8E6C9",
    "En proceso": "BBDEFB",
    "Pendiente": "FFF9C4",
    "Bloqueado": "FFCDD2",
}


def build_export_filename(report):
    slug = slugify(report.week_label) or f"reporte-{report.pk}"
    return f"Reporte_{slug}.xlsx"


def build_report_workbook(report):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte"
    headers = ["Area", "Tarea / Actividad", "Responsable", "Estado", "Prioridad", "Observaciones"]
    worksheet.append([report.week_label])
    worksheet.append([f"{report.date_start} a {report.date_end}"])
    worksheet.append([])
    worksheet.append(headers)

    for cell in worksheet[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="263238")

    for task in report.tasks.all():
        worksheet.append(
            [task.area, task.task_name, task.responsible, task.status, task.priority, task.observations]
        )
        row_index = worksheet.max_row
        worksheet.cell(row=row_index, column=1).fill = PatternFill("solid", fgColor=AREA_COLORS[task.area])
        worksheet.cell(row=row_index, column=1).font = Font(color="FFFFFF", bold=True)
        worksheet.cell(row=row_index, column=4).fill = PatternFill("solid", fgColor=STATUS_COLORS[task.status])

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 3, 60)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

