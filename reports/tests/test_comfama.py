import json
from datetime import date
from decimal import Decimal
from io import BytesIO
from io import StringIO

from django.core.management import call_command
from django.test import TestCase
from django.test import override_settings
from openpyxl import Workbook

from reports.models import ComfamaAdMetric, ComfamaProductReference, ComfamaSale, DailyAdSpend, ProductCategory
from reports.services.comfama_import import import_comfama_ad_spend_workbook, import_comfama_sales_workbook
from reports.services.sales_dashboard import build_comfama_snapshot


class ComfamaTests(TestCase):
    def test_comfama_sale_assigns_price_from_tariff(self):
        category = ProductCategory.objects.create(name="Copa Menstrual", slug="copa-menstrual", description="Test")
        reference = ComfamaProductReference.objects.create(
            category=category,
            reference="UV-TEST",
            price_tariff_a=Decimal("63100"),
            price_tariff_b=Decimal("63600"),
        )

        sale = ComfamaSale.objects.create(sale_date=date(2026, 4, 22), tariff=ComfamaSale.Tariff.TARIFF_B, reference=reference)

        self.assertEqual(sale.sales_amount, Decimal("63600"))

    def test_comfama_snapshot_includes_sales(self):
        category = ProductCategory.objects.create(name="Copa Menstrual", slug="copa-menstrual", description="Test")
        reference = ComfamaProductReference.objects.create(category=category, reference="UV-TEST", price_tariff_a=Decimal("63100"), price_tariff_b=Decimal("63600"))
        ComfamaSale.objects.create(sale_date=date(2026, 4, 22), tariff=ComfamaSale.Tariff.TARIFF_A, reference=reference)

        snapshot = build_comfama_snapshot({"date_start": "2026-04-01", "date_end": "2026-04-30"})

        self.assertEqual(snapshot["kpis"]["sales_total"], 63100.0)
        self.assertEqual(snapshot["kpis"]["purchases"], 1)
        self.assertEqual(snapshot["kpis"]["average_ticket"], 63100.0)
        self.assertEqual(snapshot["daily_series"][0]["average_ticket"], 63100.0)

    def test_comfama_ad_spend_import_creates_category_metrics(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Hoja2"
        sheet.append(["Fecha", "Producto", "CPL", "Inversion", "Conversaciones"])
        sheet.append([date(2026, 4, 26), "Copa menstrual", 500, 10000, 20])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = import_comfama_ad_spend_workbook(payload, "comfama-test.xlsx", end_date=date(2026, 4, 26))

        self.assertEqual(result["created"], 1)
        self.assertEqual(result["metric_created"], 1)
        self.assertEqual(DailyAdSpend.objects.count(), 1)
        metric = ComfamaAdMetric.objects.get()
        self.assertEqual(metric.metric_date, date(2026, 4, 26))
        self.assertEqual(metric.category.slug, "copa-menstrual")
        self.assertEqual(metric.spend_amount, Decimal("10000.00"))
        self.assertEqual(metric.conversations, 20)
        self.assertEqual(metric.cpl, Decimal("500.00"))

    def test_comfama_sales_import_supports_current_header_order(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Hoja1"
        sheet.append(["TARIFA", "SKU", "FECHA"])
        sheet.append(["T-A", "UV-BCO-002-A-SAL   T-A ", date(2026, 4, 28)])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = import_comfama_sales_workbook(payload, "ventas-comfama.xlsx")

        self.assertEqual(result["created_sales"], 1)
        sale = ComfamaSale.objects.get()
        self.assertEqual(sale.sale_date, date(2026, 4, 28))
        self.assertEqual(sale.tariff, ComfamaSale.Tariff.TARIFF_A)
        self.assertEqual(sale.reference.reference, "UV-BCO-002-A-SAL T-A")

    def test_comfama_sales_import_recovers_panty_reference_without_initial_u(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Hoja1"
        sheet.append(["Fecha", "Tarifa", "SKU"])
        sheet.append([date(2026, 1, 8), "T-A", "V-BPM-7006-ABU-M"])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = import_comfama_sales_workbook(payload, "ventas-comfama.xlsx")

        self.assertEqual(result["created_sales"], 1)
        sale = ComfamaSale.objects.get()
        self.assertEqual(sale.reference.category.slug, "panties-menstruales")
        self.assertEqual(sale.sales_amount, Decimal("48000"))

    def test_comfama_sales_import_replaces_date_range_across_source_names(self):
        category = ProductCategory.objects.create(name="Disco Menstrual", slug="disco-menstrual", description="Test")
        reference = ComfamaProductReference.objects.create(
            category=category,
            reference="UV-BCO-003-DISCO",
            price_tariff_a=Decimal("63100"),
            price_tariff_b=Decimal("63600"),
        )
        ComfamaSale.objects.create(
            sale_date=date(2026, 4, 29),
            tariff=ComfamaSale.Tariff.TARIFF_A,
            reference=reference,
            source_file="ventas-comfama.xlsx",
            source_row=2,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Hoja1"
        sheet.append(["Fecha", "SKU"])
        sheet.append([date(2026, 4, 29), "UV-BCO-003-DISCO"])
        sheet.append([date(2026, 4, 29), "UV-BCO-003-DISCO"])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = import_comfama_sales_workbook(payload, "comfama-update.xlsx", end_date=date(2026, 4, 29))

        self.assertEqual(result["deleted_sales"], 1)
        self.assertEqual(result["created_sales"], 2)
        self.assertEqual(ComfamaSale.objects.count(), 2)
        self.assertEqual(ComfamaSale.objects.filter(source_file="comfama-update.xlsx").count(), 2)

    @override_settings(
        WOOCOMMERCE_CO_BASE_URL="",
        WOOCOMMERCE_MX_BASE_URL="",
        ONEDRIVE_WHATSAPP_FILE_PATH="",
        ONEDRIVE_ECUADOR_FILE_PATH="",
        ONEDRIVE_SHARED_SALES_FILE_PATH="",
        ONEDRIVE_SHARED_COMFAMA_FILE_PATH="axis/ventas-comfama.xlsx",
    )
    def test_history_sync_includes_onedrive_comfama_sales_once(self):
        output = StringIO()
        call_command(
            "sync_axis_history_range",
            "--date-from",
            "2026-05-01",
            "--date-to",
            "2026-05-03",
            "--uva-sales",
            "--dry-run",
            stdout=output,
        )
        tasks = json.loads(output.getvalue())["tasks"]
        comfama_tasks = [task for task in tasks if task["source"] == "onedrive-comfama-sales"]

        self.assertEqual(len(comfama_tasks), 1)
        self.assertEqual(comfama_tasks[0]["command"], ["fetch_onedrive_comfama_sales"])
