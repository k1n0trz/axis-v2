import json
from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from unittest.mock import patch

from django.core.management import call_command
from django.test import TestCase, override_settings
from openpyxl import Workbook

from reports.models import AwnInternationalFollowerMetric, Country
from reports.management.commands.sync_axis_daily_data import Command as SyncAxisDailyDataCommand
from reports.services.sales_dashboard import build_awn_international_snapshot, ensure_uva_catalogs


class AwnInternationalTests(TestCase):
    def setUp(self):
        ensure_uva_catalogs()

    def _build_workbook(self, file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Hoja1"
        sheet.append(["País", "Fecha", "Visitas al perfil de instagram", "Nuevos seguidores", "Inversión", "CPR (COP)", "CPS (Costo por seguidor) COP"])
        sheet.append(["Ecuador", date(2026, 4, 1), 500, 10, 40000, 80, 4000])
        sheet.append(["México", date(2026, 4, 1), 420, 12, 36000, 85.71, 3000])
        workbook.save(file_path)

    def test_import_followers_workbook(self):
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
            temp_path = Path(handle.name)
        try:
            self._build_workbook(temp_path)
            call_command("import_awn_followers", str(temp_path))

            self.assertEqual(AwnInternationalFollowerMetric.objects.count(), 2)
            ecuador = Country.objects.get(code="EC")
            row = AwnInternationalFollowerMetric.objects.get(country=ecuador, metric_date=date(2026, 4, 1))
            self.assertEqual(row.instagram_profile_visits, 500)
            self.assertEqual(row.new_followers, 10)
            self.assertEqual(row.spend_amount, Decimal("40000"))
        finally:
            temp_path.unlink(missing_ok=True)

    def test_awareness_workbook_preserves_existing_meta_spend(self):
        mexico = Country.objects.get(code="MX")
        AwnInternationalFollowerMetric.objects.create(
            country=mexico,
            metric_date=date(2026, 5, 26),
            instagram_profile_visits=0,
            new_followers=0,
            spend_amount=Decimal("35424"),
            cpr=Decimal("0"),
            cps=Decimal("0"),
            source_file="meta-ads",
        )

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
            temp_path = Path(handle.name)
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Seguidores Awn Internacional"
            sheet.append(["Pais", "Fecha", "Visitas Al Perfil De Instagram", "Seguidores Nuevos", "Cpr", "Cps"])
            sheet.append(["Mexico", date(2026, 5, 26), 147, 36, 242, 984])
            workbook.save(temp_path)

            call_command("import_awn_followers", str(temp_path), "--sheet", "Seguidores Awn Internacional")
            row = AwnInternationalFollowerMetric.objects.get(country=mexico, metric_date=date(2026, 5, 26))

            self.assertEqual(row.instagram_profile_visits, 147)
            self.assertEqual(row.new_followers, 36)
            self.assertEqual(row.spend_amount, Decimal("35424"))
            self.assertEqual(row.cps, Decimal("984"))
        finally:
            temp_path.unlink(missing_ok=True)

    @override_settings(
        ONEDRIVE_AWARENESS_FILE_PATH="axis/seguidores-awn.xlsx",
        ONEDRIVE_CLIENT_ID="client",
        ONEDRIVE_CLIENT_SECRET="secret",
        ONEDRIVE_TENANT_ID="tenant",
        ONEDRIVE_REFRESH_TOKEN="refresh",
        ONEDRIVE_USER_ID="user-id",
    )
    def test_fetch_onedrive_awn_followers_preserves_meta_spend(self):
        ecuador = Country.objects.get(code="EC")
        AwnInternationalFollowerMetric.objects.create(
            country=ecuador,
            metric_date=date(2026, 5, 26),
            spend_amount=Decimal("50700"),
            source_file="meta-ads",
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Seguidores Awn Internacional"
        sheet.append(["Pais", "Fecha", "Visitas Al Perfil De Instagram", "Seguidores Nuevos", "Cpr", "Cps"])
        sheet.append(["Ecuador", date(2026, 5, 26), 474, 13, 107, 3903])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        with patch("reports.management.commands.fetch_onedrive_awn_followers.onedrive") as onedrive_mock:
            onedrive_mock.refresh_access_token.return_value = {"access_token": "token"}
            onedrive_mock.download_file_content_by_path.return_value = buffer
            output = StringIO()
            call_command("fetch_onedrive_awn_followers", stdout=output)

        row = AwnInternationalFollowerMetric.objects.get(country=ecuador, metric_date=date(2026, 5, 26))
        self.assertEqual(row.instagram_profile_visits, 474)
        self.assertEqual(row.new_followers, 13)
        self.assertEqual(row.spend_amount, Decimal("50700"))
        self.assertEqual(row.cpr, Decimal("106.96"))
        self.assertEqual(row.cps, Decimal("3900.00"))
        self.assertIn("seguidores-awn.xlsx", row.source_file)

    @override_settings(
        WOOCOMMERCE_CO_BASE_URL="",
        WOOCOMMERCE_MX_BASE_URL="",
        ONEDRIVE_WHATSAPP_FILE_PATH="",
        ONEDRIVE_ECUADOR_FILE_PATH="",
        ONEDRIVE_SHARED_SALES_FILE_PATH="",
        ONEDRIVE_SHARED_COMFAMA_FILE_PATH="",
        ONEDRIVE_GOOGLE_ADS_FILE_PATH="axis/google-ads.xlsx",
        ONEDRIVE_AWARENESS_FILE_PATH="axis/seguidores-awn.xlsx",
        META_CO_ACCOUNT_ID="",
        META_MX_ACCOUNT_ID="",
        META_EC_ACCOUNT_ID="",
        GOOGLE_ADS_CO_CUSTOMER_ID="",
        GOOGLE_ADS_MX_CUSTOMER_ID="",
        GOOGLE_ADS_EC_CUSTOMER_ID="",
        SHOPIFY_BALI_SHOP_DOMAIN="",
        META_REPORTS_IMAP_HOST="",
        META_REPORTS_IMAP_USERNAME="",
        META_REPORTS_IMAP_PASSWORD="",
    )
    def test_daily_sync_includes_awareness_workbook_once(self):
        tasks = SyncAxisDailyDataCommand()._build_tasks_for_dates(
            [date(2026, 5, 26), date(2026, 5, 27)],
            {"meta_rules": "docs/mappings/meta-category-rules.example.json", "google_rules": "docs/mappings/google-category-rules.example.json"},
        )
        names = [task["name"] for task in tasks]

        self.assertEqual(names.count("OneDrive Awareness Awn"), 1)

    @override_settings(
        ONEDRIVE_AWARENESS_FILE_PATH="axis/seguidores-awn.xlsx",
        ONEDRIVE_GOOGLE_ADS_FILE_PATH="",
        META_CO_ACCOUNT_ID="123",
        META_MX_ACCOUNT_ID="456",
        META_EC_ACCOUNT_ID="789",
        GOOGLE_ADS_CO_CUSTOMER_ID="",
        GOOGLE_ADS_MX_CUSTOMER_ID="",
        GOOGLE_ADS_EC_CUSTOMER_ID="",
        SHOPIFY_BALI_SHOP_DOMAIN="",
        WOOCOMMERCE_CO_BASE_URL="",
        WOOCOMMERCE_MX_BASE_URL="",
        ONEDRIVE_WHATSAPP_FILE_PATH="",
        ONEDRIVE_ECUADOR_FILE_PATH="",
        ONEDRIVE_SHARED_SALES_FILE_PATH="",
        ONEDRIVE_SHARED_COMFAMA_FILE_PATH="",
    )
    def test_history_sync_includes_awareness_for_update_button(self):
        output = StringIO()
        call_command(
            "sync_axis_history_range",
            "--date-from",
            "2026-05-21",
            "--date-to",
            "2026-05-27",
            "--uva-ads",
            "--dry-run",
            stdout=output,
        )
        tasks = json.loads(output.getvalue())["tasks"]
        awareness_tasks = [task for task in tasks if task["source"] == "onedrive-awareness-awn"]

        self.assertEqual(len(awareness_tasks), 1)
        self.assertEqual(awareness_tasks[0]["command"], ["fetch_onedrive_awn_followers", "--end-date", "2026-05-27"])

    def test_awn_snapshot_aggregates_followers(self):
        ec = Country.objects.get(code="EC")
        mx = Country.objects.get(code="MX")
        AwnInternationalFollowerMetric.objects.create(country=ec, metric_date=date(2026, 4, 1), instagram_profile_visits=500, new_followers=10, spend_amount=Decimal("40000"), cpr=Decimal("80"), cps=Decimal("4000"))
        AwnInternationalFollowerMetric.objects.create(country=mx, metric_date=date(2026, 4, 1), instagram_profile_visits=420, new_followers=12, spend_amount=Decimal("36000"), cpr=Decimal("85.71"), cps=Decimal("3000"))

        snapshot = build_awn_international_snapshot({"date_start": "2026-04-01", "date_end": "2026-04-30"})

        self.assertEqual(snapshot["kpis"]["profile_visits"], 920)
        self.assertEqual(snapshot["kpis"]["new_followers"], 22)
        self.assertEqual(snapshot["kpis"]["spend_total"], 76000.0)
