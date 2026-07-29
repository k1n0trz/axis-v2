from datetime import date
from decimal import Decimal
from io import StringIO
from tempfile import NamedTemporaryFile
import json
import os

from django.contrib import admin
from django.core.management import call_command
from django.test import TestCase, override_settings
from openpyxl import Workbook

from reports.admin import MarketplaceSaleAdmin
from reports.integrations.mercadolibre import MercadoLibreClient
from reports.models import DailyChannelSale, MarketplaceSale
from reports.services.falabella_sync import sync_orders_for_day
from reports.services.mercadolibre_sync import sync_orders_for_day as sync_meli_orders_for_day
from reports.services.sales_dashboard import ensure_marketplace_catalogs


class MercadoLibreClientTests(TestCase):
    def test_day_bounds_use_colombia_business_day(self):
        client = MercadoLibreClient('client', 'secret', seller_id='123', local_timezone='America/Bogota')
        start, end = client.day_bounds(date(2026, 6, 3))
        self.assertEqual(start, '2026-06-03T05:00:00Z')
        self.assertEqual(end, '2026-06-04T04:59:59Z')


class FalabellaSyncTests(TestCase):
    def setUp(self):
        ensure_marketplace_catalogs()

    def test_orders_sync_is_scoped_to_marketplace_falabella(self):
        class FakeClient:
            def iter_orders_for_day(self, target_date):
                return iter([
                    {'Status': 'delivered', 'Price': '18070161', 'Items': {'Item': [{'Quantity': 2}, {'Quantity': 1}]}},
                    {'Status': 'cancelled', 'Price': '89910', 'Items': {'Item': {'Quantity': 1}}},
                ])
        payload = sync_orders_for_day(date(2026, 6, 25), client=FakeClient())
        sale = DailyChannelSale.objects.get(sale_date=date(2026, 6, 25), channel__slug='falabella')
        self.assertEqual(payload['sales_amount'], Decimal('18070161'))
        self.assertEqual(sale.business_unit.slug, 'marketplace')
        self.assertEqual(sale.order_count, 1)
        self.assertEqual(sale.units, 3)
        self.assertEqual(sale.source_file, 'falabella-api')



class MarketplaceImportProtectionTests(TestCase):
    def setUp(self):
        self.catalogs = ensure_marketplace_catalogs()
        self.business_unit = self.catalogs["business_unit"]
        self.country = self.catalogs["countries"]["CO"]
        self.mercadolibre = self.catalogs["channels"]["mercado-libre"]
        self.falabella = self.catalogs["channels"]["falabella"]

    def test_admin_spend_import_preserves_existing_api_sales(self):
        DailyChannelSale.objects.create(
            business_unit=self.business_unit,
            country=self.country,
            channel=self.falabella,
            sale_date=date(2026, 6, 10),
            sales_amount=Decimal("900000"),
            spend_amount=Decimal("0"),
            order_count=4,
            units=5,
            source_file="falabella-api",
        )
        model_admin = MarketplaceSaleAdmin(MarketplaceSale, admin.site)
        defaults = model_admin.build_import_defaults("karen.xlsx", "", "120000", "", "", "Solo inversion")
        model_admin.save_import_row(
            {
                "business_unit": self.business_unit,
                "country": self.country,
                "channel": self.falabella,
                "sale_date": date(2026, 6, 10),
            },
            defaults,
        )
        sale = DailyChannelSale.objects.get(channel=self.falabella, sale_date=date(2026, 6, 10))
        self.assertEqual(sale.sales_amount, Decimal("900000"))
        self.assertEqual(sale.spend_amount, Decimal("120000"))
        self.assertEqual(sale.order_count, 4)
        self.assertEqual(sale.units, 5)

    def test_weekly_spend_import_from_june_preserves_api_sales(self):
        DailyChannelSale.objects.create(
            business_unit=self.business_unit,
            country=self.country,
            channel=self.mercadolibre,
            sale_date=date(2026, 6, 1),
            sales_amount=Decimal("700000"),
            spend_amount=Decimal("0"),
            order_count=2,
            units=3,
            source_file="mercadolibre-api",
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=3, column=1, value="1 jun - 7 jun")
        sheet.cell(row=3, column=3, value=70000)
        temp = NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp.close()
        try:
            workbook.save(temp.name)
            call_command("import_marketplace_weekly", temp.name, "--year", "2026")
        finally:
            os.unlink(temp.name)
        sale = DailyChannelSale.objects.get(channel=self.mercadolibre, sale_date=date(2026, 6, 1))
        self.assertEqual(sale.sales_amount, Decimal("700000"))
        self.assertEqual(sale.spend_amount, Decimal("10000.00"))
        self.assertEqual(sale.order_count, 2)
        self.assertEqual(sale.units, 3)

    def test_falabella_sales_sync_preserves_existing_spend(self):
        DailyChannelSale.objects.create(
            business_unit=self.business_unit,
            country=self.country,
            channel=self.falabella,
            sale_date=date(2026, 6, 25),
            sales_amount=Decimal("0"),
            spend_amount=Decimal("793061"),
            order_count=0,
            units=0,
            source_file="karen.xlsx",
        )

        class FakeClient:
            def iter_orders_for_day(self, target_date):
                return iter([{"Status": "delivered", "Price": "18070161", "Items": {"Item": [{"Quantity": 2}]}}])

        sync_orders_for_day(date(2026, 6, 25), client=FakeClient())
        sale = DailyChannelSale.objects.get(channel=self.falabella, sale_date=date(2026, 6, 25))
        self.assertEqual(sale.sales_amount, Decimal("18070161"))
        self.assertEqual(sale.spend_amount, Decimal("793061"))



    def test_falabella_zero_api_response_preserves_existing_sales(self):
        DailyChannelSale.objects.create(
            business_unit=self.business_unit,
            country=self.country,
            channel=self.falabella,
            sale_date=date(2026, 6, 26),
            sales_amount=Decimal("500000"),
            spend_amount=Decimal("13351"),
            order_count=2,
            units=2,
            source_file="falabella-api",
        )

        class EmptyClient:
            def iter_orders_for_day(self, target_date):
                return iter([])

        payload = sync_orders_for_day(date(2026, 6, 26), client=EmptyClient())
        sale = DailyChannelSale.objects.get(channel=self.falabella, sale_date=date(2026, 6, 26))
        self.assertTrue(payload["preserved_existing_sales"])
        self.assertEqual(sale.sales_amount, Decimal("500000"))
        self.assertEqual(sale.spend_amount, Decimal("13351"))
        self.assertEqual(sale.order_count, 2)
        self.assertEqual(sale.units, 2)

    def test_mercadolibre_zero_api_response_preserves_existing_sales(self):
        DailyChannelSale.objects.create(
            business_unit=self.business_unit,
            country=self.country,
            channel=self.mercadolibre,
            sale_date=date(2026, 6, 26),
            sales_amount=Decimal("600000"),
            spend_amount=Decimal("0"),
            order_count=3,
            units=4,
            source_file="mercadolibre-api",
        )

        class EmptyClient:
            def iter_orders_for_day(self, target_date):
                return iter([])

        payload = sync_meli_orders_for_day(date(2026, 6, 26), client=EmptyClient())
        sale = DailyChannelSale.objects.get(channel=self.mercadolibre, sale_date=date(2026, 6, 26))
        self.assertTrue(payload["preserved_existing_sales"])
        self.assertEqual(sale.sales_amount, Decimal("600000"))
        self.assertEqual(sale.order_count, 3)
        self.assertEqual(sale.units, 4)
class MarketplaceHistoryCommandTests(TestCase):
    @override_settings(
        WOOCOMMERCE_CO_BASE_URL='https://uva.example',
        SHOPIFY_BALI_SHOP_DOMAIN='bali.myshopify.com',
        MERCADOLIBRE_CLIENT_ID='meli-client',
        MERCADOLIBRE_CLIENT_SECRET='meli-secret',
        FALABELLA_USER_ID='daniela.garcia@balisexstore.com',
        FALABELLA_API_KEY='falabella-key',
    )
    def test_marketplace_only_history_does_not_include_uva_or_bali(self):
        output = StringIO()
        call_command('sync_axis_history_range', '--date-from', '2026-06-03', '--date-to', '2026-06-04', '--marketplace', '--dry-run', stdout=output)
        payload = json.loads(output.getvalue())
        sources = {task['source'] for task in payload['tasks']}
        self.assertEqual(sources, {'mercadolibre-marketplace', 'falabella-marketplace'})
        self.assertEqual(payload['task_count'], 4)
