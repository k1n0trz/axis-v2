from datetime import date
from io import BytesIO, StringIO
import json
from unittest.mock import patch
from decimal import Decimal

from django.core.management import call_command
from django.test import TestCase, override_settings
from openpyxl import Workbook

from reports.models import BaliDailyMetric, DailyAdSpend, DailyProductCategoryMetric, ProductCategory
from reports.management.commands.sync_axis_daily_data import Command as SyncAxisDailyDataCommand
from reports.services.google_ads_import import import_google_ads_workbook, parse_decimal
from reports.services.sales_dashboard import build_bali_snapshot, ensure_bali_catalogs, ensure_uva_catalogs


class GoogleAdsWorkbookImportTests(TestCase):
    def setUp(self):
        ensure_uva_catalogs()
        ensure_bali_catalogs()

    def test_import_google_ads_workbook_handles_uva_and_bali_sheets(self):
        workbook = Workbook()
        uva = workbook.active
        uva.title = "uva"
        uva.append(["Marca", "Pais", "Categoria", "Fecha", "Cpa Google Ads", "Inversion Google Ads"])
        uva.append(["Uva", "Colombia", "Copa Menstrual", date(2026, 5, 11), "24607.00", "332633.00"])
        uva.append(["Uva", "Colombia", "Panties Menstruales", date(2026, 5, 11), "35834.00", "95556.00"])
        uva.append(["Uva", "Colombia", "Bolas Kegel", date(2026, 5, 11), "14140.00", "28280.00"])
        uva.append(["Uva", "Mexico", "Copa Menstrual", date(2026, 5, 11), "85716.00", "85716.00"])

        bali = workbook.create_sheet("bali")
        bali.append(["Marca", "Fecha", "Visitas Registradas", "Ventas Web", "Pedidos Web", "Inversion Google Ads", "Compras Google Ads", "Conversaciones Whatsapp", "Cpa"])
        bali.append(["Bali", date(2026, 5, 11), 1639, "1444190.00", 10, "367447.00", 5, 16, "81655.00"])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = import_google_ads_workbook(payload, "google-ads.xlsx")

        self.assertEqual(result["uva_category_rows"], 4)
        self.assertEqual(result["uva_daily_spend_rows"], 2)
        self.assertEqual(result["bali_rows"], 1)

        self.assertEqual(DailyProductCategoryMetric.objects.filter(country__code="CO", metric_date=date(2026, 5, 11)).count(), 3)
        self.assertEqual(
            DailyProductCategoryMetric.objects.get(country__code="CO", metric_date=date(2026, 5, 11), category__slug="bolas-kegel-uva").spend_google,
            Decimal("28280.00"),
        )
        self.assertEqual(DailyAdSpend.objects.get(country__code="CO", spend_date=date(2026, 5, 11), ad_platform__slug="google-ads").spend_amount, Decimal("456469.00"))
        self.assertEqual(DailyAdSpend.objects.get(country__code="MX", spend_date=date(2026, 5, 11), ad_platform__slug="google-ads").spend_amount, Decimal("85716.00"))
        self.assertEqual(BaliDailyMetric.objects.get(metric_date=date(2026, 5, 11)).google_spend_amount, Decimal("367447.00"))

    def test_import_google_ads_workbook_accepts_case_and_accented_headers(self):
        workbook = Workbook()
        uva = workbook.active
        uva.title = "Uva"
        uva.append(["Marca", "Pa\u00eds", "Categor\u00eda", "Fecha", "CPA Google Ads", "Inversi\u00f3n Google Ads"])
        uva.append(["Uva", "Colombia", "Cubrepezones sin adhesivos", date(2026, 6, 4), "$ 10.000,00", "$ 35.000,00"])

        bali = workbook.create_sheet("Bali")
        bali.append(["Marca", "Fecha", "Inversi\u00f3n Google Ads", "Compras Google Ads", "Conversaciones WhatsApp", "CPA"])
        bali.append(["Bali", date(2026, 6, 4), "$ 371.598,00", 1, 20, "$ 371.598,00"])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = import_google_ads_workbook(payload, "google-ads.xlsx")

        self.assertEqual(result["uva_category_rows"], 1)
        self.assertEqual(result["uva_daily_spend_rows"], 1)
        self.assertEqual(result["bali_rows"], 1)
        self.assertEqual(
            DailyProductCategoryMetric.objects.get(
                country__code="CO",
                metric_date=date(2026, 6, 4),
                category__slug="cubrepezones-sin-adhesivo",
            ).spend_google,
            Decimal("35000.00"),
        )
        self.assertEqual(
            DailyAdSpend.objects.get(country__code="CO", spend_date=date(2026, 6, 4), ad_platform__slug="google-ads").spend_amount,
            Decimal("35000.00"),
        )

    def test_bali_google_only_sheet_preserves_shopify_fields(self):
        catalogs = ensure_bali_catalogs()
        BaliDailyMetric.objects.create(
            business_unit=catalogs["business_unit"],
            country=catalogs["country"],
            metric_date=date(2026, 5, 12),
            sessions=1500,
            web_sales_amount=Decimal("600000.00"),
            web_order_count=4,
            source_file="shopifyql",
        )

        workbook = Workbook()
        uva = workbook.active
        uva.title = "uva"
        uva.append(["Marca", "Pais", "Categoria", "Fecha", "Cpa Google Ads", "Inversion Google Ads"])

        bali = workbook.create_sheet("bali")
        bali.append(["Marca", "Fecha", "Inversion Google Ads", "Compras Google Ads", "Conversaciones Whatsapp", "Cpa"])
        bali.append(["Bali", date(2026, 5, 12), "371598.00", 1, 20, "371598.00"])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = import_google_ads_workbook(payload, "google-ads.xlsx")
        metric = BaliDailyMetric.objects.get(metric_date=date(2026, 5, 12))

        self.assertEqual(result["bali_rows"], 1)
        self.assertEqual(metric.sessions, 1500)
        self.assertEqual(metric.web_sales_amount, Decimal("600000.00"))
        self.assertEqual(metric.web_order_count, 4)
        self.assertEqual(metric.google_spend_amount, Decimal("371598.00"))
        self.assertEqual(metric.google_attributed_orders, 1)
        self.assertEqual(metric.whatsapp_conversations, 20)

    def test_bali_legacy_sheet_does_not_overwrite_shopifyql_fields(self):
        catalogs = ensure_bali_catalogs()
        BaliDailyMetric.objects.create(
            business_unit=catalogs["business_unit"],
            country=catalogs["country"],
            metric_date=date(2026, 5, 12),
            sessions=1500,
            web_sales_amount=Decimal("600000.00"),
            web_order_count=4,
            source_file="shopifyql",
        )

        workbook = Workbook()
        uva = workbook.active
        uva.title = "uva"
        uva.append(["Marca", "Pais", "Categoria", "Fecha", "Cpa Google Ads", "Inversion Google Ads"])

        bali = workbook.create_sheet("bali")
        bali.append(["Marca", "Fecha", "Visitas Registradas", "Ventas Web", "Pedidos Web", "Inversion Google Ads", "Compras Google Ads", "Conversaciones Whatsapp", "Cpa"])
        bali.append(["Bali", date(2026, 5, 12), 0, "999999.00", 99, "371598.00", 1, 20, "371598.00"])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        import_google_ads_workbook(payload, "google-ads.xlsx")
        metric = BaliDailyMetric.objects.get(metric_date=date(2026, 5, 12))

        self.assertEqual(metric.sessions, 1500)
        self.assertEqual(metric.web_sales_amount, Decimal("600000.00"))
        self.assertEqual(metric.web_order_count, 4)
        self.assertEqual(metric.google_spend_amount, Decimal("371598.00"))
        self.assertIn("shopifyql", metric.source_file)
        self.assertIn("google-ads.xlsx", metric.source_file)

    def test_bali_import_keeps_source_file_within_database_limit(self):
        catalogs = ensure_bali_catalogs()
        BaliDailyMetric.objects.create(
            business_unit=catalogs["business_unit"],
            country=catalogs["country"],
            metric_date=date(2026, 6, 4),
            sessions=1200,
            web_sales_amount=Decimal("4760000.00"),
            web_order_count=25,
            source_file="shopifyql; " + "; ".join(f"old-source-{index:02d}" for index in range(30)),
        )

        workbook = Workbook()
        workbook.active.title = "uva"
        workbook.active.append(["Marca", "Pais", "Categoria", "Fecha", "Cpa Google Ads", "Inversion Google Ads"])
        bali = workbook.create_sheet("bali")
        bali.append(["Marca", "Fecha", "Inversion Google Ads", "Compras Google Ads", "Conversaciones Whatsapp", "Cpa"])
        bali.append(["Bali", "04/06/2026", 392.003, 7, 20, 56.273])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        import_google_ads_workbook(payload, "google-ads.xlsx")
        metric = BaliDailyMetric.objects.get(metric_date=date(2026, 6, 4))

        self.assertLessEqual(len(metric.source_file), 255)
        self.assertIn("shopifyql", metric.source_file)
        self.assertIn("google-ads.xlsx", metric.source_file)

    def test_bali_onedrive_sheet_imports_day_first_june_rows_for_chart(self):
        catalogs = ensure_bali_catalogs()
        BaliDailyMetric.objects.create(
            business_unit=catalogs["business_unit"],
            country=catalogs["country"],
            metric_date=date(2026, 6, 4),
            sessions=1200,
            web_sales_amount=Decimal("4760000.00"),
            web_order_count=25,
            source_file="shopifyql",
        )

        workbook = Workbook()
        workbook.active.title = "metadata"
        bali = workbook.create_sheet("bali")
        bali.append(["Marca", "Fecha", "Inversion Google Ads", "Compras Google Ads", "Conversaciones Whatsapp", "Cpa"])
        bali.append(["Bali", "04/06/2026", 392.003, 7, 20, 56.273])
        bali.append(["Bali", "03/06/2026", 378.431, 6, 18, 65.589])
        bali.append(["Bali", "02/06/2026", 386.057, 10, 22, 39.953])
        bali.append(["Bali", "01/06/2026", 387.830, 12, 23, 32.319])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        result = import_google_ads_workbook(payload, "google-ads.xlsx")

        self.assertEqual(result["bali_rows"], 4)
        imported = {
            metric.metric_date: metric
            for metric in BaliDailyMetric.objects.filter(metric_date__range=(date(2026, 6, 1), date(2026, 6, 4)))
        }
        self.assertEqual(imported[date(2026, 6, 4)].google_spend_amount, Decimal("392003.00"))
        self.assertEqual(imported[date(2026, 6, 4)].google_attributed_orders, 7)
        self.assertEqual(imported[date(2026, 6, 4)].whatsapp_conversations, 20)
        self.assertEqual(imported[date(2026, 6, 4)].cpa, Decimal("56273.00"))
        self.assertEqual(imported[date(2026, 6, 1)].google_spend_amount, Decimal("387830.00"))
        self.assertIn("shopifyql", imported[date(2026, 6, 4)].source_file)
        self.assertIn("google-ads.xlsx", imported[date(2026, 6, 4)].source_file)

        snapshot = build_bali_snapshot(
            {"date_start": "2026-06-01", "date_end": "2026-06-04", "business_unit": "bali", "country": "CO"},
            include_comparison=False,
        )
        series_by_label = {row["label"]: row for row in snapshot["daily_series"]}
        self.assertEqual(series_by_label["2026-06-04"]["spend"], 392003.0)

    def test_google_workbook_preserves_existing_meta_fields_for_same_category_day(self):
        catalogs = ensure_uva_catalogs()
        category, _ = ProductCategory.objects.get_or_create(
            slug="bolas-kegel-uva",
            defaults={"name": "Bolas Kegel Uva"},
        )
        DailyProductCategoryMetric.objects.create(
            business_unit=catalogs["business_unit"],
            country=catalogs["countries"]["CO"],
            category=category,
            metric_date=date(2026, 5, 24),
            cpa_meta=Decimal("27597.00"),
            spend_meta=Decimal("27597.00"),
            source_file="meta-ads-api",
        )

        workbook = Workbook()
        uva = workbook.active
        uva.title = "uva"
        uva.append(["Marca", "Pais", "Categoria", "Fecha", "Cpa Google Ads", "Inversion Google Ads"])
        uva.append(["Uva", "Colombia", "Bolas Kegel", date(2026, 5, 24), "41966.00", "74421.00"])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        import_google_ads_workbook(payload, "google-ads.xlsx")
        metric = DailyProductCategoryMetric.objects.get(category__slug="bolas-kegel-uva", metric_date=date(2026, 5, 24))

        self.assertEqual(metric.spend_meta, Decimal("27597.00"))
        self.assertEqual(metric.cpa_meta, Decimal("27597.00"))
        self.assertEqual(metric.spend_google, Decimal("74421.00"))
        self.assertEqual(metric.cpa_google, Decimal("41966.00"))
        self.assertEqual(metric.total_spend, Decimal("102018.00"))

    def test_parse_decimal_accepts_currency_and_latin_formats(self):
        self.assertEqual(parse_decimal("$ 371.598,00"), Decimal("371598.00"))
        self.assertEqual(parse_decimal("COP 371,598.00"), Decimal("371598.00"))
        self.assertEqual(parse_decimal("371598.00"), Decimal("371598.00"))
        self.assertEqual(parse_decimal(392.003), Decimal("392003"))
        self.assertEqual(parse_decimal(56.273), Decimal("56273"))
        self.assertEqual(parse_decimal(392.03), Decimal("392.03"))

    @override_settings(
        GOOGLE_ADS_DEVELOPER_TOKEN="dev-token",
        GOOGLE_ADS_CLIENT_ID="client-id",
        GOOGLE_ADS_CLIENT_SECRET="client-secret",
        GOOGLE_ADS_REFRESH_TOKEN="refresh-token",
        GOOGLE_ADS_CO_CUSTOMER_ID="701-524-5415",
    )
    @patch("reports.management.commands.fetch_google_ads.GoogleAdsClient")
    def test_fetch_google_ads_api_groups_uva_campaigns_by_rules(self, client_class):
        client = client_class.return_value
        client.search.return_value = [
            {
                "results": [
                    {
                        "campaign": {"name": "11/06/26 | Ventas | Search | Hidratante"},
                        "customer": {"currencyCode": "COP"},
                        "metrics": {"costMicros": "26545000000", "conversions": "2"},
                    },
                    {
                        "campaign": {"name": "21/06/2024 | Ventas | PMax | CO"},
                        "customer": {"currencyCode": "COP"},
                        "metrics": {"costMicros": "205666000000", "conversions": "7.25"},
                    },
                    {
                        "campaign": {"name": "Brand awareness sin categoria"},
                        "customer": {"currencyCode": "COP"},
                        "metrics": {"costMicros": "999000000", "conversions": "1"},
                    },
                ]
            }
        ]
        output = StringIO()

        call_command(
            "fetch_google_ads",
            "--date",
            "2026-06-24",
            "--country",
            "CO",
            "--rules",
            "docs/mappings/google-category-rules.example.json",
            stdout=output,
        )

        payload = json.loads(output.getvalue())
        metrics = {row["category_slug"]: row for row in payload["category_metrics"]}
        self.assertEqual(payload["daily_spend"]["spend_amount"], "232211")
        self.assertEqual(metrics["hidratante-intimo-uva"]["cpa_google"], "13272.5")
        self.assertEqual(metrics["copa-menstrual"]["spend_google"], "205666")

    @override_settings(
        GOOGLE_ADS_DEVELOPER_TOKEN="dev-token",
        GOOGLE_ADS_CLIENT_ID="client-id",
        GOOGLE_ADS_CLIENT_SECRET="client-secret",
        GOOGLE_ADS_REFRESH_TOKEN="refresh-token",
        GOOGLE_ADS_MX_CUSTOMER_ID="614-371-5017",
    )
    @patch("reports.management.commands.fetch_google_ads.GoogleAdsClient")
    def test_fetch_google_ads_api_defaults_mx_campaigns_to_copa(self, client_class):
        client = client_class.return_value
        client.search.return_value = [
            {
                "results": [
                    {
                        "campaign": {"name": "Ventas | Search | MX"},
                        "customer": {"currencyCode": "COP"},
                        "metrics": {"costMicros": "156124000000", "conversions": "1"},
                    },
                ]
            }
        ]
        output = StringIO()

        call_command(
            "fetch_google_ads",
            "--date",
            "2026-06-24",
            "--country",
            "MX",
            "--rules",
            "docs/mappings/google-category-rules.example.json",
            stdout=output,
        )

        payload = json.loads(output.getvalue())
        self.assertEqual(payload["daily_spend"]["spend_amount"], "156124")
        self.assertEqual(payload["category_metrics"][0]["category_slug"], "copa-menstrual")

    @override_settings(
        GOOGLE_ADS_DEVELOPER_TOKEN="dev-token",
        GOOGLE_ADS_CLIENT_ID="client-id",
        GOOGLE_ADS_CLIENT_SECRET="client-secret",
        GOOGLE_ADS_REFRESH_TOKEN="refresh-token",
        GOOGLE_ADS_BALI_CUSTOMER_ID="404-209-3126",
        GOOGLE_ADS_BALI_WHATSAPP_CONVERSION_NAME="Balisexstore - GA4 (web) boton_de_whatsapp",
    )
    @patch("reports.management.commands.fetch_google_ads.GoogleAdsClient")
    def test_fetch_google_ads_api_updates_bali_whatsapp_conversion(self, client_class):
        catalogs = ensure_bali_catalogs()
        BaliDailyMetric.objects.create(
            business_unit=catalogs["business_unit"],
            country=catalogs["country"],
            metric_date=date(2026, 6, 24),
            sessions=1200,
            web_sales_amount=Decimal("2000000"),
            web_order_count=12,
            source_file="google-ads.xlsx; shopifyql",
        )
        client = client_class.return_value
        client.search.side_effect = [
            [
                {
                    "results": [
                        {
                            "campaign": {"name": "Bali Store PMax"},
                            "customer": {"currencyCode": "COP"},
                            "metrics": {"costMicros": "326000000000", "conversions": "0"},
                        }
                    ]
                }
            ],
            [
                {
                    "results": [
                        {
                            "segments": {"conversionActionName": "Balisexstore - GA4 (web) boton_de_whatsapp"},
                            "metrics": {"allConversions": "652"},
                        },
                        {
                            "segments": {"conversionActionName": "Otra conversion"},
                            "metrics": {"allConversions": "100"},
                        },
                    ]
                }
            ],
        ]
        output = StringIO()

        call_command("fetch_google_ads", "--date", "2026-06-24", "--country", "CO", "--business-unit", "bali", stdout=output)

        payload = json.loads(output.getvalue())
        self.assertEqual(payload["daily_spend"]["spend_amount"], "326000")
        self.assertEqual(payload["bali_metric"]["sessions"], 1200)
        self.assertEqual(payload["bali_metric"]["web_sales_amount"], "2000000.00")
        self.assertEqual(payload["bali_metric"]["whatsapp_conversations"], 652)
        self.assertEqual(payload["bali_metric"]["cpa"], "500")
        self.assertEqual(payload["bali_metric"]["source_file"], "shopifyql; google-ads-api")

    @override_settings(
        ONEDRIVE_GOOGLE_ADS_FILE_PATH="axis/google-ads.xlsx",
        GOOGLE_ADS_CO_CUSTOMER_ID="701-524-5415",
        GOOGLE_ADS_MX_CUSTOMER_ID="614-371-5017",
        GOOGLE_ADS_EC_CUSTOMER_ID="638-560-0284",
        GOOGLE_ADS_BALI_CUSTOMER_ID="404-209-3126",
        SHOPIFY_BALI_SHOP_DOMAIN="bali.example.myshopify.com",
        WOOCOMMERCE_CO_BASE_URL="",
        WOOCOMMERCE_MX_BASE_URL="",
        ONEDRIVE_WHATSAPP_FILE_PATH="",
        ONEDRIVE_SHARED_SALES_FILE_PATH="",
        ONEDRIVE_ECUADOR_FILE_PATH="",
        ONEDRIVE_SHARED_COMFAMA_FILE_PATH="",
        META_CO_ACCOUNT_ID="",
        META_MX_ACCOUNT_ID="",
        META_EC_ACCOUNT_ID="",
        MERCADOLIBRE_CLIENT_ID="",
        MERCADOLIBRE_CLIENT_SECRET="",
        ONEDRIVE_AWARENESS_FILE_PATH="",
        META_REPORTS_IMAP_HOST="",
    )
    def test_daily_sync_prefers_onedrive_workbook_over_google_ads_api(self):
        command = SyncAxisDailyDataCommand()
        tasks = command._build_tasks(date(2026, 6, 24), {"meta_rules": "meta.json", "google_rules": "google.json"})
        names = [task["name"] for task in tasks]

        self.assertIn("OneDrive Google Ads Workbook", names)
        self.assertNotIn("Google Ads Colombia", names)
        self.assertNotIn("Google Ads Mexico", names)
        self.assertNotIn("Google Ads Ecuador", names)
        self.assertNotIn("Google Ads Bali", names)

