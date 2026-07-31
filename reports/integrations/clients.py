import json
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from html.parser import HTMLParser
from io import BytesIO
from urllib.parse import parse_qs, parse_qsl, urlencode, urlparse, urlunparse

import msal
import requests
from openpyxl import load_workbook


def decimalize(value, default="0"):
    if value in (None, ""):
        return Decimal(default)
    return Decimal(str(value))


def day_bounds(target_date):
    start = datetime.combine(target_date, time.min, tzinfo=timezone.utc)
    end = start + timedelta(days=1)
    return start.isoformat(), end.isoformat()


def load_json_mapping(path):
    if not path:
        return {}
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def match_rule(value, rules, default=None):
    normalized = str(value or "").strip().lower()
    for rule in rules or []:
        needle = str(rule.get("match", "")).strip().lower()
        if needle and needle in normalized:
            return rule.get("category") or rule.get("value")
    return default


class BaseAPIClient:
    SENSITIVE_QUERY_KEYS = {
        "access_token",
        "client_secret",
        "consumer_key",
        "consumer_secret",
        "key",
        "refresh_token",
        "token",
    }

    def __init__(self, base_url="", headers=None, timeout=45):
        self.base_url = str(base_url or "").strip().rstrip("/")
        self.timeout = timeout
        self.session = requests.Session()
        if headers:
            self.session.headers.update(headers)

    def _safe_url(self, url):
        parsed = urlparse(url)
        query = urlencode(
            [
                (key, "***" if key.lower() in self.SENSITIVE_QUERY_KEYS else value)
                for key, value in parse_qsl(parsed.query, keep_blank_values=True)
            ],
            doseq=True,
        )
        return urlunparse(parsed._replace(query=query))

    def _raise_for_status(self, response):
        if response.ok:
            return
        status = getattr(response, "status_code", "")
        reason = getattr(response, "reason", "")
        safe_url = self._safe_url(getattr(response, "url", ""))
        raise RuntimeError(f"{status} {reason} for url: {safe_url}".strip())

    def get_json(self, path, params=None, headers=None):
        url = path if path.startswith("http") else f"{self.base_url}{path}"
        response = self.session.get(url, params=params, headers=headers, timeout=self.timeout)
        self._raise_for_status(response)
        return response.json()

    def get_response(self, path, params=None, headers=None):
        url = path if path.startswith("http") else f"{self.base_url}{path}"
        return self.session.get(url, params=params, headers=headers, timeout=self.timeout)

    def post_json(self, path, payload=None, headers=None):
        url = path if path.startswith("http") else f"{self.base_url}{path}"
        response = self.session.post(url, json=payload or {}, headers=headers, timeout=self.timeout)
        self._raise_for_status(response)
        if response.content:
            return response.json()
        return {}


class _IframeSrcParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.src = ""

    def handle_starttag(self, tag, attrs):
        if self.src or tag.lower() != "iframe":
            return
        attr_map = dict(attrs)
        self.src = attr_map.get("src", "")


class WooCommerceClient(BaseAPIClient):
    def __init__(self, base_url, consumer_key, consumer_secret, timeout=45):
        super().__init__(base_url=base_url, timeout=timeout)
        self.consumer_key = str(consumer_key or "").strip()
        self.consumer_secret = str(consumer_secret or "").strip()

    def iter_orders_for_day(self, target_date, statuses=None, after=None, before=None):
        if not after or not before:
            after, before = day_bounds(target_date)
        page = 1
        while True:
            payload = self.get_json(
                "/wp-json/wc/v3/orders",
                params={
                    "after": after,
                    "before": before,
                    "per_page": 100,
                    "page": page,
                    "status": ",".join(statuses or ["processing", "completed"]),
                    "consumer_key": self.consumer_key,
                    "consumer_secret": self.consumer_secret,
                },
            )
            if not payload:
                break
            for item in payload:
                yield item
            page += 1

    def get_sales_report_for_day(self, target_date):
        day = target_date.isoformat()
        payload = self.get_json(
            "/wp-json/wc/v3/reports/sales",
            params={
                "date_min": day,
                "date_max": day,
                "consumer_key": self.consumer_key,
                "consumer_secret": self.consumer_secret,
            },
        )
        if isinstance(payload, list) and payload:
            return payload[0]
        return {}


class MicrosoftGraphClient:
    graph_root = "https://graph.microsoft.com/v1.0"

    def __init__(self, tenant_id, client_id, client_secret):
        authority = f"https://login.microsoftonline.com/{tenant_id}"
        self.app = msal.ConfidentialClientApplication(
            client_id,
            authority=authority,
            client_credential=client_secret,
        )
        self._token = None

    def access_token(self):
        if self._token:
            return self._token
        result = self.app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
        token = result.get("access_token")
        if not token:
            raise RuntimeError(result.get("error_description") or "No fue posible obtener token de Microsoft Graph.")
        self._token = token
        return token

    def _headers(self):
        return {"Authorization": f"Bearer {self.access_token()}"}

    def download_file_by_path(self, user_id, drive_path):
        safe_path = drive_path.strip("/")
        url = f"{self.graph_root}/users/{user_id}/drive/root:/{safe_path}:/content"
        response = requests.get(url, headers=self._headers(), timeout=60)
        response.raise_for_status()
        return BytesIO(response.content)

    def get_user_drive(self, user_id):
        response = requests.get(
            f"{self.graph_root}/users/{user_id}/drive",
            headers=self._headers(),
            timeout=30,
        )
        response.raise_for_status()
        return response.json()

    def list_children(self, user_id, item_path=""):
        clean_path = item_path.strip("/")
        if clean_path:
            url = f"{self.graph_root}/users/{user_id}/drive/root:/{clean_path}:/children"
        else:
            url = f"{self.graph_root}/users/{user_id}/drive/root/children"
        response = requests.get(url, headers=self._headers(), timeout=30)
        response.raise_for_status()
        return response.json().get("value", [])

    def search_drive(self, user_id, query):
        response = requests.get(
            f"{self.graph_root}/users/{user_id}/drive/root/search(q='{query}')",
            headers=self._headers(),
            timeout=30,
        )
        response.raise_for_status()
        return response.json().get("value", [])

    def workbook_rows(self, user_id, drive_path, sheet_name=None):
        buffer = self.download_file_by_path(user_id, drive_path)
        workbook = load_workbook(filename=buffer, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name] if sheet_name else workbook.active
            headers = [str(cell or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                yield dict(zip(headers, row))
        finally:
            workbook.close()


class MetaAdsClient(BaseAPIClient):
    def __init__(self, access_token, api_version="v20.0", timeout=45):
        super().__init__(base_url=f"https://graph.facebook.com/{api_version}", timeout=timeout)
        self.access_token = access_token

    def _raise_meta_error(self, response, action):
        detail = ""
        try:
            payload = response.json()
            error = payload.get("error") or {}
            detail = error.get("message") or error.get("error_user_msg") or ""
        except (ValueError, AttributeError):
            detail = ""
        status = getattr(response, "status_code", None)
        status_label = f"HTTP {status}" if status else "error HTTP"
        message = f"Meta Ads devolvio {status_label} al {action}."
        if detail:
            message = f"{message} Detalle: {detail}"
        raise RuntimeError(message)

    def _get_meta_json(self, path, params=None, action="consultar Meta Ads"):
        url = path if path.startswith("http") else f"{self.base_url}{path}"
        response = self.session.get(url, params=params, timeout=self.timeout)
        if not response.ok:
            self._raise_meta_error(response, action)
        return response.json()

    def get_campaign_insights(self, account_id, target_date, level="adset"):
        payload = []
        url = f"/act_{account_id}/insights"
        base_params = {
            "access_token": self.access_token,
            "level": level,
            "time_range": json.dumps({"since": target_date.isoformat(), "until": target_date.isoformat()}),
            "limit": 500,
        }
        field_sets = [
            [
                "campaign_id",
                "campaign_name",
                "adset_id",
                "adset_name",
                "spend",
                "actions",
                "cost_per_action_type",
                "account_currency",
            ],
            [
                "campaign_id",
                "campaign_name",
                "adset_id",
                "adset_name",
                "spend",
                "actions",
                "account_currency",
            ],
            [
                "campaign_id",
                "campaign_name",
                "adset_id",
                "adset_name",
                "spend",
                "account_currency",
            ],
        ]
        last_error = None
        params = None
        for fields in field_sets:
            params = dict(base_params)
            params["fields"] = ",".join(fields)
            response = self.get_response(url, params=params)
            if response.ok:
                break
            last_error = response.text
        else:
            raise RuntimeError(
                "Meta Ads devolvio un error al consultar insights. "
                f"Cuenta: {account_id}. Nivel: {level}. Respuesta: {last_error or 'sin detalle'}"
            )

        while True:
            response = self.get_json(url, params=params)
            payload.extend(response.get("data", []))
            next_url = response.get("paging", {}).get("next")
            if not next_url:
                break
            url = next_url
            params = None
        return payload

    def get_geo_insights(self, account_id, target_date, level="account", breakdown="region"):
        payload = []
        url = f"/act_{account_id}/insights"
        params = {
            "access_token": self.access_token,
            "level": level,
            "breakdowns": breakdown,
            "fields": ",".join([
                "spend",
                "impressions",
                "reach",
                "clicks",
                "inline_link_clicks",
                "actions",
                "action_values",
                "account_currency",
            ]),
            "time_range": json.dumps({"since": target_date.isoformat(), "until": target_date.isoformat()}),
            "limit": 500,
        }
        while True:
            response = self._get_meta_json(url, params=params, action=f"consultar breakdown geografico {breakdown}")
            payload.extend(response.get("data", []))
            next_url = response.get("paging", {}).get("next")
            if not next_url:
                break
            url = next_url
            params = None
        return payload

    def get_active_ads(self, account_id, limit=None, date_start=None, date_end=None, max_records=500):
        return self._get_active_ads(account_id, limit=limit, date_start=date_start, date_end=date_end, max_records=max_records)

    def _get_active_ads(self, account_id, limit=None, date_start=None, date_end=None, include_insights=True, max_records=500):
        target_count = int(limit) if limit else int(max_records or 500)
        page_size = min(max(target_count if limit else 100, 1), 100)
        base_fields = [
            "id",
            "name",
            "status",
            "effective_status",
            "created_time",
            "updated_time",
            "campaign{id,name,status,effective_status}",
            "adset{id,name,status,effective_status}",
        ]
        creative_field_sets = [
            "creative{id,name,title,body,thumbnail_url,image_url,object_url,link_url,call_to_action_type,object_story_spec,asset_feed_spec}",
            "creative{id,name,title,body,thumbnail_url,image_url,object_url,link_url,call_to_action_type}",
            "creative{id,name,thumbnail_url,image_url}",
            "creative{id,name}",
        ]

        def collect(field_list, with_insights, size):
            payload = []
            url = f"/act_{account_id}/ads"
            fields = list(field_list)
            if with_insights and date_start and date_end:
                time_range = json.dumps({"since": date_start.isoformat(), "until": date_end.isoformat()})
                fields.append(
                    "insights.time_range("
                    + time_range
                    + "){spend,impressions,reach,clicks,inline_link_clicks,ctr,cpc,cpm,actions,action_values,cost_per_action_type,purchase_roas,website_purchase_roas}"
                )
            params = {
                "access_token": self.access_token,
                "effective_status": json.dumps(["ACTIVE"]),
                "fields": ",".join(fields),
                "limit": size,
            }
            while True:
                response = self._get_meta_json(url, params=params, action="consultar anuncios activos")
                payload.extend(response.get("data", []))
                if len(payload) >= target_count:
                    return payload[:target_count]
                next_url = response.get("paging", {}).get("next")
                if not next_url:
                    break
                url = next_url
                params = None
            return payload

        # Cuando Meta responde "Please reduce the amount of data you're asking for",
        # lo que hay que reducir es el tamano de pagina, no los insights. La version
        # anterior reintentaba sin insights con el MISMO limit: la peticion pasaba y
        # el panel se quedaba con todas las metricas en cero, en silencio.
        attempts = []
        rich_fields = [*base_fields, creative_field_sets[0]]
        lean_fields = [*base_fields, creative_field_sets[-1]]
        if include_insights and date_start and date_end:
            for fields, size in ((rich_fields, page_size), (lean_fields, 25), (lean_fields, 10)):
                if (fields, True, size) not in attempts:
                    attempts.append((fields, True, size))
        # Sin insights se degrada el detalle del creativo, como antes.
        for creative_fields in creative_field_sets:
            attempts.append(([*base_fields, creative_fields], False, page_size))

        last_error = None
        for fields, with_insights, size in attempts:
            try:
                return collect(fields, with_insights, size)
            except RuntimeError as exc:
                last_error = exc
                continue
        if last_error:
            raise last_error
        return []

    def get_ad_preview_iframe_src(self, ad_id, ad_format="INSTAGRAM_STANDARD"):
        response = self._get_meta_json(
            f"/{ad_id}/previews",
            params={
                "access_token": self.access_token,
                "ad_format": ad_format,
            },
            action="consultar preview del anuncio",
        )
        body = ((response.get("data") or [{}])[0] or {}).get("body") or ""
        parser = _IframeSrcParser()
        parser.feed(body)
        return parser.src

    def get_ad_images_by_hashes(self, account_id, hashes):
        unique_hashes = [value for value in dict.fromkeys(hashes or []) if value]
        if not unique_hashes:
            return {}

        images = {}
        for index in range(0, len(unique_hashes), 50):
            chunk = unique_hashes[index : index + 50]
            response = self._get_meta_json(
                f"/act_{account_id}/adimages",
                params={
                    "access_token": self.access_token,
                    "hashes": json.dumps(chunk),
                    "fields": "hash,url,width,height",
                    "limit": len(chunk),
                },
                action="consultar imagenes de anuncios",
            )
            for row in response.get("data", []):
                image_hash = row.get("hash")
                if image_hash:
                    images[image_hash] = row
        return images


class GoogleAdsClient(BaseAPIClient):
    token_url = "https://oauth2.googleapis.com/token"

    def __init__(self, developer_token, client_id, client_secret, refresh_token, login_customer_id=None, timeout=45):
        super().__init__(base_url="https://googleads.googleapis.com", timeout=timeout)
        self.developer_token = str(developer_token or "").strip()
        self.client_id = str(client_id or "").strip()
        self.client_secret = str(client_secret or "").strip()
        self.refresh_token = str(refresh_token or "").strip()
        self.login_customer_id = str(login_customer_id or "").strip()
        self._access_token = None

    def access_token(self):
        if self._access_token:
            return self._access_token
        response = requests.post(
            self.token_url,
            data={
                "client_id": self.client_id,
                "client_secret": self.client_secret,
                "refresh_token": self.refresh_token,
                "grant_type": "refresh_token",
            },
            timeout=self.timeout,
        )
        response.raise_for_status()
        self._access_token = response.json()["access_token"]
        return self._access_token

    def search(self, customer_id, query):
        headers = {
            "Authorization": f"Bearer {self.access_token()}",
            "developer-token": self.developer_token,
        }
        if self.login_customer_id:
            headers["login-customer-id"] = self.login_customer_id
        return self.post_json(
            f"/v24/customers/{customer_id}/googleAds:searchStream",
            payload={"query": query},
            headers=headers,
        )


class ShopifyClient(BaseAPIClient):
    def __init__(self, shop_domain, access_token, api_version="2024-10", timeout=45):
        normalized_domain = str(shop_domain or "").strip()
        normalized_domain = normalized_domain.replace("https://", "").replace("http://", "").strip("/")
        self.graphql_url = f"https://{normalized_domain}/admin/api/{api_version}/graphql.json"
        super().__init__(
            base_url=f"https://{normalized_domain}/admin/api/{api_version}",
            headers={"X-Shopify-Access-Token": access_token},
            timeout=timeout,
        )

    def get_orders_for_day(self, target_date):
        after, before = day_bounds(target_date)
        page_info = None
        orders = []
        while True:
            params = {
                "status": "any",
                "created_at_min": after,
                "created_at_max": before,
                "limit": 250,
                "fields": "id,total_price,line_items,created_at,current_total_price",
            }
            if page_info:
                params = {"limit": 250, "page_info": page_info}

            url = f"{self.base_url}/orders.json"
            response = self.session.get(url, params=params, timeout=self.timeout)
            response.raise_for_status()
            orders.extend(response.json().get("orders", []))

            link_header = response.headers.get("Link", "")
            if 'rel="next"' not in link_header:
                break

            page_info = None
            for part in link_header.split(","):
                if 'rel="next"' not in part:
                    continue
                segment = part.split(";")[0].strip().strip("<>")
                parsed = urlparse(segment)
                query = parse_qs(parsed.query)
                page_info = query.get("page_info", [None])[0]
                break

            if not page_info:
                break
        return orders

    def shopifyql_query(self, query):
        response = self.session.post(
            self.graphql_url,
            json={"query": "query runShopifyQL($query: String!) { shopifyqlQuery(query: $query) { tableData { columns { name displayName dataType } rows } parseErrors } }", "variables": {"query": query}},
            timeout=self.timeout,
        )
        response.raise_for_status()
        payload = response.json()
        if payload.get("errors"):
            raise RuntimeError(f"Shopify GraphQL devolvio errores: {payload['errors']}")
        result = payload.get("data", {}).get("shopifyqlQuery", {})
        parse_errors = result.get("parseErrors") or []
        if parse_errors:
            raise RuntimeError(f"ShopifyQL devolvio errores de parseo: {parse_errors}")
        return result.get("tableData", {})

    def product_images_by_title(self):
        graphql = """
            query ProductImages($cursor: String) {
              products(first: 250, after: $cursor) {
                nodes {
                  title
                  featuredMedia {
                    preview {
                      image {
                        url
                      }
                    }
                  }
                }
                pageInfo {
                  hasNextPage
                  endCursor
                }
              }
            }
        """
        images = {}
        cursor = None
        while True:
            response = self.session.post(
                self.graphql_url,
                json={"query": graphql, "variables": {"cursor": cursor}},
                timeout=self.timeout,
            )
            response.raise_for_status()
            payload = response.json()
            if payload.get("errors"):
                raise RuntimeError(f"Shopify GraphQL devolvio errores: {payload['errors']}")
            products = payload.get("data", {}).get("products", {})
            for product in products.get("nodes") or []:
                title = str(product.get("title") or "").strip()
                image = (((product.get("featuredMedia") or {}).get("preview") or {}).get("image") or {})
                url = str(image.get("url") or "").strip()
                if title and url:
                    images[title.casefold()] = url
            page_info = products.get("pageInfo") or {}
            if not page_info.get("hasNextPage"):
                break
            cursor = page_info.get("endCursor")
            if not cursor:
                break
        return images


class ExchangeRateClient(BaseAPIClient):
    def __init__(self, base_url, api_key="", timeout=30):
        super().__init__(base_url=base_url, timeout=timeout)
        self.api_key = api_key

    def convert(self, from_currency, to_currency, amount, target_date=None):
        from_currency = str(from_currency or "").upper()
        to_currency = str(to_currency or "").upper()
        if from_currency == to_currency:
            return decimalize(amount)

        base_url = (self.base_url or "").lower()
        if "exchangerate-api.com" in base_url:
            if not self.api_key:
                raise RuntimeError("EXCHANGE_RATE_API_KEY es obligatorio para exchangerate-api.com")
            normalized_base = self.base_url.rstrip("/")
            try:
                response = self.session.get(
                    f"{normalized_base}/{self.api_key}/pair/{from_currency}/{to_currency}/{amount}",
                    timeout=self.timeout,
                )
                response.raise_for_status()
                payload = response.json()
                result = payload.get("conversion_result")
                if result is None:
                    raise RuntimeError(f"Respuesta inesperada del proveedor de divisas: {payload}")
                return decimalize(result)
            except Exception:
                return self._convert_with_open_er_api(from_currency, to_currency, amount)

        params = {
            "from": from_currency,
            "to": to_currency,
            "amount": str(amount),
        }
        if self.api_key:
            params["access_key"] = self.api_key
        if target_date:
            params["date"] = target_date.isoformat()
        payload = self.get_json("/convert", params=params)
        if "result" in payload:
            return decimalize(payload["result"])
        info = payload.get("info", {})
        rate = decimalize(info.get("rate", "0"))
        return decimalize(amount) * rate

    def _convert_with_open_er_api(self, from_currency, to_currency, amount):
        response = self.session.get(
            f"https://open.er-api.com/v6/latest/{from_currency}",
            timeout=self.timeout,
        )
        response.raise_for_status()
        payload = response.json()
        rate = payload.get("rates", {}).get(to_currency)
        if rate is None:
            raise RuntimeError(f"Fallback de divisas no devolvio tasa {from_currency}->{to_currency}.")
        return decimalize(amount) * decimalize(rate)
