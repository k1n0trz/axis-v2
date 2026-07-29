from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.contrib import admin
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.db.models import Q, Sum
from django import forms
from django.contrib import messages
from django.template.response import TemplateResponse
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import path
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook, load_workbook

from django.contrib.auth.admin import UserAdmin as DjangoUserAdmin
from django.contrib.auth.models import User
from django.core.exceptions import ObjectDoesNotExist, PermissionDenied

from .admin_forms import AxisUserChangeForm, AxisUserCreationForm, BusinessUnitAdminForm
from .models import (
    AdPlatform,
    AgendaTask,
    Attachment,
    AwnInternationalFollowerMetric,
    BaliCommunityWebcamMetric,
    BaliDailyMetric,
    BaliPhysicalStoreSale,
    BaliWhatsAppSale,
    BusinessUnit,
    Channel,
    ComfamaAdMetric,
    ComfamaProductReference,
    ComfamaSale,
    Country,
    DailyAdSpend,
    DailyChannelSale,
    DailyGeoAdMetric,
    DailyProductCategoryMetric,
    DailyProductCategorySale,
    ExportJob,
    ImportJob,
    InsightAchievement,
    JobTitle,
    MarketplaceSale,
    MarketplaceProductInventory,
    MetricRecord,
    OperationalGoalTask,
    OperationalGoalTaskAttachment,
    Product,
    ProductCategory,
    RoasTrafficLightSetting,
    SalesTarget,
    SalesTransaction,
    Task,
    UserTask,
    UserTaskAttachment,
    UserTaskLink,
    UserProfile,
    Website,
    WebsiteHealthCheck,
    WeeklyReport,
    WeeklyTask,
)
from .sanitizers import sanitize_rich_text
from .services.comfama_import import import_comfama_ad_spend_workbook, import_comfama_sales_workbook
from .utils.admin_export import export_queryset_to_excel, get_excel_response


admin.site.site_header = "Helti"
admin.site.site_title = "Helti"
admin.site.index_title = "Panel de administracion"

class ExportExcelMixin:
    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path("exportar-excel-universal/", self.admin_site.admin_view(self.export_excel_universal_view), name=f"{self.opts.app_label}_{self.opts.model_name}_export_universal"),
        ]
        return my_urls + urls

    def export_excel_universal_view(self, request):
        # get_changelist_instance no valida permisos por si mismo (a diferencia
        # de changelist_view), y todos los usuarios de Axis son is_staff.
        if not self.has_view_permission(request):
            raise PermissionDenied
        cl = self.get_changelist_instance(request)
        queryset = cl.get_queryset(request)
        wb = export_queryset_to_excel(queryset, self.model)
        filename = f"{self.model._meta.model_name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        return get_excel_response(wb, filename)

class AxisModelAdmin(ExportExcelMixin, admin.ModelAdmin):
    pass

@admin.action(description="Exportar seleccionados a Excel")
def export_as_excel_action(modeladmin, request, queryset):
    from .utils.admin_export import export_queryset_to_excel, get_excel_response
    wb = export_queryset_to_excel(queryset, modeladmin.model)
    filename = f"{modeladmin.model._meta.model_name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    return get_excel_response(wb, filename)

admin.site.add_action(export_as_excel_action)

MARKETPLACE_GROUP = "Marketplace"
BALI_WHATSAPP_GROUP = "Bali WhatsApp"

MODELS_EXCLUDE_FOR_MARKETPLACE = [
    'weeklyreport', 'task', 'agendatask', 'weeklytask',
    'comfamasale', 'comfamaadmetric', 'comfamaproductreference',
    'balidailymetric', 'balicpwhatsappale',
    'awninternationalfollowermetric', 'productcategorymetric',
    'dailyproductcategorymetric', 'dailyproductcategorysale',
    'importjob', 'exportjob', 'userprofile', 'adplatform', 'jobtitle',
]


def get_admin_site_markov(request):
    return is_marketplace_restricted_user(request.user)


def is_marketplace_restricted_user(user):
    return user.is_authenticated and not user.is_superuser and user.groups.filter(name=MARKETPLACE_GROUP).exists()


def is_bali_whatsapp_restricted_user(user):
    return user.is_authenticated and not user.is_superuser and user.groups.filter(name=BALI_WHATSAPP_GROUP).exists()


def is_operational_task_employee(user):
    if not user.is_authenticated or user.is_superuser:
        return False
    if UserProfile.objects.filter(manager=user).exists():
        return False
    return (
        is_marketplace_restricted_user(user)
        or is_bali_whatsapp_restricted_user(user)
        or user.has_perm("reports.view_operationalgoaltask")
    )


def axis_admin_index_view(self, request, context):
    if is_marketplace_restricted_user(request.user):
        app_list = []
        for app in context.get('app_list', []):
            app_models = []
            for model in app.get('models', []):
                if model.get('object_url') not in ['/', None]:
                    model_name = model.get('model_url', '').split('/')[-2] if model.get('model_url') else ''
                    if model_name and model_name not in MODELS_EXCLUDE_FOR_MARKETPLACE:
                        app_models.append(model)
            if app_models:
                app['models'] = app_models
                app_list.append(app)
        context['app_list'] = app_list
    return context


def set_admin_field_labels(model, labels):
    for field_name, label in labels.items():
        try:
            model._meta.get_field(field_name).verbose_name = label
        except Exception:
            continue


def set_admin_model_label(model, singular, plural):
    model._meta.verbose_name = singular
    model._meta.verbose_name_plural = plural


ADMIN_MODEL_LABELS = {
    AgendaTask: ("Agenda", "Agenda"),
    Channel: ("Canal", "Canales"),
    JobTitle: ("Cargo", "Cargos"),
    ProductCategory: ("Categoria de producto", "Categorias de producto"),
    AdPlatform: ("Fuente de pauta", "Fuentes de pauta"),
    DailyAdSpend: ("Inversion diaria", "Inversiones diarias"),
    BusinessUnit: ("Marca", "Marcas"),
    MetricRecord: ("Registro de metrica", "Registros de metricas"),
    DailyProductCategoryMetric: ("Metrica diaria por categoria", "Metricas diarias por categoria"),
    Country: ("Pais", "Paises"),
    ComfamaAdMetric: ("Pauta Uva Comfama", "Pauta Uva Comfama"),
    Product: ("Producto", "Productos"),
    ComfamaProductReference: ("Referencia Comfama", "Referencias Comfama"),
    AwnInternationalFollowerMetric: ("Seguidores Awn Internacional", "Seguidores Awn Internacional"),
    BaliCommunityWebcamMetric: ("Comunidad Webcam Bali", "Comunidad Webcam Bali"),
    BaliDailyMetric: ("Metrica diaria Bali", "Metricas diarias Bali"),
    BaliWhatsAppSale: ("WhatsApp Bali", "WhatsApp Bali"),
    BaliPhysicalStoreSale: ("Tienda fisica Bali", "Tienda fisica Bali"),
    ComfamaSale: ("Venta Uva Comfama", "Ventas Uva Comfama"),
    DailyChannelSale: ("Venta diaria", "Ventas diarias"),
    DailyProductCategorySale: ("Venta diaria por categoria y canal", "Ventas diarias por categoria y canal"),
    MarketplaceSale: ("Ventas Marketplace", "Ventas Marketplace"),
    OperationalGoalTask: ("Meta operativa", "Metas operativas"),
    OperationalGoalTaskAttachment: ("Adjunto de meta operativa", "Adjuntos de metas operativas"),
    UserTask: ("Tarea", "Tareas"),
    UserTaskAttachment: ("Adjunto de tarea", "Adjuntos de tareas"),
    UserTaskLink: ("Enlace de tarea", "Enlaces de tareas"),
    RoasTrafficLightSetting: ("Semaforo ROAS", "Semaforo ROAS"),
    SalesTarget: ("Meta de venta", "Metas de venta"),
    Website: ("Web", "Webs"),
    WebsiteHealthCheck: ("Chequeo de web", "Chequeos de webs"),
}


for admin_model, (singular_label, plural_label) in ADMIN_MODEL_LABELS.items():
    set_admin_model_label(admin_model, singular_label, plural_label)


COMMON_LABELS = {
    "name": "Nombre",
    "slug": "Identificador",
    "description": "Descripcion",
    "display_order": "Orden",
    "is_active": "Activo",
    "created_at": "Creado",
    "updated_at": "Actualizado",
    "business_unit": "Marca",
    "country": "Pais",
    "channel": "Canal",
    "category": "Categoria",
    "product": "Producto",
    "source_type": "Tipo de fuente",
    "source_file": "Archivo fuente",
    "source_row": "Fila fuente",
    "notes": "Notas",
}


ADMIN_FIELD_LABELS = {
    WeeklyReport: {"week_label": "Semana", "date_start": "Fecha inicio", "date_end": "Fecha fin", "notes": "Notas"},
    Task: {"report": "Reporte", "area": "Area", "task_name": "Tarea", "responsible": "Responsable", "status": "Estado", "priority": "Prioridad", "observations": "Observaciones"},
    BusinessUnit: {"channels": "Canales", "countries": "Paises", "products": "Productos"},
    JobTitle: {"is_leadership_role": "Cargo de liderazgo"},
    UserProfile: {"user": "Usuario", "phone_number": "Celular", "job_title": "Cargo", "photo": "Foto", "role": "Cargo", "manager": "Jefe", "business_units": "Marcas"},
    Country: {"code": "Codigo", "business_units": "Marcas"},
    Channel: {"logo": "Logo", "parent": "Canal padre"},
    ProductCategory: {"image": "Imagen"},
    MetricRecord: {
        "record_id": "ID de registro",
        "subchannel": "Subcanal",
        "campaign_type": "Tipo de campana",
        "source": "Fuente",
        "period_type": "Periodo",
        "period_label": "Etiqueta del periodo",
        "date_start": "Fecha inicio",
        "date_end": "Fecha fin",
        "metric_name": "Metrica",
        "metric_value": "Valor",
        "currency": "Moneda",
        "value_origin": "Origen del valor",
    },
    DailyProductCategoryMetric: {
        "metric_date": "Fecha",
        "cpa_meta": "CPA Meta Ads",
        "cpa_google": "CPA Google Ads",
        "spend_meta": "Inversion Meta Ads",
        "spend_google": "Inversion Google Ads",
        "total_spend": "Inversion total",
        "sales_amount": "Ventas",
    },
    DailyProductCategorySale: {
        "sale_date": "Fecha",
        "sales_amount": "Ventas",
        "original_amount": "Valor original",
        "original_currency": "Moneda original",
        "exchange_rate": "Tasa de cambio",
        "quantity": "Cantidad",
    },
    ComfamaProductReference: {"reference": "Referencia", "price_tariff_a": "Precio tarifa A", "price_tariff_b": "Precio tarifa B", "is_inferred": "Inferida"},
    ComfamaSale: {"sale_date": "Fecha", "tariff": "Tarifa", "reference": "Referencia", "sales_amount": "Venta", "source_row": "Fila fuente"},
    ComfamaAdMetric: {"metric_date": "Fecha", "cpl": "CPL", "spend_amount": "Inversion", "conversations": "Conversaciones"},
    AwnInternationalFollowerMetric: {
        "metric_date": "Fecha",
        "instagram_profile_visits": "Visitas al perfil de Instagram",
        "new_followers": "Seguidores nuevos",
        "spend_amount": "Inversion",
        "cpr": "CPR",
        "cps": "CPS",
    },
    BaliDailyMetric: {
        "metric_date": "Fecha",
        "sessions": "Sesiones registradas",
        "web_sales_amount": "Ventas web",
        "web_order_count": "Pedidos web",
        "google_spend_amount": "Inversion Google Ads",
        "google_attributed_orders": "Compras Google Ads",
        "whatsapp_conversations": "Conversaciones WhatsApp",
        "cpa": "CPA",
    },
    BaliCommunityWebcamMetric: {
        "metric_date": "Fecha",
        "new_subscribers": "Suscritos nuevos",
        "subscribers": "Suscritos acumulados",
        "story_screenshot": "Pantallazo story 9:16",
    },
    Product: {"category": "Categoria"},
    DailyChannelSale: {"sale_date": "Fecha", "sales_amount": "Ventas", "order_count": "Pedidos", "spend_amount": "Inversion", "units": "Unidades"},
    MarketplaceSale: {"sale_date": "Fecha", "sales_amount": "Ventas", "order_count": "Pedidos", "spend_amount": "Inversion", "units": "Unidades"},
    BaliPhysicalStoreSale: {"sale_date": "Fecha", "sales_amount": "Ventas"},
    OperationalGoalTask: {"sales_target": "Meta", "assigned_by": "Asignada por", "assigned_to": "Asignada a", "title": "Accion operativa", "description": "Instrucciones", "goal_completion_percent": "% cumplimiento de meta", "due_date": "Fecha limite", "status": "Estado", "employee_response": "Comentarios o enlaces del empleado"},
    OperationalGoalTaskAttachment: {"task": "Meta operativa", "uploaded_by": "Subido por", "file": "Archivo", "label": "Etiqueta"},
    UserTask: {"created_by": "Creada por", "assigned_to": "Asignada a", "title": "Titulo", "description": "Descripcion", "links": "Enlaces", "due_date": "Fecha de cumplimiento", "due_time": "Hora de cumplimiento", "registered_hours": "Horas registradas", "status": "Estado"},
    UserTaskAttachment: {"task": "Tarea", "uploaded_by": "Subido por", "file": "Archivo", "label": "Etiqueta"},
    UserTaskLink: {"task": "Tarea", "url": "Enlace", "label": "Etiqueta"},
    RoasTrafficLightSetting: {"name": "Nombre", "green_min": "Verde desde", "yellow_min": "Amarillo desde", "is_active": "Activo"},
    SalesTarget: {"user": "Usuario", "date_start": "Fecha inicio", "date_end": "Fecha fin", "target_amount": "Meta de ventas", "is_active": "Activa"},
    AdPlatform: {},
    DailyAdSpend: {"ad_platform": "Fuente de pauta", "spend_date": "Fecha", "spend_amount": "Inversion"},
    AgendaTask: {
        "title": "Titulo",
        "description": "Descripcion",
        "created_by": "Creada por",
        "assigned_to": "Asignada a",
        "due_at": "Fecha limite",
        "reminder_at": "Fecha de recordatorio",
        "reminder_enabled": "Recordatorio activo",
        "reminder_sent_at": "Recordatorio enviado",
        "status": "Estado",
    },
    WeeklyTask: {
        "task_id": "ID de tarea",
        "week_label": "Semana",
        "date_start": "Fecha inicio",
        "date_end": "Fecha fin",
        "area": "Area",
        "task_name": "Tarea",
        "status": "Estado",
        "priority": "Prioridad",
        "task_type": "Tipo de tarea",
        "impact": "Impacto",
        "result": "Resultado",
        "related_metric": "Metrica relacionada",
        "attachment_ref": "Referencia de archivo",
    },
    Attachment: {
        "attachment_ref": "Referencia de archivo",
        "period_label": "Periodo",
        "task": "Tarea",
        "file_name": "Nombre del archivo",
        "file_type": "Tipo de archivo",
        "uploaded_file": "Archivo cargado",
        "file_path_or_url": "Ruta o URL",
        "tags": "Etiquetas",
        "comment": "Comentario",
    },
    SalesTransaction: {
        "product_name": "Producto",
        "origin": "Origen",
        "sale_date": "Fecha",
        "quantity": "Cantidad",
        "sale_value": "Valor de venta",
        "shipping_value": "Envio",
        "source_sheet": "Hoja fuente",
    },
    ImportJob: {
        "file_name": "Archivo",
        "status": "Estado",
        "summary": "Resumen",
        "critical_errors": "Errores criticos",
        "warnings": "Alertas",
        "preview_payload": "Vista previa",
        "completed_at": "Completado",
    },
    ExportJob: {"file_name": "Archivo", "export_scope": "Alcance", "status": "Estado", "filters": "Filtros", "completed_at": "Completado"},
}


for admin_model, labels in ADMIN_FIELD_LABELS.items():
    set_admin_field_labels(admin_model, {**COMMON_LABELS, **labels})


class ExcelImportForm(forms.Form):
    excel_file = forms.FileField(label="Archivo Excel")
    end_date = forms.DateField(label="Fecha maxima", required=False, widget=forms.DateInput(attrs={"type": "date"}))
    sheet_name = forms.CharField(label="Hoja", required=False)
    replace_source = forms.BooleanField(label="Reemplazar ventas ya importadas de este mismo archivo", required=False, initial=True)


class MarketplaceExcelImportForm(forms.Form):
    excel_file = forms.FileField(label="Archivo Excel")


def _parse_excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or value == "":
        return None
    text = str(value).strip()
    for date_format in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha no valida: {text}")


def _parse_decimal(value, default=Decimal("0")):
    if value is None or value == "":
        return default
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace("$", "").replace("COP", "").replace("cop", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Numero no valido: {value}") from exc


def _parse_int(value, default=0):
    if value is None or value == "":
        return default
    return int(_parse_decimal(value, Decimal(default)))



def _has_import_value(value):
    return value is not None and str(value).strip() != ""

class DropdownListFilter(admin.SimpleListFilter):
    template = "admin/dropdown_filter.html"


class DailySaleExcelAdminMixin:
    excel_sheet_title = "Ventas"
    excel_filename_prefix = "ventas"
    excel_description = "Carga un Excel con columnas: Fecha, Ventas, Pedidos y Notas."
    excel_headers = ("Fecha", "Ventas", "Pedidos", "Notas")
    excel_widths = (14, 16, 12, 36)

    def get_excel_urls(self):
        return [
            path("importar-excel/", self.admin_site.admin_view(self.import_excel), name=f"{self.opts.app_label}_{self.opts.model_name}_import"),
            path("exportar-excel/", self.admin_site.admin_view(self.export_excel), name=f"{self.opts.app_label}_{self.opts.model_name}_export"),
            path("plantilla-excel/", self.admin_site.admin_view(self.template_excel), name=f"{self.opts.app_label}_{self.opts.model_name}_template"),
        ]

    def _workbook_response(self, workbook, filename):
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    def _set_excel_widths(self, sheet):
        for index, width in enumerate(self.excel_widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width

    def _build_template_workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.excel_sheet_title
        sheet.append(list(self.excel_headers))
        for row in self.get_template_rows():
            sheet.append(row)
        self._set_excel_widths(sheet)
        return workbook

    def get_template_rows(self):
        return [[timezone.localdate(), 0, "", "Ejemplo: observacion opcional"]]

    def template_excel(self, request):
        return self._workbook_response(self._build_template_workbook(), f"plantilla_{self.excel_filename_prefix}.xlsx")

    def export_excel(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.excel_sheet_title
        sheet.append(list(self.excel_headers))
        for row in self.get_queryset(request).order_by("-sale_date"):
            sheet.append(self.serialize_sale_row(row))
        self._set_excel_widths(sheet)
        return self._workbook_response(workbook, f"{self.excel_filename_prefix}.xlsx")

    def serialize_sale_row(self, row):
        return [row.sale_date, row.sales_amount, row.order_count, row.notes]

    def parse_import_row(self, row):
        raw_date, sales, orders, notes = (list(row) + [None] * 4)[:4]
        return raw_date, None, sales, None, orders, None, notes

    def save_import_row(self, lookup, defaults):
        return DailyChannelSale.objects.update_or_create(**lookup, defaults=defaults)

    def import_excel(self, request):
        form = MarketplaceExcelImportForm(request.POST or None, request.FILES or None)
        if request.method == "POST" and form.is_valid():
            uploaded_file = form.cleaned_data["excel_file"]
            try:
                workbook = load_workbook(uploaded_file, data_only=True)
            except Exception as exc:
                messages.error(request, f"No fue posible leer el Excel: {exc}")
                return redirect("../")
            sheet = workbook.active
            created = 0
            updated = 0
            skipped = 0
            errors = []
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                parsed_row = self.parse_import_row(row)
                if len(parsed_row) == 8:
                    raw_date, country_name, channel_name, sales, spend, orders, units, notes = parsed_row
                else:
                    raw_date, channel_name, sales, spend, orders, units, notes = parsed_row
                    country_name = None
                if not raw_date:
                    skipped += 1
                    continue
                try:
                    sale_date = _parse_excel_date(raw_date)
                    defaults = self.build_import_defaults(uploaded_file.name, sales, spend, orders, units, notes)
                    if country_name is None:
                        lookup = self.build_import_lookup(channel_name, sale_date)
                    else:
                        lookup = self.build_import_lookup(country_name, channel_name, sale_date)
                except ValueError as exc:
                    errors.append(f"Fila {row_number}: {exc}")
                    skipped += 1
                    continue
                if not lookup:
                    skipped += 1
                    continue
                sale, was_created = self.save_import_row(lookup, defaults)
                if was_created:
                    created += 1
                else:
                    updated += 1
            messages.success(request, f"Importacion completada. Creadas: {created}, actualizadas: {updated}, omitidas: {skipped}.")
            for error in errors[:5]:
                messages.warning(request, error)
            if len(errors) > 5:
                messages.warning(request, f"Hay {len(errors) - 5} errores adicionales omitidos en el resumen.")
            return redirect("../")
        context = {
            **self.admin_site.each_context(request),
            "title": f"Importar {self.model._meta.verbose_name_plural}",
            "form": form,
            "opts": self.model._meta,
            "description": self.excel_description,
        }
        return TemplateResponse(request, "admin/reports/import_excel.html", context)


class ProductCategorySaleCountryFilter(DropdownListFilter):
    title = "Pais"
    parameter_name = "country"

    def lookups(self, request, model_admin):
        countries = Country.objects.filter(daily_product_category_sales__isnull=False).distinct().order_by("display_order", "name")
        return [(country.id, country.name) for country in countries]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(country_id=self.value())
        return queryset


class DailyChannelSaleCountryFilter(DropdownListFilter):
    title = "Pais"
    parameter_name = "country"

    def lookups(self, request, model_admin):
        queryset = Country.objects.filter(daily_channel_sales__isnull=False)
        business_unit_id = request.GET.get("business_unit")
        if business_unit_id:
            queryset = queryset.filter(daily_channel_sales__business_unit_id=business_unit_id)
        countries = queryset.distinct().order_by("display_order", "name")
        return [(country.id, country.name) for country in countries]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(country_id=self.value())
        return queryset


class ProductCategorySaleChannelFilter(DropdownListFilter):
    title = "Canal"
    parameter_name = "channel"

    def lookups(self, request, model_admin):
        queryset = Channel.objects.filter(daily_product_category_sales__isnull=False)
        country_id = request.GET.get("country")
        if country_id:
            queryset = queryset.filter(daily_product_category_sales__country_id=country_id)
        channels = queryset.distinct().order_by("business_unit__display_order", "display_order", "name")
        return [(channel.id, channel.name) for channel in channels]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(channel_id=self.value())
        return queryset


class OperationalGoalTaskAttachmentInline(admin.TabularInline):
    model = OperationalGoalTaskAttachment
    extra = 1
    fields = ("file", "label", "uploaded_by", "created_at")
    readonly_fields = ("uploaded_by", "created_at")

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser or is_operational_task_employee(request.user)

    def save_new(self, form, commit=True):
        obj = super().save_new(form, commit=False)
        request = getattr(self, "_request", None)
        if request and request.user.is_authenticated:
            obj.uploaded_by = request.user
        if commit:
            obj.save()
            form.save_m2m()
        return obj

    def save_existing(self, form, obj, commit=True):
        obj = super().save_existing(form, obj, commit=False)
        if not obj.uploaded_by_id:
            request = getattr(self, "_request", None)
            if request and request.user.is_authenticated:
                obj.uploaded_by = request.user
        if commit:
            obj.save()
            form.save_m2m()
        return obj


class UserTaskAttachmentInline(admin.TabularInline):
    model = UserTaskAttachment
    extra = 1
    fields = ("file", "label", "uploaded_by", "created_at")
    readonly_fields = ("uploaded_by", "created_at")

    def save_new(self, form, commit=True):
        obj = super().save_new(form, commit=False)
        request = getattr(self, "_request", None)
        if request and request.user.is_authenticated:
            obj.uploaded_by = request.user
        if commit:
            obj.save()
            form.save_m2m()
        return obj

    def save_existing(self, form, obj, commit=True):
        obj = super().save_existing(form, obj, commit=False)
        if not obj.uploaded_by_id:
            request = getattr(self, "_request", None)
            if request and request.user.is_authenticated:
                obj.uploaded_by = request.user
        if commit:
            obj.save()
            form.save_m2m()
        return obj


class UserTaskLinkInline(admin.TabularInline):
    model = UserTaskLink
    extra = 1
    fields = ("url", "label", "created_at")
    readonly_fields = ("created_at",)


class AdminGuideMixin(ExportExcelMixin):
    guide_text = ""

    @admin.display(description="Guia")
    def admin_guide(self, obj):
        return format_html(
            '<div style="max-width:760px; color:#dce8ff; background:rgba(20,29,46,.74); border:1px solid rgba(154,168,194,.14); border-radius:12px; padding:12px 14px; line-height:1.5;">{}</div>',
            self.guide_text,
        )

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if is_marketplace_restricted_user(request.user):
            qs = qs.filter(business_unit__slug='marketplace')
        return qs

    def has_view_permission(self, request, obj=None):
        if is_marketplace_restricted_user(request.user):
            return request.user.has_perm(f'reports.view_{self.opts.model_name}')
        return super().has_view_permission(request, obj)

    def has_add_permission(self, request):
        if is_marketplace_restricted_user(request.user):
            return request.user.has_perm(f'reports.add_{self.opts.model_name}')
        return super().has_add_permission(request)

    def has_change_permission(self, request, obj=None):
        if is_marketplace_restricted_user(request.user):
            return request.user.has_perm(f'reports.change_{self.opts.model_name}')
        return super().has_change_permission(request, obj)


class TaskInline(admin.TabularInline):
    model = Task
    extra = 0
    fields = ("area", "task_name", "responsible", "status", "priority", "observations")


@admin.register(WeeklyReport)
class WeeklyReportAdmin(AxisModelAdmin):
    list_display = ("week_label", "date_start", "date_end", "created_at", "task_count")
    list_filter = ("date_start", "date_end")
    search_fields = ("week_label", "notes")
    inlines = [TaskInline]

    @admin.display(description="Tareas")
    def task_count(self, obj):
        return obj.tasks.count()


@admin.register(Task)
class TaskAdmin(AxisModelAdmin):
    list_display = ("task_name", "area", "responsible", "status", "priority", "report")
    list_filter = ("area", "status", "priority")
    search_fields = ("task_name", "responsible")
    autocomplete_fields = ("report",)


@admin.register(BusinessUnit)
class BusinessUnitAdmin(AxisModelAdmin):
    form = BusinessUnitAdminForm
    list_display = ("name", "slug", "display_order", "is_active", "channel_count")
    list_filter = ("is_active",)
    list_editable = ("display_order", "is_active")
    search_fields = ("name", "description")
    fieldsets = (
        (
            None,
            {
                "fields": ("name", "slug", "description", "display_order", "is_active", "channels", "countries", "products"),
            },
        ),
    )

    @admin.display(description="Canales")
    def channel_count(self, obj):
        return obj.channels.count()


@admin.register(Country)
class CountryAdmin(AxisModelAdmin):
    list_display = ("name", "code", "display_order", "is_active", "brand_count")
    list_editable = ("display_order", "is_active")
    search_fields = ("name", "code")
    filter_horizontal = ("business_units",)

    @admin.display(description="Marcas")
    def brand_count(self, obj):
        return obj.business_units.count()


@admin.register(Channel)
class ChannelAdmin(AxisModelAdmin):
    list_display = ("name", "business_unit", "parent", "has_logo", "display_order", "is_active")
    list_filter = ("business_unit", "is_active")
    search_fields = ("name", "slug", "description")
    fieldsets = (
        (
            None,
            {
                "fields": ("name", "slug", "description", "logo", "business_unit", "parent", "display_order", "is_active"),
            },
        ),
    )

    @admin.display(description="Logo")
    def has_logo(self, obj):
        return "Si" if obj.logo else "No"

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        if is_marketplace_restricted_user(request.user):
            return queryset.filter(business_unit__slug="marketplace")
        return queryset

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if is_marketplace_restricted_user(request.user):
            if db_field.name == "business_unit":
                kwargs["queryset"] = BusinessUnit.objects.filter(slug="marketplace")
            if db_field.name == "parent":
                kwargs["queryset"] = Channel.objects.filter(business_unit__slug="marketplace")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        if is_marketplace_restricted_user(request.user):
            marketplace = BusinessUnit.objects.filter(slug="marketplace").first()
            if marketplace:
                initial["business_unit"] = marketplace.pk
        return initial


@admin.register(ProductCategory)
class ProductCategoryAdmin(AxisModelAdmin):
    list_display = ("name", "has_image", "is_active", "product_count", "updated_at")
    list_filter = ("is_active",)
    search_fields = ("name", "description", "slug")
    fieldsets = (
        (
            None,
            {
                "fields": ("name", "slug", "description", "image", "is_active"),
            },
        ),
    )

    @admin.display(description="Productos")
    def product_count(self, obj):
        return obj.products.count()

    @admin.display(description="Imagen")
    def has_image(self, obj):
        return "Si" if obj.image else "No"


@admin.register(Product)
class ProductAdmin(AxisModelAdmin):
    list_display = ("name", "category", "business_unit", "display_order", "is_active")
    list_filter = ("category", "business_unit", "is_active")
    search_fields = ("name", "slug", "category__name")


@admin.register(JobTitle)
class JobTitleAdmin(AxisModelAdmin):
    list_display = ("name", "is_leadership_role", "is_active", "updated_at")
    list_filter = ("is_leadership_role", "is_active")
    search_fields = ("name", "description")


@admin.register(AdPlatform)
class AdPlatformAdmin(AxisModelAdmin):
    list_display = ("name", "slug", "is_active", "updated_at")
    list_filter = ("is_active",)
    search_fields = ("name", "description")


@admin.register(MetricRecord)
class MetricRecordAdmin(AxisModelAdmin):
    list_display = ("record_id", "business_unit", "channel", "country", "product", "campaign_type", "metric_name", "metric_value", "currency", "period_type", "period_label")
    list_filter = ("business_unit", "channel", "country", "product", "campaign_type", "metric_name", "period_type", "currency", "value_origin")
    search_fields = ("record_id", "period_label", "source", "notes")
    autocomplete_fields = ("business_unit", "country", "channel", "product")


@admin.register(AgendaTask)
class AgendaTaskAdmin(AxisModelAdmin):
    list_display = ("title", "assigned_to", "created_by", "due_at", "reminder_enabled", "status", "created_at")
    list_filter = ("status", "reminder_enabled", "created_at", "due_at")
    search_fields = ("title", "description", "assigned_to__username", "created_by__username", "assigned_to__email")
    readonly_fields = ("created_at", "updated_at", "created_by")

    def get_fieldsets(self, request, obj=None):
        fields = ("title", "description", "assigned_to", "due_at", "status", "reminder_enabled", "reminder_at")
        if obj:
            fields += ("created_at", "updated_at", "created_by")
        return ((None, {"fields": fields}),)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "assigned_to" and not request.user.is_superuser:
            report_ids = list(UserProfile.objects.filter(manager=request.user).values_list("user_id", flat=True))
            kwargs["queryset"] = User.objects.filter(id__in=report_ids + [request.user.id]).order_by("username")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        if not request.user.is_superuser and not UserProfile.objects.filter(manager=request.user).exists():
            initial["assigned_to"] = request.user.pk
        return initial

    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.created_by = request.user
            if not form.cleaned_data.get("assigned_to"):
                obj.assigned_to = request.user
        super().save_model(request, obj, form, change)


@admin.register(DailyChannelSale)
class DailyChannelSaleAdmin(AdminGuideMixin, AxisModelAdmin):
    guide_text = "Registra ventas diarias de marketplaces: Mercado Libre, Falabella, Rappi, Farmatodo."
    list_display = ("sale_date", "business_unit", "country", "channel", "sales_amount", "order_count", "units", "spend_amount", "source_type")
    list_filter = ("business_unit", DailyChannelSaleCountryFilter, "channel", "sale_date")
    search_fields = ("country__name", "country__code", "channel__name", "notes")
    readonly_fields = ("admin_guide",)
    ordering = ("-sale_date",)

    def get_model_perms(self, request):
        if is_marketplace_restricted_user(request.user):
            return {}
        return super().get_model_perms(request)

    def get_fieldsets(self, request, obj=None):
        context_fields = ("sale_date", "channel")
        if not is_marketplace_restricted_user(request.user):
            context_fields = ("business_unit", "country", *context_fields)
        return (
            ("Guia de uso", {"fields": ("admin_guide",)}),
            ("Fecha y canal", {"fields": context_fields}),
            ("Ventas", {"fields": ("sales_amount", "order_count", "units")}),
            ("Trazabilidad", {"fields": ("source_type", "source_file", "notes")}),
        )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if is_marketplace_restricted_user(request.user):
            if db_field.name == "channel":
                kwargs["queryset"] = Channel.objects.filter(is_active=True, business_unit__slug="marketplace").order_by("display_order", "name")
            if db_field.name == "business_unit":
                kwargs["queryset"] = BusinessUnit.objects.filter(slug="marketplace")
            if db_field.name == "country":
                kwargs["queryset"] = Country.objects.filter(is_active=True, business_units__slug="marketplace").distinct().order_by("display_order", "name")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        formfield = super().formfield_for_dbfield(db_field, request, **kwargs)
        if db_field.name in {"order_count", "units"} and formfield:
            formfield.required = False
        return formfield

    def save_model(self, request, obj, form, change):
        if is_marketplace_restricted_user(request.user):
            marketplace = BusinessUnit.objects.filter(slug="marketplace").first()
            if marketplace:
                obj.business_unit = marketplace
            if not obj.country_id:
                obj.country = Country.objects.filter(code="CO", business_units=marketplace).first() or Country.objects.filter(business_units=marketplace).first()
        obj.order_count = obj.order_count or 0
        obj.units = obj.units or 0
        super().save_model(request, obj, form, change)


@admin.register(MarketplaceSale)
class MarketplaceSaleAdmin(DailySaleExcelAdminMixin, AxisModelAdmin):
    change_list_template = "admin/reports/marketplacesale/change_list.html"
    excel_sheet_title = "Ventas Marketplace"
    excel_filename_prefix = "ventas_marketplace"
    excel_description = "Carga un Excel con columnas: Fecha, Pais, Canal, Ventas, Inversion, Pedidos, Unidades y Notas."
    excel_headers = ("Fecha", "Pais", "Canal", "Ventas", "Inversion", "Pedidos", "Unidades", "Notas")
    excel_widths = (14, 16, 18, 16, 16, 12, 12, 36)
    list_display = ("sale_date", "country", "channel", "sales_amount", "spend_amount", "order_count", "units", "source_type")
    list_filter = ("country", "channel", "sale_date", "source_type")
    search_fields = ("channel__name", "notes", "source_file")
    date_hierarchy = "sale_date"
    ordering = ("-sale_date", "channel__display_order")
    readonly_fields = ("business_unit_display",)
    fieldsets = (
        ("Dato diario", {"fields": ("sale_date", "business_unit_display", "country", "channel")}),
        ("Metricas", {"fields": ("sales_amount", "spend_amount", "order_count", "units")}),
        ("Notas", {"fields": ("notes",)}),
    )

    def has_delete_permission(self, request, obj=None):
        if is_marketplace_restricted_user(request.user):
            return request.user.has_perm("reports.delete_marketplacesale")
        return super().has_delete_permission(request, obj)

    def get_urls(self):
        urls = super().get_urls()
        return self.get_excel_urls() + urls

    def _marketplace_catalogs(self):
        business_unit = BusinessUnit.objects.filter(slug="marketplace").first()
        countries = Country.objects.filter(is_active=True, business_units=business_unit) if business_unit else Country.objects.none()
        return business_unit, countries

    def get_template_rows(self):
        return [
            [timezone.localdate(), "Colombia", "Mercadolibre", 0, 0, "", "", "Ejemplo: observacion opcional"],
            [timezone.localdate(), "Ecuador", "Mercadolibre", 0, 0, "", "", ""],
            [timezone.localdate(), "Colombia", "Falabella", 0, 0, "", "", ""],
        ]

    def serialize_sale_row(self, row):
        return [row.sale_date, row.country.name, row.channel.name, row.sales_amount, row.spend_amount, row.order_count, row.units, row.notes]

    def parse_import_row(self, row):
        values = (list(row) + [None] * 8)[:8]
        if len([value for value in values if value not in (None, "")]) <= 7 and str(values[1] or "").strip().lower() not in {"colombia", "ecuador", "co", "ec"}:
            raw_date, channel_name, sales, spend, orders, units, notes, _unused = values
            return raw_date, "Colombia", channel_name, sales, spend, orders, units, notes
        return values

    def build_import_lookup(self, country_name, channel_name, sale_date):
        business_unit, countries = self._marketplace_catalogs()
        if not business_unit:
            raise ValueError("No existe el catalogo base de Marketplace.")
        country_text = str(country_name or "Colombia").strip()
        country = countries.filter(code__iexact=country_text).first() or countries.filter(name__iexact=country_text).first()
        if not country:
            raise ValueError(f"El pais '{country_text}' no esta habilitado para Marketplace.")
        channel_text = str(channel_name or "").strip()
        channel_slug = slugify(channel_text)
        channel_aliases = {channel_slug}
        if channel_slug == "mercadolibre":
            channel_aliases.add("mercado-libre")
        channel = (
            Channel.objects.filter(business_unit=business_unit, is_active=True, slug__in=channel_aliases).first()
            or Channel.objects.filter(business_unit=business_unit, is_active=True, name__iexact=channel_text).first()
        )
        if not channel:
            return {}
        return {
            "business_unit": business_unit,
            "country": country,
            "channel": channel,
            "sale_date": sale_date,
        }

    def build_import_defaults(self, filename, sales, spend, orders, units, notes):
        return {
            "sales_amount": _parse_decimal(sales) if _has_import_value(sales) else None,
            "spend_amount": _parse_decimal(spend) if _has_import_value(spend) else None,
            "order_count": _parse_int(orders) if _has_import_value(orders) else None,
            "units": _parse_int(units) if _has_import_value(units) else None,
            "notes": notes or "",
            "source_type": DailyChannelSale.SourceType.IMPORTED,
            "source_file": filename,
        }

    def save_import_row(self, lookup, defaults):
        sale, was_created = DailyChannelSale.objects.get_or_create(
            **lookup,
            defaults={
                "sales_amount": defaults["sales_amount"] or Decimal("0"),
                "spend_amount": defaults["spend_amount"] or Decimal("0"),
                "order_count": defaults["order_count"] or 0,
                "units": defaults["units"] or 0,
                "notes": defaults["notes"],
                "source_type": defaults["source_type"],
                "source_file": defaults["source_file"],
            },
        )
        update_fields = ["notes", "source_type", "source_file", "updated_at"]
        for field in ("sales_amount", "spend_amount", "order_count", "units"):
            value = defaults[field]
            if value is None:
                continue
            setattr(sale, field, value)
            update_fields.append(field)
        sale.notes = defaults["notes"]
        sale.source_type = defaults["source_type"]
        sale.source_file = defaults["source_file"]
        sale.save(update_fields=update_fields)
        return sale, was_created

    @admin.display(description="Marca")
    def business_unit_display(self, obj):
        if obj and obj.business_unit_id:
            return obj.business_unit.name
        return "Marketplaces"

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .filter(business_unit__slug="marketplace")
            .select_related("business_unit", "country", "channel")
        )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "channel":
            kwargs["queryset"] = Channel.objects.filter(is_active=True, business_unit__slug="marketplace").order_by("display_order", "name")
        if db_field.name == "business_unit":
            kwargs["queryset"] = BusinessUnit.objects.filter(slug="marketplace")
        if db_field.name == "country":
            kwargs["queryset"] = Country.objects.filter(is_active=True, business_units__slug="marketplace").distinct().order_by("display_order", "name")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        formfield = super().formfield_for_dbfield(db_field, request, **kwargs)
        if db_field.name in {"order_count", "units"} and formfield:
            formfield.required = False
        return formfield

    def save_model(self, request, obj, form, change):
        marketplace = BusinessUnit.objects.filter(slug="marketplace").first()
        if marketplace:
            obj.business_unit = marketplace
        if not obj.country_id:
            obj.country = Country.objects.filter(code="CO", business_units=marketplace).first() or Country.objects.filter(business_units=marketplace).first()
        obj.order_count = obj.order_count or 0
        obj.units = obj.units or 0
        obj.source_type = DailyChannelSale.SourceType.MANUAL
        obj.source_file = ""
        super().save_model(request, obj, form, change)


@admin.register(MarketplaceProductInventory)
class MarketplaceProductInventoryAdmin(AxisModelAdmin):
    list_display = ("item_id", "title", "sku", "status", "available_quantity", "health_status", "last_synced_at")
    list_filter = ("marketplace", "status", "health_status", "last_synced_at")
    search_fields = ("item_id", "title", "sku", "gtin", "brand", "model")
    readonly_fields = ("raw_payload", "created_at", "updated_at", "last_synced_at")
    ordering = ("health_status", "status", "title")


@admin.register(SalesTarget)
class SalesTargetAdmin(AxisModelAdmin):
    list_display = ("user", "business_unit", "channel", "target_amount", "date_start", "date_end", "is_active")
    list_filter = ("business_unit", "channel", "is_active", "date_start", "date_end")
    search_fields = ("user__username", "user__first_name", "user__last_name", "channel__name", "notes")
    autocomplete_fields = ("user", "business_unit", "channel")


@admin.register(RoasTrafficLightSetting)
class RoasTrafficLightSettingAdmin(AxisModelAdmin):
    list_display = ("name", "green_min", "yellow_min", "is_active", "updated_at")
    list_editable = ("green_min", "yellow_min", "is_active")
    list_filter = ("is_active",)
    fieldsets = (
        (
            None,
            {
                "fields": ("name", "green_min", "yellow_min", "is_active"),
                "description": "Verde aplica por encima del umbral verde. Amarillo aplica entre amarillo y verde. Rojo aplica por debajo del umbral amarillo.",
            },
        ),
    )


@admin.register(OperationalGoalTask)
class OperationalGoalTaskAdmin(AxisModelAdmin):
    list_display = ("title", "assigned_to", "assigned_by", "sales_target_display", "goal_completion_percent", "status", "due_date", "updated_at")
    list_filter = ("status", "due_date", "sales_target__business_unit", "sales_target__channel")
    search_fields = ("title", "description", "employee_response", "assigned_to__username", "assigned_to__first_name", "assigned_to__last_name")
    autocomplete_fields = ("sales_target", "assigned_to")
    readonly_fields = ("assigned_by", "employee_response_preview", "created_at", "updated_at", "completed_at")
    inlines = [OperationalGoalTaskAttachmentInline]
    fieldsets = (
        ("Meta y responsable", {"fields": ("sales_target", "assigned_to", "assigned_by")}),
        ("Tarea", {"fields": ("title", "description", "goal_completion_percent", "due_date", "status")}),
        ("Respuesta del empleado", {"fields": ("employee_response_preview",)}),
        ("Auditoria", {"fields": ("created_at", "updated_at", "completed_at")}),
    )

    def get_formsets_with_inlines(self, request, obj=None):
        for inline in self.get_inline_instances(request, obj):
            inline._request = request
            yield inline.get_formset(request, obj), inline

    def get_fields(self, request, obj=None):
        if is_operational_task_employee(request.user):
            return ("title", "sales_target_display", "description", "goal_completion_percent", "status", "employee_response", "due_date", "assigned_by", "created_at", "updated_at", "completed_at")
        return super().get_fields(request, obj)

    def get_fieldsets(self, request, obj=None):
        if is_operational_task_employee(request.user):
            return (
                ("Tarea asignada", {"fields": ("title", "sales_target_display", "description", "goal_completion_percent", "due_date", "assigned_by")}),
                ("Gestion de la tarea", {"fields": ("status", "employee_response")}),
                ("Auditoria", {"fields": ("created_at", "updated_at", "completed_at")}),
            )
        return super().get_fieldsets(request, obj)

    def get_readonly_fields(self, request, obj=None):
        if is_operational_task_employee(request.user):
            return ("title", "sales_target_display", "description", "goal_completion_percent", "due_date", "assigned_by", "created_at", "updated_at", "completed_at")
        return super().get_readonly_fields(request, obj)

    def _direct_report_ids(self, request):
        if not request.user.is_authenticated:
            return []
        return list(UserProfile.objects.filter(manager=request.user).values_list("user_id", flat=True))

    def _can_manage_operational_tasks(self, request):
        return request.user.is_authenticated and (
            request.user.is_superuser
            or bool(self._direct_report_ids(request))
            or is_operational_task_employee(request.user)
        )

    def get_model_perms(self, request):
        if self._can_manage_operational_tasks(request):
            return {"add": True, "change": True, "delete": False, "view": True}
        return {}

    def has_module_permission(self, request):
        return self._can_manage_operational_tasks(request)

    def has_view_permission(self, request, obj=None):
        return self._can_manage_operational_tasks(request)

    def has_add_permission(self, request):
        if is_operational_task_employee(request.user):
            return False
        return self._can_manage_operational_tasks(request)

    def has_change_permission(self, request, obj=None):
        if is_operational_task_employee(request.user):
            return obj is None or obj.assigned_to_id == request.user.id
        return self._can_manage_operational_tasks(request)

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser

    def get_queryset(self, request):
        queryset = super().get_queryset(request).select_related("assigned_by", "assigned_to", "sales_target", "sales_target__business_unit", "sales_target__channel")
        if is_operational_task_employee(request.user):
            return queryset.filter(assigned_to=request.user)
        if request.user.is_superuser:
            return queryset
        report_ids = self._direct_report_ids(request)
        return queryset.filter(assigned_to_id__in=report_ids)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if is_operational_task_employee(request.user):
            kwargs["queryset"] = User.objects.none() if db_field.name == "assigned_to" else kwargs.get("queryset")
            return super().formfield_for_foreignkey(db_field, request, **kwargs)
        if not request.user.is_superuser:
            report_ids = self._direct_report_ids(request)
            if db_field.name == "assigned_to":
                kwargs["queryset"] = User.objects.filter(id__in=report_ids).order_by("first_name", "username")
            if db_field.name == "sales_target":
                kwargs["queryset"] = SalesTarget.objects.filter(user_id__in=report_ids, is_active=True).select_related("business_unit", "channel", "user")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def save_model(self, request, obj, form, change):
        obj.employee_response = sanitize_rich_text(obj.employee_response)
        if is_operational_task_employee(request.user):
            original = OperationalGoalTask.objects.get(pk=obj.pk)
            original.status = obj.status
            original.employee_response = obj.employee_response
            if obj.status == OperationalGoalTask.Status.COMPLETED and not original.completed_at:
                original.completed_at = timezone.now()
            original.save(update_fields=["status", "employee_response", "completed_at", "updated_at"])
            return
        if not obj.pk:
            obj.assigned_by = request.user
        if obj.sales_target_id:
            obj.assigned_to = obj.sales_target.user
        super().save_model(request, obj, form, change)

    @admin.display(description="Meta")
    def sales_target_display(self, obj):
        if not obj.sales_target_id:
            return ""
        channel = obj.sales_target.channel.name if obj.sales_target.channel_id else "General"
        return f"{obj.sales_target.business_unit} · {channel}"

    @admin.display(description="Comentarios o enlaces del empleado")
    def employee_response_preview(self, obj):
        if not obj or not obj.employee_response:
            return "Sin respuesta registrada."
        return format_html('<div style="max-width:760px; line-height:1.5;">{}</div>', mark_safe(sanitize_rich_text(obj.employee_response)))

    class Media:
        js = ("admin/operational_task_editor.js",)
        css = {"all": ("admin/operational_task_editor.css",)}


@admin.register(InsightAchievement)
class InsightAchievementAdmin(AxisModelAdmin):
    list_display = ("title", "user", "business_unit", "channel", "month", "achievement_type", "metric_value", "delta_percent")
    list_filter = ("month", "achievement_type", "business_unit", "channel")
    search_fields = ("title", "description", "user__username", "user__first_name", "user__last_name")
    readonly_fields = ("user", "sales_target", "business_unit", "channel", "month", "achievement_type", "title", "description", "metric_value", "delta_percent", "created_at", "updated_at")

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser


class UserTaskExportDateForm(forms.Form):
    desde = forms.DateField(label="Desde", required=False, widget=forms.DateInput(attrs={"type": "date"}))
    hasta = forms.DateField(label="Hasta", required=False, widget=forms.DateInput(attrs={"type": "date"}))


@admin.register(UserTask)
class UserTaskAdmin(AxisModelAdmin):
    change_list_template = "admin/reports/usertask/change_list.html"
    list_display = ("title", "assigned_to_display", "created_by", "status", "due_date", "due_time", "registered_hours", "attachment_count", "link_count", "updated_at")
    list_filter = ("status", "due_date", "assigned_to")
    search_fields = ("title", "description", "assigned_to__username", "assigned_to__first_name", "assigned_to__last_name", "created_by__username")
    autocomplete_fields = ("assigned_to",)
    readonly_fields = ("created_by", "created_at", "updated_at", "completed_at")
    inlines = [UserTaskAttachmentInline, UserTaskLinkInline]
    fieldsets = (
        ("Responsable", {"fields": ("assigned_to", "created_by"), "description": "Si no eliges responsable, la tarea queda asignada a ti."}),
        ("Tarea", {"fields": ("title", "description", "links", "due_date", "due_time", "registered_hours", "status")}),
        ("Auditoria", {"fields": ("created_at", "updated_at", "completed_at")}),
    )

    def get_formsets_with_inlines(self, request, obj=None):
        for inline in self.get_inline_instances(request, obj):
            inline._request = request
            yield inline.get_formset(request, obj), inline

    def _direct_report_ids(self, request):
        if not request.user.is_authenticated:
            return []
        return list(UserProfile.objects.filter(manager=request.user).values_list("user_id", flat=True))

    def _visible_user_ids(self, request):
        if request.user.is_superuser:
            return None
        return [request.user.id, *self._direct_report_ids(request)]

    def get_model_perms(self, request):
        if request.user.is_authenticated and request.user.is_staff:
            return {"add": True, "change": True, "delete": True, "view": True}
        return {}

    def has_module_permission(self, request):
        return request.user.is_authenticated and request.user.is_staff

    def has_view_permission(self, request, obj=None):
        return request.user.is_authenticated and request.user.is_staff

    def has_add_permission(self, request):
        return request.user.is_authenticated and request.user.is_staff

    def has_change_permission(self, request, obj=None):
        if obj is None or request.user.is_superuser:
            return request.user.is_authenticated and request.user.is_staff
        return obj.assigned_to_id == request.user.id or obj.created_by_id == request.user.id or obj.assigned_to_id in self._direct_report_ids(request)

    def has_delete_permission(self, request, obj=None):
        if obj is None:
            return request.user.is_authenticated and request.user.is_staff
        if request.user.is_superuser:
            return True
        return obj.created_by_id == request.user.id or obj.assigned_to_id == request.user.id or obj.assigned_to_id in self._direct_report_ids(request)

    def get_queryset(self, request):
        queryset = super().get_queryset(request).select_related("assigned_to", "created_by").prefetch_related("attachments", "task_links")
        visible_ids = self._visible_user_ids(request)
        if visible_ids is None:
            return queryset
        return queryset.filter(Q(assigned_to_id__in=visible_ids) | Q(created_by=request.user)).distinct()

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "assigned_to" and not request.user.is_superuser:
            visible_ids = [request.user.id, *self._direct_report_ids(request)]
            kwargs["queryset"] = User.objects.filter(id__in=visible_ids).order_by("first_name", "username")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.created_by = request.user
        if not obj.assigned_to_id:
            obj.assigned_to = request.user
        if obj.status == UserTask.Status.COMPLETED and not obj.completed_at:
            obj.completed_at = timezone.now()
        if obj.status != UserTask.Status.COMPLETED:
            obj.completed_at = None
        super().save_model(request, obj, form, change)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("exportar-excel/", self.admin_site.admin_view(self.export_excel), name="reports_usertask_export"),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["export_date_form"] = UserTaskExportDateForm(request.GET or None)
        return super().changelist_view(request, extra_context=extra_context)

    def export_excel(self, request):
        form = UserTaskExportDateForm(request.GET)
        queryset = self.get_queryset(request).order_by("due_date", "assigned_to__first_name", "title")
        if form.is_valid():
            if form.cleaned_data.get("desde"):
                queryset = queryset.filter(due_date__gte=form.cleaned_data["desde"])
            if form.cleaned_data.get("hasta"):
                queryset = queryset.filter(due_date__lte=form.cleaned_data["hasta"])

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Tareas"
        sheet.append(["Titulo", "Descripcion", "Enlaces", "Asignada a", "Creada por", "Fecha cumplimiento", "Horas registradas", "Estado", "Adjuntos", "Creada", "Actualizada"])
        for task in queryset:
            link_values = list(task.task_links.values_list("url", flat=True))
            if task.links:
                link_values = [*link_values, *[line.strip() for line in task.links.splitlines() if line.strip()]]
            sheet.append([
                task.title,
                task.description,
                "\n".join(link_values),
                task.assigned_to.get_full_name() or task.assigned_to.username if task.assigned_to_id else "",
                task.created_by.get_full_name() or task.created_by.username,
                task.due_date,
                float(task.registered_hours or 0),
                task.get_status_display(),
                task.attachments.count(),
                timezone.localtime(task.created_at).strftime("%Y-%m-%d %H:%M"),
                timezone.localtime(task.updated_at).strftime("%Y-%m-%d %H:%M"),
            ])
        for column, width in {"A": 32, "B": 55, "C": 48, "D": 22, "E": 22, "F": 18, "G": 16, "H": 16, "I": 10, "J": 18, "K": 18}.items():
            sheet.column_dimensions[column].width = width
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="tareas.xlsx"'
        workbook.save(response)
        return response

    @admin.display(description="Adjuntos")
    def attachment_count(self, obj):
        return obj.attachments.count()

    @admin.display(description="Enlaces")
    def link_count(self, obj):
        inline_links = obj.task_links.count()
        text_links = len([line for line in (obj.links or "").splitlines() if line.strip()])
        return inline_links + text_links

    @admin.display(description="Asignada a")
    def assigned_to_display(self, obj):
        if obj.assigned_to_id:
            return obj.assigned_to.get_full_name() or obj.assigned_to.username
        return obj.created_by.get_full_name() or obj.created_by.username

    class Media:
        js = ("admin/operational_task_editor.js",)
        css = {"all": ("admin/operational_task_editor.css",)}


@admin.register(DailyAdSpend)
class DailyAdSpendAdmin(AdminGuideMixin, AxisModelAdmin):
    guide_text = "Aqui se registra la inversion diaria por plataforma. Actualiza a diario si el gasto no viene por importacion."
    change_list_template = "admin/reports/dailyadspend/change_list.html"
    list_display = ("spend_date", "business_unit", "country", "ad_platform", "spend_amount", "source_type")
    list_filter = ("business_unit", "country", "ad_platform", "source_type", "spend_date")
    search_fields = ("business_unit__name", "country__name", "ad_platform__name", "notes", "source_file")
    fieldsets = (
        ("Guia de uso", {"fields": ("admin_guide",)}),
        ("Dato diario", {"fields": ("business_unit", "country", "ad_platform", "spend_date", "spend_amount")}),
        ("Trazabilidad", {"fields": ("source_type", "source_file", "notes")}),
    )
    readonly_fields = ("admin_guide",)


@admin.register(DailyGeoAdMetric)
class DailyGeoAdMetricAdmin(AxisModelAdmin):
    list_display = ("metric_date", "business_unit", "country", "ad_platform", "geo_level", "location_name", "impressions", "purchases", "spend_amount")
    list_filter = ("business_unit", "country", "ad_platform", "geo_level", "source_type", "metric_date")
    search_fields = ("location_name", "location_key", "platform_location_id", "notes", "source_file")
    readonly_fields = ("created_at", "updated_at")
    fieldsets = (
        ("Contexto", {"fields": ("business_unit", "country", "ad_platform", "metric_date", "geo_level", "location_name", "location_key", "platform_location_id")}),
        ("Metricas", {"fields": ("impressions", "reach", "clicks", "purchases", "conversion_value", "spend_amount")}),
        ("Trazabilidad", {"fields": ("source_type", "source_file", "notes", "created_at", "updated_at")}),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("importar-comfama/", self.admin_site.admin_view(self.import_comfama_ad_spend), name="reports_dailyadspend_import_comfama"),
        ]
        return custom_urls + urls

    def import_comfama_ad_spend(self, request):
        form = ExcelImportForm(request.POST or None, request.FILES or None, initial={"sheet_name": "Hoja2"})
        form.fields["replace_source"].widget = forms.HiddenInput()
        if request.method == "POST" and form.is_valid():
            uploaded_file = form.cleaned_data["excel_file"]
            result = import_comfama_ad_spend_workbook(
                uploaded_file,
                uploaded_file.name,
                sheet_name=form.cleaned_data["sheet_name"] or "Hoja2",
                end_date=form.cleaned_data["end_date"],
            )
            messages.success(
                request,
                f"Inversion Comfama importada como marca {result['business_unit']}. Fechas: {result['dates']}. "
                f"Gastos diarios creados: {result['created']}, actualizados: {result['updated']}. "
                f"Metricas por categoria creadas: {result['metric_created']}, actualizadas: {result['metric_updated']}.",
            )
            form = ExcelImportForm(initial={"sheet_name": "Hoja2"})
            form.fields["replace_source"].widget = forms.HiddenInput()

        context = {
            **self.admin_site.each_context(request),
            "title": "Importar pauta Meta Ads Comfama",
            "form": form,
            "opts": self.model._meta,
            "description": "Carga la hoja dos del Excel de Comfama. Se importan Fecha, Producto, CPL, Inversion y Conversaciones para el modulo Comfama y la inversion diaria consolidada.",
        }
        return TemplateResponse(request, "admin/reports/import_excel.html", context)


@admin.register(Website)
class WebsiteAdmin(AxisModelAdmin):
    list_display = ("name", "country_label", "platform", "stage", "monitor_enabled", "url", "display_order", "updated_at")
    list_filter = ("platform", "stage", "monitor_enabled")
    search_fields = ("name", "country_label", "url", "slug")
    prepopulated_fields = {"slug": ("name", "country_label")}
    fields = ("name", "slug", "country_label", "url", "platform", "stage", "logo", "business_unit", "monitor_enabled", "display_order", "notes")

    def has_module_permission(self, request):
        if is_marketplace_restricted_user(request.user) or is_bali_whatsapp_restricted_user(request.user):
            return False
        return super().has_module_permission(request)


@admin.register(WebsiteHealthCheck)
class WebsiteHealthCheckAdmin(AxisModelAdmin):
    list_display = (
        "website",
        "checked_at",
        "overall_status",
        "availability_status",
        "http_status",
        "response_time_ms",
        "ssl_days_remaining",
        "performance_score",
        "accessibility_score",
        "security_headers_display",
    )
    list_filter = ("overall_status", "availability_status", "pagespeed_status", "website__platform")
    search_fields = ("website__name", "website__country_label", "final_url", "error_message")
    readonly_fields = (
        "website",
        "checked_at",
        "overall_status",
        "availability_status",
        "http_status",
        "response_time_ms",
        "final_url",
        "page_title",
        "platform_detected",
        "is_https",
        "ssl_valid",
        "ssl_expires_at",
        "ssl_days_remaining",
        "security_headers_score",
        "security_headers_total",
        "missing_security_headers",
        "pagespeed_status",
        "performance_score",
        "accessibility_score",
        "best_practices_score",
        "seo_score",
        "first_contentful_paint_ms",
        "largest_contentful_paint_ms",
        "speed_index_ms",
        "total_blocking_time_ms",
        "cumulative_layout_shift",
        "products_visible_status",
        "products_visible_count",
        "products_in_stock_count",
        "products_out_of_stock_count",
        "raw_payload",
        "error_message",
        "created_at",
        "updated_at",
    )

    @admin.display(description="Headers")
    def security_headers_display(self, obj):
        return f"{obj.security_headers_score}/{obj.security_headers_total}"

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_module_permission(self, request):
        if is_marketplace_restricted_user(request.user) or is_bali_whatsapp_restricted_user(request.user):
            return False
        return super().has_module_permission(request)


@admin.register(BaliDailyMetric)
class BaliDailyMetricAdmin(AdminGuideMixin, AxisModelAdmin):
    guide_text = "Este es el modulo diario principal de Bali. Aqui actualizas fecha, sesiones, ventas web, pedidos web, inversion Google Ads, compras atribuibles a Google Ads, conversaciones a WhatsApp y CPA."
    list_display = (
        "metric_date",
        "sessions",
        "web_sales_amount",
        "web_order_count",
        "average_ticket_display",
        "google_spend_amount",
        "google_attributed_orders",
        "whatsapp_conversations",
        "conversion_rate_display",
        "cpa",
        "source_type",
    )
    list_filter = ("source_type", "metric_date")
    search_fields = ("notes", "source_file")
    date_hierarchy = "metric_date"
    fieldsets = (
        ("Guia de uso", {"fields": ("admin_guide",)}),
        ("Dato diario Bali", {"fields": ("metric_date", "sessions", "web_sales_amount", "web_order_count", "google_spend_amount", "google_attributed_orders", "whatsapp_conversations", "cpa")}),
        ("Indicadores automaticos", {"fields": ("average_ticket_display", "conversion_rate_display")}),
        ("Trazabilidad", {"fields": ("source_type", "source_file", "source_row", "notes")}),
    )
    readonly_fields = ("admin_guide", "average_ticket_display", "conversion_rate_display")

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.filter(business_unit__slug="bali", country__code="CO").select_related("business_unit", "country")

    def save_model(self, request, obj, form, change):
        bali = BusinessUnit.objects.filter(slug="bali").first()
        colombia = Country.objects.filter(code="CO").first()
        if bali:
            obj.business_unit = bali
        if colombia:
            obj.country = colombia
        super().save_model(request, obj, form, change)

    @admin.display(description="Ticket promedio")
    def average_ticket_display(self, obj):
        return round(float(obj.web_sales_amount / obj.web_order_count), 2) if obj.web_order_count else 0

    @admin.display(description="% conversion web")
    def conversion_rate_display(self, obj):
        return f"{round((obj.web_order_count / obj.sessions) * 100, 2)}%" if obj.sessions else "0%"


@admin.register(BaliCommunityWebcamMetric)
class BaliCommunityWebcamMetricAdmin(AdminGuideMixin, AxisModelAdmin):
    guide_text = "Registra aqui el crecimiento diario del canal de WhatsApp Comunidad Webcam de Bali. No maneja pauta, ROAS ni ventas; solo suscritos y el pantallazo story 9:16 de la vista actual del canal."
    list_display = ("metric_date", "new_subscribers", "subscribers", "story_screenshot_status", "updated_at")
    list_filter = ("metric_date",)
    search_fields = ("notes",)
    date_hierarchy = "metric_date"
    fieldsets = (
        ("Guia de uso", {"fields": ("admin_guide",)}),
        ("Comunidad Webcam", {"fields": ("metric_date", "new_subscribers", "subscribers", "story_screenshot")}),
        ("Notas", {"fields": ("notes",)}),
    )
    readonly_fields = ("admin_guide",)

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.filter(business_unit__slug="bali", country__code="CO").select_related("business_unit", "country")

    def save_model(self, request, obj, form, change):
        bali = BusinessUnit.objects.filter(slug="bali").first()
        colombia = Country.objects.filter(code="CO").first()
        if bali:
            obj.business_unit = bali
        if colombia:
            obj.country = colombia
        super().save_model(request, obj, form, change)

    @admin.display(description="Pantallazo")
    def story_screenshot_status(self, obj):
        if not obj.story_screenshot:
            return "Sin imagen"
        return format_html('<a href="{}" target="_blank">Ver story</a>', obj.story_screenshot.url)


@admin.register(BaliWhatsAppSale)
class BaliWhatsAppSaleAdmin(DailySaleExcelAdminMixin, AdminGuideMixin, AxisModelAdmin):
    change_list_template = "admin/reports/baliwhatsappsale/change_list.html"
    guide_text = "Usa este modulo para registrar las ventas diarias de WhatsApp Bali. Si el dato llega consolidado, el sistema puede distribuirlo por dia en la importacion inicial."
    excel_sheet_title = "WhatsApp Bali"
    excel_filename_prefix = "ventas_whatsapp_bali"
    excel_description = "Carga un Excel con columnas: Fecha, Ventas, Pedidos y Notas. El canal se asigna automaticamente a WhatsApp Bali."
    excel_headers = ("Fecha", "Ventas", "Pedidos", "Notas")
    excel_widths = (14, 16, 12, 36)
    list_display = ("sale_date", "sales_amount", "order_count", "average_ticket_display", "source_type")
    list_filter = ("source_type", "sale_date")
    search_fields = ("notes", "source_file")
    date_hierarchy = "sale_date"
    fieldsets = (
        ("Guia de uso", {"fields": ("admin_guide",)}),
        ("Venta diaria WhatsApp Bali", {"fields": ("sale_date", "sales_amount", "order_count", "average_ticket_display")}),
        ("Trazabilidad", {"fields": ("source_type", "source_file", "notes")}),
    )
    readonly_fields = ("admin_guide", "average_ticket_display")

    def has_delete_permission(self, request, obj=None):
        if is_bali_whatsapp_restricted_user(request.user):
            return request.user.has_perm("reports.delete_baliwhatsappsale")
        return super().has_delete_permission(request, obj)

    def get_urls(self):
        urls = super().get_urls()
        return self.get_excel_urls() + urls

    def get_template_rows(self):
        return [[timezone.localdate(), 0, "", "Ejemplo: venta diaria WhatsApp Bali"]]

    def build_import_lookup(self, channel_name, sale_date):
        bali = BusinessUnit.objects.filter(slug="bali").first()
        colombia = Country.objects.filter(code="CO").first()
        whatsapp = Channel.objects.filter(business_unit__slug="bali", slug="bali-whatsapp").first()
        if not bali or not colombia or not whatsapp:
            raise ValueError("No existe el catalogo base Bali / Colombia / WhatsApp Bali.")
        return {
            "business_unit": bali,
            "country": colombia,
            "channel": whatsapp,
            "sale_date": sale_date,
        }

    def build_import_defaults(self, filename, sales, spend, orders, units, notes):
        return {
            "sales_amount": _parse_decimal(sales),
            "spend_amount": 0,
            "order_count": _parse_int(orders),
            "units": _parse_int(orders),
            "notes": notes or "",
            "source_type": DailyChannelSale.SourceType.IMPORTED,
            "source_file": filename,
        }

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.filter(business_unit__slug="bali", channel__slug="bali-whatsapp", country__code="CO").select_related("business_unit", "country", "channel")

    def save_model(self, request, obj, form, change):
        bali = BusinessUnit.objects.filter(slug="bali").first()
        colombia = Country.objects.filter(code="CO").first()
        whatsapp = Channel.objects.filter(business_unit__slug="bali", slug="bali-whatsapp").first()
        if bali:
            obj.business_unit = bali
        if colombia:
            obj.country = colombia
        if whatsapp:
            obj.channel = whatsapp
        super().save_model(request, obj, form, change)

    @admin.display(description="Ticket promedio")
    def average_ticket_display(self, obj):
        return round(float(obj.sales_amount / obj.order_count), 2) if obj.order_count else 0


@admin.register(BaliPhysicalStoreSale)
class BaliPhysicalStoreSaleAdmin(AdminGuideMixin, AxisModelAdmin):
    guide_text = "Registra la fecha, ventas, visitantes y pedidos de la tienda fisica Bali. El dashboard calcula conversion y ticket promedio automaticamente."
    list_display = ("sale_date", "sales_amount", "visitors_display", "orders_display", "conversion_display", "average_ticket_display", "source_type")
    list_filter = ("sale_date", "source_type")
    date_hierarchy = "sale_date"
    search_fields = ("notes", "source_file")
    fieldsets = (
        ("Guia de uso", {"fields": ("admin_guide",)}),
        ("Venta tienda fisica", {"fields": ("sale_date", "sales_amount", "units", "order_count")}),
        ("Trazabilidad", {"fields": ("source_type", "source_file", "notes")}),
    )
    readonly_fields = ("admin_guide",)

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.filter(business_unit__slug="bali", channel__slug="bali-tienda-fisica", country__code="CO").select_related("business_unit", "country", "channel")

    def save_model(self, request, obj, form, change):
        bali = BusinessUnit.objects.filter(slug="bali").first()
        colombia = Country.objects.filter(code="CO").first()
        channel = Channel.objects.filter(business_unit__slug="bali", slug="bali-tienda-fisica").first()
        if bali:
            obj.business_unit = bali
        if colombia:
            obj.country = colombia
        if channel:
            obj.channel = channel
        obj.spend_amount = 0
        obj.source_type = DailyChannelSale.SourceType.MANUAL
        obj.source_file = ""
        super().save_model(request, obj, form, change)

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        field = super().formfield_for_dbfield(db_field, request, **kwargs)
        if db_field.name == "units":
            field.label = "Visitantes a la tienda"
        if db_field.name == "order_count":
            field.label = "Pedidos"
        return field

    @admin.display(description="Visitantes a la tienda")
    def visitors_display(self, obj):
        return obj.units

    @admin.display(description="Pedidos")
    def orders_display(self, obj):
        return obj.order_count

    @admin.display(description="Conversion")
    def conversion_display(self, obj):
        return f"{(float(obj.order_count or 0) / float(obj.units or 1)) * 100:.2f}%" if obj.units else "0.00%"

    @admin.display(description="Ticket promedio")
    def average_ticket_display(self, obj):
        return round(float(obj.sales_amount / obj.order_count), 2) if obj.order_count else 0


@admin.register(DailyProductCategoryMetric)
class DailyProductCategoryMetricAdmin(AdminGuideMixin, AxisModelAdmin):
    guide_text = "Este modulo se actualiza a diario para las categorias de producto de Colombia, Ecuador y Mexico. Aqui van CPA y gasto por categoria. Las ventas se toman del modulo Ventas diarias por categoria y canal."
    list_display = ("metric_date", "category", "country", "sales_from_channel_display", "average_ticket_display", "total_spend", "cpa_meta", "cpa_google", "source_type")
    list_filter = ("business_unit", "country", "category", "source_type", "metric_date")
    search_fields = ("category__name", "notes", "source_file")
    fieldsets = (
        ("Guia de uso", {"fields": ("admin_guide",)}),
        ("Contexto", {"fields": ("business_unit", "country", "category", "metric_date")}),
        ("Metricas de pauta", {"fields": ("cpa_meta", "cpa_google", "spend_meta", "spend_google", "total_spend")}),
        ("Trazabilidad", {"fields": ("source_type", "source_file", "notes")}),
    )
    readonly_fields = ("admin_guide", "total_spend")

    def _category_sale_totals(self, obj):
        return DailyProductCategorySale.objects.filter(
            business_unit=obj.business_unit,
            country=obj.country,
            category=obj.category,
            sale_date=obj.metric_date,
        ).aggregate(
            sales_total=Sum("sales_amount"),
            quantity_total=Sum("quantity"),
        )

    @admin.display(description="Ventas")
    def sales_from_channel_display(self, obj):
        totals = self._category_sale_totals(obj)
        return totals["sales_total"] or 0

    @admin.display(description="Ticket promedio")
    def average_ticket_display(self, obj):
        totals = self._category_sale_totals(obj)
        sales_total = totals["sales_total"] or 0
        quantity_total = totals["quantity_total"] or 0
        return round(float(sales_total / quantity_total), 2) if quantity_total else 0


@admin.register(DailyProductCategorySale)
class DailyProductCategorySaleAdmin(AdminGuideMixin, AxisModelAdmin):
    guide_text = "Este registro guarda ventas diarias por categoria y canal. Actualiza a diario cuando cambien las ventas web o WhatsApp por categoria."
    list_display = ("sale_date", "category", "business_unit", "country", "channel", "sales_amount", "quantity", "average_ticket_display", "source_type")
    list_filter = (ProductCategorySaleCountryFilter, ProductCategorySaleChannelFilter, "category", "source_type", "sale_date")
    search_fields = ("category__name", "channel__name", "notes", "source_file")
    list_select_related = ("business_unit", "country", "channel", "category")
    date_hierarchy = "sale_date"
    fieldsets = (
        ("Guia de uso", {"fields": ("admin_guide",)}),
        ("Contexto", {"fields": ("business_unit", "country", "channel", "category", "sale_date")}),
        ("Valores", {"fields": ("sales_amount", "quantity", "average_ticket_display")}),
        ("Trazabilidad", {"fields": ("source_type", "source_file", "notes")}),
    )
    readonly_fields = ("admin_guide", "average_ticket_display")

    def save_model(self, request, obj, form, change):
        obj.original_amount = obj.sales_amount
        obj.original_currency = "COP"
        obj.exchange_rate = 1
        super().save_model(request, obj, form, change)

    @admin.display(description="Ticket promedio")
    def average_ticket_display(self, obj):
        return round(float(obj.sales_amount / obj.quantity), 2) if obj.quantity else 0


@admin.register(ComfamaProductReference)
class ComfamaProductReferenceAdmin(AxisModelAdmin):
    list_display = ("reference", "category", "price_tariff_a", "price_tariff_b", "is_inferred", "is_active")
    list_filter = ("category", "is_inferred", "is_active")
    list_editable = ("price_tariff_a", "price_tariff_b", "is_active")
    search_fields = ("reference", "category__name", "notes")
    autocomplete_fields = ("category",)


@admin.register(ComfamaSale)
class ComfamaSaleAdmin(AdminGuideMixin, AxisModelAdmin):
    guide_text = "Actualiza este modulo a diario con la fecha, la tarifa y la referencia vendida. El valor de venta se calcula automaticamente desde la referencia Comfama."
    change_list_template = "admin/reports/comfamasale/change_list.html"
    list_display = ("sale_date", "reference", "category_display", "tariff", "sales_amount", "average_ticket_display", "source_file")
    list_filter = ("sale_date", "tariff", "reference__category")
    search_fields = ("reference__reference", "reference__category__name", "source_file", "notes")
    autocomplete_fields = ("reference",)
    fieldsets = (
        ("Guia de uso", {"fields": ("admin_guide",)}),
        ("Venta diaria", {"fields": ("sale_date", "tariff", "reference", "sales_amount", "average_ticket_display")}),
        ("Trazabilidad", {"fields": ("source_file", "source_row", "notes")}),
    )
    readonly_fields = ("admin_guide", "sales_amount", "average_ticket_display")

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("importar-excel/", self.admin_site.admin_view(self.import_sales), name="reports_comfamasale_import"),
        ]
        return custom_urls + urls

    def import_sales(self, request):
        form = ExcelImportForm(request.POST or None, request.FILES or None, initial={"sheet_name": "Hoja1", "replace_source": True})
        if request.method == "POST" and form.is_valid():
            uploaded_file = form.cleaned_data["excel_file"]
            result = import_comfama_sales_workbook(
                uploaded_file,
                uploaded_file.name,
                sheet_name=form.cleaned_data["sheet_name"] or "Hoja1",
                end_date=form.cleaned_data["end_date"],
                replace_source=form.cleaned_data["replace_source"],
            )
            messages.success(
                request,
                "Ventas Comfama importadas. "
                f"Referencias creadas: {result['created_refs']}, actualizadas: {result['updated_refs']}, inferidas: {result['inferred_refs']}. "
                f"Ventas creadas: {result['created_sales']}, actualizadas: {result['updated_sales']}, eliminadas: {result['deleted_sales']}, omitidas: {result['skipped_sales']}.",
            )
            form = ExcelImportForm(initial={"sheet_name": "Hoja1", "replace_source": True})

        context = {
            **self.admin_site.each_context(request),
            "title": "Importar ventas Comfama",
            "form": form,
            "opts": self.model._meta,
            "description": "Carga la hoja uno con Fecha, Tarifa y Referencia. La tabla de precios del mismo archivo actualiza las referencias y el valor de venta se calcula automaticamente.",
        }
        return TemplateResponse(request, "admin/reports/import_excel.html", context)

    @admin.display(description="Categoria")
    def category_display(self, obj):
        return obj.reference.category if obj.reference_id else ""

    @admin.display(description="Ticket promedio")
    def average_ticket_display(self, obj):
        return round(float(obj.sales_amount), 2) if obj.sales_amount else 0


@admin.register(ComfamaAdMetric)
class ComfamaAdMetricAdmin(AdminGuideMixin, AxisModelAdmin):
    guide_text = "Aqui se actualiza la pauta diaria de Comfama por categoria: gasto, conversaciones y CPL."
    list_display = ("metric_date", "category", "spend_amount", "conversations", "cpl", "source_file")
    list_filter = ("category", "metric_date")
    search_fields = ("category__name", "source_file", "notes")
    autocomplete_fields = ("category",)
    fieldsets = (
        ("Guia de uso", {"fields": ("admin_guide",)}),
        ("Dato diario", {"fields": ("metric_date", "category", "spend_amount", "conversations", "cpl")}),
        ("Trazabilidad", {"fields": ("source_file", "notes")}),
    )
    readonly_fields = ("admin_guide",)


@admin.register(AwnInternationalFollowerMetric)
class AwnInternationalFollowerMetricAdmin(AdminGuideMixin, AxisModelAdmin):
    guide_text = "Este modulo se actualiza a diario para las campanas de seguidores de Instagram en Ecuador y Mexico. Registra visitas al perfil, seguidores nuevos, inversion, CPR y CPS."
    list_display = ("metric_date", "country", "instagram_profile_visits", "new_followers", "spend_amount", "cpr", "cps", "source_type")
    list_filter = ("country", "source_type", "metric_date")
    search_fields = ("country__name", "source_file", "notes")
    fieldsets = (
        ("Guia de uso", {"fields": ("admin_guide",)}),
        ("Dato diario", {"fields": ("country", "metric_date", "instagram_profile_visits", "new_followers", "spend_amount", "cpr", "cps")}),
        ("Trazabilidad", {"fields": ("source_type", "source_file", "source_row", "notes")}),
    )
    readonly_fields = ("admin_guide",)


@admin.register(WeeklyTask)
class WeeklyTaskAdmin(AxisModelAdmin):
    list_display = ("task_id", "week_label", "business_unit", "channel", "area", "status", "priority", "impact")
    list_filter = ("business_unit", "channel", "area", "status", "priority", "impact")
    search_fields = ("task_id", "week_label", "task_name", "notes", "result")
    autocomplete_fields = ("business_unit", "country", "channel", "related_metric")


@admin.register(Attachment)
class AttachmentAdmin(AxisModelAdmin):
    list_display = ("attachment_ref", "file_name", "file_type", "business_unit", "channel", "period_label", "tags")
    list_filter = ("file_type", "business_unit", "channel")
    search_fields = ("attachment_ref", "file_name", "comment", "description", "tags", "file_path_or_url")
    autocomplete_fields = ("business_unit", "country", "channel", "task")


@admin.register(SalesTransaction)
class SalesTransactionAdmin(AxisModelAdmin):
    list_display = ("sale_date", "product_name", "origin", "country", "channel", "quantity", "sale_value", "shipping_value")
    list_filter = ("business_unit", "country", "channel", "sale_date")
    search_fields = ("product_name", "origin", "source_file", "source_sheet")
    autocomplete_fields = ("business_unit", "country", "channel", "product")


@admin.register(ImportJob)
class ImportJobAdmin(AxisModelAdmin):
    list_display = ("file_name", "status", "critical_errors", "warnings", "created_at", "completed_at")
    list_filter = ("status",)
    search_fields = ("file_name", "summary")


@admin.register(ExportJob)
class ExportJobAdmin(AxisModelAdmin):
    list_display = ("file_name", "export_scope", "status", "created_at", "completed_at")
    list_filter = ("status", "export_scope")
    search_fields = ("file_name",)


@admin.register(UserProfile)
class UserProfileAdmin(AxisModelAdmin):
    list_display = ("user", "phone_number", "job_title", "brand_count", "updated_at")
    search_fields = ("user__username", "user__email", "phone_number", "job_title")
    filter_horizontal = ("business_units",)
    autocomplete_fields = ("user",)

    @admin.display(description="Marcas")
    def brand_count(self, obj):
        return obj.business_units.count()


admin.site.unregister(User)


@admin.register(User)
class UserAdmin(ExportExcelMixin, DjangoUserAdmin):
    add_form = AxisUserCreationForm
    form = AxisUserChangeForm
    list_display = ("username", "email", "first_name", "last_name", "role_display", "manager_display", "brand_access_display", "is_staff", "is_superuser", "is_active")
    search_fields = ("username", "first_name", "last_name", "email", "profile__phone_number", "profile__job_title", "profile__role__name")
    fieldsets = (
        (None, {"fields": ("username", "password")}),
        ("Informacion personal", {"fields": ("first_name", "last_name", "email")}),
        ("Informacion laboral", {"fields": ("phone_number", "role", "manager", "business_units")}),
        ("Permisos", {"fields": ("is_active", "is_staff", "is_superuser", "groups", "user_permissions")}),
        ("Fechas importantes", {"fields": ("last_login", "date_joined")}),
    )
    add_fieldsets = (
        (
            None,
            {
                "classes": ("wide",),
                "fields": ("username", "first_name", "last_name", "email", "phone_number", "role", "manager", "business_units", "password1", "password2"),
            },
        ),
        (
            "Permisos",
            {
                "classes": ("wide",),
                "fields": ("is_active", "is_staff", "is_superuser", "groups", "user_permissions"),
            },
        ),
    )

    class Media:
        js = ("admin/password_toggle.js",)
        css = {"all": ("admin/password_toggle.css",)}

    def get_form(self, request, obj=None, change=False, **kwargs):
        form = super().get_form(request, obj=obj, change=change, **kwargs)
        if obj is None and "is_staff" in form.base_fields:
            form.base_fields["is_staff"].initial = True
        return form

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.select_related("profile").prefetch_related("profile__business_units")

    def _save_profile_fields(self, user, form):
        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.phone_number = form.cleaned_data.get("phone_number", "")
        profile.role = form.cleaned_data.get("role")
        profile.job_title = profile.role.name if profile.role else ""
        profile.manager = form.cleaned_data.get("manager")
        profile.save()
        profile.business_units.set(form.cleaned_data.get("business_units") or [])

    def save_model(self, request, obj, form, change):
        if obj.is_superuser and not obj.is_staff:
            obj.is_staff = True
        super().save_model(request, obj, form, change)
        self._save_profile_fields(obj, form)

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        self._save_profile_fields(form.instance, form)

    @admin.display(description="Cargo")
    def role_display(self, obj):
        try:
            profile = obj.profile
        except ObjectDoesNotExist:
            profile = None
        if not profile:
            return ""
        return profile.role.name if profile.role else profile.job_title

    @admin.display(description="Jefe")
    def manager_display(self, obj):
        try:
            profile = obj.profile
        except ObjectDoesNotExist:
            profile = None
        if not profile or not profile.manager:
            return ""
        return profile.manager.get_full_name() or profile.manager.username

    @admin.display(description="Marcas")
    def brand_access_display(self, obj):
        try:
            profile = obj.profile
        except ObjectDoesNotExist:
            profile = None
        if not profile:
            return ""
        return ", ".join(profile.business_units.values_list("name", flat=True))


for hidden_model in (Attachment, ExportJob, ImportJob, SalesTransaction, Task, UserProfile, WeeklyReport, WeeklyTask):
    try:
        admin.site.unregister(hidden_model)
    except admin.sites.NotRegistered:
        pass


REPORTS_ADMIN_GROUPS = [
    (
        "Catalogos y operacion",
        {
            "AgendaTask",
            "BusinessUnit",
            "Country",
            "Channel",
            "ProductCategory",
            "Product",
            "AdPlatform",
            "MetricRecord",
            "RoasTrafficLightSetting",
            "JobTitle",
        },
    ),
    (
        "Metas",
        {
            "SalesTarget",
            "OperationalGoalTask",
        },
    ),
    (
        "Tareas",
        {
            "UserTask",
        },
    ),
    (
        "Webs",
        {
            "Website",
            "WebsiteHealthCheck",
        },
    ),
    (
        "Reportes Uva",
        {
            "DailyChannelSale",
            "DailyAdSpend",
            "DailyProductCategoryMetric",
            "DailyProductCategorySale",
            "AwnInternationalFollowerMetric",
        },
    ),
    (
        "Reportes Comfama",
        {
            "ComfamaProductReference",
            "ComfamaSale",
            "ComfamaAdMetric",
        },
    ),
    (
        "Reportes Bali",
        {
            "BaliCommunityWebcamMetric",
            "BaliDailyMetric",
            "BaliWhatsAppSale",
            "BaliPhysicalStoreSale",
        },
    ),
]

REPORTS_ADMIN_GROUP_ORDER = {
    object_name: group_index
    for group_index, (_, object_names) in enumerate(REPORTS_ADMIN_GROUPS)
    for object_name in object_names
}

_original_admin_get_app_list = admin.site.get_app_list


def grouped_reports_admin_app_list(request, app_label=None):
    app_list = _original_admin_get_app_list(request, app_label)
    if app_label and app_label != "reports":
        return app_list

    if is_marketplace_restricted_user(request.user):
        marketplace_models = [
            "Channel",
            "MarketplaceSale",
            "OperationalGoalTask",
            "UserTask",
        ]
        filtered_app_list = []
        for app in app_list:
            if app["app_label"] != "reports":
                continue
            models = [m for m in app["models"] if m.get("object_name") in marketplace_models]
            if models:
                report_models = [m for m in models if m.get("object_name") not in {"OperationalGoalTask", "UserTask"}]
                goal_models = [m for m in models if m.get("object_name") == "OperationalGoalTask"]
                task_models = [m for m in models if m.get("object_name") == "UserTask"]
                if report_models:
                    filtered_app_list.append({**app, "name": "Reportes Marketplace", "models": report_models})
                if goal_models:
                    filtered_app_list.append({**app, "name": "Metas", "models": goal_models})
                if task_models:
                    filtered_app_list.append({**app, "name": "Tareas", "models": task_models})
        return filtered_app_list

    if is_bali_whatsapp_restricted_user(request.user):
        bali_models = [
            "BaliCommunityWebcamMetric",
            "BaliWhatsAppSale",
            "BaliPhysicalStoreSale",
            "OperationalGoalTask",
            "UserTask",
        ]
        filtered_app_list = []
        for app in app_list:
            if app["app_label"] != "reports":
                continue
            models = [m for m in app["models"] if m.get("object_name") in bali_models]
            if models:
                report_models = [m for m in models if m.get("object_name") not in {"OperationalGoalTask", "UserTask"}]
                goal_models = [m for m in models if m.get("object_name") == "OperationalGoalTask"]
                task_models = [m for m in models if m.get("object_name") == "UserTask"]
                if report_models:
                    filtered_app_list.append({**app, "name": "Reportes Bali", "models": report_models})
                if goal_models:
                    filtered_app_list.append({**app, "name": "Metas", "models": goal_models})
                if task_models:
                    filtered_app_list.append({**app, "name": "Tareas", "models": task_models})
        return filtered_app_list

    grouped_app_list = []
    for app in app_list:
        if app["app_label"] != "reports":
            grouped_app_list.append(app)
            continue

        grouped_models = {name: [] for name, _ in REPORTS_ADMIN_GROUPS}
        ungrouped_models = []
        for model in app["models"]:
            group_index = REPORTS_ADMIN_GROUP_ORDER.get(model["object_name"])
            if group_index is None:
                ungrouped_models.append(model)
            else:
                grouped_models[REPORTS_ADMIN_GROUPS[group_index][0]].append(model)

        for group_name, _ in REPORTS_ADMIN_GROUPS:
            models = grouped_models[group_name]
            if models:
                grouped_app_list.append(
                    {
                        **app,
                        "name": group_name,
                        "app_label": f"reports_{group_name.lower().replace(' ', '_')}",
                        "app_url": "",
                        "models": models,
                    }
                )

        if ungrouped_models:
            grouped_app_list.append({**app, "name": "Otros reportes", "app_label": "reports_otros", "app_url": "", "models": ungrouped_models})

    return grouped_app_list


admin.site.get_app_list = grouped_reports_admin_app_list
