from datetime import date, datetime
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.models import Attachment, BusinessUnit, Channel, Country, ExportJob, ImportJob, MetricRecord, Product, WeeklyTask

REQUIRED_SHEETS = ["metrics", "tasks", "attachments", "catalogs", "instructions"]
METRICS_COLUMNS = [
    "record_id", "business_unit", "country", "channel", "subchannel", "product", "campaign_type", "period_type", "period_label", "year", "date_start", "date_end", "metric_name", "metric_value", "currency", "source", "notes",
]
METRICS_REQUIRED_COLUMNS = ["record_id", "business_unit", "period_type", "period_label", "date_start", "date_end", "metric_name", "metric_value", "currency"]
TASKS_COLUMNS = [
    "task_id", "week_label", "date_start", "date_end", "area", "business_unit", "country", "channel", "task_name", "status", "priority", "task_type", "impact", "result", "notes", "related_metric", "attachment_ref",
]
TASKS_REQUIRED_COLUMNS = ["task_id", "week_label", "date_start", "date_end", "area", "task_name", "status", "priority", "task_type", "impact"]
ATTACHMENT_COLUMNS = [
    "attachment_ref", "business_unit", "country", "channel", "period_label", "task_id", "file_name", "file_type", "file_path_or_url", "description", "tags", "comment",
]
ATTACHMENTS_REQUIRED_COLUMNS = ["attachment_ref", "file_name"]
METRIC_ALIASES = {
    "sales": MetricRecord.MetricName.SALES_TOTAL,
    "ventas": MetricRecord.MetricName.SALES_TOTAL,
    "sales_total": MetricRecord.MetricName.SALES_TOTAL,
    "sales_month": MetricRecord.MetricName.SALES_MONTH,
    "ventas_mes": MetricRecord.MetricName.SALES_MONTH,
    "sales_web": MetricRecord.MetricName.SALES_WEB,
    "sales_whatsapp": MetricRecord.MetricName.SALES_WHATSAPP,
    "sales_marketplace": MetricRecord.MetricName.SALES_MARKETPLACE,
    "investment": MetricRecord.MetricName.INVESTMENT,
    "inversion": MetricRecord.MetricName.INVESTMENT,
    "inversión": MetricRecord.MetricName.INVESTMENT,
    "ad_spend": MetricRecord.MetricName.AD_SPEND,
    "ad_spend_by_country": MetricRecord.MetricName.AD_SPEND_BY_COUNTRY,
    "investment_by_product": MetricRecord.MetricName.INVESTMENT_BY_PRODUCT,
    "cpa": MetricRecord.MetricName.CPA,
    "cpa_weekly": MetricRecord.MetricName.CPA_WEEKLY,
    "cpa_monthly": MetricRecord.MetricName.CPA_MONTHLY,
    "cpa_by_product": MetricRecord.MetricName.CPA_BY_PRODUCT,
    "cpl": MetricRecord.MetricName.CPL,
    "cpl_weekly": MetricRecord.MetricName.CPL_WEEKLY,
    "cpl_monthly": MetricRecord.MetricName.CPL_MONTHLY,
    "cpl_by_campaign": MetricRecord.MetricName.CPL_BY_CAMPAIGN,
    "roas": MetricRecord.MetricName.ROAS,
    "messages": MetricRecord.MetricName.MESSAGES,
    "clicks": MetricRecord.MetricName.CLICKS,
    "mensajes": MetricRecord.MetricName.MESSAGES,
    "chats": MetricRecord.MetricName.MESSAGES,
    "conversations": MetricRecord.MetricName.MESSAGES,
    "conversaciones": MetricRecord.MetricName.MESSAGES,
    "purchases": MetricRecord.MetricName.PURCHASES,
    "compras": MetricRecord.MetricName.PURCHASES,
    "closed_deals": MetricRecord.MetricName.CLOSED_DEALS,
    "conversion_rate": MetricRecord.MetricName.CONVERSION_RATE,
    "close_rate": MetricRecord.MetricName.CLOSE_RATE,
    "average_ticket": MetricRecord.MetricName.AVERAGE_TICKET,
    "orders": MetricRecord.MetricName.ORDERS,
    "units": MetricRecord.MetricName.UNITS,
    "utility": MetricRecord.MetricName.UTILITY,
    "operational_profit": MetricRecord.MetricName.OPERATIONAL_PROFIT,
}
CHANNEL_ALIASES = {
    ("uva", "web", ""): "ecommerce-uva",
    ("uva", "ecommerce", ""): "ecommerce-uva",
    ("uva", "whatsapp", "CO"): "whatsapp-uva-co",
    ("uva", "whatsapp", "EC"): "whatsapp-uva-ec",
    ("uva", "whatsapp colombia", ""): "whatsapp-uva-co",
    ("uva", "whatsapp ecuador", ""): "whatsapp-uva-ec",
    ("bali", "web", ""): "bali-web",
    ("bali", "whatsapp", ""): "bali-whatsapp",
    ("bali", "tienda fisica", ""): "bali-tienda-fisica",
    ("bali", "tienda física", ""): "bali-tienda-fisica",
    ("marketplace", "mercado libre", ""): "mercado-libre",
    ("marketplace", "falabella", ""): "falabella",
    ("marketplace", "rappi", ""): "rappi",
    ("marketplace", "farmatodo", ""): "farmatodo",
}
CURRENCY_ALIASES = {"CO": "COP", "COP": "COP", "USD": "USD", "MXN": "MXN", "EUR": "EUR"}
FILE_TYPE_ALIASES = {
    "pdf": Attachment.FileType.PDF,
    "ppt": Attachment.FileType.PRESENTATION,
    "pptx": Attachment.FileType.PRESENTATION,
    "presentation": Attachment.FileType.PRESENTATION,
    "png": Attachment.FileType.IMAGE,
    "jpg": Attachment.FileType.IMAGE,
    "jpeg": Attachment.FileType.IMAGE,
    "image": Attachment.FileType.IMAGE,
    "excel": Attachment.FileType.EXCEL,
    "xlsx": Attachment.FileType.EXCEL,
    "xls": Attachment.FileType.EXCEL,
    "document": Attachment.FileType.DOCUMENT,
    "doc": Attachment.FileType.DOCUMENT,
    "docx": Attachment.FileType.DOCUMENT,
}
BUSINESS_UNIT_ALIASES = {
    "uva": "uva",
    "bali": "bali",
    "marketplace": "marketplace",
}
COUNTRY_ALIASES = {
    "co": "CO",
    "cop": "CO",
    "colombia": "CO",
    "ec": "EC",
    "ecuador": "EC",
    "mx": "MX",
    "mexico": "MX",
    "méxico": "MX",
    "es": "ES",
    "espana": "ES",
    "españa": "ES",
    "pa": "PA",
    "panama": "PA",
    "panamá": "PA",
}
PERIOD_TYPE_ALIASES = {
    "weekly": MetricRecord.PeriodType.WEEKLY,
    "week": MetricRecord.PeriodType.WEEKLY,
    "semanal": MetricRecord.PeriodType.WEEKLY,
    "monthly": MetricRecord.PeriodType.MONTHLY,
    "month": MetricRecord.PeriodType.MONTHLY,
    "mensual": MetricRecord.PeriodType.MONTHLY,
}
CAMPAIGN_TYPE_ALIASES = {
    "meta ads": MetricRecord.CampaignType.META_ADS,
    "meta": MetricRecord.CampaignType.META_ADS,
    "google ads": MetricRecord.CampaignType.GOOGLE_ADS,
    "google": MetricRecord.CampaignType.GOOGLE_ADS,
    "tiktok ads": MetricRecord.CampaignType.TIKTOK_ADS,
    "tik tok ads": MetricRecord.CampaignType.TIKTOK_ADS,
    "comfama": MetricRecord.CampaignType.COMFAMA_UVA,
    "comfama uva": MetricRecord.CampaignType.COMFAMA_UVA,
    "rappi ads": MetricRecord.CampaignType.RAPPI_ADS,
    "falabella ads": MetricRecord.CampaignType.FALABELLA_ADS,
    "mercado libre ads": MetricRecord.CampaignType.MERCADO_LIBRE_ADS,
    "mercadolibre ads": MetricRecord.CampaignType.MERCADO_LIBRE_ADS,
    "whatsapp campaign": MetricRecord.CampaignType.WHATSAPP_CAMPAIGNS,
    "whatsapp campaigns": MetricRecord.CampaignType.WHATSAPP_CAMPAIGNS,
    "sellerchat": MetricRecord.CampaignType.SELLERCHAT,
}


def allowed_catalogs():
    return {
        "business_units": list(BusinessUnit.objects.filter(is_active=True).values_list("name", flat=True)),
        "countries": list(Country.objects.filter(is_active=True).values_list("code", flat=True)),
        "channels": list(Channel.objects.filter(is_active=True).values_list("name", flat=True)),
        "products": list(Product.objects.filter(is_active=True).values_list("name", flat=True)),
        "campaign_types": [choice for choice, _ in MetricRecord.CampaignType.choices],
        "period_types": [choice for choice, _ in MetricRecord.PeriodType.choices],
        "metric_names": sorted(set(METRIC_ALIASES.keys())),
        "canonical_metric_names": [choice for choice, _ in MetricRecord.MetricName.choices],
        "task_areas": [choice for choice, _ in WeeklyTask.Area.choices],
        "task_statuses": [choice for choice, _ in WeeklyTask.Status.choices],
        "task_priorities": [choice for choice, _ in WeeklyTask.Priority.choices],
        "task_types": [choice for choice, _ in WeeklyTask.TaskType.choices],
        "task_impacts": [choice for choice, _ in WeeklyTask.Impact.choices],
        "file_types": [choice for choice, _ in Attachment.FileType.choices],
    }


def _header_style(row):
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")


def _autosize(sheet):
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 3, 40)


def _serialize_preview_value(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _parse_date_value(value):
    if isinstance(value, date):
        return value
    if not value:
        return None
    return date.fromisoformat(str(value))


def _normalize_currency(value):
    normalized = CURRENCY_ALIASES.get(str(value or "").strip().upper(), str(value or "").strip().upper())
    return normalized or "COP"


def _normalize_file_type(value, file_name=""):
    raw = str(value or "").strip().lower()
    if raw in FILE_TYPE_ALIASES:
        return FILE_TYPE_ALIASES[raw]
    if "." in file_name:
        ext = file_name.rsplit(".", 1)[-1].lower()
        return FILE_TYPE_ALIASES.get(ext, Attachment.FileType.DOCUMENT)
    return Attachment.FileType.DOCUMENT


def build_master_workbook(filters=None, include_data=True, template_only=False):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "metrics"
    sheet.append(METRICS_COLUMNS)
    _header_style(sheet[1])

    if include_data and not template_only:
        from reports.services.analytics import apply_attachment_filters, apply_metric_filters, apply_task_filters

        metric_qs = apply_metric_filters(MetricRecord.objects.select_related("business_unit", "country", "channel", "product"), filters or {})
        for record in metric_qs:
            sheet.append([
                record.record_id,
                record.business_unit.name,
                record.country.code if record.country else "",
                record.channel.name if record.channel else "",
                record.subchannel,
                record.product.name if record.product else "",
                record.campaign_type,
                record.period_type,
                record.period_label,
                record.date_start.year,
                record.date_start.isoformat(),
                record.date_end.isoformat(),
                record.metric_name,
                float(record.metric_value),
                _normalize_currency(record.currency),
                record.source,
                record.notes,
            ])

        tasks_sheet = workbook.create_sheet("tasks")
        tasks_sheet.append(TASKS_COLUMNS)
        _header_style(tasks_sheet[1])
        task_qs = apply_task_filters(WeeklyTask.objects.select_related("business_unit", "country", "channel", "related_metric"), filters or {})
        for task in task_qs:
            tasks_sheet.append([
                task.task_id,
                task.week_label,
                task.date_start.isoformat(),
                task.date_end.isoformat(),
                task.area,
                task.business_unit.name if task.business_unit else "",
                task.country.code if task.country else "",
                task.channel.name if task.channel else "",
                task.task_name,
                task.status,
                task.priority,
                task.task_type,
                task.impact,
                task.result,
                task.notes,
                task.related_metric.record_id if task.related_metric else "",
                task.attachment_ref,
            ])

        attachments_sheet = workbook.create_sheet("attachments")
        attachments_sheet.append(ATTACHMENT_COLUMNS)
        _header_style(attachments_sheet[1])
        attachment_qs = apply_attachment_filters(Attachment.objects.select_related("business_unit", "country", "channel", "task"), filters or {})
        for attachment in attachment_qs:
            attachments_sheet.append([
                attachment.attachment_ref,
                attachment.business_unit.name if attachment.business_unit else "",
                attachment.country.code if attachment.country else "",
                attachment.channel.name if attachment.channel else "",
                attachment.period_label,
                attachment.task.task_id if attachment.task else "",
                attachment.file_name,
                attachment.file_type,
                attachment.file_link,
                attachment.description,
                attachment.tags,
                attachment.comment,
            ])
    else:
        tasks_sheet = workbook.create_sheet("tasks")
        tasks_sheet.append(TASKS_COLUMNS)
        _header_style(tasks_sheet[1])
        attachments_sheet = workbook.create_sheet("attachments")
        attachments_sheet.append(ATTACHMENT_COLUMNS)
        _header_style(attachments_sheet[1])

    catalogs_sheet = workbook.create_sheet("catalogs")
    catalogs_sheet.append(["catalog", "value"])
    _header_style(catalogs_sheet[1])
    for catalog_name, values in allowed_catalogs().items():
        for value in values:
            catalogs_sheet.append([catalog_name, value])

    instructions = workbook.create_sheet("instructions")
    instructions.append(["sheet", "instruction"])
    _header_style(instructions[1])
    for row in [
        ("metrics", "La hoja metrics acepta aliases en ingles y espanol. chats y conversations se normalizan a messages."),
        ("metrics", "source y notes son opcionales. year tambien es opcional y se usa solo como apoyo visual."),
        ("metrics", "ROAS, conversion_rate y close_rate pueden importarse. Si existen, el sistema prioriza el valor importado."),
        ("tasks", "Las tareas siempre se cargan por semana. task_id debe ser unico."),
        ("attachments", "Puedes cargar links o archivos. description y tags son opcionales."),
        ("catalogs", "Usar estos valores para validaciones y dropdowns."),
    ]:
        instructions.append(list(row))

    for current_sheet in workbook.worksheets:
        _autosize(current_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _rows_from_sheet(sheet):
    header = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append((index, {header[position]: _serialize_preview_value(values[position]) for position in range(len(header))}))
    return header, rows


def _validate_required_columns(header, expected):
    return [column for column in expected if column not in header]


def _lookup_business_unit(value):
    if not value:
        return None
    raw_value = str(value).strip()
    alias = BUSINESS_UNIT_ALIASES.get(raw_value.lower(), raw_value.lower())
    return BusinessUnit.objects.filter(slug=alias).first() or BusinessUnit.objects.filter(name__iexact=raw_value).first()


def _lookup_country(value):
    if not value:
        return None
    raw_value = str(value).strip()
    alias = COUNTRY_ALIASES.get(raw_value.lower(), raw_value.upper())
    return Country.objects.filter(code__iexact=alias).first() or Country.objects.filter(name__iexact=raw_value).first()


def _lookup_channel(value, business_unit_value=None, country_value=None):
    if not value:
        return None
    business_unit = business_unit_value if isinstance(business_unit_value, BusinessUnit) else _lookup_business_unit(business_unit_value)
    country = country_value if isinstance(country_value, Country) else _lookup_country(country_value)
    raw_value = str(value).strip()
    raw_key = raw_value.lower()
    alias_slug = None
    if business_unit:
        alias_slug = CHANNEL_ALIASES.get((business_unit.slug, raw_key, country.code if country else ""))
        if not alias_slug:
            alias_slug = CHANNEL_ALIASES.get((business_unit.slug, raw_key, ""))

    if business_unit and alias_slug:
        channel = Channel.objects.filter(business_unit=business_unit, slug=alias_slug).first()
        if channel:
            return channel

    queryset = Channel.objects.all()
    if business_unit:
        queryset = queryset.filter(business_unit=business_unit)
    return queryset.filter(slug=raw_key).first() or queryset.filter(name__iexact=raw_value).first()


def _lookup_product(value, business_unit_value=None):
    if not value:
        return None
    queryset = Product.objects.all()
    business_unit = business_unit_value if isinstance(business_unit_value, BusinessUnit) else _lookup_business_unit(business_unit_value)
    if business_unit:
        queryset = queryset.filter(business_unit=business_unit)
    raw_value = str(value).strip()
    raw_slug = raw_value.lower().replace(" ", "-")
    return queryset.filter(slug=raw_slug).first() or queryset.filter(name__iexact=raw_value).first()


def _normalize_period_type(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return ""
    return PERIOD_TYPE_ALIASES.get(raw_value.lower(), raw_value.lower())


def _normalize_campaign_type(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return ""
    normalized = CAMPAIGN_TYPE_ALIASES.get(raw_value.lower(), raw_value)
    return normalized if normalized in dict(MetricRecord.CampaignType.choices) else ""


def _normalize_metric_name(value):
    return METRIC_ALIASES.get(str(value or "").strip().lower(), "")


def preview_master_import(file_obj, file_name=""):
    workbook = load_workbook(file_obj, data_only=True)
    preview = {"file_name": file_name, "critical_errors": [], "warnings": [], "sheet_stats": {}, "metrics": [], "tasks": [], "attachments": []}

    missing_sheets = [sheet_name for sheet_name in REQUIRED_SHEETS if sheet_name not in workbook.sheetnames]
    if missing_sheets:
        preview["critical_errors"].append({"sheet": "workbook", "row": 0, "message": f"Faltan hojas obligatorias: {', '.join(missing_sheets)}"})
        return preview

    metrics_sheet = workbook["metrics"]
    metrics_header, metric_rows = _rows_from_sheet(metrics_sheet)
    missing_columns = _validate_required_columns(metrics_header, METRICS_REQUIRED_COLUMNS)
    if missing_columns:
        preview["critical_errors"].append({"sheet": "metrics", "row": 1, "message": f"Faltan columnas: {', '.join(missing_columns)}"})
    metric_ids = set()
    for row_number, row in metric_rows:
        row_errors = []
        row_warnings = []
        skip_row = False
        record_id = str(row.get("record_id") or "").strip()
        metric_name = _normalize_metric_name(row.get("metric_name"))
        business_unit = _lookup_business_unit(row.get("business_unit"))
        period_type = _normalize_period_type(row.get("period_type"))
        date_start = _parse_date_value(row.get("date_start"))
        date_end = _parse_date_value(row.get("date_end"))
        if not record_id:
            row_errors.append("record_id es obligatorio")
        elif record_id in metric_ids:
            row_errors.append("record_id duplicado en el archivo")
        metric_ids.add(record_id)
        if row.get("business_unit") and not business_unit:
            row_errors.append("business_unit no existe en catalogos")
        if row.get("country") and not _lookup_country(row.get("country")):
            row_errors.append("country no existe en catalogos")
        if row.get("channel") and not _lookup_channel(row.get("channel"), row.get("business_unit"), row.get("country")):
            row_errors.append("channel no existe para la unidad seleccionada")
        if row.get("product") and not _lookup_product(row.get("product"), row.get("business_unit")):
            row_errors.append("product no existe para la unidad seleccionada")
        if not period_type:
            row_warnings.append("Fila omitida: no tiene period_type valido.")
            skip_row = True
        if not date_start or not date_end:
            row_warnings.append("Fila omitida: no tiene rango de fechas completo.")
            skip_row = True
        campaign_type = _normalize_campaign_type(row.get("campaign_type"))
        if row.get("campaign_type") and not campaign_type:
            row_warnings.append("campaign_type no homologado. Se importara vacio.")
        if not metric_name:
            row_errors.append("metric_name invalido")
        currency = _normalize_currency(row.get("currency"))
        if currency != "COP":
            row_warnings.append(f"La moneda de origen es {currency}. El dashboard asume valores comparables o previamente normalizados a COP.")
        row["metric_name"] = metric_name
        row["campaign_type"] = campaign_type
        row["period_type"] = period_type
        row["date_start"] = date_start.isoformat() if date_start else ""
        row["date_end"] = date_end.isoformat() if date_end else ""
        row["currency"] = currency
        preview["metrics"].append({"row_number": row_number, "data": row, "errors": row_errors, "warnings": row_warnings, "skip": skip_row})
        for error in row_errors:
            preview["critical_errors"].append({"sheet": "metrics", "row": row_number, "message": error})
        for warning in row_warnings:
            preview["warnings"].append({"sheet": "metrics", "row": row_number, "message": warning})

    tasks_sheet = workbook["tasks"]
    tasks_header, task_rows = _rows_from_sheet(tasks_sheet)
    missing_columns = _validate_required_columns(tasks_header, TASKS_REQUIRED_COLUMNS)
    if missing_columns:
        preview["critical_errors"].append({"sheet": "tasks", "row": 1, "message": f"Faltan columnas: {', '.join(missing_columns)}"})
    task_ids = set()
    for row_number, row in task_rows:
        row_errors = []
        task_id = str(row.get("task_id") or "").strip()
        if not task_id:
            row_errors.append("task_id es obligatorio")
        elif task_id in task_ids:
            row_errors.append("task_id duplicado en el archivo")
        task_ids.add(task_id)
        if row.get("business_unit") and not _lookup_business_unit(row.get("business_unit")):
            row_errors.append("business_unit no existe en catalogos")
        if row.get("country") and not _lookup_country(row.get("country")):
            row_errors.append("country no existe en catalogos")
        if row.get("channel") and not _lookup_channel(row.get("channel"), row.get("business_unit"), row.get("country")):
            row_errors.append("channel no existe para la unidad seleccionada")
        if str(row.get("status") or "") not in dict(WeeklyTask.Status.choices):
            row_errors.append("status invalido")
        if str(row.get("priority") or "") not in dict(WeeklyTask.Priority.choices):
            row_errors.append("priority invalido")
        if str(row.get("task_type") or "") not in dict(WeeklyTask.TaskType.choices):
            row_errors.append("task_type invalido")
        if str(row.get("impact") or "") not in dict(WeeklyTask.Impact.choices):
            row_errors.append("impact invalido")
        preview["tasks"].append({"row_number": row_number, "data": row, "errors": row_errors})
        for error in row_errors:
            preview["critical_errors"].append({"sheet": "tasks", "row": row_number, "message": error})

    attachments_sheet = workbook["attachments"]
    attachments_header, attachment_rows = _rows_from_sheet(attachments_sheet)
    missing_columns = _validate_required_columns(attachments_header, ATTACHMENTS_REQUIRED_COLUMNS)
    if missing_columns:
        preview["critical_errors"].append({"sheet": "attachments", "row": 1, "message": f"Faltan columnas: {', '.join(missing_columns)}"})
    attachment_ids = set()
    for row_number, row in attachment_rows:
        row_errors = []
        attachment_ref = str(row.get("attachment_ref") or "").strip()
        if not attachment_ref:
            row_errors.append("attachment_ref es obligatorio")
        elif attachment_ref in attachment_ids:
            row_errors.append("attachment_ref duplicado en el archivo")
        attachment_ids.add(attachment_ref)
        file_type = _normalize_file_type(row.get("file_type"), str(row.get("file_name") or ""))
        row["file_type"] = file_type
        preview["attachments"].append({"row_number": row_number, "data": row, "errors": row_errors})
        for error in row_errors:
            preview["critical_errors"].append({"sheet": "attachments", "row": row_number, "message": error})

    preview["sheet_stats"] = {
        "metrics": {"rows": len(metric_rows), "errors": sum(1 for item in preview["metrics"] if item["errors"]), "warnings": sum(len(item.get("warnings", [])) for item in preview["metrics"])} ,
        "tasks": {"rows": len(task_rows), "errors": sum(1 for item in preview["tasks"] if item["errors"])} ,
        "attachments": {"rows": len(attachment_rows), "errors": sum(1 for item in preview["attachments"] if item["errors"])} ,
    }
    return preview


def _to_string(value):
    return "" if value is None else str(value).strip()


def commit_master_import(preview_payload):
    job = ImportJob.objects.create(
        file_name=preview_payload.get("file_name") or f"import-{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx",
        status=ImportJob.Status.COMPLETED,
        critical_errors=len(preview_payload.get("critical_errors", [])),
        warnings=len(preview_payload.get("warnings", [])),
        preview_payload=preview_payload,
        summary="Importacion completada",
        completed_at=timezone.now(),
    )

    if preview_payload.get("critical_errors"):
        job.status = ImportJob.Status.FAILED
        job.summary = "Importacion bloqueada por errores criticos"
        job.save(update_fields=["status", "summary"])
        return job

    metric_lookup = {record.record_id: record for record in MetricRecord.objects.all()}
    task_lookup = {task.task_id: task for task in WeeklyTask.objects.all()}

    for row in preview_payload.get("metrics", []):
        if row["errors"] or row.get("skip"):
            continue
        data = row["data"]
        business_unit = _lookup_business_unit(data.get("business_unit"))
        defaults = {
            "business_unit": business_unit,
            "country": _lookup_country(data.get("country")),
            "channel": _lookup_channel(data.get("channel"), business_unit, data.get("country")),
            "subchannel": _to_string(data.get("subchannel")),
            "product": _lookup_product(data.get("product"), business_unit),
            "campaign_type": _normalize_campaign_type(data.get("campaign_type")),
            "source": _to_string(data.get("source")),
            "period_type": _normalize_period_type(data.get("period_type")),
            "period_label": _to_string(data.get("period_label")),
            "date_start": _parse_date_value(data.get("date_start")),
            "date_end": _parse_date_value(data.get("date_end")),
            "metric_name": _normalize_metric_name(data.get("metric_name")),
            "metric_value": data.get("metric_value") or 0,
            "currency": _normalize_currency(data.get("currency")),
            "value_origin": MetricRecord.ValueOrigin.IMPORTED,
            "notes": _to_string(data.get("notes")),
        }
        metric, _ = MetricRecord.objects.update_or_create(record_id=_to_string(data.get("record_id")), defaults=defaults)
        metric_lookup[metric.record_id] = metric

    for row in preview_payload.get("tasks", []):
        if row["errors"] or row.get("skip"):
            continue
        data = row["data"]
        business_unit = _lookup_business_unit(data.get("business_unit"))
        defaults = {
            "week_label": _to_string(data.get("week_label")),
            "date_start": _parse_date_value(data.get("date_start")),
            "date_end": _parse_date_value(data.get("date_end")),
            "area": _to_string(data.get("area")),
            "business_unit": business_unit,
            "country": _lookup_country(data.get("country")),
            "channel": _lookup_channel(data.get("channel"), business_unit, data.get("country")),
            "task_name": _to_string(data.get("task_name")),
            "status": _to_string(data.get("status")),
            "priority": _to_string(data.get("priority")),
            "task_type": _to_string(data.get("task_type")),
            "impact": _to_string(data.get("impact")),
            "result": _to_string(data.get("result")),
            "notes": _to_string(data.get("notes")),
            "related_metric": metric_lookup.get(_to_string(data.get("related_metric"))),
            "attachment_ref": _to_string(data.get("attachment_ref")),
        }
        task, _ = WeeklyTask.objects.update_or_create(task_id=_to_string(data.get("task_id")), defaults=defaults)
        task_lookup[task.task_id] = task

    for row in preview_payload.get("attachments", []):
        if row["errors"] or row.get("skip"):
            continue
        data = row["data"]
        business_unit = _lookup_business_unit(data.get("business_unit"))
        defaults = {
            "business_unit": business_unit,
            "country": _lookup_country(data.get("country")),
            "channel": _lookup_channel(data.get("channel"), business_unit, data.get("country")),
            "period_label": _to_string(data.get("period_label")),
            "task": task_lookup.get(_to_string(data.get("task_id"))),
            "file_name": _to_string(data.get("file_name")),
            "file_type": _normalize_file_type(data.get("file_type"), _to_string(data.get("file_name"))),
            "file_path_or_url": _to_string(data.get("file_path_or_url")),
            "description": _to_string(data.get("description")),
            "tags": _to_string(data.get("tags")),
            "comment": _to_string(data.get("comment")),
        }
        Attachment.objects.update_or_create(attachment_ref=_to_string(data.get("attachment_ref")), defaults=defaults)

    return job


def create_export_job(scope, filters):
    return ExportJob.objects.create(file_name=f"reporte-master-{timezone.now().strftime('%Y%m%d%H%M%S')}.xlsx", export_scope=scope, filters=filters, status=ExportJob.Status.COMPLETED, completed_at=timezone.now())




