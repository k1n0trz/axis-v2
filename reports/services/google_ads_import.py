from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import PurePosixPath
import re

from openpyxl import load_workbook

from reports.integrations.axis_sync import AxisSyncService
from reports.integrations.schema import AdSpendRecord, BaliMetricRecord, CategoryMetricRecord
from reports.models import BaliDailyMetric
from reports.services.sales_dashboard import category_slug_from_product_name, normalize_text, parse_excel_date


COUNTRY_CODE_BY_NAME = {
    "colombia": "CO",
    "ecuador": "EC",
    "mexico": "MX",
    "méxico": "MX",
}

COUNTRY_LABEL_BY_CODE = {
    "CO": "Colombia",
    "EC": "Ecuador",
    "MX": "Mexico",
}


def _resolve_sheet(workbook, requested_name):
    requested = normalize_text(requested_name)
    for sheet_name in workbook.sheetnames:
        if normalize_text(sheet_name) == requested:
            return workbook[sheet_name]
    return None


def _header_map(headers):
    return {normalize_text(value): index for index, value in enumerate(headers or []) if str(value or "").strip()}


def _require_headers(header_map, aliases_by_key, sheet_name):
    resolved = {}
    missing = []
    for key, aliases in aliases_by_key.items():
        index = None
        for alias in aliases:
            if normalize_text(alias) in header_map:
                index = header_map[normalize_text(alias)]
                break
        if index is None:
            missing.append(aliases[0])
        else:
            resolved[key] = index
    if missing:
        raise ValueError(f"La hoja '{sheet_name}' no tiene columnas requeridas: {', '.join(missing)}.")
    return resolved


def _merge_source_files(existing_source, new_source, max_length=255):
    parts = []
    for candidate in (existing_source, new_source):
        for part in str(candidate or "").split(";"):
            cleaned = part.strip()
            if cleaned and cleaned not in parts:
                parts.append(cleaned)

    if not parts:
        return ""

    merged = "; ".join(parts)
    if len(merged) <= max_length:
        return merged

    prioritized = []
    for part in parts:
        if "shopifyql" in part.lower() and part not in prioritized:
            prioritized.append(part)
    if new_source and new_source not in prioritized:
        prioritized.append(str(new_source))
    for part in reversed(parts):
        if part not in prioritized:
            prioritized.append(part)

    kept = []
    for part in prioritized:
        candidate = "; ".join(kept + [part])
        if len(candidate) <= max_length:
            kept.append(part)
    return "; ".join(kept)[:max_length]


def parse_decimal(value):
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        number = value
        raw_number = format(value, "f")
        if abs(number) < 1000 and "." in raw_number and len(raw_number.rsplit(".", 1)[-1]) == 3:
            return Decimal(raw_number.replace(".", ""))
        return number
    if isinstance(value, (int, float)):
        raw_number = str(value)
        number = Decimal(raw_number)
        if abs(number) < 1000 and "." in raw_number and len(raw_number.rsplit(".", 1)[-1]) == 3:
            return Decimal(raw_number.replace(".", ""))
        return number

    raw = str(value or "").strip()
    raw = raw.replace("\xa0", "").replace(" ", "")
    raw = re.sub(r"[^\d,.\-]", "", raw)
    if not raw:
        return Decimal("0")

    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        parts = raw.split(",")
        raw = raw.replace(".", "")
        raw = raw.replace(",", ".") if len(parts[-1]) in (1, 2) else raw.replace(",", "")
    elif "." in raw:
        parts = raw.split(".")
        if len(parts) > 2 or (len(parts) == 2 and len(parts[-1]) == 3 and len(parts[0]) <= 3):
            raw = raw.replace(".", "")

    try:
        return Decimal(raw)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


def parse_cop_amount(value):
    amount = parse_decimal(value)
    if not isinstance(value, (Decimal, int, float)):
        return amount

    raw_number = format(value, "f") if isinstance(value, Decimal) else str(value)
    if "." in raw_number and 0 < abs(amount) < 1000:
        decimal_places = len(raw_number.rsplit(".", 1)[-1])
        if decimal_places in (2, 3):
            return (amount * Decimal("1000")).quantize(Decimal("0.01"))
    return amount


def _country_code(value):
    return COUNTRY_CODE_BY_NAME.get(normalize_text(value), "")


def import_google_ads_workbook(file_obj, source_name, uva_sheet="uva", bali_sheet="bali"):
    workbook = load_workbook(file_obj, data_only=True, read_only=True)
    sync = AxisSyncService()

    try:
        category_records = []
        spend_by_country_date = defaultdict(Decimal)
        bali_records = []

        uva_resolved_sheet = _resolve_sheet(workbook, uva_sheet)
        if uva_resolved_sheet:
            sheet = uva_resolved_sheet
            headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            columns = _require_headers(
                _header_map(headers),
                {
                    "country": ("Pais", "País", "Country"),
                    "category": ("Categoria", "Categoría", "Producto", "Product Category"),
                    "date": ("Fecha", "Date"),
                    "cpa_google": ("Cpa Google Ads", "CPA Google Ads", "CPA Google"),
                    "spend_google": ("Inversion Google Ads", "Inversión Google Ads", "Gasto Google Ads", "Coste", "Costo"),
                },
                sheet.title,
            )

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                country_code = _country_code(row[columns["country"]])
                category_name = str(row[columns["category"]] or "").strip()
                metric_date = parse_excel_date(row[columns["date"]])
                if not country_code or not category_name or not metric_date:
                    continue

                spend_google = parse_cop_amount(row[columns["spend_google"]])
                cpa_google = parse_cop_amount(row[columns["cpa_google"]]) or None
                category_slug = category_slug_from_product_name(category_name)

                category_records.append(
                    CategoryMetricRecord(
                        business_unit_slug="uva",
                        country_code=country_code,
                        category_slug=category_slug,
                        category_name=category_name,
                        metric_date=metric_date,
                        cpa_google=cpa_google,
                        spend_google=spend_google,
                        source_file=source_name,
                        notes=f"Importado desde hoja '{sheet.title}' de Google Ads.",
                    )
                )
                spend_by_country_date[(country_code, metric_date)] += spend_google

        ad_spend_records = [
            AdSpendRecord(
                business_unit_slug="uva",
                country_code=country_code,
                ad_platform_slug="google-ads",
                spend_date=spend_date,
                spend_amount=spend_amount,
                source_file=source_name,
                notes=f"Importado desde hoja '{uva_resolved_sheet.title if uva_resolved_sheet else uva_sheet}' de Google Ads.",
            )
            for (country_code, spend_date), spend_amount in sorted(spend_by_country_date.items())
        ]

        bali_resolved_sheet = _resolve_sheet(workbook, bali_sheet)
        if bali_resolved_sheet:
            sheet = bali_resolved_sheet
            headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            header_map = _header_map(headers)
            required_google = {
                "date": ("Fecha", "Date"),
                "spend_google": ("Inversion Google Ads", "Inversión Google Ads", "Gasto Google Ads", "Coste", "Costo"),
                "orders_google": ("Compras Google Ads", "Compras", "Conversiones Google Ads"),
                "whatsapp_conversations": ("Conversaciones Whatsapp", "Conversaciones WhatsApp", "Conversaciones"),
                "cpa": ("Cpa", "CPA"),
            }
            columns = _require_headers(header_map, required_google, sheet.title)
            full_columns = {}
            has_full_bali_format = all(normalize_text(alias) in header_map for alias in ("Visitas Registradas", "Ventas Web", "Pedidos Web"))
            if has_full_bali_format:
                full_columns = _require_headers(
                    header_map,
                    {
                        "sessions": ("Visitas Registradas", "Sesiones"),
                        "web_sales": ("Ventas Web",),
                        "web_orders": ("Pedidos Web",),
                    },
                    sheet.title,
                )
            has_google_only_bali_format = not has_full_bali_format

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                metric_date = parse_excel_date(row[columns["date"]])
                if not metric_date:
                    continue

                existing_metric = BaliDailyMetric.objects.filter(
                    business_unit__slug="bali",
                    country__code="CO",
                    metric_date=metric_date,
                ).first()
                preserve_shopify_fields = bool(
                    existing_metric and "shopifyql" in str(existing_metric.source_file or "").lower()
                )
                if has_full_bali_format:
                    if preserve_shopify_fields:
                        sessions = existing_metric.sessions
                        web_sales_amount = existing_metric.web_sales_amount
                        web_order_count = existing_metric.web_order_count
                    else:
                        sessions = int(parse_decimal(row[full_columns["sessions"]]))
                        web_sales_amount = parse_cop_amount(row[full_columns["web_sales"]])
                        web_order_count = int(parse_decimal(row[full_columns["web_orders"]]))
                    google_spend_amount = parse_cop_amount(row[columns["spend_google"]])
                    google_attributed_orders = int(parse_decimal(row[columns["orders_google"]]))
                    whatsapp_conversations = int(parse_decimal(row[columns["whatsapp_conversations"]]))
                    cpa = parse_cop_amount(row[columns["cpa"]])
                else:
                    sessions = existing_metric.sessions if existing_metric else 0
                    web_sales_amount = existing_metric.web_sales_amount if existing_metric else Decimal("0")
                    web_order_count = existing_metric.web_order_count if existing_metric else 0
                    google_spend_amount = parse_cop_amount(row[columns["spend_google"]])
                    google_attributed_orders = int(parse_decimal(row[columns["orders_google"]]))
                    whatsapp_conversations = int(parse_decimal(row[columns["whatsapp_conversations"]]))
                    cpa = parse_cop_amount(row[columns["cpa"]])

                source_file = source_name
                if existing_metric and (has_google_only_bali_format or preserve_shopify_fields) and existing_metric.source_file:
                    source_file = _merge_source_files(existing_metric.source_file, source_name)

                bali_records.append(
                    BaliMetricRecord(
                        business_unit_slug="bali",
                        country_code="CO",
                        metric_date=metric_date,
                        sessions=sessions,
                        web_sales_amount=web_sales_amount,
                        web_order_count=web_order_count,
                        google_spend_amount=google_spend_amount,
                        google_attributed_orders=google_attributed_orders,
                        whatsapp_conversations=whatsapp_conversations,
                        cpa=cpa,
                        source_file=source_file,
                        notes=(
                            f"Importado desde hoja '{sheet.title}' de Google Ads. "
                            "Campos Shopify preservados desde Analytics cuando ya existe una carga shopifyql."
                            if preserve_shopify_fields
                            else f"Importado desde hoja '{sheet.title}' de Google Ads. "
                            "Campos Shopify preservados desde la API cuando el workbook usa formato solo Google Ads."
                        ),
                    )
                )

        category_stats = sync.sync_category_metrics(category_records) if category_records else {"created": 0, "updated": 0}
        spend_stats = sync.sync_ad_spends(ad_spend_records) if ad_spend_records else {"created": 0, "updated": 0}
        bali_stats = sync.sync_bali_metrics(bali_records) if bali_records else {"created": 0, "updated": 0}

        return {
            "uva_category_rows": len(category_records),
            "uva_daily_spend_rows": len(ad_spend_records),
            "bali_rows": len(bali_records),
            "category_stats": category_stats,
            "spend_stats": spend_stats,
            "bali_stats": bali_stats,
            "countries": sorted({COUNTRY_LABEL_BY_CODE.get(code, code) for code, _ in spend_by_country_date.keys()}),
        }
    finally:
        workbook.close()


def source_name_from_drive_path(path):
    return PurePosixPath(path).name
