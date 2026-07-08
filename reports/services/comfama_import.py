from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation
import re

from django.db import transaction
from openpyxl import load_workbook

from reports.models import AdPlatform, BusinessUnit, ComfamaAdMetric, ComfamaProductReference, ComfamaSale, Country, DailyAdSpend, ProductCategory
from reports.services.sales_dashboard import category_slug_from_product_name, normalize_text, parse_excel_date


REFERENCE_INFERENCE = (
    ("UV-BPM", "Panty Menstrual", "panties-menstruales", Decimal("48000"), Decimal("68300")),
    ("V-BPM", "Panty Menstrual", "panties-menstruales", Decimal("48000"), Decimal("68300")),
    ("UV-BCO-003", "Disco Menstrual", "disco-menstrual", Decimal("63100"), Decimal("63600")),
    ("UV-BCO", "Copa Menstrual UVA", "copa-menstrual", Decimal("63100"), Decimal("63600")),
)

CANONICAL_COMFAMA_REFERENCES = (
    ("Copa Menstrual UVA", "copa-menstrual", "UV-BCO-002-A-SAL T-A", Decimal("63100"), Decimal("63600")),
    ("Copa Menstrual UVA", "copa-menstrual", "UV-BCO-002-B-MOR T-B", Decimal("63100"), Decimal("63600")),
    ("Copa Menstrual UVA", "copa-menstrual", "UV-BCO-002-0-ROS T-0", Decimal("63100"), Decimal("63600")),
    ("Disco Menstrual", "disco-menstrual", "UV-BCO-003-DISCO", Decimal("63100"), Decimal("63600")),
    ("Panty Menstrual", "panties-menstruales", "UV-BPM-7006-LEV-S", Decimal("48000"), Decimal("68300")),
    ("Panty Menstrual", "panties-menstruales", "UV-BPM-7006-LEV-L", Decimal("48000"), Decimal("68300")),
    ("Panty Menstrual", "panties-menstruales", "UV-BPM-7006-LEV-M", Decimal("48000"), Decimal("68300")),
    ("Panty Menstrual", "panties-menstruales", "UV-BPM-7006-MOD-S", Decimal("48000"), Decimal("68300")),
    ("Panty Menstrual", "panties-menstruales", "UV-BPM-7006-MOD-L", Decimal("48000"), Decimal("68300")),
    ("Panty Menstrual", "panties-menstruales", "UV-BPM-7006-MOD-M", Decimal("48000"), Decimal("68300")),
)


def parse_decimal(value):
    raw = str(value or "").strip().replace(",", "")
    if not raw:
        return Decimal("0")
    try:
        return Decimal(raw)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


def clean_ref(value):
    return " ".join(str(value or "").strip().split())


def _column_value(row, index):
    if index is None or index >= len(row):
        return None
    return row[index]


def _sales_column_indexes(sheet):
    fallback = {"date": 0, "tariff": 1, "reference": 2}
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        return fallback

    normalized = {normalize_text(value): index for index, value in enumerate(header) if value}
    date_index = normalized.get("fecha")
    tariff_index = normalized.get("tarifa")
    reference_index = normalized.get("sku", normalized.get("referencia"))
    if date_index is None or reference_index is None:
        return fallback
    return {"date": date_index, "tariff": tariff_index, "reference": reference_index}


def _tariff_from_values(raw_tariff, reference_value):
    tariff = str(raw_tariff or "").strip()
    valid_tariffs = dict(ComfamaSale.Tariff.choices)
    if tariff in valid_tariffs:
        return tariff

    normalized_reference = clean_ref(reference_value).upper()
    if re.search(r"\bT-B\b", normalized_reference):
        return ComfamaSale.Tariff.TARIFF_B
    return ComfamaSale.Tariff.TARIFF_A


def category_for_label(label):
    slug = category_slug_from_product_name(label)
    category = ProductCategory.objects.filter(slug=slug).first()
    if category:
        return category
    return ProductCategory.objects.create(name=str(label).strip(), slug=slug, description="Categoria importada desde modulo Comfama.")


def ensure_comfama_product_references():
    created = 0
    updated = 0
    for category_name, category_slug, reference_code, tariff_a, tariff_b in CANONICAL_COMFAMA_REFERENCES:
        category = ProductCategory.objects.filter(slug=category_slug).first() or ProductCategory.objects.create(
            name=category_name,
            slug=category_slug,
            description="Categoria importada desde modulo Comfama.",
        )
        _, was_created = ComfamaProductReference.objects.update_or_create(
            reference=clean_ref(reference_code),
            defaults={
                "category": category,
                "price_tariff_a": tariff_a,
                "price_tariff_b": tariff_b,
                "is_inferred": False,
                "is_active": True,
                "notes": "Referencia canonica de Comfama cargada desde Helti.",
            },
        )
        if was_created:
            created += 1
        else:
            updated += 1
    return {"created": created, "updated": updated, "count": len(CANONICAL_COMFAMA_REFERENCES)}


def infer_reference(reference):
    normalized = clean_ref(reference).upper().rstrip(",")
    for prefix, category_name, category_slug, tariff_a, tariff_b in REFERENCE_INFERENCE:
        if normalized.startswith(prefix):
            category = ProductCategory.objects.filter(slug=category_slug).first() or ProductCategory.objects.create(
                name=category_name,
                slug=category_slug,
                description="Categoria importada desde modulo Comfama.",
            )
            return category, tariff_a, tariff_b
    return None, Decimal("0"), Decimal("0")


def import_comfama_sales_workbook(file_obj, source_name, sheet_name="Hoja1", end_date=None, replace_source=True):
    end_date = end_date or date.max
    workbook = load_workbook(file_obj, data_only=True, read_only=True)
    created_refs = 0
    updated_refs = 0
    inferred_refs = 0
    created_sales = 0
    updated_sales = 0
    skipped_sales = 0
    deleted_sales = 0

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"No existe la hoja '{sheet_name}'. Hojas disponibles: {', '.join(workbook.sheetnames)}")

        sheet = workbook[sheet_name]
        with transaction.atomic():
            ensure_comfama_product_references()
            current_category = ""
            for row in sheet.iter_rows(min_row=2, values_only=True):
                category_label = row[4] if len(row) > 4 else ""
                reference = row[5] if len(row) > 5 else ""
                tariff_a = row[6] if len(row) > 6 else 0
                tariff_b = row[7] if len(row) > 7 else 0
                if category_label:
                    current_category = str(category_label).strip()
                if reference and (tariff_a or tariff_b):
                    category = category_for_label(current_category or reference)
                    _, was_created = ComfamaProductReference.objects.update_or_create(
                        reference=clean_ref(reference),
                        defaults={
                            "category": category,
                            "price_tariff_a": parse_decimal(tariff_a),
                            "price_tariff_b": parse_decimal(tariff_b),
                            "is_inferred": False,
                            "is_active": True,
                            "notes": f"Importado desde tabla de precios de {source_name}.",
                        },
                    )
                    if was_created:
                        created_refs += 1
                    else:
                        updated_refs += 1

            if replace_source:
                sales_columns = _sales_column_indexes(sheet)
                imported_dates = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    sale_date = parse_excel_date(_column_value(row, sales_columns["date"]))
                    if sale_date and sale_date <= end_date:
                        imported_dates.append(sale_date)
                if imported_dates:
                    deleted_sales, _ = ComfamaSale.objects.exclude(source_file="").filter(
                        sale_date__gte=min(imported_dates),
                        sale_date__lte=max(imported_dates),
                    ).delete()
                else:
                    deleted_sales, _ = ComfamaSale.objects.filter(source_file=source_name).delete()

            valid_tariffs = dict(ComfamaSale.Tariff.choices)
            sales_columns = _sales_column_indexes(sheet)
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                sale_date = parse_excel_date(_column_value(row, sales_columns["date"]))
                reference_value = clean_ref(_column_value(row, sales_columns["reference"]))
                tariff = _tariff_from_values(_column_value(row, sales_columns["tariff"]), reference_value)
                if not sale_date or sale_date > end_date or tariff not in valid_tariffs or not reference_value:
                    skipped_sales += 1
                    continue

                reference = ComfamaProductReference.objects.filter(reference=reference_value).first()
                if not reference:
                    category, tariff_a, tariff_b = infer_reference(reference_value)
                    if not category:
                        skipped_sales += 1
                        continue
                    reference, was_created = ComfamaProductReference.objects.update_or_create(
                        reference=reference_value,
                        defaults={
                            "category": category,
                            "price_tariff_a": tariff_a,
                            "price_tariff_b": tariff_b,
                            "is_inferred": True,
                            "is_active": True,
                            "notes": f"Referencia inferida desde ventas de {source_name}; validar precio si aplica.",
                        },
                    )
                    if was_created:
                        inferred_refs += 1

                sale, was_created = ComfamaSale.objects.update_or_create(
                    source_file=source_name,
                    source_row=row_number,
                    defaults={
                        "sale_date": sale_date,
                        "tariff": tariff,
                        "reference": reference,
                        "notes": "Importado desde ventas Comfama.",
                    },
                )
                sale.save()
                if was_created:
                    created_sales += 1
                else:
                    updated_sales += 1
    finally:
        workbook.close()

    return {
        "created_refs": created_refs,
        "updated_refs": updated_refs,
        "inferred_refs": inferred_refs,
        "created_sales": created_sales,
        "updated_sales": updated_sales,
        "deleted_sales": deleted_sales,
        "skipped_sales": skipped_sales,
    }


def ensure_comfama_ad_catalogs():
    business_unit, _ = BusinessUnit.objects.update_or_create(
        slug="comfama-uva",
        defaults={"name": "Comfama Uva", "display_order": 2, "is_active": True},
    )
    country = Country.objects.filter(code__iexact="CO").first() or Country.objects.create(code="CO", name="Colombia", display_order=1, is_active=True)
    platform, _ = AdPlatform.objects.update_or_create(name="Meta Ads", defaults={"slug": "meta-ads", "is_active": True})
    return business_unit, country, platform


def import_comfama_ad_spend_workbook(file_obj, source_name, sheet_name="Hoja2", end_date=None):
    end_date = end_date or date.max
    business_unit, country, platform = ensure_comfama_ad_catalogs()
    workbook = load_workbook(file_obj, data_only=True, read_only=True)
    spend_by_date = defaultdict(lambda: {"spend": Decimal("0"), "conversations": 0})
    metrics_by_date_category = defaultdict(lambda: {"spend": Decimal("0"), "conversations": 0, "cpl": Decimal("0"), "category": None})

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"No existe la hoja '{sheet_name}'. Hojas disponibles: {', '.join(workbook.sheetnames)}")

        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            spend_date = parse_excel_date(row[0] if len(row) > 0 else None)
            if not spend_date or spend_date > end_date:
                continue
            category_label = str(row[1] or "").strip() if len(row) > 1 else ""
            spend = parse_decimal(row[3] if len(row) > 3 else 0)
            conversations = int(parse_decimal(row[4] if len(row) > 4 else 0))
            cpl = parse_decimal(row[2] if len(row) > 2 else 0)

            spend_by_date[spend_date]["spend"] += spend
            spend_by_date[spend_date]["conversations"] += conversations

            if category_label:
                category = category_for_label(category_label)
                key = (spend_date, category.id)
                metrics_by_date_category[key]["category"] = category
                metrics_by_date_category[key]["spend"] += spend
                metrics_by_date_category[key]["conversations"] += conversations
                metrics_by_date_category[key]["cpl"] = cpl
    finally:
        workbook.close()

    created = 0
    updated = 0
    metric_created = 0
    metric_updated = 0
    with transaction.atomic():
        for spend_date, values in sorted(spend_by_date.items()):
            _, was_created = DailyAdSpend.objects.update_or_create(
                business_unit=business_unit,
                country=country,
                ad_platform=platform,
                spend_date=spend_date,
                defaults={
                    "spend_amount": values["spend"],
                    "source_type": DailyAdSpend.SourceType.IMPORTED,
                    "source_file": source_name,
                    "notes": f"Importado desde pauta Meta Ads Comfama. Conversaciones registradas en archivo: {values['conversations']}.",
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        for (metric_date, _category_id), values in sorted(metrics_by_date_category.items(), key=lambda item: (item[0][0], item[1]["category"].name)):
            conversations = values["conversations"]
            cpl = _safe_cpl(values["spend"], conversations, values["cpl"])
            _, was_created = ComfamaAdMetric.objects.update_or_create(
                metric_date=metric_date,
                category=values["category"],
                defaults={
                    "cpl": cpl,
                    "spend_amount": values["spend"],
                    "conversations": conversations,
                    "source_file": source_name,
                    "notes": "Importado desde hoja de gastos, CPL y mensajes Comfama.",
                },
            )
            if was_created:
                metric_created += 1
            else:
                metric_updated += 1

    return {
        "created": created,
        "updated": updated,
        "dates": len(spend_by_date),
        "metric_created": metric_created,
        "metric_updated": metric_updated,
        "metric_rows": len(metrics_by_date_category),
        "business_unit": business_unit.name,
    }


def _safe_cpl(spend, conversations, fallback):
    if conversations:
        return spend / Decimal(conversations)
    return fallback
