from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import logging
import unicodedata

from django.conf import settings
from django.core.cache import cache
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from django.utils.text import slugify

from reports.integrations.clients import MetaAdsClient
from reports.query_cache import memoize_per_request
from reports.models import AdPlatform, AwnInternationalFollowerMetric, BaliCommunityWebcamMetric, BaliDailyMetric, BaliWebProductDailyMetric, BusinessUnit, Channel, ComfamaAdMetric, ComfamaSale, Country, DailyAdSpend, DailyChannelSale, DailyGeoAdMetric, DailyProductCategoryMetric, DailyProductCategorySale, MarketplaceProductInventory, Product, ProductCategory, SalesTransaction

logger = logging.getLogger(__name__)

ZERO = Decimal("0")
COLOMBIA_VAT_DIVISOR = Decimal("1.19")
MONEY_QUANT = Decimal("0.01")
ECUADOR_USD_TO_COP_RATE = Decimal("3700")
MEXICO_MXN_TO_COP_RATE = Decimal("200")


def _setting_int(name, default):
    try:
        return int(getattr(settings, name, default))
    except (TypeError, ValueError):
        return int(default)


def uva_exchange_rate_for_country(country_code, currency, fallback_rate=None):
    normalized_country = str(country_code or "").upper()
    normalized_currency = str(currency or "").upper()
    if normalized_country == "EC" and normalized_currency == "USD":
        return ECUADOR_USD_TO_COP_RATE
    if normalized_country == "MX" and normalized_currency == "MXN":
        return MEXICO_MXN_TO_COP_RATE
    return Decimal(str(fallback_rate if fallback_rate is not None else 1))


def remove_colombia_vat(value):
    amount = Decimal(str(value or 0))
    if not amount:
        return ZERO
    return (amount / COLOMBIA_VAT_DIVISOR).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


UVA_SHEET_CONFIG = {
    "colombia": {
        "country_code": "CO",
        "columns": {
            "product_name": "PRODUCTO",
            "origin": "ORIGEN",
            "sale_date": "FECHA",
            "quantity": "CANTIDAD",
            "sale_value": "VALOR",
            "shipping_value": "ENVIO",
        },
    },
    "ecuador": {
        "country_code": "EC",
        "columns": {
            "product_name": "PRODUCTO",
            "origin": "CENTRO DE COSTOS",
            "sale_date": "FECHA",
            "quantity": "CANTIDAD",
            "sale_value": "VALOR",
            "shipping_value": "ENVIO",
        },
    },
}
CHANNEL_ALIASES = {
    "pagina web": "ecommerce-uva",
    "paginaweb": "ecommerce-uva",
}
UVA_WHATSAPP_BY_COUNTRY = {
    "CO": ("WhatsApp Colombia", "whatsapp-uva-co"),
    "EC": ("WhatsApp Ecuador", "whatsapp-uva-ec"),
}
PRODUCT_CATEGORY_ALIASES = {
    "copa menstrual uva": "copa-menstrual",
    "bolas kegel uva": "bolas-kegel-uva",
    "kits copa uva": "kits",
    "panties menstruales": "panties-menstruales",
    "dilatadores": "dilatadores-vaginales",
    "disco menstrual": "disco-menstrual",
    "hidratante intimo uva": "hidratante-intimo-uva",
    "hidratante intimo": "hidratante-intimo-uva",
    "hidratante íntimo uva": "hidratante-intimo-uva",
    "hidratante íntimo": "hidratante-intimo-uva",
    "hidratante": "hidratante-intimo-uva",
    "cubrepezones sin adhesivo": "cubrepezones",
    "cubrepezones sin adhesivos": "cubrepezones",
    "cubrepezon sin adhesivo": "cubrepezones",
    "cubrepezon sin adhesivos": "cubrepezones",
    "cubrepezones": "cubrepezones",
    "otros uva": "otros-uva",
    "otros": "otros-uva",
}
PRODUCT_CATEGORY_KEYWORDS = (
    ("kit cuidate", "kits"),
    ("kit bienestar", "kits"),
    ("panty", "panties-menstruales"),
    ("panties", "panties-menstruales"),
    ("calzon", "panties-menstruales"),
    ("calzones", "panties-menstruales"),
    ("lubricante", "lubricantes"),
    ("bolas kegel", "bolas-kegel-uva"),
    ("kegel", "bolas-kegel-uva"),
    ("copa menstrual", "copa-menstrual"),
    ("limpiador de copa", "copa-menstrual"),
    ("limpiador", "copa-menstrual"),
    ("disco menstrual", "disco-menstrual"),
    ("hidratante intimo", "hidratante-intimo-uva"),
    ("hidratante íntimo", "hidratante-intimo-uva"),
    ("hidratante", "hidratante-intimo-uva"),
    ("sin adhesivos", "cubrepezones"),
    ("sin adhesivo", "cubrepezones"),
    ("cubrepezones uva", "cubrepezones"),
    ("cubrepezones ultradelgados", "cubrepezones"),
    ("cubrepezones sin adhesivos", "cubrepezones"),
    ("cubrepezones sin adhesivo", "cubrepezones"),
    ("cubrepezon sin adhesivos", "cubrepezones"),
    ("cubrepezon sin adhesivo", "cubrepezones"),
    ("cubrepezon", "cubrepezones"),
    ("cubrepezones", "cubrepezones"),
    ("pezonera", "cubrepezones"),
    ("esterilizador electrico", "esterilizador"),
    ("esterilizador eléctrico", "esterilizador"),
    ("vaso esterilizador", "vaso-esterilizador-copa-uva"),
    ("dilatador", "dilatadores-vaginales"),
    ("dilatadores", "dilatadores-vaginales"),
    ("jabon", "higiene-intima"),
    ("jabon intimo", "higiene-intima"),
    ("kit", "kits"),
)
PRODUCT_CATEGORY_IMAGE_FALLBACKS = {
    "bolas-kegel-uva": "https://copauva.com/wp-content/uploads/2026/05/FOTO-18.jpg",
    "copa-menstrual": "https://copauva.com/wp-content/uploads/2023/01/A-Espanol-Principal-1.png",
    "cubrepezones": "https://copauva.com/wp-content/uploads/2026/05/CUBRE-P-NUDE.jpg",
    "dilatadores-vaginales": "https://copauva.com/wp-content/uploads/2025/09/1-espa__ol.jpg",
    "disco-menstrual": "https://copauva.com/wp-content/uploads/2023/08/Disco-menstrual12.jpg",
    "esterilizador": "https://copauva.com/wp-content/uploads/2024/04/Esterelizador-UVA5-2.jpg",
    "hidratante-intimo-uva": "/static/reports/products/hidratante-intimo-uva.jpg",
    "higiene-intima": "https://copauva.com/wp-content/uploads/2024/10/DSC04764-1.jpg",
    "kits": "https://copauva.com/wp-content/uploads/2025/08/KIT-COPA_Mesa-de-trabajo-1.jpg",
    "lubricantes": "https://copauva.com/wp-content/uploads/2025/12/Diseno-sin-titulo-20.jpg",
    "panties-menstruales": "https://copauva.com/wp-content/uploads/2025/11/7-1.jpg",
    "vaso-esterilizador-copa-uva": "https://copauva.com/wp-content/uploads/2023/06/Uva.jpg",
}
UVA_PRODUCT_CATEGORY_SLUGS = frozenset(
    {
        "bolas-kegel-uva",
        "copa-menstrual",
        "cubrepezones",
        "dilatadores-vaginales",
        "disco-menstrual",
        "esterilizador",
        "hidratante-intimo-uva",
        "higiene-intima",
        "kits",
        "lubricantes",
        "ofertas",
        "otros-uva",
        "panties-menstruales",
        "vaso-esterilizador-copa-uva",
    }
)
UVA_EXCLUDED_PRODUCT_KEYWORDS = (
    "anesty",
    "camtoyz",
    "dildo",
    "elixir",
    "erotic",
    "lovense",
    "masturbador",
    "plug",
    "real men",
    "vibrador",
)


def subtract_one_month(today):
    if today.month == 1:
        year = today.year - 1
        month = 12
    else:
        year = today.year
        month = today.month - 1
    day = min(today.day, monthrange(year, month)[1])
    return date(year, month, day)


def normalize_text(value):
    raw = str(value or "").strip()
    normalized = unicodedata.normalize("NFKD", raw)
    return "".join(char for char in normalized if not unicodedata.combining(char)).strip().lower()


def parse_excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return from_excel(value).date()
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def parse_decimal(value):
    raw = str(value or "").strip().replace(",", "")
    if not raw:
        return ZERO
    try:
        return Decimal(raw)
    except (InvalidOperation, TypeError, ValueError):
        return ZERO


def parse_quantity(value):
    number = parse_decimal(value)
    return int(number) if number > 0 else 0


def ensure_uva_catalogs():
    business_unit, _ = BusinessUnit.objects.update_or_create(name="Uva", defaults={"slug": "uva", "display_order": 1, "is_active": True})
    countries = {}
    for index, (code, name) in enumerate((("CO", "Colombia"), ("EC", "Ecuador"), ("MX", "Mexico"), ("ES", "Espana")), start=1):
        country = Country.objects.filter(code__iexact=code).first() or Country.objects.filter(name__iexact=name).first()
        if country:
            country.code = code
            country.name = name
            country.display_order = index
            country.is_active = True
            country.save(update_fields=["code", "name", "display_order", "is_active"])
            countries[code] = country
        else:
            countries[code] = Country.objects.create(code=code, name=name, display_order=index, is_active=True)
    channels = {}
    channel_specs = [("Web", "ecommerce-uva")]
    channel_specs.extend(UVA_WHATSAPP_BY_COUNTRY.values())
    for index, (name, slug) in enumerate(channel_specs, start=1):
        channel = Channel.objects.filter(business_unit=business_unit, slug=slug).first()
        if channel:
            channel.name = name
            channel.display_order = index
            channel.is_active = True
            channel.save(update_fields=["name", "display_order", "is_active"])
            channels[slug] = channel
            continue
        channels[slug], _ = Channel.objects.update_or_create(
            name=name,
            business_unit=business_unit,
            defaults={"slug": slug, "display_order": index, "is_active": True},
        )
    return {"business_unit": business_unit, "countries": countries, "channels": channels}


def ensure_ad_platform_catalogs():
    platforms = {}
    for name in ("Meta Ads", "Google Ads"):
        platform, _ = AdPlatform.objects.update_or_create(
            name=name,
            defaults={"slug": slugify(name), "is_active": True},
        )
        platforms[platform.slug] = platform
    return platforms


def ensure_bali_catalogs():
    business_unit, _ = BusinessUnit.objects.update_or_create(name="Bali", defaults={"slug": "bali", "display_order": 2, "is_active": True})
    country = Country.objects.filter(code__iexact="CO").first() or Country.objects.filter(name__iexact="Colombia").first()
    if country:
        country.code = "CO"
        country.name = "Colombia"
        country.display_order = country.display_order or 1
        country.is_active = True
        country.save(update_fields=["code", "name", "display_order", "is_active"])
    else:
        country = Country.objects.create(code="CO", name="Colombia", display_order=1, is_active=True)
    country.business_units.add(business_unit)

    channels = {}
    for index, (name, slug) in enumerate((("Web", "bali-web"), ("WhatsApp", "bali-whatsapp"), ("Comunidad Webcam", "bali-community-webcam"), ("Tienda Fisica", "bali-tienda-fisica")), start=1):
        channel, _ = Channel.objects.update_or_create(
            business_unit=business_unit,
            slug=slug,
            defaults={"name": name, "display_order": index, "is_active": True},
        )
        channels[slug] = channel
    return {"business_unit": business_unit, "country": country, "channels": channels}


def ensure_marketplace_catalogs():
    business_unit, _ = BusinessUnit.objects.update_or_create(
        slug="marketplace",
        defaults={"name": "Marketplace", "display_order": 3, "is_active": True},
    )
    countries = {}
    for index, (code, name) in enumerate((("CO", "Colombia"), ("EC", "Ecuador")), start=1):
        country = Country.objects.filter(code__iexact=code).first() or Country.objects.filter(name__iexact=name).first()
        if country:
            country.code = code
            country.name = name
            country.display_order = country.display_order or index
            country.is_active = True
            country.save(update_fields=["code", "name", "display_order", "is_active"])
        else:
            country = Country.objects.create(code=code, name=name, display_order=index, is_active=True)
        country.business_units.add(business_unit)
        countries[code] = country

    channels = {}
    for index, (name, slug, aliases) in enumerate(
        (
            ("Mercadolibre", "mercado-libre", ("mercadolibre",)),
            ("Falabella", "falabella", ()),
            ("Rappi", "rappi", ()),
            ("Farmatodo", "farmatodo", ()),
        ),
        start=1,
    ):
        channel = Channel.objects.filter(business_unit=business_unit, slug__in=(slug, *aliases)).first()
        if channel:
            channel.name = name
            channel.slug = slug
            channel.display_order = index
            channel.is_active = True
            channel.save(update_fields=["name", "slug", "display_order", "is_active"])
        else:
            channel = Channel.objects.create(
                business_unit=business_unit,
                slug=slug,
                name=name,
                display_order=index,
                is_active=True,
            )
        channels[channel.slug] = channel
    return {"business_unit": business_unit, "countries": countries, "channels": channels}


def category_slug_from_product_name(name):
    normalized = normalize_text(name)
    if normalized in PRODUCT_CATEGORY_ALIASES:
        return PRODUCT_CATEGORY_ALIASES[normalized]
    for keyword, slug in PRODUCT_CATEGORY_KEYWORDS:
        if keyword in normalized:
            return slug
    for source, slug in PRODUCT_CATEGORY_ALIASES.items():
        if normalized == source or normalized in source or source in normalized:
            return slug
    return slugify(normalized)


def is_uva_category_slug_allowed(slug):
    return str(slug or "").strip() in UVA_PRODUCT_CATEGORY_SLUGS


def uva_category_slug_from_product_name(name, category_map=None):
    normalized = normalize_text(name)
    if any(keyword in normalized for keyword in UVA_EXCLUDED_PRODUCT_KEYWORDS):
        return ""
    mapped_slug = (category_map or {}).get(str(name or "").strip())
    slug = mapped_slug or category_slug_from_product_name(name)
    return slug if is_uva_category_slug_allowed(slug) else ""


def _category_scope_slugs(filters):
    if normalize_text(filters.get("business_unit")) == "uva":
        return UVA_PRODUCT_CATEGORY_SLUGS
    return None


@memoize_per_request
def product_category_metrics(filters, limit=None):
    queryset = DailyProductCategoryMetric.objects.select_related("business_unit", "country", "category")
    if filters.get("date_start"):
        queryset = queryset.filter(metric_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(metric_date__lte=filters["date_end"])
    if filters.get("business_unit"):
        queryset = queryset.filter(business_unit__slug=filters["business_unit"])
    if filters.get("country"):
        queryset = queryset.filter(country__code=filters["country"])
    category_scope = _category_scope_slugs(filters)
    if category_scope is not None:
        queryset = queryset.filter(category__slug__in=category_scope)
    if limit:
        queryset = queryset[:limit]
    return list(queryset)


@memoize_per_request
def product_category_channel_sales(filters, channel_slug=None, limit=None):
    queryset = DailyProductCategorySale.objects.select_related("business_unit", "country", "channel", "category")
    if filters.get("date_start"):
        queryset = queryset.filter(sale_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(sale_date__lte=filters["date_end"])
    if filters.get("business_unit"):
        queryset = queryset.filter(business_unit__slug=filters["business_unit"])
    if filters.get("country"):
        queryset = queryset.filter(country__code=filters["country"])
    if channel_slug:
        queryset = queryset.filter(channel__slug=channel_slug)
    elif filters.get("channel"):
        queryset = queryset.filter(channel__slug=filters["channel"])
    category_scope = _category_scope_slugs(filters)
    if category_scope is not None:
        queryset = queryset.filter(category__slug__in=category_scope)
    if limit:
        queryset = queryset[:limit]
    return list(queryset)


def _sale_channel_group(channel):
    if not channel:
        return "Sin canal"
    slug = normalize_text(getattr(channel, "slug", ""))
    name = normalize_text(getattr(channel, "name", ""))
    if "whatsapp" in slug or "whatsapp" in name:
        return "WhatsApp"
    if slug in {"ecommerce-uva", "web"} or "web" in slug or "ecommerce" in slug or "web" in name or "pagina" in name:
        return "Web"
    return channel.name


def _combined_direct_sales(filters, limit=None):
    daily_rows = daily_channel_sales(filters)
    raw_category_rows = product_category_channel_sales(filters)
    category_keys = {
        (row.sale_date, row.business_unit_id, row.country_id, row.channel_id)
        for row in raw_category_rows
    }
    prefer_category_channels = {"WhatsApp"}
    daily_rows = [
        row
        for row in daily_rows
        if not (
            _sale_channel_group(row.channel) in prefer_category_channels
            and (row.sale_date, row.business_unit_id, row.country_id, row.channel_id) in category_keys
        )
    ]
    daily_keys = {
        (row.sale_date, row.business_unit_id, row.country_id, row.channel_id)
        for row in daily_rows
    }
    category_rows = [
        row
        for row in raw_category_rows
        if (row.sale_date, row.business_unit_id, row.country_id, row.channel_id) not in daily_keys
    ]
    rows = [*daily_rows, *category_rows]
    rows.sort(key=lambda item: item.sale_date, reverse=True)
    return rows[:limit] if limit else rows


def _safe_ratio(numerator, denominator):
    return (numerator / denominator) if denominator else ZERO


def _daily_order_counts_by_channel(filters):
    counts = defaultdict(int)
    for row in daily_channel_sales(filters):
        count_value = row.order_count or 0
        if not count_value:
            continue
        counts[_sale_channel_group(row.channel)] += count_value
    return counts


def _aggregate_platform_cpa(rows, spend_attr, cpa_attr):
    total_spend = ZERO
    acquisitions = ZERO
    for row in rows:
        spend_value = getattr(row, spend_attr) or ZERO
        cpa_value = getattr(row, cpa_attr) or ZERO
        if spend_value > 0:
            total_spend += spend_value
        if spend_value > 0 and cpa_value and cpa_value > 0:
            acquisitions += spend_value / cpa_value
    return _safe_ratio(total_spend, acquisitions), total_spend


def build_uva_category_snapshot(filters):
    rows = product_category_metrics(filters)
    comparison = None
    if filters.get("compare_mode") == "previous_period":
        comparison = _previous_period_filters(filters)

    grouped = defaultdict(list)
    for row in rows:
        grouped[row.category_id].append(row)

    previous_grouped = defaultdict(list)
    if comparison:
        for row in product_category_metrics(comparison):
            previous_grouped[row.category_id].append(row)

    category_sales_rows = product_category_channel_sales(filters)
    sales_grouped = defaultdict(list)
    for row in category_sales_rows:
        sales_grouped[row.category_id].append(row)

    whatsapp_rows = [
        row
        for row in category_sales_rows
        if row.channel and normalize_text(row.channel.slug).startswith("whatsapp")
    ]
    whatsapp_grouped = defaultdict(list)
    for row in whatsapp_rows:
        whatsapp_grouped[row.category_id].append(row)

    previous_sales_grouped = defaultdict(list)
    previous_whatsapp_grouped = defaultdict(list)
    if comparison:
        for row in product_category_channel_sales(comparison):
            previous_sales_grouped[row.category_id].append(row)
            if row.channel and normalize_text(row.channel.slug).startswith("whatsapp"):
                previous_whatsapp_grouped[row.category_id].append(row)

    cards = []
    category_ids = set(grouped) | set(sales_grouped) | set(whatsapp_grouped)
    for category_id in category_ids:
        current_rows = grouped.get(category_id, [])
        sales_rows = sales_grouped.get(category_id, [])
        if not current_rows and sales_rows:
            category = sales_rows[0].category
        elif not current_rows and whatsapp_grouped.get(category_id):
            category = whatsapp_grouped[category_id][0].category
        elif current_rows:
            category = current_rows[0].category
        else:
            continue
        current_rows = sorted(current_rows, key=lambda item: item.metric_date)
        current_cpa_meta, current_spend_meta = _aggregate_platform_cpa(current_rows, "spend_meta", "cpa_meta")
        current_cpa_google, current_spend_google = _aggregate_platform_cpa(current_rows, "spend_google", "cpa_google")
        web_sales_rows = [row for row in sales_rows if _sale_channel_group(row.channel) == "Web"]
        whatsapp_sales_rows = [row for row in sales_rows if _sale_channel_group(row.channel) == "WhatsApp"]
        current_sales = sum((row.sales_amount for row in web_sales_rows), ZERO)
        current_web_units = sum((row.quantity for row in web_sales_rows), 0)
        if not current_sales and current_rows:
            current_sales = sum((row.sales_amount for row in current_rows), ZERO)
        current_total_spend = sum((row.total_spend for row in current_rows), ZERO)
        current_whatsapp_sales = sum((row.sales_amount for row in whatsapp_sales_rows), ZERO)
        current_whatsapp_units = sum((row.quantity for row in whatsapp_sales_rows), 0)
        current_quantity = sum((row.quantity for row in sales_rows), 0)
        current_total_sales = current_sales + current_whatsapp_sales
        current_roas = _safe_ratio(current_total_sales, current_total_spend)
        current_average_ticket = _safe_ratio(current_total_sales, Decimal(current_quantity))
        current_web_average_ticket = _safe_ratio(sum((row.sales_amount for row in web_sales_rows), ZERO), Decimal(current_web_units))
        current_whatsapp_average_ticket = _safe_ratio(current_whatsapp_sales, Decimal(current_whatsapp_units))

        previous_rows = previous_grouped.get(category_id, [])
        previous_sales_rows = previous_sales_grouped.get(category_id, [])
        previous_cpa_meta, _ = _aggregate_platform_cpa(previous_rows, "spend_meta", "cpa_meta")
        previous_cpa_google, _ = _aggregate_platform_cpa(previous_rows, "spend_google", "cpa_google")
        previous_web_sales_rows = [row for row in previous_sales_rows if _sale_channel_group(row.channel) == "Web"]
        previous_whatsapp_sales_rows = [row for row in previous_sales_rows if _sale_channel_group(row.channel) == "WhatsApp"]
        previous_sales = sum((row.sales_amount for row in previous_web_sales_rows), ZERO)
        if not previous_sales and previous_rows:
            previous_sales = sum((row.sales_amount for row in previous_rows), ZERO)
        previous_total_spend = sum((row.total_spend for row in previous_rows), ZERO)
        previous_whatsapp_sales = sum((row.sales_amount for row in previous_whatsapp_sales_rows), ZERO)
        previous_quantity = sum((row.quantity for row in previous_sales_rows), 0)
        previous_total_sales = previous_sales + previous_whatsapp_sales
        previous_roas = _safe_ratio(previous_total_sales, previous_total_spend)
        previous_average_ticket = _safe_ratio(previous_total_sales, Decimal(previous_quantity))

        cards.append(
            {
                "category_id": category_id,
                "name": category.name,
                "image_url": category_image_url(category),
                "image_fallback_url": category_fallback_image_url(category),
                "sales_total": float(current_sales),
                "web_units": current_web_units,
                "whatsapp_sales_total": float(current_whatsapp_sales),
                "whatsapp_units": current_whatsapp_units,
                "combined_sales_total": float(current_total_sales),
                "quantity_total": current_quantity,
                "spend_total": float(current_total_spend),
                "spend_meta": float(current_spend_meta),
                "spend_google": float(current_spend_google),
                "cpa_meta": round(float(current_cpa_meta), 2) if current_cpa_meta else 0,
                "cpa_google": round(float(current_cpa_google), 2) if current_cpa_google else 0,
                "roas": round(float(current_roas), 2) if current_roas else 0,
                "average_ticket": round(float(current_average_ticket), 2) if current_average_ticket else 0,
                "web_average_ticket": round(float(current_web_average_ticket), 2) if current_web_average_ticket else 0,
                "whatsapp_average_ticket": round(float(current_whatsapp_average_ticket), 2) if current_whatsapp_average_ticket else 0,
                "comparison": {
                    "sales_total": _comparison_payload(float(current_sales), float(previous_sales)),
                    "whatsapp_sales_total": _comparison_payload(float(current_whatsapp_sales), float(previous_whatsapp_sales)),
                    "cpa_meta": _comparison_payload(round(float(current_cpa_meta), 2) if current_cpa_meta else 0, round(float(previous_cpa_meta), 2) if previous_cpa_meta else 0),
                    "cpa_google": _comparison_payload(round(float(current_cpa_google), 2) if current_cpa_google else 0, round(float(previous_cpa_google), 2) if previous_cpa_google else 0),
                    "roas": _comparison_payload(round(float(current_roas), 2) if current_roas else 0, round(float(previous_roas), 2) if previous_roas else 0),
                    "average_ticket": _comparison_payload(round(float(current_average_ticket), 2) if current_average_ticket else 0, round(float(previous_average_ticket), 2) if previous_average_ticket else 0),
                },
            }
        )

    cards.sort(key=lambda item: item["sales_total"] + item["whatsapp_sales_total"], reverse=True)
    total_sales = sum(Decimal(str(item["combined_sales_total"])) for item in cards)
    total_web_sales = sum(Decimal(str(item["sales_total"])) for item in cards)
    total_whatsapp_sales = sum(Decimal(str(item["whatsapp_sales_total"])) for item in cards)
    total_spend = sum(Decimal(str(item["spend_total"])) for item in cards)
    total_quantity = sum(item["quantity_total"] for item in cards)
    total_web_units = sum(item["web_units"] for item in cards)
    total_whatsapp_units = sum(item["whatsapp_units"] for item in cards)
    daily_order_counts = _daily_order_counts_by_channel(filters)
    total_web_orders = daily_order_counts.get("Web", 0) or total_web_units
    total_whatsapp_orders = daily_order_counts.get("WhatsApp", 0) or len([row for row in category_sales_rows if _sale_channel_group(row.channel) == "WhatsApp"]) or total_whatsapp_units
    total_orders = total_web_orders + total_whatsapp_orders
    average_roas = _safe_ratio(total_sales, total_spend)
    average_ticket = _safe_ratio(total_sales, Decimal(total_orders))
    web_average_ticket = _safe_ratio(total_web_sales, Decimal(total_web_orders))
    whatsapp_average_ticket = _safe_ratio(total_whatsapp_sales, Decimal(total_whatsapp_orders))
    profitable_cards = [item for item in cards if item["roas"] >= 3]
    top_sales = cards[0] if cards else None
    top_roas = max(cards, key=lambda item: item["roas"]) if cards else None
    top_whatsapp = max(cards, key=lambda item: item["whatsapp_sales_total"]) if cards else None

    profitability_rows = [
        {"label": item["name"], "value": round(item["roas"], 2), "sales": item["combined_sales_total"], "spend": item["spend_total"]}
        for item in cards
        if item["spend_total"] > 0 and item["combined_sales_total"] > 0
    ]
    profitability_rows.sort(key=lambda item: item["value"], reverse=True)

    insights = []
    if top_sales:
        share = (Decimal(str(top_sales["combined_sales_total"])) / total_sales * 100) if total_sales else ZERO
        insights.append(f"{top_sales['name']} lidera ventas por categoria con {format_cop(top_sales['combined_sales_total'])}, equivalente al {round(float(share), 1)}% del total categorizado.")
    if top_roas and top_roas["roas"]:
        insights.append(f"{top_roas['name']} es la categoria con mejor ROAS del filtro: {top_roas['roas']:.2f}.")
    if top_whatsapp and top_whatsapp["whatsapp_sales_total"]:
        insights.append(f"WhatsApp Detal aporta mas en {top_whatsapp['name']}: {format_cop(top_whatsapp['whatsapp_sales_total'])} y {top_whatsapp['whatsapp_units']} unidades.")
    if total_spend and average_roas < 3:
        insights.append("El ROAS promedio por categorias esta por debajo de 3.0; conviene revisar inversion y mix de productos.")

    return {
        "cards": cards,
        "card_count": len(cards),
        "kpis": {
            "combined_sales_total": float(total_sales),
            "web_sales_total": float(total_web_sales),
            "whatsapp_sales_total": float(total_whatsapp_sales),
            "spend_total": float(total_spend),
            "average_roas": round(float(average_roas), 2) if average_roas else 0,
            "average_ticket": round(float(average_ticket), 2) if average_ticket else 0,
            "web_average_ticket": round(float(web_average_ticket), 2) if web_average_ticket else 0,
            "whatsapp_average_ticket": round(float(whatsapp_average_ticket), 2) if whatsapp_average_ticket else 0,
            "web_order_count": total_web_orders,
            "whatsapp_order_count": total_whatsapp_orders,
            "profitable_count": len(profitable_cards),
        },
        "profitability_json": profitability_rows,
        "insights": insights[:4],
        "insight_cards": _decorate_insights(insights[:4]),
    }


def build_uva_product_detail(filters, category_id):
    try:
        category = ProductCategory.objects.get(pk=category_id)
    except (ProductCategory.DoesNotExist, ValueError, TypeError):
        return None
    if normalize_text(filters.get("business_unit")) == "uva" and not is_uva_category_slug_allowed(category.slug):
        return None

    metric_rows = [row for row in product_category_metrics(filters) if row.category_id == category.id]
    sale_rows = [row for row in product_category_channel_sales(filters) if row.category_id == category.id]
    if not metric_rows and not sale_rows:
        return None

    daily_map = defaultdict(
        lambda: {
            "sales": ZERO,
            "web_sales": ZERO,
            "whatsapp_sales": ZERO,
            "spend": ZERO,
            "spend_meta": ZERO,
            "spend_google": ZERO,
            "metric_sales": ZERO,
            "units": 0,
            "web_units": 0,
            "whatsapp_units": 0,
        }
    )
    for row in metric_rows:
        values = daily_map[row.metric_date]
        values["spend"] += row.total_spend
        values["spend_meta"] += row.spend_meta
        values["spend_google"] += row.spend_google
        if row.sales_amount:
            values["metric_sales"] += row.sales_amount
    for row in sale_rows:
        values = daily_map[row.sale_date]
        channel_group = _sale_channel_group(row.channel)
        values["sales"] += row.sales_amount
        values["units"] += row.quantity or 0
        if channel_group == "WhatsApp":
            values["whatsapp_sales"] += row.sales_amount
            values["whatsapp_units"] += row.quantity or 0
        else:
            values["web_sales"] += row.sales_amount
            values["web_units"] += row.quantity or 0

    daily_series = []
    for metric_date, values in sorted(daily_map.items(), key=lambda item: item[0]):
        if not values["sales"] and values["metric_sales"]:
            values["sales"] = values["metric_sales"]
        roas = _safe_ratio(values["sales"], values["spend"])
        ticket = _safe_ratio(values["sales"], Decimal(values["units"]))
        daily_series.append(
            {
                "label": metric_date.isoformat(),
                "sales": float(values["sales"]),
                "web_sales": float(values["web_sales"]),
                "whatsapp_sales": float(values["whatsapp_sales"]),
                "spend": float(values["spend"]),
                "spend_meta": float(values["spend_meta"]),
                "spend_google": float(values["spend_google"]),
                "units": values["units"],
                "web_units": values["web_units"],
                "whatsapp_units": values["whatsapp_units"],
                "roas": round(float(roas), 2) if roas else 0,
                "average_ticket": round(float(ticket), 2) if ticket else 0,
            }
        )

    sales_total = sum((Decimal(str(item["sales"])) for item in daily_series), ZERO)
    web_sales_total = sum((Decimal(str(item["web_sales"])) for item in daily_series), ZERO)
    whatsapp_sales_total = sum((Decimal(str(item["whatsapp_sales"])) for item in daily_series), ZERO)
    spend_total = sum((Decimal(str(item["spend"])) for item in daily_series), ZERO)
    spend_meta = sum((Decimal(str(item["spend_meta"])) for item in daily_series), ZERO)
    spend_google = sum((Decimal(str(item["spend_google"])) for item in daily_series), ZERO)
    units_total = sum((item["units"] for item in daily_series), 0)
    roas = _safe_ratio(sales_total, spend_total)
    average_ticket = _safe_ratio(sales_total, Decimal(units_total))
    best_day = max(daily_series, key=lambda item: item["sales"], default=None)
    best_roas_day = max([item for item in daily_series if item["spend"] > 0], key=lambda item: item["roas"], default=None)
    cpa_meta, _ = _aggregate_platform_cpa(metric_rows, "spend_meta", "cpa_meta")
    cpa_google, _ = _aggregate_platform_cpa(metric_rows, "spend_google", "cpa_google")

    insights = []
    if sales_total:
        insights.append(f"{category.name} genera {format_cop(sales_total)} y {units_total} unidades en el rango filtrado.")
    if spend_total:
        insights.append(f"La inversion asociada al producto es {format_cop(spend_total)}, con ROAS de {float(roas):.2f}.")
    if whatsapp_sales_total:
        share = _safe_ratio(whatsapp_sales_total, sales_total) * Decimal("100")
        insights.append(f"WhatsApp aporta {format_cop(whatsapp_sales_total)} ({float(share):.1f}% de las ventas del producto).")
    if best_day:
        insights.append(f"El mejor dia por ventas fue {best_day['label']} con {format_cop(best_day['sales'])}.")
    if spend_total and roas < Decimal("3"):
        insights.append("El ROAS esta por debajo de 3.0; conviene revisar pauta, precio y mix de canales.")
    elif best_roas_day and best_roas_day["roas"] >= 3:
        insights.append(f"El mejor dia por eficiencia fue {best_roas_day['label']} con ROAS {best_roas_day['roas']:.2f}.")

    return {
        "type": "uva",
        "key": str(category.id),
        "title": category.name,
        "subtitle": "Detalle por categoria de producto Uva",
        "image_url": category_image_url(category),
        "image_fallback_url": category_fallback_image_url(category),
        "daily_series": daily_series,
        "stats": [
            {"label": "Ventas totales", "value": float(sales_total), "kind": "money"},
            {"label": "Inversion", "value": float(spend_total), "kind": "money"},
            {"label": "ROAS", "value": round(float(roas), 2) if roas else 0, "kind": "ratio"},
            {"label": "Unidades", "value": units_total, "kind": "number"},
            {"label": "Ticket promedio", "value": round(float(average_ticket), 2) if average_ticket else 0, "kind": "money"},
            {"label": "Ventas Web", "value": float(web_sales_total), "kind": "money"},
            {"label": "Ventas WhatsApp", "value": float(whatsapp_sales_total), "kind": "money"},
            {"label": "CPA Meta", "value": round(float(cpa_meta), 2) if cpa_meta else 0, "kind": "money"},
            {"label": "CPA Google", "value": round(float(cpa_google), 2) if cpa_google else 0, "kind": "money"},
            {"label": "Inversion Meta", "value": float(spend_meta), "kind": "money"},
            {"label": "Inversion Google", "value": float(spend_google), "kind": "money"},
        ],
        "insights": insights[:6],
        "allocation_note": "Uva usa la inversion diaria registrada por categoria/producto en Meta y Google Ads.",
    }


def comfama_sales(filters):
    queryset = ComfamaSale.objects.select_related("reference", "reference__category")
    if filters.get("date_start"):
        queryset = queryset.filter(sale_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(sale_date__lte=filters["date_end"])
    return list(queryset)


def comfama_ad_metrics(filters):
    queryset = ComfamaAdMetric.objects.select_related("category")
    if filters.get("date_start"):
        queryset = queryset.filter(metric_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(metric_date__lte=filters["date_end"])
    return list(queryset)


def comfama_daily_ad_spends(filters):
    queryset = DailyAdSpend.objects.select_related("business_unit", "country", "ad_platform").filter(
        business_unit__slug="comfama-uva",
        country__code="CO",
        ad_platform__slug="meta-ads",
    )
    if filters.get("date_start"):
        queryset = queryset.filter(spend_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(spend_date__lte=filters["date_end"])
    return list(queryset)


def build_comfama_snapshot(filters):
    if filters.get("country") and filters["country"] != "CO":
        sales_rows = []
        ad_rows = []
        daily_spend_rows = []
    else:
        sales_rows = comfama_sales(filters)
        ad_rows = comfama_ad_metrics(filters)
        daily_spend_rows = comfama_daily_ad_spends(filters)
    sales_total = sum((row.sales_amount for row in sales_rows), ZERO)
    category_spend_total = sum((row.spend_amount for row in ad_rows), ZERO)
    daily_spend_total = sum((row.spend_amount for row in daily_spend_rows), ZERO)
    spend_total = daily_spend_total or category_spend_total
    conversations_total = sum((row.conversations for row in ad_rows), 0)
    purchases_total = len(sales_rows)
    average_cpl = _safe_ratio(spend_total, Decimal(conversations_total))
    roas = _safe_ratio(sales_total, spend_total)
    conversion_rate = (Decimal(purchases_total) / Decimal(conversations_total)) if conversations_total else ZERO
    average_ticket = _safe_ratio(sales_total, Decimal(purchases_total))

    category_rows = {}
    for sale in sales_rows:
        label = sale.reference.category.name
        category_rows.setdefault(label, {"label": label, "sales": ZERO, "spend": ZERO, "conversations": 0, "purchases": 0})
        category_rows[label]["sales"] += sale.sales_amount
        category_rows[label]["purchases"] += 1
    for row in ad_rows:
        label = row.category.name
        category_rows.setdefault(label, {"label": label, "sales": ZERO, "spend": ZERO, "conversations": 0, "purchases": 0})
        category_rows[label]["spend"] += row.spend_amount
        category_rows[label]["conversations"] += row.conversations

    categories = []
    for values in category_rows.values():
        values["sales"] = float(values["sales"])
        values["spend"] = float(values["spend"])
        values["cpl"] = round(float(Decimal(str(values["spend"])) / Decimal(values["conversations"])), 2) if values["conversations"] else 0
        values["roas"] = round(float(Decimal(str(values["sales"])) / Decimal(str(values["spend"]))), 2) if values["spend"] else 0
        values["conversion_rate"] = round(float(Decimal(values["purchases"]) / Decimal(values["conversations"])), 4) if values["conversations"] else 0
        values["average_ticket"] = round(float(Decimal(str(values["sales"])) / Decimal(values["purchases"])), 2) if values["purchases"] else 0
        categories.append(values)
    categories.sort(key=lambda item: item["sales"], reverse=True)

    daily = defaultdict(lambda: {"sales": ZERO, "spend": ZERO, "conversations": 0, "purchases": 0})
    for sale in sales_rows:
        daily[sale.sale_date.isoformat()]["sales"] += sale.sales_amount
        daily[sale.sale_date.isoformat()]["purchases"] += 1
    for row in ad_rows:
        daily[row.metric_date.isoformat()]["conversations"] += row.conversations

    if daily_spend_rows:
        for row in daily_spend_rows:
            daily[row.spend_date.isoformat()]["spend"] += row.spend_amount
    else:
        for row in ad_rows:
            daily[row.metric_date.isoformat()]["spend"] += row.spend_amount
    daily_series = [
        {
            "label": label,
            "sales": float(values["sales"]),
            "spend": float(values["spend"]),
            "conversations": values["conversations"],
            "purchases": values["purchases"],
            "cpl": round(float(values["spend"] / Decimal(values["conversations"])), 2) if values["conversations"] else None,
            "average_ticket": round(float(values["sales"] / Decimal(values["purchases"])), 2) if values["purchases"] else 0,
        }
        for label, values in sorted(daily.items(), key=lambda item: item[0])
    ]

    insights = []
    if categories:
        top_sales = categories[0]
        insights.append(f"{top_sales['label']} lidera ventas Comfama con {format_cop(top_sales['sales'])}.")
        top_roas = max(categories, key=lambda item: item["roas"])
        if top_roas["roas"]:
            insights.append(f"{top_roas['label']} tiene el mejor ROAS del modulo Comfama: {top_roas['roas']:.2f}.")
    if conversations_total:
        insights.append(f"Meta Ads genero {conversations_total} conversaciones con CPL promedio de {format_cop(average_cpl)}.")
    if conversion_rate:
        insights.append(f"La conversion estimada de conversacion a compra es {float(conversion_rate) * 100:.1f}%.")

    return {
        "kpis": {
            "sales_total": float(sales_total),
            "spend_total": float(spend_total),
            "conversations": conversations_total,
            "purchases": purchases_total,
            "average_cpl": round(float(average_cpl), 2) if average_cpl else 0,
            "roas": round(float(roas), 2) if roas else 0,
            "conversion_rate": round(float(conversion_rate), 4) if conversion_rate else 0,
            "average_ticket": round(float(average_ticket), 2) if average_ticket else 0,
        },
        "categories": categories,
        "daily_series": daily_series,
        "insights": insights[:4],
        "insight_cards": _decorate_insights(insights[:4]),
    }


def awn_international_metrics(filters):
    queryset = AwnInternationalFollowerMetric.objects.select_related("country")
    if filters.get("date_start"):
        queryset = queryset.filter(metric_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(metric_date__lte=filters["date_end"])
    if filters.get("country"):
        queryset = queryset.filter(country__code=filters["country"])
    return list(queryset)


def build_awn_international_snapshot(filters):
    valid_countries = {"EC", "MX"}
    if filters.get("country") and filters["country"] not in valid_countries:
        rows = []
    else:
        rows = awn_international_metrics(filters)

    visits_total = sum((row.instagram_profile_visits for row in rows), 0)
    followers_total = sum((row.new_followers for row in rows), 0)
    spend_total = sum((row.spend_amount for row in rows), ZERO)
    average_cpr = _safe_ratio(spend_total, Decimal(visits_total))
    average_cps = _safe_ratio(spend_total, Decimal(followers_total))
    follow_rate = _safe_ratio(Decimal(followers_total), Decimal(visits_total))

    country_rows = defaultdict(lambda: {"visits": 0, "followers": 0, "spend": ZERO})
    daily_rows = defaultdict(lambda: {"visits": 0, "followers": 0, "spend": ZERO})
    for row in rows:
        country_rows[row.country.name]["visits"] += row.instagram_profile_visits
        country_rows[row.country.name]["followers"] += row.new_followers
        country_rows[row.country.name]["spend"] += row.spend_amount
        daily_rows[row.metric_date.isoformat()]["visits"] += row.instagram_profile_visits
        daily_rows[row.metric_date.isoformat()]["followers"] += row.new_followers
        daily_rows[row.metric_date.isoformat()]["spend"] += row.spend_amount

    countries = []
    for label, values in sorted(country_rows.items(), key=lambda item: item[1]["followers"], reverse=True):
        cps = _safe_ratio(values["spend"], Decimal(values["followers"]))
        cpr = _safe_ratio(values["spend"], Decimal(values["visits"]))
        follow_rate_country = _safe_ratio(Decimal(values["followers"]), Decimal(values["visits"]))
        countries.append(
            {
                "label": label,
                "visits": values["visits"],
                "followers": values["followers"],
                "spend": float(values["spend"]),
                "cpr": round(float(cpr), 2) if cpr else 0,
                "cps": round(float(cps), 2) if cps else 0,
                "follow_rate": round(float(follow_rate_country), 4) if follow_rate_country else 0,
            }
        )

    daily_series = []
    for label, values in sorted(daily_rows.items(), key=lambda item: item[0]):
        cps = _safe_ratio(values["spend"], Decimal(values["followers"]))
        cpr = _safe_ratio(values["spend"], Decimal(values["visits"]))
        daily_series.append(
            {
                "label": label,
                "visits": values["visits"],
                "followers": values["followers"],
                "spend": float(values["spend"]),
                "cpr": round(float(cpr), 2) if cpr else 0,
                "cps": round(float(cps), 2) if cps else 0,
            }
        )

    insights = []
    if countries:
        top_followers = countries[0]
        insights.append(f"{top_followers['label']} lidera la captacion con {top_followers['followers']} seguidores nuevos.")
        most_efficient = min([item for item in countries if item["followers"] > 0], key=lambda item: item["cps"], default=None)
        if most_efficient:
            insights.append(f"{most_efficient['label']} tiene el mejor costo por seguidor: {format_cop(most_efficient['cps'])}.")
    if visits_total:
        insights.append(f"La tasa de conversion de visita a seguidor es {float(follow_rate) * 100:.1f}% en el rango filtrado.")
    if spend_total:
        insights.append(f"La inversion acumulada en seguidores suma {format_cop(spend_total)} entre Ecuador y Mexico.")

    return {
        "kpis": {
            "profile_visits": visits_total,
            "new_followers": followers_total,
            "spend_total": float(spend_total),
            "average_cpr": round(float(average_cpr), 2) if average_cpr else 0,
            "average_cps": round(float(average_cps), 2) if average_cps else 0,
            "follow_rate": round(float(follow_rate), 4) if follow_rate else 0,
        },
        "countries": countries,
        "daily_series": daily_series,
        "insights": insights[:4],
        "insight_cards": _decorate_insights(insights[:4]),
    }


def bali_daily_metrics(filters):
    queryset = BaliDailyMetric.objects.select_related("business_unit", "country")
    if filters.get("date_start"):
        queryset = queryset.filter(metric_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(metric_date__lte=filters["date_end"])
    return list(queryset.filter(business_unit__slug="bali").order_by("metric_date"))


def bali_whatsapp_sales(filters):
    queryset = DailyChannelSale.objects.select_related("business_unit", "country", "channel")
    if filters.get("date_start"):
        queryset = queryset.filter(sale_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(sale_date__lte=filters["date_end"])
    return list(queryset.filter(business_unit__slug="bali", channel__slug="bali-whatsapp").order_by("sale_date"))


def bali_physical_store_sales(filters):
    queryset = DailyChannelSale.objects.select_related("business_unit", "country", "channel")
    if filters.get("date_start"):
        queryset = queryset.filter(sale_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(sale_date__lte=filters["date_end"])
    return list(queryset.filter(business_unit__slug="bali", channel__slug="bali-tienda-fisica").order_by("sale_date"))


def bali_community_webcam_metrics(filters):
    queryset = BaliCommunityWebcamMetric.objects.select_related("business_unit", "country")
    if filters.get("date_start"):
        queryset = queryset.filter(metric_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(metric_date__lte=filters["date_end"])
    return list(queryset.filter(business_unit__slug="bali", country__code="CO").order_by("metric_date"))


def bali_web_product_metrics(filters):
    queryset = BaliWebProductDailyMetric.objects.select_related("business_unit", "country")
    if filters.get("date_start"):
        queryset = queryset.filter(metric_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(metric_date__lte=filters["date_end"])
    return list(queryset.filter(business_unit__slug="bali", country__code="CO"))


def _marketplace_inventory_key(marketplace):
    marketplace = normalize_text(marketplace or "mercadolibre")
    return {
        "mercado-libre": "mercadolibre",
        "mercadolibre": "mercadolibre",
        "falabella": "falabella",
    }.get(marketplace, marketplace)


def _marketplace_channel_slug(marketplace):
    marketplace = _marketplace_inventory_key(marketplace)
    return {"mercadolibre": "mercado-libre", "falabella": "falabella"}.get(marketplace, marketplace)


def _marketplace_product_share(item, marketplace):
    qs = MarketplaceProductInventory.objects.filter(marketplace=marketplace)
    total_sold = sum((row.sold_quantity or 0 for row in qs), 0)
    if total_sold > 0:
        return Decimal(item.sold_quantity or 0) / Decimal(total_sold)
    total_items = qs.count() or 1
    return Decimal("1") / Decimal(total_items)


def _shopify_reference_for_marketplace_product(filters, item):
    title_tokens = [token for token in normalize_text(item.title).split("-") if len(token) >= 4]
    if not title_tokens:
        return None
    bali_filters = dict(filters)
    bali_filters["business_unit"] = "bali"
    bali_filters["country"] = "CO"
    rows = bali_web_product_metrics(bali_filters)
    scored = []
    for row in rows:
        normalized = normalize_text(row.product_title)
        score = sum(1 for token in title_tokens if token in normalized)
        if score:
            scored.append((score, row.product_title, row))
    if not scored:
        return None
    best_title = sorted(scored, key=lambda item_score: item_score[0], reverse=True)[0][1]
    matched_rows = [row for row in rows if row.product_title == best_title]
    sales_total = sum((row.total_sales for row in matched_rows), ZERO)
    units_total = sum((row.net_items_sold for row in matched_rows), 0)
    return {"title": best_title, "sales": sales_total, "units": units_total}


def build_marketplace_product_detail(filters, marketplace, item_id):
    marketplace = _marketplace_inventory_key(marketplace)
    channel_slug = _marketplace_channel_slug(marketplace)
    try:
        item = MarketplaceProductInventory.objects.get(marketplace=marketplace, item_id=item_id)
    except MarketplaceProductInventory.DoesNotExist:
        return None

    product_share = _marketplace_product_share(item, marketplace)
    marketplace_filters = dict(filters)
    marketplace_filters["business_unit"] = "marketplace"
    marketplace_filters["channel"] = channel_slug
    sales_rows = [row for row in daily_channel_sales(marketplace_filters) if normalize_text(row.channel.slug) == channel_slug]
    spend_rows = daily_ad_spends(marketplace_filters)
    spend_by_date = defaultdict(lambda: ZERO)
    for row in spend_rows:
        spend_by_date[row.spend_date] += row.spend_amount or ZERO

    daily_map = defaultdict(lambda: {"sales": ZERO, "spend": ZERO, "orders": 0, "units": 0})
    for row in sales_rows:
        values = daily_map[row.sale_date]
        values["sales"] += (row.sales_amount or ZERO) * product_share
        values["orders"] += int(round((row.order_count or 0) * float(product_share)))
        values["units"] += int(round((row.units or 0) * float(product_share)))
    for metric_date, spend in spend_by_date.items():
        daily_map[metric_date]["spend"] += spend * product_share

    daily_series = []
    for metric_date, values in sorted(daily_map.items(), key=lambda entry: entry[0]):
        if not values["units"] and values["orders"]:
            values["units"] = values["orders"]
        roas = _safe_ratio(values["sales"], values["spend"])
        ticket = _safe_ratio(values["sales"], Decimal(values["orders"] or values["units"] or 0))
        daily_series.append(
            {
                "label": metric_date.isoformat(),
                "sales": float(values["sales"]),
                "spend": float(values["spend"]),
                "units": values["units"],
                "orders": values["orders"],
                "roas": round(float(roas), 2) if roas else 0,
                "average_ticket": round(float(ticket), 2) if ticket else 0,
            }
        )

    sales_total = sum((Decimal(str(row["sales"])) for row in daily_series), ZERO)
    spend_total = sum((Decimal(str(row["spend"])) for row in daily_series), ZERO)
    units_total = sum((row["units"] for row in daily_series), 0) or (item.sold_quantity or 0)
    orders_total = sum((row["orders"] for row in daily_series), 0)
    roas = _safe_ratio(sales_total, spend_total)
    ticket = _safe_ratio(sales_total, Decimal(orders_total or units_total or 0))
    inventory_value = (item.price or ZERO) * Decimal(item.available_quantity or 0)
    best_day = max(daily_series, key=lambda row: row["sales"], default=None)
    shopify_reference = _shopify_reference_for_marketplace_product(filters, item)

    insights = []
    insights.append(f"{item.title} tiene {item.available_quantity} unidades disponibles y {item.sold_quantity} vendidas registradas en {item.marketplace.title()}.")
    if sales_total:
        insights.append(f"La venta atribuida del periodo es {format_cop(sales_total)}, con ROAS estimado de {float(roas):.2f}.")
    if spend_total:
        insights.append(f"La inversion Google Ads atribuida al producto es {format_cop(spend_total)} segun la participacion del producto dentro del canal.")
    if item.health_status != MarketplaceProductInventory.HealthStatus.OK:
        messages = ", ".join(item.warning_messages or []) or "requiere revision de publicacion"
        insights.append(f"Prioridad operativa: {messages}.")
    if item.available_quantity <= 0:
        insights.append("Sin stock disponible: no conviene empujar pauta hasta recuperar inventario o pausar la publicacion.")
    elif item.sold_quantity and item.available_quantity < max(3, item.sold_quantity * 0.15):
        insights.append("Stock sensible frente a ventas historicas: revisar reposicion antes de escalar presupuesto.")
    if best_day:
        insights.append(f"El mejor dia atribuido fue {best_day['label']} con {format_cop(best_day['sales'])}.")
    if shopify_reference:
        insights.append(f"Referencia Shopify relacionada: {shopify_reference['title']} suma {format_cop(shopify_reference['sales'])} y {shopify_reference['units']} unidades en Bali.")
    else:
        insights.append("No hay match claro contra Shopify por nombre/SKU; conviene normalizar SKU/GTIN para comparar Marketplace vs Shopify producto a producto.")

    return {
        "type": "marketplace",
        "key": item.item_id,
        "title": item.title,
        "subtitle": f"Detalle de producto {item.marketplace.title()}",
        "image_url": item.thumbnail_url,
        "daily_series": daily_series,
        "stats": [
            {"label": "Ventas atribuidas", "value": float(sales_total), "kind": "money"},
            {"label": "Inversion Google", "value": float(spend_total), "kind": "money"},
            {"label": "ROAS estimado", "value": round(float(roas), 2) if roas else 0, "kind": "ratio"},
            {"label": "Ticket estimado", "value": round(float(ticket), 2) if ticket else 0, "kind": "money"},
            {"label": "Disponible", "value": item.available_quantity or 0, "kind": "number"},
            {"label": "Vendido", "value": item.sold_quantity or 0, "kind": "number"},
            {"label": "Precio", "value": float(item.price or 0), "kind": "money"},
            {"label": "Valor inventario", "value": float(inventory_value), "kind": "money"},
            {"label": "Estado", "value": item.status or "Sin estado", "kind": "text"},
            {"label": "Revision", "value": item.get_health_status_display(), "kind": "text"},
        ],
        "insights": insights[:7],
        "allocation_note": "Marketplace aun no guarda ventas por SKU diario; Axis atribuye ventas e inversion del canal proporcionalmente al vendido acumulado de cada publicacion y cruza Shopify por coincidencia de nombre cuando existe.",
    }


def build_bali_product_detail(filters, product_title):
    product_title = str(product_title or "").strip()
    if not product_title:
        return None

    product_rows = [row for row in bali_web_product_metrics(filters) if row.product_title == product_title]
    if not product_rows:
        return None

    all_product_rows = bali_web_product_metrics(filters)
    product_sales_by_date = defaultdict(lambda: ZERO)
    all_sales_by_date = defaultdict(lambda: ZERO)
    for row in all_product_rows:
        all_sales_by_date[row.metric_date] += row.total_sales
    for row in product_rows:
        product_sales_by_date[row.metric_date] += row.total_sales

    metric_rows = bali_daily_metrics(filters)
    spend_by_date = {row.metric_date: row.google_spend_amount for row in metric_rows}
    web_sales_by_date = {row.metric_date: row.web_sales_amount for row in metric_rows}

    image_url = ""
    daily_map = defaultdict(
        lambda: {
            "units": 0,
            "gross_sales": ZERO,
            "discounts": ZERO,
            "returns": ZERO,
            "net_sales": ZERO,
            "sales": ZERO,
            "spend": ZERO,
        }
    )
    for row in product_rows:
        values = daily_map[row.metric_date]
        values["units"] += row.net_items_sold
        values["gross_sales"] += row.gross_sales
        values["discounts"] += row.discounts
        values["returns"] += row.returns
        values["net_sales"] += row.net_sales
        values["sales"] += row.total_sales
        if row.product_image_url and not image_url:
            image_url = row.product_image_url

    for metric_date, values in daily_map.items():
        daily_product_total = all_sales_by_date.get(metric_date) or web_sales_by_date.get(metric_date) or ZERO
        if daily_product_total:
            values["spend"] = (spend_by_date.get(metric_date, ZERO) or ZERO) * values["sales"] / daily_product_total

    daily_series = []
    for metric_date, values in sorted(daily_map.items(), key=lambda item: item[0]):
        sales_with_vat = values["sales"]
        sales = sales_with_vat
        roas = _safe_ratio(sales, values["spend"])
        ticket = _safe_ratio(sales, Decimal(values["units"]))
        daily_series.append(
            {
                "label": metric_date.isoformat(),
                "sales": float(sales),
                "sales_with_vat": float(sales_with_vat),
                "gross_sales": float(values["gross_sales"]),
                "discounts": float(values["discounts"]),
                "returns": float(values["returns"]),
                "net_sales": float(values["net_sales"]),
                "spend": float(values["spend"]),
                "units": values["units"],
                "roas": round(float(roas), 2) if roas else 0,
                "average_ticket": round(float(ticket), 2) if ticket else 0,
            }
        )

    sales_total = sum((Decimal(str(item["sales"])) for item in daily_series), ZERO)
    sales_total_with_vat = sum((Decimal(str(item["sales_with_vat"])) for item in daily_series), ZERO)
    gross_sales = sum((Decimal(str(item["gross_sales"])) for item in daily_series), ZERO)
    discounts = sum((Decimal(str(item["discounts"])) for item in daily_series), ZERO)
    returns = sum((Decimal(str(item["returns"])) for item in daily_series), ZERO)
    net_sales = sum((Decimal(str(item["net_sales"])) for item in daily_series), ZERO)
    spend_total = sum((Decimal(str(item["spend"])) for item in daily_series), ZERO)
    units_total = sum((item["units"] for item in daily_series), 0)
    all_period_product_sales = sum(all_sales_by_date.values(), ZERO)
    roas = _safe_ratio(sales_total, spend_total)
    unit_value = _safe_ratio(sales_total, Decimal(units_total))
    sales_share = _safe_ratio(sales_total, all_period_product_sales)
    best_day = max(daily_series, key=lambda item: item["sales"], default=None)
    best_roas_day = max([item for item in daily_series if item["spend"] > 0], key=lambda item: item["roas"], default=None)

    insights = []
    insights.append(f"{product_title} vende {units_total} unidades y genera {format_cop(sales_total)} en el rango filtrado.")
    if spend_total:
        insights.append(f"La inversion atribuida es {format_cop(spend_total)} y el ROAS estimado del producto es {float(roas):.2f}.")
    if sales_share:
        insights.append(f"Representa el {float(sales_share) * 100:.1f}% de las ventas web con detalle de producto.")
    if best_day:
        insights.append(f"El mejor dia por ventas fue {best_day['label']} con {format_cop(best_day['sales'])}.")
    if discounts or returns:
        insights.append(f"Ajustes del periodo: {format_cop(discounts)} en descuentos y {format_cop(returns)} en devoluciones.")
    if spend_total and roas < Decimal("3"):
        insights.append("El ROAS estimado esta por debajo de 3.0; conviene revisar pauta, precio o disponibilidad.")
    elif best_roas_day and best_roas_day["roas"] >= 3:
        insights.append(f"El mejor dia por eficiencia fue {best_roas_day['label']} con ROAS {best_roas_day['roas']:.2f}.")

    return {
        "type": "bali",
        "key": product_title,
        "title": product_title,
        "subtitle": "Detalle de producto Web Shopify",
        "image_url": image_url,
        "daily_series": daily_series,
        "stats": [
            {"label": "Ventas totales", "value": float(sales_total), "kind": "money"},
            {"label": "Total Shopify", "value": float(sales_total_with_vat), "kind": "money"},
            {"label": "Inversion atribuida", "value": float(spend_total), "kind": "money"},
            {"label": "ROAS estimado", "value": round(float(roas), 2) if roas else 0, "kind": "ratio"},
            {"label": "Unidades", "value": units_total, "kind": "number"},
            {"label": "Valor por unidad", "value": round(float(unit_value), 2) if unit_value else 0, "kind": "money"},
            {"label": "Participacion web", "value": round(float(sales_share), 4) if sales_share else 0, "kind": "percent"},
            {"label": "Ventas netas", "value": float(net_sales), "kind": "money"},
            {"label": "Ventas brutas", "value": float(gross_sales), "kind": "money"},
            {"label": "Descuentos", "value": float(discounts), "kind": "money"},
            {"label": "Devoluciones", "value": float(returns), "kind": "money"},
        ],
        "insights": insights[:6],
        "allocation_note": "En Bali, Google Ads no llega por producto; la inversion se atribuye proporcionalmente a la participacion diaria de ventas del producto dentro del total Shopify.",
    }


def _distribute_integer_total(total, periods):
    if periods <= 0:
        return []
    base = total // periods
    remainder = total % periods
    return [base + (1 if index < remainder else 0) for index in range(periods)]


def _distribute_decimal_total(total, periods):
    if periods <= 0:
        return []
    cents = int((total * 100).quantize(Decimal("1")))
    distributed_cents = _distribute_integer_total(cents, periods)
    return [Decimal(value) / Decimal("100") for value in distributed_cents]


def build_bali_snapshot(filters, include_comparison=True):
    metrics = bali_daily_metrics(filters)
    whatsapp_rows = bali_whatsapp_sales(filters)
    physical_rows = bali_physical_store_sales(filters)
    community_rows = bali_community_webcam_metrics(filters)
    web_product_rows = bali_web_product_metrics(filters)

    daily_map = defaultdict(
        lambda: {
            "web_sales": ZERO,
            "whatsapp_sales": ZERO,
            "physical_sales": ZERO,
            "spend": ZERO,
            "sessions": 0,
            "web_orders": 0,
            "whatsapp_orders": 0,
            "google_orders": 0,
            "conversations": 0,
            "cpa": ZERO,
        }
    )

    for row in metrics:
        values = daily_map[row.metric_date.isoformat()]
        values["web_sales"] += row.web_sales_amount
        values["spend"] += row.google_spend_amount
        values["sessions"] += row.sessions
        values["web_orders"] += row.web_order_count
        values["google_orders"] += row.google_attributed_orders
        values["conversations"] += row.whatsapp_conversations
        values["cpa"] = row.cpa or values["cpa"]

    for row in whatsapp_rows:
        values = daily_map[row.sale_date.isoformat()]
        values["whatsapp_sales"] += row.sales_amount
        values["whatsapp_orders"] += row.order_count

    for row in physical_rows:
        values = daily_map[row.sale_date.isoformat()]
        values["physical_sales"] += row.sales_amount
        values["physical_visitors"] = values.get("physical_visitors", 0) + (row.units or 0)
        values["physical_orders"] = values.get("physical_orders", 0) + (row.order_count or 0)

    daily_series = []
    for label, values in sorted(daily_map.items(), key=lambda item: item[0]):
        web_sales = values["web_sales"]
        whatsapp_sales = remove_colombia_vat(values["whatsapp_sales"])
        physical_sales = values["physical_sales"]
        total_sales = web_sales + whatsapp_sales + physical_sales
        total_sales_with_vat = values["web_sales"] + values["whatsapp_sales"] + physical_sales
        physical_orders = values.get("physical_orders", 0)
        physical_visitors = values.get("physical_visitors", 0)
        total_orders = values["web_orders"] + values["whatsapp_orders"] + physical_orders
        daily_series.append(
            {
                "label": label,
                "web_sales": float(web_sales),
                "web_sales_with_vat": float(values["web_sales"]),
                "whatsapp_sales": float(whatsapp_sales),
                "whatsapp_sales_with_vat": float(values["whatsapp_sales"]),
                "physical_sales": float(physical_sales),
                "sales_total": float(total_sales),
                "sales_total_with_vat": float(total_sales_with_vat),
                "spend": float(values["spend"]),
                "sessions": values["sessions"],
                "visits": values["sessions"],
                "web_orders": values["web_orders"],
                "whatsapp_orders": values["whatsapp_orders"],
                "physical_orders": physical_orders,
                "physical_visitors": physical_visitors,
                "orders_total": total_orders,
                "google_orders": values["google_orders"],
                "conversations": values["conversations"],
                "roas": round(float(total_sales / values["spend"]), 2) if values["spend"] else 0,
                "average_ticket": round(float(_safe_ratio(total_sales, Decimal(total_orders))), 2) if total_orders else 0,
                "conversion_rate": round(float(_safe_ratio(Decimal(values["web_orders"]), Decimal(values["sessions"]))), 4) if values["sessions"] else 0,
                "whatsapp_conversion_rate": round(float(_safe_ratio(Decimal(values["whatsapp_orders"]), Decimal(values["conversations"]))), 4) if values["conversations"] else 0,
                "physical_conversion_rate": round(float(_safe_ratio(Decimal(physical_orders), Decimal(physical_visitors))), 4) if physical_visitors else 0,
                "physical_average_ticket": round(float(_safe_ratio(physical_sales, Decimal(physical_orders))), 2) if physical_orders else 0,
                "physical_sales_per_visitor": round(float(_safe_ratio(physical_sales, Decimal(physical_visitors))), 2) if physical_visitors else 0,
                "cpa": round(float(values["cpa"]), 2) if values["cpa"] else 0,
            }
        )

    web_sales_total_with_vat = sum((row.web_sales_amount for row in metrics), ZERO)
    whatsapp_sales_total_with_vat = sum((row.sales_amount for row in whatsapp_rows), ZERO)
    web_sales_total = web_sales_total_with_vat
    whatsapp_sales_total = remove_colombia_vat(whatsapp_sales_total_with_vat)
    physical_sales_total = sum((row.sales_amount for row in physical_rows), ZERO)
    sales_total = web_sales_total + whatsapp_sales_total + physical_sales_total
    sales_total_with_vat = web_sales_total_with_vat + whatsapp_sales_total_with_vat + physical_sales_total
    spend_total = sum((row.google_spend_amount for row in metrics), ZERO)
    sessions_total = sum((row.sessions for row in metrics), 0)
    web_orders_total = sum((row.web_order_count for row in metrics), 0)
    whatsapp_orders_total = sum((row.order_count for row in whatsapp_rows), 0)
    physical_orders_total = sum((row.order_count for row in physical_rows), 0)
    physical_visitors_total = sum((row.units for row in physical_rows), 0)
    orders_total = web_orders_total + whatsapp_orders_total + physical_orders_total
    google_orders_total = sum((row.google_attributed_orders for row in metrics), 0)
    conversations_total = sum((row.whatsapp_conversations for row in metrics), 0)
    web_sessions_measured = not (metrics and web_orders_total > 0 and sessions_total == 0)
    web_provisional_dates = sorted(
        {
            row.metric_date
            for row in metrics
            if (
                "orders-api" in str(row.source_file or "").lower()
                and "shopifyql" not in str(row.source_file or "").lower()
            )
            or (row.web_order_count > 0 and row.sessions == 0)
        }
    )
    web_analytics_provisional = bool(web_provisional_dates)
    if not web_sessions_measured:
        web_analytics_provisional = True
    web_analytics_mixed = bool(web_provisional_dates) and len(web_provisional_dates) < len(metrics)
    average_ticket = _safe_ratio(sales_total, Decimal(orders_total))
    average_ticket_with_vat = _safe_ratio(sales_total_with_vat, Decimal(orders_total))
    web_average_ticket = _safe_ratio(web_sales_total, Decimal(web_orders_total))
    web_average_ticket_with_vat = _safe_ratio(web_sales_total_with_vat, Decimal(web_orders_total))
    whatsapp_average_ticket = _safe_ratio(whatsapp_sales_total, Decimal(whatsapp_orders_total))
    whatsapp_average_ticket_with_vat = _safe_ratio(whatsapp_sales_total_with_vat, Decimal(whatsapp_orders_total))
    roas = _safe_ratio(sales_total, spend_total)
    web_roas = _safe_ratio(web_sales_total, spend_total)
    conversion_rate = _safe_ratio(Decimal(web_orders_total), Decimal(sessions_total))
    whatsapp_conversion_rate = _safe_ratio(Decimal(whatsapp_orders_total), Decimal(conversations_total))
    physical_conversion_rate = _safe_ratio(Decimal(physical_orders_total), Decimal(physical_visitors_total))
    physical_average_ticket = _safe_ratio(physical_sales_total, Decimal(physical_orders_total))
    physical_sales_per_visitor = _safe_ratio(physical_sales_total, Decimal(physical_visitors_total))
    overall_conversion_rate = _safe_ratio(Decimal(orders_total), Decimal(sessions_total + conversations_total + physical_visitors_total))
    average_cpa = _safe_ratio(spend_total, Decimal(google_orders_total))
    period_days = len(daily_series)
    average_daily_orders = _safe_ratio(Decimal(orders_total), Decimal(period_days))
    average_daily_sales = _safe_ratio(sales_total, Decimal(period_days))
    average_daily_sales_with_vat = _safe_ratio(sales_total_with_vat, Decimal(period_days))
    web_products_map = defaultdict(
        lambda: {
            "units": 0,
            "gross_sales": ZERO,
            "discounts": ZERO,
            "returns": ZERO,
            "net_sales": ZERO,
            "total_sales": ZERO,
            "image_url": "",
        }
    )
    for row in web_product_rows:
        product = web_products_map[row.product_title]
        product["units"] += row.net_items_sold
        product["gross_sales"] += row.gross_sales
        product["discounts"] += row.discounts
        product["returns"] += row.returns
        product["net_sales"] += row.net_sales
        product["total_sales"] += row.total_sales
        if row.product_image_url and not product["image_url"]:
            product["image_url"] = row.product_image_url
    top_web_products = []
    for index, (title, values) in enumerate(
        sorted(
            web_products_map.items(),
            key=lambda item: (item[1]["units"], item[1]["total_sales"]),
            reverse=True,
        )[:20],
        start=1,
    ):
        total_sales_with_vat = values["total_sales"]
        total_sales = total_sales_with_vat
        share = _safe_ratio(total_sales, web_sales_total)
        unit_value = _safe_ratio(total_sales, Decimal(values["units"])) if values["units"] else ZERO
        top_web_products.append(
            {
                "rank": index,
                "title": title,
                "units": values["units"],
                "total_sales": float(total_sales),
                "total_sales_with_vat": float(total_sales_with_vat),
                "gross_sales": float(values["gross_sales"]),
                "discounts": float(values["discounts"]),
                "returns": float(values["returns"]),
                "net_sales": float(values["net_sales"]),
                "unit_value": round(float(unit_value), 2) if unit_value else 0,
                "sales_share": round(float(share), 4) if share else 0,
                "image_url": values["image_url"],
            }
        )

    channels = [
        {"label": "Web", "value": float(web_sales_total), "value_with_vat": float(web_sales_total_with_vat), "orders": web_orders_total},
        {"label": "WhatsApp", "value": float(whatsapp_sales_total), "value_with_vat": float(whatsapp_sales_total_with_vat), "orders": whatsapp_orders_total},
        {"label": "Tienda Fisica", "value": float(physical_sales_total), "orders": physical_orders_total},
    ]
    web_daily = [
        {
            "label": row.metric_date.isoformat(),
            "sales": float(row.web_sales_amount),
            "sales_with_vat": float(row.web_sales_amount),
            "orders": row.web_order_count,
            "sessions": row.sessions,
            "visits": row.sessions,
            "spend": float(row.google_spend_amount),
            "google_orders": row.google_attributed_orders,
            "conversations": row.whatsapp_conversations,
            "roas": round(float(_safe_ratio(row.web_sales_amount, row.google_spend_amount)), 2) if row.google_spend_amount else 0,
            "conversion_rate": round(float(_safe_ratio(Decimal(row.web_order_count), Decimal(row.sessions))), 4) if row.sessions else 0,
            "average_ticket": round(float(_safe_ratio(row.web_sales_amount, Decimal(row.web_order_count))), 2) if row.web_order_count else 0,
            "cpa": round(float(row.cpa), 2) if row.cpa else 0,
        }
        for row in metrics
    ]
    whatsapp_daily = [
        {
            "label": row.sale_date.isoformat(),
            "sales": float(remove_colombia_vat(row.sales_amount)),
            "sales_with_vat": float(row.sales_amount),
            "orders": row.order_count,
            "average_ticket": round(float(_safe_ratio(remove_colombia_vat(row.sales_amount), Decimal(row.order_count))), 2) if row.order_count else 0,
        }
        for row in whatsapp_rows
    ]
    physical_daily = [
        {
            "label": row.sale_date.isoformat(),
            "sales": float(row.sales_amount),
            "visitors": row.units or 0,
            "orders": row.order_count or 0,
            "conversion_rate": round(float(_safe_ratio(Decimal(row.order_count or 0), Decimal(row.units or 0))), 4) if row.units else 0,
            "average_ticket": round(float(_safe_ratio(row.sales_amount, Decimal(row.order_count or 0))), 2) if row.order_count else 0,
            "sales_per_visitor": round(float(_safe_ratio(row.sales_amount, Decimal(row.units or 0))), 2) if row.units else 0,
        }
        for row in physical_rows
    ]
    community_daily = [
        {
            "label": row.metric_date.isoformat(),
            "new_subscribers": row.new_subscribers,
            "subscribers": row.subscribers,
        }
        for row in community_rows
    ]
    community_latest = community_rows[-1] if community_rows else None
    community_total_new = sum((row.new_subscribers for row in community_rows), 0)
    community_days = len(community_rows)
    community_average_new = _safe_ratio(Decimal(community_total_new), Decimal(community_days))
    community_story_url = ""
    latest_story_metric = (
        BaliCommunityWebcamMetric.objects.select_related("business_unit", "country")
        .filter(business_unit__slug="bali", country__code="CO")
        .exclude(story_screenshot="")
        .order_by("-updated_at", "-metric_date")
        .first()
    )
    if latest_story_metric and latest_story_metric.story_screenshot:
        try:
            storage = latest_story_metric.story_screenshot.storage
            if not hasattr(storage, "exists") or storage.exists(latest_story_metric.story_screenshot.name):
                community_story_url = latest_story_metric.story_screenshot.url
        except (OSError, ValueError):
            community_story_url = ""
    community_insights = []
    if community_latest:
        community_insights.append(f"Comunidad Webcam cierra el rango con {community_latest.subscribers} suscritos acumulados.")
    if community_total_new:
        community_insights.append(f"En el periodo filtrado entraron {community_total_new} suscritos nuevos, sin inversion publicitaria.")
    if community_days:
        community_insights.append(f"El promedio organico diario es de {float(community_average_new):.1f} suscritos nuevos.")
    if not community_story_url:
        community_insights.append("Puedes cargar el pantallazo story 9:16 desde el admin para verlo junto a la grafica.")

    summary_insights = []
    if sales_total:
        summary_insights.append(f"Bali acumula {format_cop(sales_total)} en el rango filtrado, con ROAS consolidado de {float(roas):.2f}.")
    if web_sessions_measured and (sessions_total or conversations_total):
        summary_insights.append(f"La conversion general de Bali entre Web y WhatsApp es {float(overall_conversion_rate) * 100:.2f}% en el rango filtrado.")
    if period_days:
        summary_insights.append(f"El promedio diario es {format_cop(average_daily_sales)} y {float(average_daily_orders):.1f} pedidos por dia.")
    if physical_rows:
        summary_insights.append(f"Tienda Fisica aporta {format_cop(physical_sales_total)} en ventas del periodo.")

    web_insights = []
    if web_sales_total:
        web_insights.append(f"La web de Bali genera {format_cop(web_sales_total)} y {web_orders_total} pedidos en el rango filtrado.")
    if web_sessions_measured and sessions_total:
        web_insights.append(f"La conversion web del periodo es {float(conversion_rate) * 100:.2f}% sobre {sessions_total} sesiones.")
    if spend_total:
        web_insights.append(f"Google Ads invierte {format_cop(spend_total)} para sostener el canal web, con ROAS de {float(web_roas):.2f}.")
    if web_orders_total:
        web_insights.append(f"El ticket promedio web se ubica en {format_cop(web_average_ticket)}.")

    web_recommendations = []
    top_web_product = top_web_products[0] if top_web_products else None
    if spend_total and web_roas < Decimal("2.5"):
        web_recommendations.append(
            {
                "title": "Revisar eficiencia Google Ads",
                "message": f"La inversion web es {format_cop(spend_total)} y el ROAS Shopify queda en {float(web_roas):.2f}.",
                "action": "Cruzar terminos/campanas con productos de mayor ticket antes de subir presupuesto.",
            }
        )
    elif spend_total and web_roas >= Decimal("3"):
        web_recommendations.append(
            {
                "title": "Escalar con control",
                "message": f"El canal web sostiene ROAS {float(web_roas):.2f} con {web_orders_total} pedidos.",
                "action": "Priorizar presupuesto en campanas que empujen los productos top de Shopify.",
            }
        )
    if top_web_product:
        web_recommendations.append(
            {
                "title": "Producto ancla Shopify",
                "message": f"{top_web_product['title']} lidera con {format_cop(top_web_product['total_sales'])} y {top_web_product['units']} unidades.",
                "action": "Usarlo como referencia para copies, bundles o campanas de remarketing.",
            }
        )
    no_order_spend_days = [row for row in web_daily if row.get("spend") and not row.get("orders")]
    if no_order_spend_days:
        web_recommendations.append(
            {
                "title": "Dias con inversion sin pedidos",
                "message": f"Hay {len(no_order_spend_days)} dias con gasto Google Ads y cero pedidos Shopify.",
                "action": "Revisar disponibilidad, landing y busquedas de esos dias antes de repetir pauta.",
            }
        )
    if web_sessions_measured and sessions_total and conversion_rate < Decimal("0.01"):
        web_recommendations.append(
            {
                "title": "Conversion web baja",
                "message": f"La web convierte {float(conversion_rate) * 100:.2f}% sobre {sessions_total} sesiones.",
                "action": "Auditar friccion de checkout y coherencia producto-anuncio en las campanas activas.",
            }
        )

    whatsapp_insights = []
    if whatsapp_sales_total:
        whatsapp_insights.append(f"WhatsApp aporta {format_cop(whatsapp_sales_total)} y {whatsapp_orders_total} pedidos en el periodo.")
    if conversations_total:
        whatsapp_insights.append(f"WhatsApp registra {conversations_total} conversaciones y convierte {float(whatsapp_conversion_rate) * 100:.2f}% a pedido.")
    if whatsapp_orders_total:
        whatsapp_insights.append(f"El ticket promedio de WhatsApp es {format_cop(whatsapp_average_ticket)}.")
    if whatsapp_rows:
        whatsapp_insights.append("Las ventas historicas de WhatsApp quedaron prorrateadas por dia para conservar consistencia en el dashboard.")

    physical_insights = []
    if physical_sales_total:
        physical_insights.append(f"Tienda Fisica aporta {format_cop(physical_sales_total)} con {physical_orders_total} pedidos en el periodo.")
    if physical_visitors_total:
        physical_insights.append(f"Tienda Fisica convierte {float(physical_conversion_rate) * 100:.2f}% sobre {physical_visitors_total} visitantes registrados.")
    if physical_orders_total:
        physical_insights.append(f"El ticket promedio de tienda es {format_cop(physical_average_ticket)} y la venta por visitante llega a {format_cop(physical_sales_per_visitor)}.")

    snapshot = {
        "data_quality": {
            "web_sessions_measured": web_sessions_measured,
            "web_analytics_provisional": web_analytics_provisional,
            "web_analytics_mixed": web_analytics_mixed,
            "web_provisional_dates": [row.strftime("%d/%m/%Y") for row in web_provisional_dates],
        },
        "kpis": {
            "sales_total": float(sales_total),
            "sales_total_with_vat": float(sales_total_with_vat),
            "web_sales_total": float(web_sales_total),
            "web_sales_total_with_vat": float(web_sales_total_with_vat),
            "whatsapp_sales_total": float(whatsapp_sales_total),
            "whatsapp_sales_total_with_vat": float(whatsapp_sales_total_with_vat),
            "physical_sales_total": float(physical_sales_total),
            "physical_visitors_total": physical_visitors_total,
            "physical_orders_total": physical_orders_total,
            "physical_conversion_rate": round(float(physical_conversion_rate), 4) if physical_conversion_rate else 0,
            "physical_average_ticket": round(float(physical_average_ticket), 2) if physical_average_ticket else 0,
            "physical_sales_per_visitor": round(float(physical_sales_per_visitor), 2) if physical_sales_per_visitor else 0,
            "spend_total": float(spend_total),
            "sessions_total": sessions_total,
            "visits_total": sessions_total,
            "orders_total": orders_total,
            "web_orders_total": web_orders_total,
            "whatsapp_orders_total": whatsapp_orders_total,
            "google_orders_total": google_orders_total,
            "conversations_total": conversations_total,
            "average_ticket": round(float(average_ticket), 2) if average_ticket else 0,
            "average_ticket_with_vat": round(float(average_ticket_with_vat), 2) if average_ticket_with_vat else 0,
            "web_average_ticket": round(float(web_average_ticket), 2) if web_average_ticket else 0,
            "web_average_ticket_with_vat": round(float(web_average_ticket_with_vat), 2) if web_average_ticket_with_vat else 0,
            "whatsapp_average_ticket": round(float(whatsapp_average_ticket), 2) if whatsapp_average_ticket else 0,
            "whatsapp_average_ticket_with_vat": round(float(whatsapp_average_ticket_with_vat), 2) if whatsapp_average_ticket_with_vat else 0,
            "conversion_rate": round(float(conversion_rate), 4) if conversion_rate else 0,
            "whatsapp_conversion_rate": round(float(whatsapp_conversion_rate), 4) if whatsapp_conversion_rate else 0,
            "overall_conversion_rate": round(float(overall_conversion_rate), 4) if overall_conversion_rate else 0,
            "average_daily_orders": round(float(average_daily_orders), 2) if average_daily_orders else 0,
            "average_daily_sales": round(float(average_daily_sales), 2) if average_daily_sales else 0,
            "average_daily_sales_with_vat": round(float(average_daily_sales_with_vat), 2) if average_daily_sales_with_vat else 0,
            "roas": round(float(roas), 2) if roas else 0,
            "web_roas": round(float(web_roas), 2) if web_roas else 0,
            "average_cpa": round(float(average_cpa), 2) if average_cpa else 0,
            "days_count": period_days,
        },
        "daily_series": daily_series,
        "web_daily": web_daily,
        "top_web_products": top_web_products,
        "whatsapp_daily": whatsapp_daily,
        "physical_daily": physical_daily,
        "channels": channels,
        "insights": summary_insights[:5],
        "insight_cards": _decorate_insights(summary_insights[:5]),
        "web_insights": web_insights[:4],
        "web_insight_cards": _decorate_insights(web_insights[:4]),
        "web_recommendations": web_recommendations[:4],
        "whatsapp_insights": whatsapp_insights[:4],
        "whatsapp_insight_cards": _decorate_insights(whatsapp_insights[:4]),
        "physical_insights": physical_insights[:4],
        "physical_insight_cards": _decorate_insights(physical_insights[:4]),
        "community": {
            "kpis": {
                "subscribers": community_latest.subscribers if community_latest else 0,
                "new_subscribers": community_total_new,
                "average_new_subscribers": round(float(community_average_new), 2) if community_average_new else 0,
                "days_count": community_days,
            },
            "daily_series": community_daily,
            "latest_story_url": community_story_url,
            "insights": community_insights[:4],
            "insight_cards": _decorate_insights(community_insights[:4]),
        },
    }
    snapshot["web_geo_map"] = build_bali_web_geo_map_data(filters, snapshot["kpis"])
    if include_comparison and filters.get("compare_mode") == "previous_period":
        comparison_filters = _previous_period_filters(filters)
        if comparison_filters:
            previous_snapshot = build_bali_snapshot(comparison_filters, include_comparison=False)
            snapshot["comparison"] = _summarize_bali_comparison(snapshot["kpis"], previous_snapshot["kpis"])
            snapshot["community"]["comparison"] = _summarize_bali_community_comparison(
                snapshot["community"]["kpis"],
                previous_snapshot["community"]["kpis"],
            )
            snapshot["comparison_range"] = {
                "date_start": comparison_filters["date_start"],
                "date_end": comparison_filters["date_end"],
            }
    snapshot.setdefault("comparison", {})
    snapshot["community"].setdefault("comparison", {})
    snapshot.setdefault("comparison_range", {})
    return snapshot


def import_bali_workbook(path):
    catalogs = ensure_bali_catalogs()
    platforms = ensure_ad_platform_catalogs()
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    stats = {
        "created_metrics": 0,
        "updated_metrics": 0,
        "created_whatsapp": 0,
        "updated_whatsapp": 0,
        "deleted_whatsapp": 0,
        "skipped": 0,
    }

    try:
        web_sheet = workbook["Hoja1"]
        whatsapp_sheet = workbook["Hoja2"]

        header_row = next(web_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header_map = {normalize_text(value): index for index, value in enumerate(header_row) if value not in (None, "")}
        session_header = "sesiones" if "sesiones" in header_map else ("visitas" if "visitas" in header_map else None)
        expected_headers = [
            "fecha",
            "ventas web",
            "pedidos web",
            "inversion google ads",
            "cpa",
            "conversaciones gads",
            "pedidos gads",
        ]
        missing_headers = [header for header in expected_headers if header not in header_map]
        if not session_header:
            missing_headers.append("sesiones")
        if missing_headers:
            raise ValueError(f"Hoja1 no tiene las columnas esperadas: {', '.join(missing_headers)}")

        metric_dates = []
        for row_number, row in enumerate(web_sheet.iter_rows(min_row=2, values_only=True), start=2):
            metric_date = parse_excel_date(row[header_map["fecha"]])
            if not metric_date:
                stats["skipped"] += 1
                continue
            metric_dates.append(metric_date)
            _, created = BaliDailyMetric.objects.update_or_create(
                business_unit=catalogs["business_unit"],
                country=catalogs["country"],
                metric_date=metric_date,
                defaults={
                    "sessions": parse_quantity(row[header_map[session_header]]),
                    "web_sales_amount": parse_decimal(row[header_map["ventas web"]]),
                    "web_order_count": parse_quantity(row[header_map["pedidos web"]]),
                    "google_spend_amount": parse_decimal(row[header_map["inversion google ads"]]),
                    "cpa": parse_decimal(row[header_map["cpa"]]),
                    "whatsapp_conversations": parse_quantity(row[header_map["conversaciones gads"]]),
                    "google_attributed_orders": parse_quantity(row[header_map["pedidos gads"]]),
                    "source_type": BaliDailyMetric.SourceType.IMPORTED,
                    "source_file": Path(path).name,
                    "source_row": row_number,
                    "notes": "",
                },
            )
            if created:
                stats["created_metrics"] += 1
            else:
                stats["updated_metrics"] += 1

        whatsapp_values = list(whatsapp_sheet.iter_rows(min_row=3, max_row=3, values_only=True))
        if whatsapp_values and metric_dates:
            total_orders = parse_quantity(whatsapp_values[0][0])
            total_sales = parse_decimal(whatsapp_values[0][1])
            metric_dates = sorted(set(metric_dates))
            distributed_orders = _distribute_integer_total(total_orders, len(metric_dates))
            distributed_sales = _distribute_decimal_total(total_sales, len(metric_dates))
            preserved_dates = set(metric_dates)

            stale_rows = DailyChannelSale.objects.filter(
                business_unit=catalogs["business_unit"],
                country=catalogs["country"],
                channel=catalogs["channels"]["bali-whatsapp"],
                source_file=Path(path).name,
            ).exclude(sale_date__in=preserved_dates)
            stats["deleted_whatsapp"] = stale_rows.count()
            stale_rows.delete()

            for sale_date, order_count, sales_amount in zip(metric_dates, distributed_orders, distributed_sales):
                _, created = DailyChannelSale.objects.update_or_create(
                    business_unit=catalogs["business_unit"],
                    country=catalogs["country"],
                    channel=catalogs["channels"]["bali-whatsapp"],
                    sale_date=sale_date,
                    defaults={
                        "sales_amount": sales_amount,
                        "order_count": order_count,
                        "source_type": DailyChannelSale.SourceType.IMPORTED,
                        "source_file": Path(path).name,
                        "notes": "Valores prorrateados automaticamente desde el consolidado mensual de WhatsApp Bali.",
                    },
                )
                if created:
                    stats["created_whatsapp"] += 1
                else:
                    stats["updated_whatsapp"] += 1

        for metric in BaliDailyMetric.objects.filter(business_unit=catalogs["business_unit"], country=catalogs["country"], source_file=Path(path).name):
            DailyAdSpend.objects.update_or_create(
                business_unit=catalogs["business_unit"],
                country=catalogs["country"],
                ad_platform=platforms["google-ads"],
                spend_date=metric.metric_date,
                defaults={
                    "spend_amount": metric.google_spend_amount,
                    "source_type": DailyAdSpend.SourceType.IMPORTED,
                    "source_file": Path(path).name,
                    "notes": "Sincronizado automaticamente desde BaliDailyMetric.",
                },
            )
    finally:
        workbook.close()

    return stats


def build_ecuador_snapshot(filters):
    ecuador_filters = dict(filters)
    ecuador_filters["business_unit"] = "uva"
    ecuador_filters["country"] = "EC"
    if filters.get("country") and filters["country"] != "EC":
        rows = []
    else:
        rows = product_category_channel_sales(ecuador_filters)

    sales_total = sum((row.sales_amount for row in rows), ZERO)
    usd_total = sum((row.original_amount for row in rows if row.original_currency == "USD"), ZERO)
    units_total = sum((row.quantity for row in rows), 0)
    average_exchange_rate = _safe_ratio(sales_total, usd_total)
    spend_rows = daily_ad_spends(ecuador_filters) if rows or not filters.get("country") or filters["country"] == "EC" else []
    spend_total = sum((row.spend_amount or ZERO for row in spend_rows), ZERO)
    roas = _safe_ratio(sales_total, spend_total)
    average_ticket = _safe_ratio(sales_total, Decimal(units_total))

    category_rows = defaultdict(lambda: {"sales": ZERO, "usd": ZERO, "units": 0})
    channel_rows = defaultdict(lambda: ZERO)
    daily_rows = defaultdict(lambda: {"sales": ZERO, "usd": ZERO, "units": 0, "spend": ZERO})
    for row in rows:
        category_rows[row.category.name]["sales"] += row.sales_amount
        category_rows[row.category.name]["usd"] += row.original_amount if row.original_currency == "USD" else ZERO
        category_rows[row.category.name]["units"] += row.quantity
        channel_rows[row.channel.name] += row.sales_amount
        daily_rows[row.sale_date.isoformat()]["sales"] += row.sales_amount
        daily_rows[row.sale_date.isoformat()]["usd"] += row.original_amount if row.original_currency == "USD" else ZERO
        daily_rows[row.sale_date.isoformat()]["units"] += row.quantity
    for row in spend_rows:
        daily_rows[row.spend_date.isoformat()]["spend"] += row.spend_amount

    categories = [
        {
            "label": label,
            "sales": float(values["sales"]),
            "usd": float(values["usd"]),
            "units": values["units"],
            "average_ticket": round(float(_safe_ratio(values["sales"], Decimal(values["units"]))), 2) if values["units"] else 0,
        }
        for label, values in sorted(category_rows.items(), key=lambda item: item[1]["sales"], reverse=True)
    ]
    channels = [{"label": label, "value": float(value)} for label, value in sorted(channel_rows.items(), key=lambda item: item[1], reverse=True)]
    daily = [
        {
            "label": label,
            "sales": float(values["sales"]),
            "usd": float(values["usd"]),
            "units": values["units"],
            "spend": float(values["spend"]),
            "roas": round(float(values["sales"] / values["spend"]), 2) if values["spend"] else 0,
            "average_ticket": round(float(_safe_ratio(values["sales"], Decimal(values["units"]))), 2) if values["units"] else 0,
        }
        for label, values in sorted(daily_rows.items(), key=lambda item: item[0])
    ]

    insights = []
    if categories:
        top = categories[0]
        insights.append(f"{top['label']} lidera Ecuador con {format_cop(top['sales'])} y {top['units']} unidades.")
    if usd_total:
        insights.append(f"Las ventas Ecuador suman USD {float(usd_total):,.2f} equivalentes a {format_cop(sales_total)}.")
    if average_exchange_rate:
        insights.append(f"La tasa promedio usada en el periodo es {round(float(average_exchange_rate), 2):,.2f} COP/USD.")

    return {
        "kpis": {
            "sales_total": float(sales_total),
            "usd_total": float(usd_total),
            "units": units_total,
            "average_exchange_rate": round(float(average_exchange_rate), 2) if average_exchange_rate else 0,
            "spend_total": float(spend_total),
            "roas": round(float(roas), 2) if roas else 0,
            "average_ticket": round(float(average_ticket), 2) if average_ticket else 0,
        },
        "categories": categories,
        "channels": channels,
        "daily_series": daily,
        "insights": insights[:4],
        "insight_cards": _decorate_insights(insights[:4]),
    }


def build_uva_country_snapshot(filters, country_code):
    scoped_filters = dict(filters)
    scoped_filters["business_unit"] = "uva"
    scoped_filters["country"] = country_code
    snapshot = build_sales_snapshot(scoped_filters, include_comparison=False)
    country_names = {"CO": "Colombia", "EC": "Ecuador", "MX": "Mexico"}
    return {
        "code": country_code,
        "label": country_names.get(country_code, country_code),
        "sales": snapshot["kpis"]["sales_total"],
        "spend": snapshot["kpis"]["ad_spend"],
        "roas": snapshot["kpis"]["roas"],
        "average_ticket": snapshot["kpis"].get("average_ticket", 0),
        "orders": snapshot["kpis"].get("orders", 0),
        "units": snapshot["kpis"].get("units", 0),
        "loaded_days": snapshot.get("coverage", {}).get("loaded_days", 0),
        "expected_days": snapshot.get("coverage", {}).get("expected_days", 0),
        "coverage_ratio": snapshot.get("coverage", {}).get("ratio", 0),
        "is_partial_coverage": snapshot.get("coverage", {}).get("is_partial", False),
        "coverage_message": snapshot.get("coverage", {}).get("message", ""),
        "insights": snapshot.get("insights", []),
        "insight_cards": snapshot.get("insight_cards", []),
    }


def _nested_lookup(payload, *keys):
    value = payload or {}
    for key in keys:
        if not isinstance(value, dict):
            return ""
        value = value.get(key)
    return value or ""


def _creative_image_hashes(creative):
    if not isinstance(creative, dict):
        return []
    hashes = []
    if creative.get("image_hash"):
        hashes.append(creative["image_hash"])
    asset_feed = creative.get("asset_feed_spec") or {}
    images = asset_feed.get("images") if isinstance(asset_feed, dict) else None
    if isinstance(images, list):
        for image in images:
            image_hash = (image or {}).get("hash")
            if image_hash:
                hashes.append(image_hash)
    return hashes


def _creative_image_url(creative, image_lookup=None):
    if not isinstance(creative, dict):
        return ""
    if creative.get("image_url"):
        return creative["image_url"]
    image_lookup = image_lookup or {}
    for image_hash in _creative_image_hashes(creative):
        resolved = image_lookup.get(image_hash) or {}
        if resolved.get("url"):
            return resolved["url"]
    if creative.get("thumbnail_url"):
        return creative["thumbnail_url"]
    asset_feed = creative.get("asset_feed_spec") or {}
    images = asset_feed.get("images") if isinstance(asset_feed, dict) else None
    if isinstance(images, list) and images:
        first = images[0] or {}
        return first.get("url") or first.get("hash") or ""
    return ""


def _creative_video_assets(creative):
    if not isinstance(creative, dict):
        return []

    videos = []
    story = creative.get("object_story_spec") or {}
    video_data = story.get("video_data") if isinstance(story, dict) else None
    if isinstance(video_data, dict) and (video_data.get("video_id") or video_data.get("id")):
        videos.append(video_data)

    asset_feed = creative.get("asset_feed_spec") or {}
    asset_videos = asset_feed.get("videos") if isinstance(asset_feed, dict) else None
    if isinstance(asset_videos, list):
        videos.extend(item for item in asset_videos if isinstance(item, dict))
    return videos


def _creative_video_asset(creative):
    videos = _creative_video_assets(creative)
    return videos[0] if videos else {}


def _creative_video_id(creative):
    video = _creative_video_asset(creative)
    return video.get("video_id") or video.get("id") or ""


def _creative_video_thumbnail_url(creative):
    video = _creative_video_asset(creative)
    return (
        video.get("thumbnail_url")
        or video.get("image_url")
        or video.get("picture")
        or ""
    )


def _is_technical_meta_name(value):
    raw = str(value or "").strip()
    return "{{" in raw or "}}" in raw


def _meta_ad_display_name(row, creative):
    ad_name = str((row or {}).get("name") or "").strip()
    if ad_name and not _is_technical_meta_name(ad_name):
        return ad_name
    creative_name = str((creative or {}).get("name") or "").strip()
    if creative_name and not _is_technical_meta_name(creative_name):
        return creative_name
    return _creative_text(creative, "title") or "Anuncio activo"


def _creative_text(creative, *keys):
    if not isinstance(creative, dict):
        return ""
    for key in keys:
        value = creative.get(key)
        if value:
            return str(value)
    story = creative.get("object_story_spec") or {}
    for section in ("link_data", "video_data", "photo_data"):
        block = story.get(section) if isinstance(story, dict) else None
        if not isinstance(block, dict):
            continue
        for key in keys:
            value = block.get(key)
            if value:
                return str(value)
    asset_feed = creative.get("asset_feed_spec") or {}
    for key in keys:
        plural_key = {"body": "bodies"}.get(key, f"{key}s")
        items = asset_feed.get(plural_key) if isinstance(asset_feed, dict) else None
        if isinstance(items, list) and items:
            value = (items[0] or {}).get("text") or (items[0] or {}).get(key)
            if value:
                return str(value)
    return ""


def _meta_preview_formats(row, has_video):
    normalized_name = normalize_text((row or {}).get("name"))
    if has_video and "reel" in normalized_name:
        return ("INSTAGRAM_REELS", "INSTAGRAM_STANDARD", "MOBILE_FEED_STANDARD")
    if has_video:
        return ("INSTAGRAM_STANDARD", "MOBILE_FEED_STANDARD", "DESKTOP_FEED_STANDARD")
    return ()


def _meta_row_is_comfama(row):
    creative = (row or {}).get("creative") or {}
    campaign = (row or {}).get("campaign") or {}
    adset = (row or {}).get("adset") or {}
    values = [
        (row or {}).get("name"),
        campaign.get("name"),
        adset.get("name"),
        creative.get("name"),
    ]
    return any("comfama" in normalize_text(value) for value in values)


def _meta_ad_preview_url(client, row, has_video, force_preview=False):
    ad_id = (row or {}).get("id")
    if not ad_id or not (has_video or force_preview):
        return ""
    formats = _meta_preview_formats(row, has_video) or ("INSTAGRAM_STANDARD", "MOBILE_FEED_STANDARD")
    for ad_format in formats:
        try:
            preview_url = client.get_ad_preview_iframe_src(ad_id, ad_format=ad_format)
        except Exception:
            continue
        if preview_url:
            return preview_url
    return ""


def _meta_insight_payload(row):
    insights = row.get("insights") if isinstance(row, dict) else {}
    data = insights.get("data") if isinstance(insights, dict) else None
    if isinstance(data, list) and data:
        return data[0] or {}
    return {}


def _meta_action_value(items, action_types):
    wanted = {normalize_text(item) for item in action_types}
    total = ZERO
    for item in items or []:
        if normalize_text((item or {}).get("action_type")) in wanted:
            total += parse_decimal((item or {}).get("value"))
    return total


def _meta_first_value(items, action_types):
    wanted = {normalize_text(item) for item in action_types}
    for item in items or []:
        if normalize_text((item or {}).get("action_type")) in wanted:
            return parse_decimal((item or {}).get("value"))
    return ZERO


def _meta_preferred_value(items, action_types):
    rows = items or []
    for action_type in action_types:
        wanted = normalize_text(action_type)
        for item in rows:
            if normalize_text((item or {}).get("action_type")) == wanted:
                return parse_decimal((item or {}).get("value"))
    return ZERO


def _meta_parse_created_at(value):
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
    except ValueError:
        try:
            return datetime.strptime(raw[:10], "%Y-%m-%d").date()
        except ValueError:
            return None


def _meta_ad_maturity(created_time, reference_date):
    created_date = _meta_parse_created_at(created_time)
    if not created_date:
        return {
            "level": "unknown",
            "label": "Sin dato",
            "age_days": None,
            "description": "Meta no devolvio fecha de activacion para este anuncio.",
        }
    age_days = max(0, (reference_date - created_date).days)
    if age_days < 14:
        return {
            "level": "low",
            "label": "Bajo",
            "age_days": age_days,
            "description": "Activado hace poco; leer ventas y ROAS como senales tempranas.",
        }
    if age_days < 45:
        return {
            "level": "medium",
            "label": "Medio",
            "age_days": age_days,
            "description": "Ya tiene aprendizaje inicial, pero aun puede estabilizar rendimiento.",
        }
    return {
        "level": "high",
        "label": "Alto",
        "age_days": age_days,
        "description": "Lleva varias semanas activo; comparar eficiencia contra anuncios nuevos.",
    }


def _meta_ad_metrics(row):
    insight = _meta_insight_payload(row)
    actions = insight.get("actions") or []
    action_values = insight.get("action_values") or []
    cost_per_action = insight.get("cost_per_action_type") or []
    purchase_types = (
        "offsite_conversion.fb_pixel_purchase",
        "onsite_web_purchase",
        "website_purchase",
        "purchase",
        "omni_purchase",
    )
    spend = parse_decimal(insight.get("spend"))
    impressions = parse_quantity(insight.get("impressions"))
    reach = parse_quantity(insight.get("reach"))
    clicks = parse_quantity(insight.get("clicks"))
    link_clicks = int(_meta_action_value(actions, ("link_click", "inline_link_click")) or parse_decimal(insight.get("inline_link_clicks")) or 0)
    purchases = int(_meta_preferred_value(actions, purchase_types) or 0)
    purchase_value = _meta_preferred_value(action_values, purchase_types)
    cpa_purchase = _meta_preferred_value(cost_per_action, purchase_types) or _safe_ratio(spend, Decimal(purchases))
    roas_values = insight.get("purchase_roas") or insight.get("website_purchase_roas") or []
    roas = _meta_preferred_value(roas_values, purchase_types + ("website_purchase_roas",))
    if not roas:
        roas = _safe_ratio(purchase_value, spend)
    if not purchase_value and roas and spend:
        purchase_value = roas * spend
    ctr = parse_decimal(insight.get("ctr"))
    if not ctr:
        ctr = _safe_ratio(Decimal(clicks) * Decimal("100"), Decimal(impressions))
    cpc = parse_decimal(insight.get("cpc")) or _safe_ratio(spend, Decimal(clicks))
    cpm = parse_decimal(insight.get("cpm")) or _safe_ratio(spend * Decimal("1000"), Decimal(impressions))
    frequency = _safe_ratio(Decimal(impressions), Decimal(reach))

    return {
        "spend": float(spend),
        "impressions": impressions,
        "reach": reach,
        "clicks": clicks,
        "link_clicks": link_clicks,
        "ctr": round(float(ctr), 2) if ctr else 0,
        "cpc": round(float(cpc), 2) if cpc else 0,
        "cpm": round(float(cpm), 2) if cpm else 0,
        "frequency": round(float(frequency), 2) if frequency else 0,
        "purchases": purchases,
        "purchase_value": float(purchase_value),
        "cpa_purchase": round(float(cpa_purchase), 2) if cpa_purchase else 0,
        "roas": round(float(roas), 2) if roas else 0,
    }


def _build_meta_ads_pacing_insights(ads):
    if not ads:
        return {"positive": [], "negative": []}

    def metric(ad, key):
        return parse_decimal((ad.get("metrics") or {}).get(key))

    def ad_label(ad):
        return ad.get("title") or ad.get("name") or "Anuncio activo"

    spend_ads = [ad for ad in ads if metric(ad, "spend") > 0]
    purchase_ads = [ad for ad in spend_ads if metric(ad, "purchases") > 0]
    positive = []
    negative = []
    used_positive = set()
    used_negative = set()

    def add_positive(ad, title, message):
        ad_id = ad.get("id") or ad_label(ad)
        if ad_id in used_positive or len(positive) >= 3:
            return
        used_positive.add(ad_id)
        positive.append({"title": title, "message": message})

    def add_negative(ad, title, message, recommendation):
        ad_id = ad.get("id") or ad_label(ad)
        if ad_id in used_negative or len(negative) >= 3:
            return
        used_negative.add(ad_id)
        negative.append({"title": title, "message": message, "recommendation": recommendation})

    top_purchases = sorted(purchase_ads, key=lambda ad: (metric(ad, "purchases"), metric(ad, "roas"), metric(ad, "spend")), reverse=True)
    if top_purchases:
        ad = top_purchases[0]
        add_positive(
            ad,
            "Mayor volumen de compras",
            f"{ad_label(ad)} registra {int(metric(ad, 'purchases'))} compras con ROAS {float(metric(ad, 'roas')):.2f}.",
        )

    top_roas = sorted(
        [ad for ad in purchase_ads if metric(ad, "roas") > 0],
        key=lambda ad: (metric(ad, "roas"), metric(ad, "purchases"), metric(ad, "spend")),
        reverse=True,
    )
    if top_roas:
        ad = top_roas[0]
        add_positive(
            ad,
            "Mejor eficiencia",
            f"{ad_label(ad)} lidera en ROAS con {float(metric(ad, 'roas')):.2f} sobre {format_cop(metric(ad, 'spend'))} de inversion.",
        )

    efficient_cpa = sorted(
        [ad for ad in purchase_ads if metric(ad, "cpa_purchase") > 0],
        key=lambda ad: (metric(ad, "cpa_purchase"), -metric(ad, "purchases")),
    )
    if efficient_cpa:
        ad = efficient_cpa[0]
        add_positive(
            ad,
            "CPA mas sano",
            f"{ad_label(ad)} compra a {format_cop(metric(ad, 'cpa_purchase'))} con {int(metric(ad, 'purchases'))} compras en el periodo.",
        )

    no_purchase_spend = sorted(
        [ad for ad in spend_ads if metric(ad, "purchases") == 0],
        key=lambda ad: metric(ad, "spend"),
        reverse=True,
    )
    if no_purchase_spend:
        ad = no_purchase_spend[0]
        add_negative(
            ad,
            "Gasto sin compras",
            f"{ad_label(ad)} invierte {format_cop(metric(ad, 'spend'))} sin compras registradas en el rango.",
            "Revisar segmentacion, destino y creativo antes de seguir aumentando presupuesto.",
        )

    low_roas = sorted(
        [ad for ad in purchase_ads if metric(ad, "roas") and metric(ad, "roas") < Decimal("1.5")],
        key=lambda ad: (metric(ad, "roas"), -metric(ad, "spend")),
    )
    if low_roas:
        ad = low_roas[0]
        add_negative(
            ad,
            "ROAS por debajo del punto sano",
            f"{ad_label(ad)} tiene ROAS {float(metric(ad, 'roas')):.2f} con {int(metric(ad, 'purchases'))} compras.",
            "Mantenerlo en observacion y mover presupuesto hacia piezas con ROAS y compras consistentes.",
        )

    high_cpc = sorted(
        [ad for ad in spend_ads if metric(ad, "cpc") > 0 and metric(ad, "clicks") >= 20],
        key=lambda ad: metric(ad, "cpc"),
        reverse=True,
    )
    if high_cpc:
        ad = high_cpc[0]
        add_negative(
            ad,
            "Trafico costoso",
            f"{ad_label(ad)} tiene CPC de {format_cop(metric(ad, 'cpc'))} y CTR {float(metric(ad, 'ctr')):.2f}%.",
            "Probar otro gancho visual o primer texto para bajar friccion antes del clic.",
        )

    if not positive and spend_ads:
        top_spend = max(spend_ads, key=lambda ad: metric(ad, "spend"))
        positive.append(
            {
                "title": "Lectura en progreso",
                "message": f"Hay {len(spend_ads)} anuncios activos con inversion; aun no aparece un ganador claro por compras o ROAS.",
            }
        )
        if not negative:
            add_negative(
                top_spend,
                "Sin ganador claro",
                f"{ad_label(top_spend)} concentra {format_cop(metric(top_spend, 'spend'))} de inversion.",
                "Esperar mas conversiones o redistribuir hacia creatividades con senales tempranas de compra.",
            )

    return {"positive": positive[:3], "negative": negative[:3]}


def build_uva_meta_ads_preview(filters, limit=None, comfama_scope="exclude", force_refresh=False, timeout=None):
    """Construye el panel de anuncios activos de Meta.

    `force_refresh` y `timeout` existen para el precalentamiento en segundo
    plano (ver el comando warm_meta_ads_preview): alli conviene ignorar la
    cache y esperar a Meta lo que haga falta, porque nadie esta mirando.
    """
    requested_country = (filters.get("country") or "").upper()
    country_code = requested_country or "CO"

    country = Country.objects.filter(code__iexact=country_code).first()
    country_label = country.name if country else country_code
    account_id = getattr(settings, f"META_{country_code}_ACCOUNT_ID", "")
    token = getattr(settings, "META_ACCESS_TOKEN", "")
    if not account_id or not token:
        return {
            "ads": [],
            "pacing_insights": {"positive": [], "negative": []},
            "country_code": country_code,
            "country_label": country_label,
            "requires_country": False,
            "message": f"No hay credenciales Meta configuradas para {country_label}.",
        }

    date_start = _parse_filter_date(filters.get("date_start")) or timezone.localdate()
    date_end = _parse_filter_date(filters.get("date_end")) or date_start
    cache_key = "uva-meta-ads-preview:{country}:{start}:{end}:{limit}:{scope}".format(
        country=country_code,
        start=date_start.isoformat(),
        end=date_end.isoformat(),
        limit=limit or "default",
        scope=comfama_scope,
    )
    if not force_refresh:
        cached_preview = cache.get(cache_key)
        if cached_preview is not None:
            return cached_preview

    fallback_ttl = _setting_int("META_ADS_PREVIEW_FALLBACK_CACHE_SECONDS", 120)

    client = MetaAdsClient(
        token,
        api_version=getattr(settings, "META_API_VERSION", "v20.0"),
        timeout=int(timeout) if timeout else _setting_int("META_ADS_PREVIEW_TIMEOUT", 8),
    )
    max_records = _setting_int("META_ADS_PREVIEW_MAX_RECORDS", 36)
    try:
        rows = client.get_active_ads(
            account_id,
            limit=limit,
            date_start=date_start,
            date_end=date_end,
            max_records=max_records,
        )
    except Exception:
        logger.exception("Meta Ads preview fallo para %s (%s a %s)", country_code, date_start, date_end)
        failure = {
            "ads": [],
            "pacing_insights": {"positive": [], "negative": []},
            "country_code": country_code,
            "country_label": country_label,
            "requires_country": False,
            "message": f"No fue posible cargar anuncios activos de Meta para {country_label}. Intenta actualizar de nuevo en unos minutos.",
        }
        # Cachear el fallo evita repetir el camino lento en cada request, pero
        # un precalentamiento fallido no debe borrar un panel bueno ya guardado.
        if force_refresh and cache.get(cache_key) is not None:
            logger.warning("Se conserva el preview de Meta en cache para %s tras un precalentamiento fallido", country_code)
        else:
            cache.set(cache_key, failure, fallback_ttl)
        return failure

    image_hashes = []
    for row in rows:
        image_hashes.extend(_creative_image_hashes(row.get("creative") or {}))
    try:
        image_lookup = client.get_ad_images_by_hashes(account_id, image_hashes)
    except Exception:
        logger.warning("No se pudieron resolver imagenes de Meta para %s", country_code, exc_info=True)
        image_lookup = {}

    ads = []
    max_preview_fetches = _setting_int("META_ADS_PREVIEW_MAX_IFRAMES", 8)
    for row in rows:
        is_comfama = _meta_row_is_comfama(row)
        if comfama_scope == "exclude" and is_comfama:
            continue
        if comfama_scope == "only" and not is_comfama:
            continue

        creative = row.get("creative") or {}
        campaign = row.get("campaign") or {}
        adset = row.get("adset") or {}
        video_id = _creative_video_id(creative)
        display_name = _meta_ad_display_name(row, creative)
        normalized_ad_text = " ".join(
            normalize_text(value)
            for value in (
                row.get("name"),
                display_name,
                creative.get("name"),
                campaign.get("name"),
                adset.get("name"),
            )
            if value
        )
        looks_like_video = any(marker in normalized_ad_text for marker in ("reel", "video", "story"))
        has_video = bool(video_id or looks_like_video)
        headline = _creative_text(creative, "title")
        destination_url = (
            creative.get("link_url")
            or creative.get("object_url")
            or _nested_lookup(creative, "object_story_spec", "link_data", "link")
        )
        maturity = _meta_ad_maturity(row.get("created_time") or "", date_end)
        ads.append(
            {
                "id": row.get("id", ""),
                "name": display_name,
                "raw_name": row.get("name") or creative.get("name") or "",
                "created_time": row.get("created_time") or "",
                "updated_time": row.get("updated_time") or "",
                "maturity": maturity,
                "maturity_level": maturity["level"],
                "maturity_label": maturity["label"],
                "maturity_description": maturity["description"],
                "maturity_age_days": maturity["age_days"],
                "campaign_name": campaign.get("name") or "",
                "adset_name": adset.get("name") or "",
                "status": row.get("effective_status") or row.get("status") or "ACTIVE",
                "image_url": _creative_video_thumbnail_url(creative) or _creative_image_url(creative, image_lookup=image_lookup),
                "media_kind": "video" if has_video else "image",
                "video_id": video_id,
                "preview_url": _meta_ad_preview_url(
                    client,
                    row,
                    has_video,
                    force_preview=looks_like_video or (is_comfama and not has_video),
                )
                if len(ads) < max_preview_fetches
                else "",
                "title": display_name,
                "headline": headline,
                "body": _creative_text(creative, "body", "message"),
                "cta": creative.get("call_to_action_type") or "",
                "destination_url": destination_url,
                "metrics": _meta_ad_metrics(row),
            }
        )
        if limit and len(ads) >= int(limit):
            break

    ads.sort(key=lambda item: item.get("created_time") or "", reverse=True)

    preview = {
        "ads": ads,
        "pacing_insights": _build_meta_ads_pacing_insights(ads),
        "country_code": country_code,
        "country_label": country_label,
        "date_start": date_start.isoformat(),
        "date_end": date_end.isoformat(),
        "requires_country": False,
        "message": "" if ads else f"No se encontraron anuncios activos en la cuenta Meta de {country_label}.",
    }
    # Un resultado vacio tambien se cachea (con TTL corto): antes cada request
    # sin anuncios repetia la ronda completa de llamadas a Meta.
    ttl = _setting_int("META_ADS_PREVIEW_CACHE_SECONDS", 900) if ads else fallback_ttl
    cache.set(cache_key, preview, ttl)
    return preview



UVA_GEO_COUNTRY_MAPS = {
    "CO": {"label": "Colombia", "geojson": "reports/maps/CO-adm1.min.geojson"},
    "EC": {"label": "Ecuador", "geojson": "reports/maps/EC-adm1.min.geojson"},
    "MX": {"label": "Mexico", "geojson": "reports/maps/MX-adm1.min.geojson"},
}

GEO_LOCATION_ALIASES = {
    "bogota": "bogota",
    "bogota-d-c": "bogota",
    "bogota-dc": "bogota",
    "distrito-capital": "bogota",
    "capital-district": "bogota",
    "bogota-capital-district": "bogota",
    "distrito-especial": "bogota",
    "mexico-city": "distrito-federal",
    "ciudad-de-mexico": "distrito-federal",
    "cdmx": "distrito-federal",
    "distrito-federal": "distrito-federal",
    "coahuila": "coahuila-de-zaragoza",
    "estado-de-mexico": "mexico",
    "edomex": "mexico",
    "nuevo-leon": "nuevo-leon",
    "manabi": "manabi",
    "santo-domingo": "santo-domingo-de-los-tsachilas",
    "santo-domingo-de-los-tsachilas": "santo-domingo-de-los-tsachilas",
}


# Google Ads y Meta devuelven la misma zona con y sin calificativo ("Guayas" y
# "Guayas Province"), lo que antes generaba dos claves distintas y partia las
# metricas de una region en dos filas.
GEO_NAME_PREFIXES = (
    "departamento-del-",
    "departamento-de-",
    "provincia-del-",
    "provincia-de-",
    "estado-del-",
    "estado-de-",
    "region-del-",
    "region-de-",
    "province-of-",
    "department-of-",
    "state-of-",
)
GEO_NAME_SUFFIXES = (
    "-departamento",
    "-department",
    "-provincia",
    "-province",
    "-estado",
    "-state",
    "-region",
)


def _strip_geo_qualifiers(key):
    current = key
    while True:
        previous = current
        for prefix in GEO_NAME_PREFIXES:
            if current.startswith(prefix) and len(current) > len(prefix):
                current = current[len(prefix):]
                break
        for suffix in GEO_NAME_SUFFIXES:
            if current.endswith(suffix) and len(current) > len(suffix):
                current = current[: -len(suffix)]
                break
        if current == previous:
            return current or key


def geo_location_key(value):
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = "".join(char for char in normalized if not unicodedata.combining(char))
    key = slugify(ascii_value)
    if key in GEO_LOCATION_ALIASES:
        return GEO_LOCATION_ALIASES[key]
    stripped = _strip_geo_qualifiers(key)
    return GEO_LOCATION_ALIASES.get(stripped, stripped)


def _geo_metric_payload(row):
    return {
        "impressions": int(row.get("impressions", 0)),
        "reach": int(row.get("reach", 0)),
        "clicks": int(row.get("clicks", 0)),
        "purchases": float(row.get("purchases", 0)),
        "conversion_value": float(row.get("conversion_value", 0)),
        "spend": float(row.get("spend", 0)),
    }


def _empty_geo_totals():
    return {
        "impressions": 0,
        "reach": 0,
        "clicks": 0,
        "purchases": 0,
        "conversion_value": 0,
        "spend": 0,
    }


def _static_map_url(path):
    return f"{settings.STATIC_URL.rstrip('/')}/{path.lstrip('/')}"


def json_dumps_safe(value):
    import json

    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def build_bali_web_geo_map_data(filters, kpis):
    code = "CO"
    config = UVA_GEO_COUNTRY_MAPS[code]
    queryset = DailyGeoAdMetric.objects.select_related("ad_platform").filter(
        business_unit__slug="bali", country__code=code,
        geo_level__in=[DailyGeoAdMetric.GeoLevel.REGION, DailyGeoAdMetric.GeoLevel.CITY],
    )
    if filters.get("date_start"):
        queryset = queryset.filter(metric_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(metric_date__lte=filters["date_end"])
    region_totals = defaultdict(_empty_geo_totals)
    city_totals = defaultdict(_empty_geo_totals)
    region_names = {}
    city_names = {}
    platforms = set()
    latest_date = None
    for metric in queryset:
        key = geo_location_key(metric.location_key or metric.location_name)
        target = city_totals if metric.geo_level == DailyGeoAdMetric.GeoLevel.CITY else region_totals
        names = city_names if metric.geo_level == DailyGeoAdMetric.GeoLevel.CITY else region_names
        values = target[key]
        values["impressions"] += int(metric.impressions or 0)
        values["reach"] += int(metric.reach or 0)
        values["clicks"] += int(metric.clicks or 0)
        values["purchases"] += Decimal(str(metric.purchases or 0))
        values["conversion_value"] += Decimal(str(metric.conversion_value or 0))
        values["spend"] += Decimal(str(metric.spend_amount or 0))
        names.setdefault(key, metric.location_name)
        platforms.add(metric.ad_platform.name)
        if latest_date is None or metric.metric_date > latest_date:
            latest_date = metric.metric_date
    regions = []
    for key, values in region_totals.items():
        payload = _geo_metric_payload(values)
        row = {"key": key, "name": region_names.get(key, key.replace("-", " ").title())}
        row.update(payload)
        regions.append(row)
    regions.sort(key=lambda item: (item["impressions"], item["purchases"], item["spend"]), reverse=True)
    source_points = city_totals or region_totals
    source_names = city_names if city_totals else region_names
    points = []
    for key, values in source_points.items():
        purchases = float(values.get("purchases", 0))
        if purchases <= 0:
            continue
        payload = _geo_metric_payload(values)
        row = {"key": key, "name": source_names.get(key, key.replace("-", " ").title())}
        row.update(payload)
        points.append(row)
    points.sort(key=lambda item: (item["purchases"], item["conversion_value"], item["spend"]), reverse=True)
    totals_source = region_totals or city_totals
    totals = _empty_geo_totals()
    for values in totals_source.values():
        for metric_name in totals:
            totals[metric_name] += values.get(metric_name, 0)
    total_payload = _geo_metric_payload(totals)

    fallback_payload = _geo_metric_payload({
        "impressions": kpis.get("sessions_total", 0),
        "reach": kpis.get("sessions_total", 0),
        "clicks": 0,
        "purchases": kpis.get("web_orders_total", 0),
        "conversion_value": kpis.get("web_sales_total", 0),
        "spend": kpis.get("spend_total", 0),
    })
    display_totals = total_payload if any(total_payload.values()) else fallback_payload
    has_real_geo = bool(regions or points)
    return {
        "code": code,
        "label": config["label"],
        "geojson_url": _static_map_url(config["geojson"]),
        "regions": regions,
        "regions_json": json_dumps_safe(regions),
        "points": points[:12],
        "points_json": json_dumps_safe(points[:12]),
        "totals": display_totals,
        "totals_json": json_dumps_safe(display_totals),
        "has_real_geo": has_real_geo,
        "platforms": ", ".join(sorted(platforms)) if platforms else "Google Ads",
        "latest_date": latest_date.isoformat() if latest_date else "",
        "sessions_total": int(kpis.get("sessions_total", 0) or 0),
        "web_orders_total": float(kpis.get("web_orders_total", 0) or 0),
        "web_sales_total": float(kpis.get("web_sales_total", 0) or 0),
        "spend_total": float(kpis.get("spend_total", 0) or 0),
        "attribution": "Mapa ADM1: geoBoundaries. Metricas: Shopify y Google Ads sincronizados por Axis; no usa Analytics como fuente principal.",
    }


def build_uva_geo_map_data(filters, country_rows):
    rows_by_code = {row.get("code"): row for row in country_rows if row.get("code")}
    selected_country = filters.get("country")
    if selected_country in UVA_GEO_COUNTRY_MAPS:
        codes = [selected_country]
    else:
        codes = [code for code in ("CO", "EC", "MX") if code in rows_by_code] or ["CO", "EC", "MX"]

    maps = []
    for code in codes:
        config = UVA_GEO_COUNTRY_MAPS[code]
        row = rows_by_code.get(code, {})
        queryset = DailyGeoAdMetric.objects.select_related("ad_platform").filter(
            business_unit__slug="uva",
            country__code=code,
            geo_level__in=[DailyGeoAdMetric.GeoLevel.REGION, DailyGeoAdMetric.GeoLevel.CITY],
        )
        if filters.get("date_start"):
            queryset = queryset.filter(metric_date__gte=filters["date_start"])
        if filters.get("date_end"):
            queryset = queryset.filter(metric_date__lte=filters["date_end"])

        region_totals = defaultdict(_empty_geo_totals)
        city_totals = defaultdict(_empty_geo_totals)
        region_names = {}
        city_names = {}
        platforms = set()
        latest_date = None
        for metric in queryset:
            key = geo_location_key(metric.location_key or metric.location_name)
            target = city_totals if metric.geo_level == DailyGeoAdMetric.GeoLevel.CITY else region_totals
            names = city_names if metric.geo_level == DailyGeoAdMetric.GeoLevel.CITY else region_names
            values = target[key]
            values["impressions"] += int(metric.impressions or 0)
            values["reach"] += int(metric.reach or 0)
            values["clicks"] += int(metric.clicks or 0)
            values["purchases"] += Decimal(str(metric.purchases or 0))
            values["conversion_value"] += Decimal(str(metric.conversion_value or 0))
            values["spend"] += Decimal(str(metric.spend_amount or 0))
            names.setdefault(key, metric.location_name)
            platforms.add(metric.ad_platform.name)
            if latest_date is None or metric.metric_date > latest_date:
                latest_date = metric.metric_date

        regions = []
        for key, values in region_totals.items():
            payload = _geo_metric_payload(values)
            regions.append({"key": key, "name": region_names.get(key, key.replace("-", " ").title()), **payload})
        regions.sort(key=lambda item: (item["impressions"], item["purchases"]), reverse=True)

        source_points = city_totals or region_totals
        source_names = city_names if city_totals else region_names
        points = []
        for key, values in source_points.items():
            purchases = float(values.get("purchases", 0))
            if purchases <= 0:
                continue
            payload = _geo_metric_payload(values)
            points.append({"key": key, "name": source_names.get(key, key.replace("-", " ").title()), **payload})
        points.sort(key=lambda item: item["purchases"], reverse=True)

        totals = _empty_geo_totals()
        for values in region_totals.values():
            for metric_name in totals:
                totals[metric_name] += values.get(metric_name, 0)
        total_payload = _geo_metric_payload(totals)
        maps.append({
            "code": code,
            "label": config["label"],
            "geojson_url": _static_map_url(config["geojson"]),
            "regions": regions,
            "regions_json": json_dumps_safe(regions),
            "points": points[:12],
            "points_json": json_dumps_safe(points[:12]),
            "totals": total_payload,
            "totals_json": json_dumps_safe(total_payload),
            "sales": row.get("sales", 0),
            "spend": row.get("spend", 0),
            "roas": row.get("roas", 0),
            "has_real_geo": bool(regions or points),
            "platforms": ", ".join(sorted(platforms)) if platforms else "Google Ads / Meta Ads",
            "latest_date": latest_date.isoformat() if latest_date else "",
            "attribution": "Mapa ADM1: geoBoundaries. Metricas: Google Ads y Meta Ads sincronizados por Axis.",
        })
    return {
        "maps": maps,
        "selected_country": selected_country or "",
        "is_estimated": False,
    }

def build_copa_uva_country_comparison(filters):
    rows = [build_uva_country_snapshot(filters, code) for code in ("CO", "EC", "MX")]
    return [
        row
        for row in rows
        if row["sales"] or row["spend"] or row.get("units", 0)
    ]


def build_uva_category_country_comparison(filters):
    country_codes = ("CO", "EC", "MX")
    country_labels = {"CO": "Colombia", "EC": "Ecuador", "MX": "Mexico"}
    merged = {}

    for code in country_codes:
        scoped_filters = dict(filters)
        scoped_filters["business_unit"] = "uva"
        scoped_filters["country"] = code
        snapshot = build_uva_category_snapshot(scoped_filters)
        for card in snapshot.get("cards", []):
            bucket = merged.setdefault(
                card["category_id"],
                {
                    "category_id": card["category_id"],
                    "name": card["name"],
                    "image_url": card.get("image_url", ""),
                    "image_fallback_url": card.get("image_fallback_url", ""),
                    "countries": {},
                },
            )
            bucket["countries"][code] = {
                "code": code,
                "label": country_labels[code],
                "sales_total": card.get("sales_total", 0),
                "whatsapp_sales_total": card.get("whatsapp_sales_total", 0),
                "combined_sales_total": card.get("combined_sales_total", 0),
                "quantity_total": card.get("quantity_total", 0),
                "spend_total": card.get("spend_total", 0),
                "spend_meta": card.get("spend_meta", 0),
                "spend_google": card.get("spend_google", 0),
                "cpa_meta": card.get("cpa_meta", 0),
                "cpa_google": card.get("cpa_google", 0),
                "roas": card.get("roas", 0),
                "average_ticket": card.get("average_ticket", 0),
            }

    cards = []
    chart_rows = []
    for item in merged.values():
        total_sales = ZERO
        total_spend = ZERO
        total_units = 0
        country_rows = []
        for code in country_codes:
            country_data = item["countries"].get(
                code,
                {
                    "code": code,
                    "label": country_labels[code],
                    "sales_total": 0,
                    "whatsapp_sales_total": 0,
                    "combined_sales_total": 0,
                    "quantity_total": 0,
                    "spend_total": 0,
                    "spend_meta": 0,
                    "spend_google": 0,
                    "cpa_meta": 0,
                    "cpa_google": 0,
                    "roas": 0,
                    "average_ticket": 0,
                },
            )
            total_sales += Decimal(str(country_data["combined_sales_total"] or 0))
            total_spend += Decimal(str(country_data["spend_total"] or 0))
            total_units += int(country_data["quantity_total"] or 0)
            country_rows.append(country_data)

        overall_roas = _safe_ratio(total_sales, total_spend)
        overall_ticket = _safe_ratio(total_sales, Decimal(total_units))
        cards.append(
            {
                "category_id": item["category_id"],
                "name": item["name"],
                "image_url": item["image_url"],
                "image_fallback_url": item.get("image_fallback_url", ""),
                "countries": country_rows,
                "combined_sales_total": float(total_sales),
                "spend_total": float(total_spend),
                "quantity_total": total_units,
                "roas": round(float(overall_roas), 2) if overall_roas else 0,
                "average_ticket": round(float(overall_ticket), 2) if overall_ticket else 0,
            }
        )
        chart_rows.append(
            {
                "label": item["name"],
                "CO": item["countries"].get("CO", {}).get("combined_sales_total", 0),
                "EC": item["countries"].get("EC", {}).get("combined_sales_total", 0),
                "MX": item["countries"].get("MX", {}).get("combined_sales_total", 0),
            }
        )

    cards.sort(key=lambda item: item["combined_sales_total"], reverse=True)
    chart_rows.sort(key=lambda item: item["CO"] + item["EC"] + item["MX"], reverse=True)
    return {
        "cards": cards,
        "chart_rows": chart_rows,
        "category_count": len(cards),
    }


def format_cop(value):
    formatted = f"{float(value or 0):,.0f}".replace(",", ".")
    return f"${formatted} COP"


def category_image_url(category):
    if not category:
        return ""
    if category.image:
        try:
            return category.image.url
        except (OSError, ValueError):
            pass
    return PRODUCT_CATEGORY_IMAGE_FALLBACKS.get(category.slug, "")


def category_fallback_image_url(category):
    if not category:
        return ""
    return PRODUCT_CATEGORY_IMAGE_FALLBACKS.get(category.slug, "")


def _insight_card(message, signal="warning", title="Atencion"):
    metadata = {
        "success": {"label": "Logro", "icon": "check"},
        "warning": {"label": "Prevencion", "icon": "warning"},
        "danger": {"label": "Cuidado", "icon": "alert"},
    }
    style = metadata.get(signal, metadata["warning"])
    return {
        "message": message,
        "signal": signal if signal in metadata else "warning",
        "title": title,
        **style,
    }


def _decorate_insights(messages):
    cards = []
    for message in messages:
        normalized = normalize_text(message)
        if any(keyword in normalized for keyword in ("por debajo", "sin rango", "faltante", "sin datos", "conviene revisar")):
            cards.append(_insight_card(message, "danger", "Requiere revision"))
        elif any(keyword in normalized for keyword in ("lidera", "mejor", "aporta", "acumula", "genera", "conversion")):
            cards.append(_insight_card(message, "success", "Resultado destacado"))
        else:
            cards.append(_insight_card(message, "warning", "Seguimiento"))
    return cards


def _build_comparative_insights(snapshot):
    comparisons = snapshot.get("comparison", {})
    cards = []
    sales = comparisons.get("sales_total", {})
    spend = comparisons.get("ad_spend", {})
    roas = comparisons.get("roas", {})
    ticket = comparisons.get("average_ticket", {})

    if sales.get("has_previous") and sales.get("delta_pct") is not None:
        pct = sales["delta_pct"]
        signal = "success" if pct > 0 else ("danger" if pct <= -10 else "warning")
        verb = "aumentaron" if pct > 0 else "disminuyeron"
        cards.append(_insight_card(f"Las ventas {verb} {abs(pct):.1f}% frente al periodo anterior.", signal, "Ventas vs. periodo anterior"))
    if spend.get("has_previous") and spend.get("delta_pct") is not None:
        pct = spend["delta_pct"]
        sales_not_lower = not sales.get("has_previous") or sales.get("delta", 0) >= 0
        if pct < 0 and sales_not_lower:
            cards.append(_insight_card(f"La inversion se redujo {abs(pct):.1f}% sin reducir ventas; mejora de eficiencia.", "success", "Eficiencia de inversion"))
        elif pct > 10 and not sales_not_lower:
            cards.append(_insight_card(f"La inversion crecio {pct:.1f}% mientras las ventas no mejoraron.", "danger", "Presupuesto bajo cuidado"))
        else:
            cards.append(_insight_card(f"La inversion vario {pct:+.1f}% frente al periodo anterior.", "warning", "Control de presupuesto"))
    if roas.get("has_previous") and roas.get("delta_pct") is not None:
        pct = roas["delta_pct"]
        signal = "success" if pct > 0 else ("danger" if pct <= -10 else "warning")
        cards.append(_insight_card(f"El ROAS cambio {pct:+.1f}% frente al periodo anterior y se ubica en {snapshot['kpis']['roas']:.2f}.", signal, "Rentabilidad"))
    if ticket.get("has_previous") and ticket.get("delta_pct") is not None and ticket["delta_pct"] > 0:
        cards.append(_insight_card(f"El ticket promedio aumento {ticket['delta_pct']:.1f}% frente al periodo anterior.", "success", "Valor por pedido"))

    if snapshot.get("coverage", {}).get("is_partial"):
        cards.append(_insight_card(snapshot["coverage"]["message"], "warning", "Cobertura parcial"))
    if not cards:
        cards = _decorate_insights(snapshot.get("insights", []))
    return cards[:5]


def _parse_filter_date(value):
    return parse_excel_date(value) if value else None


def _previous_period_filters(filters):
    start = _parse_filter_date(filters.get("date_start"))
    end = _parse_filter_date(filters.get("date_end"))
    if not start or not end or start > end:
        return None
    days = (end - start).days + 1
    previous_end = start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=days - 1)
    comparison_filters = dict(filters)
    comparison_filters["date_start"] = previous_start.isoformat()
    comparison_filters["date_end"] = previous_end.isoformat()
    comparison_filters["compare_mode"] = "none"
    return comparison_filters


def _month_start(value):
    return value.replace(day=1)


def _next_month(value):
    if value.month == 12:
        return value.replace(year=value.year + 1, month=1, day=1)
    return value.replace(month=value.month + 1, day=1)


def _week_start(value):
    return value - timedelta(days=value.weekday())


def _bucket_start(value, granularity):
    if granularity == "monthly":
        return _month_start(value)
    if granularity == "weekly":
        return _week_start(value)
    return value


def _next_bucket(value, granularity):
    if granularity == "monthly":
        return _next_month(value)
    if granularity == "weekly":
        return value + timedelta(days=7)
    return value + timedelta(days=1)


def _bucket_label(value, granularity):
    if granularity == "monthly":
        return value.strftime("%b %Y").capitalize()
    if granularity == "weekly":
        end = value + timedelta(days=6)
        return f"Sem {value.isocalendar().week} | {value.strftime('%d %b')} - {end.strftime('%d %b')}"
    return value.strftime("%Y-%m-%d")


def _build_bucket_index(start, end, granularity):
    if not start or not end or start > end:
        return []
    current = _bucket_start(start, granularity)
    final = _bucket_start(end, granularity)
    buckets = []
    while current <= final:
        buckets.append(current)
        current = _next_bucket(current, granularity)
    return buckets


def _comparison_payload(current, previous):
    if previous in (None, 0):
        return {"previous": previous or 0, "delta": current, "delta_pct": None, "direction": "neutral", "has_previous": False}
    delta = current - previous
    delta_pct = (delta / previous) * 100 if previous else None
    if delta > 0:
        direction = "up"
    elif delta < 0:
        direction = "down"
    else:
        direction = "neutral"
    return {
        "previous": previous,
        "delta": delta,
        "delta_pct": round(delta_pct, 1) if delta_pct is not None else None,
        "direction": direction,
        "has_previous": True,
    }


def _summarize_comparison(current_kpis, previous_snapshot):
    if not previous_snapshot:
        return {}
    previous_kpis = previous_snapshot["kpis"]
    return {
        "sales_total": _comparison_payload(current_kpis["sales_total"], previous_kpis["sales_total"]),
        "ad_spend": _comparison_payload(current_kpis["ad_spend"], previous_kpis["ad_spend"]),
        "roas": _comparison_payload(current_kpis["roas"], previous_kpis["roas"]),
        "average_ticket": _comparison_payload(current_kpis.get("average_ticket", 0), previous_kpis.get("average_ticket", 0)),
    }


def _bali_comparison_payload(current, previous, lower_is_better=False):
    payload = _comparison_payload(current, previous)
    if lower_is_better and payload["direction"] in {"up", "down"}:
        payload["direction"] = "down" if payload["direction"] == "up" else "up"
    return payload


def _summarize_bali_comparison(current_kpis, previous_kpis):
    lower_is_better = {"spend_total", "average_cpa"}
    comparable_keys = (
        "sales_total",
        "spend_total",
        "roas",
        "average_ticket",
        "orders_total",
        "average_daily_sales",
        "physical_sales_total",
        "web_sales_total",
        "web_orders_total",
        "web_roas",
        "sessions_total",
        "conversion_rate",
        "web_average_ticket",
        "average_daily_orders",
        "whatsapp_sales_total",
        "whatsapp_orders_total",
        "conversations_total",
        "whatsapp_conversion_rate",
        "whatsapp_average_ticket",
        "physical_visitors_total",
        "physical_orders_total",
        "physical_conversion_rate",
        "physical_average_ticket",
        "physical_sales_per_visitor",
        "average_cpa",
    )
    return {
        key: _bali_comparison_payload(
            current_kpis.get(key, 0),
            previous_kpis.get(key, 0),
            lower_is_better=key in lower_is_better,
        )
        for key in comparable_keys
    }


def _summarize_bali_community_comparison(current_kpis, previous_kpis):
    return {
        key: _comparison_payload(current_kpis.get(key, 0), previous_kpis.get(key, 0))
        for key in ("subscribers", "new_subscribers", "average_new_subscribers")
    }


def _build_period_series(sales_daily_map, spend_daily_map, count_daily_map, granularity, start, end):
    bucket_totals = {}
    for bucket in _build_bucket_index(start, end, granularity):
        bucket_totals[bucket] = {"sales": ZERO, "spend": ZERO, "count": 0}
    for day_label, amount in sales_daily_map.items():
        bucket = _bucket_start(parse_excel_date(day_label), granularity)
        bucket_totals.setdefault(bucket, {"sales": ZERO, "spend": ZERO, "count": 0})
        bucket_totals[bucket]["sales"] += amount
    for day_label, amount in spend_daily_map.items():
        bucket = _bucket_start(parse_excel_date(day_label), granularity)
        bucket_totals.setdefault(bucket, {"sales": ZERO, "spend": ZERO, "count": 0})
        bucket_totals[bucket]["spend"] += amount
    for day_label, count in count_daily_map.items():
        bucket = _bucket_start(parse_excel_date(day_label), granularity)
        bucket_totals.setdefault(bucket, {"sales": ZERO, "spend": ZERO, "count": 0})
        bucket_totals[bucket]["count"] += count

    rows = []
    for bucket, values in sorted(bucket_totals.items(), key=lambda item: item[0]):
        sales_value = values["sales"]
        spend_value = values["spend"]
        count_value = values["count"]
        average_ticket = _safe_ratio(sales_value, Decimal(count_value))
        rows.append(
            {
                "label": _bucket_label(bucket, granularity),
                "sales": float(sales_value),
                "spend": float(spend_value),
                "roas": round(float(sales_value / spend_value), 2) if spend_value else 0,
                "average_ticket": round(float(average_ticket), 2) if average_ticket else 0,
            }
        )
    return rows


def _build_roas_by_unit(daily_rows, spend_rows):
    spend_by_unit = defaultdict(lambda: ZERO)
    sales_by_unit = defaultdict(lambda: ZERO)
    for row in daily_rows:
        label = row.business_unit.name if row.business_unit else "Sin unidad"
        sales_by_unit[label] += row.sales_amount
    for row in spend_rows:
        label = row.business_unit.name if row.business_unit else "Sin unidad"
        spend_by_unit[label] += row.spend_amount
    rows = []
    for label in sorted(set(sales_by_unit) | set(spend_by_unit)):
        spend_value = spend_by_unit[label]
        sales_value = sales_by_unit[label]
        rows.append({"label": label, "value": round(float(sales_value / spend_value), 2) if spend_value else 0})
    return sorted(rows, key=lambda item: item["value"], reverse=True)


def _build_snapshot_response(daily_rows, spend_rows, filters, limit, row_mode):
    remove_vat_from_sales = filters.get("business_unit") == "marketplace"
    sales_value = remove_colombia_vat if remove_vat_from_sales else (lambda value: value)

    if row_mode == "daily":
        sales_total_with_vat = sum((row.sales_amount for row in daily_rows), ZERO)
        sales_total = sum((sales_value(row.sales_amount) for row in daily_rows), ZERO)
        sales_whatsapp_with_vat = sum((row.sales_amount for row in daily_rows if row.channel and row.channel.slug.startswith("whatsapp-uva-")), ZERO)
        sales_web_with_vat = sum((row.sales_amount for row in daily_rows if row.channel and row.channel.slug == "ecommerce-uva"), ZERO)
        sales_whatsapp = sum((sales_value(row.sales_amount) for row in daily_rows if row.channel and row.channel.slug.startswith("whatsapp-uva-")), ZERO)
        sales_web = sum((sales_value(row.sales_amount) for row in daily_rows if row.channel and row.channel.slug == "ecommerce-uva"), ZERO)
        order_count = sum((getattr(row, "order_count", 0) or getattr(row, "quantity", 0) or 0) for row in daily_rows)
        direct_units = sum((getattr(row, "units", 0) or getattr(row, "quantity", 0) or 0) for row in daily_rows)
        direct_spend_total = sum((getattr(row, "spend_amount", ZERO) or ZERO) for row in daily_rows)
    else:
        sales_total_with_vat = sum((row.sale_value for row in daily_rows), ZERO)
        sales_total = sum((sales_value(row.sale_value) for row in daily_rows), ZERO)
        sales_whatsapp_with_vat = sum((row.sale_value for row in daily_rows if row.channel and row.channel.slug.startswith("whatsapp-uva-")), ZERO)
        sales_web_with_vat = sum((row.sale_value for row in daily_rows if row.channel and row.channel.slug == "ecommerce-uva"), ZERO)
        sales_whatsapp = sum((sales_value(row.sale_value) for row in daily_rows if row.channel and row.channel.slug.startswith("whatsapp-uva-")), ZERO)
        sales_web = sum((sales_value(row.sale_value) for row in daily_rows if row.channel and row.channel.slug == "ecommerce-uva"), ZERO)
        order_count = len(daily_rows)
        direct_units = 0
        direct_spend_total = ZERO

    ad_spend_total = sum((row.spend_amount for row in spend_rows), ZERO)
    spend_total = direct_spend_total or ad_spend_total
    category_sales_rows = product_category_channel_sales(filters) if row_mode == "daily" else []
    product_quantity = sum((row.quantity or 0 for row in category_sales_rows), 0) or direct_units
    product_quantity_web = sum((row.quantity or 0 for row in category_sales_rows if row.channel and row.channel.slug == "ecommerce-uva"), 0)
    product_quantity_whatsapp = sum((row.quantity or 0 for row in category_sales_rows if row.channel and row.channel.slug.startswith("whatsapp-uva-")), 0)
    daily_order_counts = _daily_order_counts_by_channel(filters) if row_mode == "daily" else defaultdict(int)
    order_count_web = daily_order_counts.get("Web", 0)
    order_count_whatsapp = daily_order_counts.get("WhatsApp", 0)
    category_order_count_whatsapp = len([row for row in category_sales_rows if row.channel and row.channel.slug.startswith("whatsapp-uva-")])
    order_count_web = order_count_web or sum((getattr(row, "order_count", 0) or 0) for row in daily_rows if row_mode == "daily" and row.channel and row.channel.slug == "ecommerce-uva")
    order_count_whatsapp = order_count_whatsapp or category_order_count_whatsapp
    ticket_order_count = (order_count_web + order_count_whatsapp) or order_count
    average_ticket_denominator = ticket_order_count
    average_ticket = _safe_ratio(sales_total, Decimal(average_ticket_denominator))
    web_average_ticket_denominator = order_count_web or product_quantity_web
    whatsapp_average_ticket_denominator = order_count_whatsapp or product_quantity_whatsapp
    web_average_ticket = _safe_ratio(sales_web, Decimal(web_average_ticket_denominator))
    whatsapp_average_ticket = _safe_ratio(sales_whatsapp, Decimal(whatsapp_average_ticket_denominator))
    sales_by_unit_totals = defaultdict(lambda: ZERO)
    sales_by_unit_totals_with_vat = defaultdict(lambda: ZERO)
    sales_by_channel_totals = defaultdict(lambda: ZERO)
    sales_by_channel_totals_with_vat = defaultdict(lambda: ZERO)
    sales_daily_map = defaultdict(lambda: ZERO)
    sales_daily_map_with_vat = defaultdict(lambda: ZERO)
    spend_daily_map = defaultdict(lambda: ZERO)
    count_daily_map = defaultdict(int)

    for row in daily_rows:
        unit_label = row.business_unit.name if row.business_unit else "Sin unidad"
        channel_label = _sale_channel_group(row.channel)
        raw_amount = row.sales_amount if row_mode == "daily" else row.sale_value
        amount = sales_value(raw_amount)
        row_date = row.sale_date.isoformat()
        count_value = (getattr(row, "order_count", 0) or getattr(row, "quantity", 0) or 0) if row_mode == "daily" else 1
        sales_by_unit_totals[unit_label] += amount
        sales_by_unit_totals_with_vat[unit_label] += raw_amount
        sales_by_channel_totals[channel_label] += amount
        sales_by_channel_totals_with_vat[channel_label] += raw_amount
        sales_daily_map[row_date] += amount
        sales_daily_map_with_vat[row_date] += raw_amount
        count_daily_map[row_date] += count_value
        if row_mode == "daily":
            spend_daily_map[row_date] += getattr(row, "spend_amount", ZERO) or ZERO

    if not direct_spend_total:
        for row in spend_rows:
            spend_daily_map[row.spend_date.isoformat()] += row.spend_amount

    range_start = _parse_filter_date(filters.get("date_start"))
    range_end = _parse_filter_date(filters.get("date_end"))
    all_dates = [parse_excel_date(label) for label in [*sales_daily_map.keys(), *spend_daily_map.keys()] if label]
    if not range_start and all_dates:
        range_start = min(all_dates)
    if not range_end and all_dates:
        range_end = max(all_dates)

    loaded_dates = set()
    if row_mode == "daily":
        loaded_dates = {row.sale_date for row in daily_rows if getattr(row, "sale_date", None)}
    else:
        loaded_dates = {row.sale_date for row in daily_rows if getattr(row, "sale_date", None)}
    expected_days = (range_end - range_start).days + 1 if range_start and range_end and range_end >= range_start else len(loaded_dates)
    loaded_days = len(loaded_dates)
    coverage_ratio = (Decimal(loaded_days) / Decimal(expected_days)) if expected_days else ZERO
    is_partial_coverage = bool(expected_days and loaded_days < expected_days)
    coverage_message = (
        f"Cobertura cargada: {loaded_days}/{expected_days} dias del rango."
        if expected_days
        else "Sin rango suficiente para medir cobertura."
    )

    time_granularity = filters.get("time_granularity") or "daily"
    period_rows = _build_period_series(sales_daily_map, spend_daily_map, count_daily_map, time_granularity, range_start, range_end)
    average_roas = round(float(sales_total / spend_total), 2) if spend_total else 0

    sales_by_unit = []
    for label, value in sorted(sales_by_unit_totals.items(), key=lambda item: item[1], reverse=True):
        row = {"label": label, "value": float(value)}
        if remove_vat_from_sales:
            row["value_with_vat"] = float(sales_by_unit_totals_with_vat[label])
        sales_by_unit.append(row)
    sales_by_channel = []
    for label, value in sorted(sales_by_channel_totals.items(), key=lambda item: item[1], reverse=True):
        row = {"label": label, "value": float(value)}
        if remove_vat_from_sales:
            row["value_with_vat"] = float(sales_by_channel_totals_with_vat[label])
        sales_by_channel.append(row)

    insights = []
    if sales_by_channel:
        top_channel = sales_by_channel[0]
        insights.append(f"{top_channel['label']} concentra {format_cop(top_channel['value'])} en ventas del periodo filtrado.")
    if range_start and range_end:
        insights.append(f"Se analizaron datos entre {range_start.strftime('%d/%m/%Y')} y {range_end.strftime('%d/%m/%Y')}.")
    if is_partial_coverage:
        insights.append(coverage_message)
    if spend_total:
        insights.append(f"La inversion total del periodo es {format_cop(spend_total)} y el ROAS consolidado es {average_roas:.2f}.")

    return {
        "rows": daily_rows[:limit],
        "spend_rows": spend_rows[:limit],
        "row_count": len(daily_rows),
        "row_mode": row_mode,
        "time_granularity": time_granularity,
        "kpis": {
            "sales_total": float(sales_total),
            **({"sales_total_with_vat": float(sales_total_with_vat)} if remove_vat_from_sales else {}),
            "sales_month": float(sales_total),
            **({"sales_month_with_vat": float(sales_total_with_vat)} if remove_vat_from_sales else {}),
            "sales_whatsapp": float(sales_whatsapp),
            **({"sales_whatsapp_with_vat": float(sales_whatsapp_with_vat)} if remove_vat_from_sales else {}),
            "sales_web": float(sales_web),
            **({"sales_web_with_vat": float(sales_web_with_vat)} if remove_vat_from_sales else {}),
            "ad_spend": float(spend_total),
            "roas": average_roas,
            "orders": ticket_order_count,
            "web_order_count": order_count_web,
            "whatsapp_order_count": order_count_whatsapp,
            "units": product_quantity,
            "product_units_web": product_quantity_web,
            "product_units_whatsapp": product_quantity_whatsapp,
            "average_ticket": round(float(average_ticket), 2) if average_ticket else 0,
            **({"average_ticket_with_vat": round(float(_safe_ratio(sales_total_with_vat, Decimal(average_ticket_denominator))), 2) if average_ticket_denominator else 0} if remove_vat_from_sales else {}),
            "web_average_ticket": round(float(web_average_ticket), 2) if web_average_ticket else 0,
            "whatsapp_average_ticket": round(float(whatsapp_average_ticket), 2) if whatsapp_average_ticket else 0,
        },
        "sales_by_unit": sales_by_unit,
        "sales_by_channel": sales_by_channel,
        "sales_by_day": [
            {
                "label": label,
                "value": float(value),
                **({"value_with_vat": float(sales_daily_map_with_vat[label])} if remove_vat_from_sales else {}),
            }
            for label, value in sorted(sales_daily_map.items(), key=lambda item: item[0])
        ],
        "spend_by_day": [{"label": label, "value": float(value)} for label, value in sorted(spend_daily_map.items(), key=lambda item: item[0])],
        "roas_by_day": [
            {"label": label, "value": round(float(sales_daily_map[label] / spend_daily_map.get(label, ZERO)), 2) if spend_daily_map.get(label, ZERO) else 0}
            for label in sorted(sales_daily_map.keys())
        ],
        "combined_series": period_rows,
        "roas_average": average_roas,
        "series_label": {"daily": "Diario", "weekly": "Semanal", "monthly": "Mensual"}.get(time_granularity, "Diario"),
        "roas_by_unit": _build_roas_by_unit(daily_rows if row_mode == "daily" else [], spend_rows),
        "insights": insights,
        "range_start": range_start.isoformat() if range_start else "",
        "range_end": range_end.isoformat() if range_end else "",
        "coverage": {
            "loaded_days": loaded_days,
            "expected_days": expected_days,
            "ratio": round(float(coverage_ratio), 4) if coverage_ratio else 0,
            "is_partial": is_partial_coverage,
            "message": coverage_message,
        },
    }


def product_for_name(name, business_unit):
    cleaned = str(name or "").strip()
    if not cleaned:
        return None
    product, _ = Product.objects.get_or_create(name=cleaned, business_unit=business_unit, defaults={"is_active": True})
    return product


def channel_from_origin(origin, catalogs):
    slug = CHANNEL_ALIASES.get(normalize_text(origin))
    return catalogs["channels"].get(slug) if slug else None


def import_uva_sales_workbook(path):
    catalogs = ensure_uva_catalogs()
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    stats = {"created": 0, "updated": 0, "skipped": 0, "sheets": []}

    try:
        for sheet in workbook.worksheets:
            config = UVA_SHEET_CONFIG.get(normalize_text(sheet.title))
            if not config:
                continue

            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            header_map = {normalize_text(value): index for index, value in enumerate(header_row) if value not in (None, "")}
            required_headers = {field: normalize_text(header) for field, header in config["columns"].items()}
            missing = [label for label in required_headers.values() if label not in header_map]
            if missing:
                raise ValueError(f"La hoja '{sheet.title}' no tiene las columnas requeridas: {', '.join(missing)}")

            processed_rows = 0
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                product_name = str(row[header_map[required_headers["product_name"]]] or "").strip()
                sale_date = parse_excel_date(row[header_map[required_headers["sale_date"]]])
                if not product_name or not sale_date:
                    stats["skipped"] += 1
                    continue

                origin = str(row[header_map[required_headers["origin"]]] or "").strip()
                channel = channel_from_origin(origin, catalogs)
                if not channel and normalize_text(origin).startswith("whatsapp"):
                    channel = catalogs["channels"].get(UVA_WHATSAPP_BY_COUNTRY[config["country_code"]][1])
                defaults = {
                    "business_unit": catalogs["business_unit"],
                    "country": catalogs["countries"][config["country_code"]],
                    "channel": channel,
                    "product": product_for_name(product_name, catalogs["business_unit"]),
                    "product_name": product_name,
                    "origin": origin,
                    "sale_date": sale_date,
                    "quantity": parse_quantity(row[header_map[required_headers["quantity"]]]),
                    "sale_value": parse_decimal(row[header_map[required_headers["sale_value"]]]),
                    "shipping_value": parse_decimal(row[header_map[required_headers["shipping_value"]]]),
                    "notes": "",
                }
                _, created = SalesTransaction.objects.update_or_create(
                    source_file=Path(path).name,
                    source_sheet=sheet.title,
                    source_row=row_number,
                    defaults=defaults,
                )
                processed_rows += 1
                if created:
                    stats["created"] += 1
                else:
                    stats["updated"] += 1

            stats["sheets"].append({"sheet": sheet.title, "rows": processed_rows})
    finally:
        workbook.close()

    return stats


def apply_sales_filters(queryset, filters):
    if filters.get("date_start"):
        queryset = queryset.filter(sale_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(sale_date__lte=filters["date_end"])
    if filters.get("business_unit"):
        queryset = queryset.filter(business_unit__slug=filters["business_unit"])
    if filters.get("country"):
        queryset = queryset.filter(country__code=filters["country"])
    if filters.get("channel"):
        queryset = queryset.filter(channel__slug=filters["channel"])
    if filters.get("product"):
        queryset = queryset.filter(product__slug=filters["product"])
    return queryset


def sales_transactions(filters, limit=None):
    queryset = SalesTransaction.objects.select_related("business_unit", "country", "channel", "product")
    queryset = apply_sales_filters(queryset, filters)
    if limit:
        queryset = queryset[:limit]
    return list(queryset)


@memoize_per_request
def daily_channel_sales(filters, limit=None):
    queryset = DailyChannelSale.objects.select_related("business_unit", "country", "channel")
    if filters.get("date_start"):
        queryset = queryset.filter(sale_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(sale_date__lte=filters["date_end"])
    if filters.get("business_unit"):
        queryset = queryset.filter(business_unit__slug=filters["business_unit"])
    if filters.get("country"):
        queryset = queryset.filter(country__code=filters["country"])
    if filters.get("channel"):
        queryset = queryset.filter(channel__slug=filters["channel"])
    if limit:
        queryset = queryset[:limit]
    return list(queryset)


@memoize_per_request
def daily_ad_spends(filters, limit=None):
    queryset = DailyAdSpend.objects.select_related("business_unit", "country", "ad_platform")
    if filters.get("date_start"):
        queryset = queryset.filter(spend_date__gte=filters["date_start"])
    if filters.get("date_end"):
        queryset = queryset.filter(spend_date__lte=filters["date_end"])
    if filters.get("business_unit"):
        queryset = queryset.filter(business_unit__slug=filters["business_unit"])
    if filters.get("country"):
        queryset = queryset.filter(country__code=filters["country"])
    if filters.get("ad_platform"):
        queryset = queryset.filter(ad_platform__slug=filters["ad_platform"])
    if limit:
        queryset = queryset[:limit]
    return list(queryset)


def build_sales_snapshot(filters, limit=100, include_comparison=True):
    spend_rows = daily_ad_spends(filters)
    daily_rows = _combined_direct_sales(filters)
    if daily_rows:
        snapshot = _build_snapshot_response(daily_rows, spend_rows, filters, limit, "daily")
    else:
        rows = sales_transactions(filters)
        snapshot = _build_snapshot_response(rows, spend_rows, filters, limit, "transaction")
        snapshot["kpis"]["units"] = sum(row.quantity for row in rows)

    previous_snapshot = None
    if include_comparison and filters.get("compare_mode") == "previous_period":
        comparison_filters = _previous_period_filters(filters)
        if comparison_filters:
            previous_snapshot = build_sales_snapshot(comparison_filters, limit=limit, include_comparison=False)
            snapshot["comparison"] = _summarize_comparison(snapshot["kpis"], previous_snapshot)
            snapshot["comparison_range"] = {
                "label": "Vs. periodo anterior",
                "date_start": comparison_filters["date_start"],
                "date_end": comparison_filters["date_end"],
            }
    if "comparison" not in snapshot:
        snapshot["comparison"] = {}
        snapshot["comparison_range"] = {}
    snapshot["insight_cards"] = _build_comparative_insights(snapshot)
    return snapshot


def build_ad_platform_performance(filters, sales_snapshot=None):
    """Inversion por plataforma y ROAS de referencia.

    `sales_snapshot` evita reconstruir un snapshot completo solo para leer
    `sales_total`: las vistas que llaman aqui ya tienen uno calculado.
    """
    spend_rows = daily_ad_spends(filters)
    if sales_snapshot is None:
        sales_snapshot = build_sales_snapshot(filters, include_comparison=False)
    sales_total = Decimal(str(sales_snapshot["kpis"]["sales_total"] or 0))
    grouped = defaultdict(lambda: ZERO)
    for row in spend_rows:
        label = row.ad_platform.name if row.ad_platform else "Sin fuente"
        grouped[label] += row.spend_amount or ZERO
    total_spend = sum(grouped.values(), ZERO)
    rows = []
    for label, spend in sorted(grouped.items(), key=lambda item: item[1], reverse=True):
        share = _safe_ratio(spend, total_spend)
        rows.append(
            {
                "label": label,
                "spend": float(spend),
                "share": round(float(share) * 100, 1) if share else 0,
                "reference_roas": round(float(_safe_ratio(sales_total, spend)), 2) if spend else 0,
            }
        )
    return rows
