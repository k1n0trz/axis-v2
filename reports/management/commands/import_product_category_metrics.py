from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from reports.models import BusinessUnit, Country, DailyProductCategoryMetric, ProductCategory
from reports.services.sales_dashboard import category_slug_from_product_name, parse_excel_date, uva_category_slug_from_product_name
from reports.utils.numbers import parse_decimal


class Command(BaseCommand):
    help = "Importa metricas diarias por categoria de producto desde un archivo Excel."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)
        parser.add_argument("--business-unit", default="uva")
        parser.add_argument("--country", default="CO")
        parser.add_argument("--end-date", default="")

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        if not file_path.exists():
            raise CommandError(f"No existe el archivo: {file_path}")

        business_unit = BusinessUnit.objects.filter(slug=options["business_unit"]).first()
        if not business_unit:
            raise CommandError(f"No existe la marca con slug '{options['business_unit']}'")

        country = Country.objects.filter(code=options["country"]).first()
        if not country:
            raise CommandError(f"No existe el pais con codigo '{options['country']}'")

        end_date = parse_excel_date(options["end_date"]) if options["end_date"] else None
        workbook = load_workbook(file_path, data_only=True, read_only=True)
        created = 0
        updated = 0
        skipped = 0

        try:
            sheet = workbook.active
            headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            expected = {
                "Fecha": 0,
                "Producto": 1,
                "CPA Meta Ads": 2,
                "CPA Google Ads": 3,
                "Inversión meta": 4,
                "Inversión Google": 5,
                "Total inversión": 6,
                "Ventas": 7,
                "Nota": 8,
            }
            if list(headers[:9]) != list(expected.keys()):
                raise CommandError("El archivo no coincide con el formato esperado para metricas por categoria.")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                metric_date = parse_excel_date(row[0])
                product_name = str(row[1] or "").strip()
                if not metric_date or not product_name:
                    skipped += 1
                    continue
                if end_date and metric_date > end_date:
                    skipped += 1
                    continue

                category_slug = (
                    uva_category_slug_from_product_name(product_name)
                    if business_unit.slug == "uva"
                    else category_slug_from_product_name(product_name)
                )
                if not category_slug:
                    skipped += 1
                    continue
                category = ProductCategory.objects.filter(slug=category_slug).first()
                if not category:
                    category, _ = ProductCategory.objects.get_or_create(
                        name=product_name,
                        defaults={"slug": category_slug, "description": f"Categoria importada desde {file_path.name}."},
                    )

                defaults = {
                    "cpa_meta": parse_decimal(row[2]) or None,
                    "cpa_google": parse_decimal(row[3]) or None,
                    "spend_meta": parse_decimal(row[4]),
                    "spend_google": parse_decimal(row[5]),
                    "total_spend": parse_decimal(row[6]),
                    "sales_amount": parse_decimal(row[7]),
                    "notes": str(row[8] or "").strip(),
                    "source_type": DailyProductCategoryMetric.SourceType.IMPORTED,
                    "source_file": file_path.name,
                }

                if not defaults["sales_amount"] and not defaults["total_spend"] and not defaults["spend_meta"] and not defaults["spend_google"]:
                    skipped += 1
                    continue

                _, was_created = DailyProductCategoryMetric.objects.update_or_create(
                    business_unit=business_unit,
                    country=country,
                    category=category,
                    metric_date=metric_date,
                    defaults=defaults,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
        finally:
            workbook.close()

        self.stdout.write(
            self.style.SUCCESS(
                f"Importacion completada. Creados: {created}. Actualizados: {updated}. Omitidos: {skipped}."
            )
        )
