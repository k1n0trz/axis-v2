from datetime import date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook

from django.contrib.auth.models import User

from reports.models import BusinessUnit, Channel, Country, DailyChannelSale, SalesTarget


MONTHS = {
    "ene": 1,
    "enero": 1,
    "feb": 2,
    "febrero": 2,
    "mar": 3,
    "marzo": 3,
    "abril": 4,
    "abri": 4,
    "abr": 4,
    "may": 5,
    "mayo": 5,
    "jun": 6,
    "junio": 6,
    "jul": 7,
    "julio": 7,
    "ago": 8,
    "agosto": 8,
    "sep": 9,
    "sept": 9,
    "septiembre": 9,
    "oct": 10,
    "octubre": 10,
    "nov": 11,
    "noviembre": 11,
    "dic": 12,
    "diciembre": 12,
}

CHANNEL_BLOCKS = [
    ("Mercadolibre", 2, 3, 4, 5),
    ("Falabella", 6, 7, 8, 9),
    ("Rappi", 10, 11, 12, 13),
    ("Farmatodo", 14, None, None, None),
]

CHANNEL_SLUG_ALIASES = {
    "Mercadolibre": ("mercado-libre", "mercadolibre"),
}

TARGET_COLUMNS = {
    "Mercadolibre": 4,
    "Falabella": 8,
    "Rappi": 12,
    "Farmatodo": 14,
}


def parse_decimal(value):
    if value in (None, ""):
        return None
    return Decimal(str(value))


def parse_int(value):
    if value in (None, ""):
        return None
    return int(Decimal(str(value)).to_integral_value(rounding=ROUND_HALF_UP))


def parse_week_range(label, year):
    parts = re.findall(r"(\d{1,2})\s*([A-Za-z]*)", str(label or ""))
    if len(parts) < 2:
        raise ValueError(f"No se pudo leer el rango de fecha: {label}")
    start_day, start_month_text = parts[0]
    end_day, end_month_text = parts[1]
    start_month = MONTHS.get(start_month_text.lower()) if start_month_text else None
    end_month = MONTHS.get(end_month_text.lower()) if end_month_text else start_month
    if not start_month:
        start_month = end_month
    if not start_month or not end_month:
        raise ValueError(f"No se pudo leer el mes del rango: {label}")
    start = date(year, start_month, int(start_day))
    end = date(year, end_month, int(end_day))
    if end < start:
        end = date(year + 1, end_month, int(end_day))
    return start, end


def date_range(start, end):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def split_decimal(total, count):
    if total is None:
        return [Decimal("0")] * count
    base = (total / Decimal(count)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    values = [base] * count
    values[-1] += total - sum(values)
    return values


def split_int(total, count):
    if total is None:
        return [0] * count
    base, remainder = divmod(total, count)
    return [base + (1 if index < remainder else 0) for index in range(count)]


def apply_marketplace_import_row(lookup, source_name, label, sales, spend, orders, units):
    sale, was_created = DailyChannelSale.objects.get_or_create(
        **lookup,
        defaults={
            "sales_amount": Decimal("0"),
            "spend_amount": Decimal("0"),
            "order_count": 0,
            "units": 0,
            "source_type": DailyChannelSale.SourceType.IMPORTED,
            "source_file": source_name,
            "notes": "",
        },
    )
    update_fields = ["source_type", "source_file", "notes", "updated_at"]
    if sales is not None:
        sale.sales_amount = sales
        update_fields.append("sales_amount")
    if spend is not None:
        sale.spend_amount = spend
        update_fields.append("spend_amount")
    if orders is not None:
        sale.order_count = orders
        update_fields.append("order_count")
    if units is not None:
        sale.units = units
        update_fields.append("units")
    sale.source_type = DailyChannelSale.SourceType.IMPORTED
    sale.source_file = source_name
    sale.notes = f"Distribuido diariamente desde consolidado semanal: {label}."
    sale.save(update_fields=update_fields)
    return was_created


class Command(BaseCommand):
    help = "Importa ventas semanales de Marketplace y las distribuye de forma diaria."

    def add_arguments(self, parser):
        parser.add_argument("path", nargs="?", default=r"C:\Users\trafficker.digital\Downloads\Bases de resultados\Axis\marketplace.xlsx")
        parser.add_argument("--year", type=int, default=2026)

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options["path"])
        if not path.exists():
            raise CommandError(f"No existe el archivo: {path}")

        business_unit, _ = BusinessUnit.objects.get_or_create(
            slug="marketplace",
            defaults={"name": "Marketplaces", "display_order": 30, "is_active": True},
        )
        country, _ = Country.objects.get_or_create(
            code="CO",
            defaults={"name": "Colombia", "display_order": 1, "is_active": True},
        )
        country.business_units.add(business_unit)

        channels = {}
        for index, (name, *_columns) in enumerate(CHANNEL_BLOCKS, start=1):
            slug_aliases = CHANNEL_SLUG_ALIASES.get(name, (slugify(name),))
            channel = Channel.objects.filter(business_unit=business_unit, slug__in=slug_aliases).first()
            if not channel:
                channel, _ = Channel.objects.get_or_create(
                    business_unit=business_unit,
                    slug=slug_aliases[0],
                    defaults={"name": name, "display_order": index, "is_active": True},
                )
            channels[name] = channel

        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        created = 0
        updated = 0
        targets_updated = 0
        source_name = path.name

        for row_index in range(3, sheet.max_row + 1):
            label = sheet.cell(row=row_index, column=1).value
            if not label:
                continue
            try:
                start, end = parse_week_range(label, options["year"])
            except ValueError:
                continue
            days = list(date_range(start, end))
            for channel_name, sales_col, spend_col, orders_col, units_col in CHANNEL_BLOCKS:
                sales = parse_decimal(sheet.cell(row=row_index, column=sales_col).value)
                spend = parse_decimal(sheet.cell(row=row_index, column=spend_col).value) if spend_col else None
                orders = parse_int(sheet.cell(row=row_index, column=orders_col).value) if orders_col else None
                units = parse_int(sheet.cell(row=row_index, column=units_col).value) if units_col else None
                if sales is None and spend is None and orders is None and units is None:
                    continue
                sales_values = split_decimal(sales, len(days)) if sales is not None else [None] * len(days)
                spend_values = split_decimal(spend, len(days)) if spend is not None else [None] * len(days)
                order_values = split_int(orders, len(days)) if orders is not None else [None] * len(days)
                unit_values = split_int(units, len(days)) if units is not None else [None] * len(days)
                for index, sale_date in enumerate(days):
                    was_created = apply_marketplace_import_row(
                        {
                            "business_unit": business_unit,
                            "country": country,
                            "channel": channels[channel_name],
                            "sale_date": sale_date,
                        },
                        source_name,
                        label,
                        sales_values[index],
                        spend_values[index],
                        order_values[index],
                        unit_values[index],
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1

        target_row = None
        for row_index in range(1, sheet.max_row + 1):
            row_values = [sheet.cell(row=row_index, column=column).value for column in range(1, sheet.max_column + 1)]
            if any(str(value).strip().lower() == "meta ventas" for value in row_values if value is not None):
                target_row = row_index
                break
        karen = User.objects.filter(username__iexact="Karen").first()
        if target_row and karen:
            for channel_name, target_col in TARGET_COLUMNS.items():
                target_amount = parse_decimal(sheet.cell(row=target_row, column=target_col).value)
                if target_amount is None:
                    continue
                SalesTarget.objects.update_or_create(
                    user=karen,
                    business_unit=business_unit,
                    channel=channels[channel_name],
                    date_start=date(options["year"], 4, 1),
                    date_end=date(options["year"], 4, 30),
                    defaults={
                        "target_amount": target_amount,
                        "is_active": True,
                        "notes": f"Meta importada desde {source_name}.",
                    },
                )
                targets_updated += 1

        self.stdout.write(self.style.SUCCESS(f"Marketplace importado. Creados: {created}. Actualizados: {updated}. Metas: {targets_updated}."))
