from collections import defaultdict
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db.models import Count, Sum
from openpyxl import load_workbook

from reports.management.commands.import_daily_web_sales import (
    first_header_index,
    normalize_header,
    parse_decimal,
    parse_quantity,
    parse_sale_date,
    split_products,
)
from reports.models import DailyChannelSale, DailyProductCategorySale


class Command(BaseCommand):
    help = "Compara un Excel de ventas web WordPress contra los totales web Colombia cargados."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path")

    def handle(self, *args, **options):
        path = Path(options["xlsx_path"]).expanduser()
        if not path.exists():
            raise CommandError(f"No existe el archivo: {path}")

        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        totals_by_date = defaultdict(lambda: {"sales": Decimal("0"), "orders": 0, "articles": 0, "rows": 0})
        product_quantity_mismatches = []
        try:
            for sheet in workbook.worksheets:
                headers = [normalize_header(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
                fecha_idx = first_header_index(headers, ("fecha",))
                ventas_idx = first_header_index(headers, ("ventas", "ventas netas", "total", "importe"))
                quantity_idx = first_header_index(headers, ("cantidad", "articulos vendidos", "unidades"))
                products_idx = first_header_index(headers, ("producto(s)",))
                if fecha_idx is None or ventas_idx is None:
                    raise CommandError("La hoja debe tener columnas Fecha y Ventas/Ventas netas.")

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    sale_date = parse_sale_date(row[fecha_idx] if len(row) > fecha_idx else None)
                    if not sale_date:
                        continue
                    sales_amount = parse_decimal(row[ventas_idx] if len(row) > ventas_idx else None)
                    articles = 1
                    if quantity_idx is not None and len(row) > quantity_idx:
                        articles = parse_quantity(row[quantity_idx]) or 1
                    if products_idx is not None and len(row) > products_idx:
                        split_quantity = sum(quantity for _, quantity in split_products(row[products_idx]))
                        if split_quantity != articles:
                            product_quantity_mismatches.append((sheet.title, row[fecha_idx], row[products_idx], articles, split_quantity))
                    totals_by_date[sale_date]["sales"] += sales_amount
                    totals_by_date[sale_date]["orders"] += 1
                    totals_by_date[sale_date]["articles"] += articles
                    totals_by_date[sale_date]["rows"] += 1
        finally:
            workbook.close()

        if not totals_by_date:
            raise CommandError("No se encontraron ventas en el archivo.")

        min_date = min(totals_by_date)
        max_date = max(totals_by_date)
        file_sales = sum((values["sales"] for values in totals_by_date.values()), Decimal("0"))
        file_orders = sum((values["orders"] for values in totals_by_date.values()), 0)
        file_articles = sum((values["articles"] for values in totals_by_date.values()), 0)
        file_rows = sum((values["rows"] for values in totals_by_date.values()), 0)

        daily = DailyChannelSale.objects.filter(
            business_unit__slug="uva",
            country__code="CO",
            channel__slug="ecommerce-uva",
            sale_date__range=(min_date, max_date),
        )
        category = DailyProductCategorySale.objects.filter(
            business_unit__slug="uva",
            country__code="CO",
            channel__slug="ecommerce-uva",
            sale_date__range=(min_date, max_date),
        )
        daily_summary = daily.aggregate(rows=Count("id"), sales=Sum("sales_amount"), orders=Sum("order_count"))
        category_summary = category.aggregate(rows=Count("id"), sales=Sum("sales_amount"), units=Sum("quantity"))

        self.stdout.write(f"Archivo: filas={file_rows}, ventas={file_sales}, pedidos={file_orders}, articulos={file_articles}, rango={min_date}..{max_date}")
        self.stdout.write(f"Base diaria web CO: {daily_summary}")
        self.stdout.write(f"Base categorias web CO: {category_summary}")
        self.stdout.write(f"Diferencia archivo - base diaria: {file_sales - (daily_summary['sales'] or Decimal('0'))}")
        daily_by_date = {row["sale_date"]: row for row in daily.values("sale_date", "sales_amount", "order_count")}
        mismatches = []
        sales_mismatches = []
        for day in sorted(totals_by_date):
            file_values = totals_by_date[day]
            db_values = daily_by_date.get(day, {})
            db_sales = db_values.get("sales_amount") or Decimal("0")
            db_orders = db_values.get("order_count") or 0
            if file_values["sales"].quantize(Decimal("0.01")) != db_sales.quantize(Decimal("0.01")):
                sales_mismatches.append((day, file_values["sales"], db_sales, file_values["sales"] - db_sales))
            if file_values["sales"].quantize(Decimal("0.01")) != db_sales.quantize(Decimal("0.01")) or file_values["orders"] != db_orders:
                mismatches.append((day, file_values["sales"], db_sales, file_values["orders"], db_orders))
        self.stdout.write(f"Fechas con diferencia: {len(mismatches)}")
        self.stdout.write(f"Primeras diferencias: {mismatches[:12]}")
        self.stdout.write(f"Fechas con diferencia de ventas: {len(sales_mismatches)}")
        self.stdout.write(f"Diferencias de ventas: {sales_mismatches[:12]}")
        self.stdout.write(f"Filas con articulos != desglose Producto(s): {len(product_quantity_mismatches)}")
        self.stdout.write(f"Primeras filas con diferencia de articulos: {product_quantity_mismatches[:8]}")
        recent = [(day, totals_by_date[day]) for day in sorted(totals_by_date)[-5:]]
        self.stdout.write(f"Archivo fechas recientes: {recent}")
