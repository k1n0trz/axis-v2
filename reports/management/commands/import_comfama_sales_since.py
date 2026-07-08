from datetime import date

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from reports.models import ComfamaProductReference, ComfamaSale
from reports.services.comfama_import import (
    _column_value,
    _sales_column_indexes,
    _tariff_from_values,
    clean_ref,
    infer_reference,
)
from reports.services.sales_dashboard import parse_excel_date


class Command(BaseCommand):
    help = "Importa solo ventas Comfama desde una fecha inicial, sin tocar pauta."

    def add_arguments(self, parser):
        parser.add_argument("--file", default="data/update-comfama.xlsx")
        parser.add_argument("--sheet", default="Hoja1")
        parser.add_argument("--start-date", default="2026-05-05")
        parser.add_argument("--source-name", default="")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        start_date = parse_excel_date(options["start_date"])
        if not start_date:
            raise CommandError("start-date no es una fecha valida.")

        source_name = options["source_name"] or options["file"].replace("\\", "/").split("/")[-1]
        workbook = load_workbook(options["file"], data_only=True, read_only=True)
        try:
            if options["sheet"] not in workbook.sheetnames:
                raise CommandError(f"No existe la hoja {options['sheet']}.")
            sheet = workbook[options["sheet"]]
            sales_columns = _sales_column_indexes(sheet)
            rows = []
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                sale_date = parse_excel_date(_column_value(row, sales_columns["date"]))
                reference_value = clean_ref(_column_value(row, sales_columns["reference"]))
                tariff = _tariff_from_values(_column_value(row, sales_columns["tariff"]), reference_value)
                if not sale_date or sale_date < start_date or not reference_value:
                    continue
                rows.append(
                    {
                        "row_number": row_number,
                        "sale_date": sale_date,
                        "reference": reference_value,
                        "tariff": tariff,
                    }
                )
        finally:
            workbook.close()

        if not rows:
            self.stdout.write(self.style.WARNING("No se encontraron ventas para importar desde la fecha indicada."))
            return

        end_date = max(row["sale_date"] for row in rows)
        if options["dry_run"]:
            self._print_summary(rows, start_date, end_date, deleted=0, created=0, updated=0, skipped=0, dry_run=True)
            return

        created = 0
        updated = 0
        skipped = 0
        with transaction.atomic():
            deleted, _ = ComfamaSale.objects.exclude(source_file="").filter(
                sale_date__gte=start_date,
                sale_date__lte=end_date,
            ).delete()

            for row in rows:
                reference = ComfamaProductReference.objects.filter(reference=row["reference"]).first()
                if not reference:
                    category, tariff_a, tariff_b = infer_reference(row["reference"])
                    if not category:
                        skipped += 1
                        continue
                    reference, was_created = ComfamaProductReference.objects.update_or_create(
                        reference=row["reference"],
                        defaults={
                            "category": category,
                            "price_tariff_a": tariff_a,
                            "price_tariff_b": tariff_b,
                            "is_inferred": True,
                            "is_active": True,
                            "notes": f"Referencia inferida desde ventas de {source_name}; validar precio si aplica.",
                        },
                    )

                sale, was_created = ComfamaSale.objects.update_or_create(
                    source_file=source_name,
                    source_row=row["row_number"],
                    defaults={
                        "sale_date": row["sale_date"],
                        "tariff": row["tariff"],
                        "reference": reference,
                        "notes": "Importado desde ventas Comfama.",
                    },
                )
                sale.save()
                created += int(was_created)
                updated += int(not was_created)

        self._print_summary(rows, start_date, end_date, deleted, created, updated, skipped, dry_run=False)

    def _print_summary(self, rows, start_date, end_date, deleted, created, updated, skipped, dry_run):
        label = "Vista previa" if dry_run else "Ventas Comfama importadas"
        self.stdout.write(f"{label}: {len(rows)} filas entre {start_date} y {end_date}.")
        self.stdout.write(f"Eliminadas: {deleted}, creadas: {created}, actualizadas: {updated}, omitidas: {skipped}.")
        by_date = {}
        for row in rows:
            by_date[row["sale_date"]] = by_date.get(row["sale_date"], 0) + 1
        for sale_date, count in sorted(by_date.items()):
            self.stdout.write(f"{sale_date}: {count} ventas")
