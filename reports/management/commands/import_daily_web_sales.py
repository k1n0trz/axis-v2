from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from reports.models import DailyChannelSale, DailyProductCategorySale, ProductCategory
from reports.services.sales_dashboard import ensure_uva_catalogs, uva_category_slug_from_product_name


PRODUCT_QTY_RE = re.compile(r"^\s*(\d+)\s*[xX\u00d7]\s*(.+?)\s*$")
CENT = Decimal("0.01")


def parse_decimal(value):
    raw = str(value or "").strip().replace(",", "")
    if not raw:
        return Decimal("0")
    try:
        return Decimal(raw)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


def parse_quantity(value):
    number = parse_decimal(value)
    return int(number) if number > 0 else 0


def parse_sale_date(value):
    if hasattr(value, "date"):
        return value.date()
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def normalize_header(value):
    raw = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    return "".join(char for char in raw if not unicodedata.combining(char))


def first_header_index(headers, aliases):
    for alias in aliases:
        if alias in headers:
            return headers.index(alias)
    return None


def category_for_product(product_name, file_name):
    slug = uva_category_slug_from_product_name(product_name)
    if not slug:
        return None
    category = ProductCategory.objects.filter(slug=slug).first()
    if category:
        return category
    return ProductCategory.objects.create(name=str(product_name).strip(), slug=slug, description=f"Categoria importada desde {file_name}.")


def split_products(raw_value):
    parts = [item.strip() for item in str(raw_value or "").split(", ") if item and str(item).strip()]
    parsed = []
    pending = ""
    for part in parts:
        if PRODUCT_QTY_RE.match(part):
            if pending:
                parsed.append(pending)
            pending = part
        elif pending:
            pending = f"{pending}, {part}"
        else:
            pending = part
    if pending:
        parsed.append(pending)

    items = []
    for item in parsed:
        match = PRODUCT_QTY_RE.match(item)
        if match:
            quantity = int(match.group(1))
            name = match.group(2).strip()
        else:
            quantity = 1
            name = item.strip()
        if name:
            items.append((name, quantity))
    return items or [("Sin categoria", 1)]


def allocate_amount_by_quantity(amount, items):
    total_items = sum(quantity for _, quantity in items) or 1
    amount_cents = int((amount * 100).to_integral_value(rounding=ROUND_HALF_UP))
    sign = -1 if amount_cents < 0 else 1
    remaining_cents = abs(amount_cents)
    allocations = []

    for index, (_, quantity) in enumerate(items):
        weighted_cents = remaining_cents * quantity
        base_cents = weighted_cents // total_items
        remainder = weighted_cents % total_items
        allocations.append({"index": index, "cents": base_cents, "remainder": remainder})

    cents_to_distribute = remaining_cents - sum(item["cents"] for item in allocations)
    for item in sorted(allocations, key=lambda value: (-value["remainder"], value["index"]))[:cents_to_distribute]:
        item["cents"] += 1

    cents_by_index = {item["index"]: item["cents"] for item in allocations}
    return [
        (product_name, quantity, Decimal(sign * cents_by_index[index]) / 100)
        for index, (product_name, quantity) in enumerate(items)
    ]


class Command(BaseCommand):
    help = "Importa ventas web diarias desde un Excel con columnas Fecha y Ventas COP."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path")
        parser.add_argument("--country", default="CO")
        parser.add_argument("--channel-slug", default="ecommerce-uva")
        parser.add_argument("--end-date")
        parser.add_argument("--replace-existing", action="store_true", help="Reemplaza ventas por categoria/canal en las fechas importadas.")

    def handle(self, *args, **options):
        path = Path(options["xlsx_path"]).expanduser()
        if not path.exists():
            raise CommandError(f"No existe el archivo: {path}")

        end_date = None
        if options.get("end_date"):
            end_date = datetime.strptime(options["end_date"], "%Y-%m-%d").date()

        catalogs = ensure_uva_catalogs()
        country = catalogs["countries"].get(options["country"])
        channel = catalogs["channels"].get(options["channel_slug"])
        if not country or not channel:
            raise CommandError("No fue posible resolver pais o canal para la importacion.")

        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        aggregated = defaultdict(lambda: {"sales_amount": Decimal("0"), "order_count": 0})
        category_aggregated = defaultdict(lambda: {"sales_amount": Decimal("0"), "quantity": 0, "products": set()})

        try:
            for sheet in workbook.worksheets:
                headers = [normalize_header(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
                fecha_idx = first_header_index(headers, ("fecha",))
                ventas_idx = first_header_index(headers, ("ventas", "ventas netas", "total", "importe"))
                quantity_idx = first_header_index(headers, ("cantidad", "articulos vendidos", "unidades"))
                if fecha_idx is None or ventas_idx is None:
                    raise CommandError("La hoja debe tener columnas Fecha y Ventas/Ventas netas.")
                products_idx = headers.index("producto(s)") if "producto(s)" in headers else None

                rows = sheet.iter_rows(min_row=2, values_only=True)
                for row in rows:
                    if not row or len(row) <= max(fecha_idx, ventas_idx) or not row[fecha_idx]:
                        continue
                    sale_date = parse_sale_date(row[fecha_idx])
                    if not sale_date:
                        continue
                    if end_date and sale_date > end_date:
                        continue
                    sale_amount = parse_decimal(row[ventas_idx])

                    if products_idx is not None and len(row) > products_idx:
                        items = split_products(row[products_idx])
                        if not items and quantity_idx is not None and len(row) > quantity_idx:
                            items = [("Sin categoria", parse_quantity(row[quantity_idx]) or 1)]
                        quantity_sign = -1 if sale_amount < 0 else 1
                        valid_order_amount = Decimal("0")
                        valid_order_has_product = False
                        for product_name, quantity, allocated_amount in allocate_amount_by_quantity(sale_amount, items):
                            category = category_for_product(product_name, path.name)
                            if not category:
                                continue
                            key = (category.id, sale_date)
                            category_aggregated[key]["category"] = category
                            category_aggregated[key]["sales_amount"] += allocated_amount
                            category_aggregated[key]["quantity"] += quantity * quantity_sign
                            category_aggregated[key]["products"].add(product_name)
                            valid_order_amount += allocated_amount
                            valid_order_has_product = True
                        if valid_order_has_product:
                            aggregated[sale_date]["sales_amount"] += valid_order_amount
                            aggregated[sale_date]["order_count"] += 1
                    else:
                        aggregated[sale_date]["sales_amount"] += sale_amount
                        aggregated[sale_date]["order_count"] += 1
        finally:
            workbook.close()

        created = 0
        updated = 0
        category_created = 0
        category_updated = 0
        category_deleted = 0
        with transaction.atomic():
            for sale_date, values in sorted(aggregated.items()):
                _, was_created = DailyChannelSale.objects.update_or_create(
                    business_unit=catalogs["business_unit"],
                    country=country,
                    channel=channel,
                    sale_date=sale_date,
                    defaults={
                        "sales_amount": values["sales_amount"],
                        "order_count": values["order_count"],
                        "source_type": DailyChannelSale.SourceType.IMPORTED,
                        "source_file": path.name,
                        "notes": "Importado desde ventas web de WordPress.",
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

            if options["replace_existing"] and category_aggregated:
                imported_dates = sorted({sale_date for _, sale_date in category_aggregated})
                category_deleted, _ = DailyProductCategorySale.objects.filter(
                    business_unit=catalogs["business_unit"],
                    country=country,
                    channel=channel,
                    sale_date__in=imported_dates,
                ).delete()

            for (_, sale_date), values in sorted(category_aggregated.items(), key=lambda item: (item[0][1], item[1]["category"].name)):
                _, was_created = DailyProductCategorySale.objects.update_or_create(
                    business_unit=catalogs["business_unit"],
                    country=country,
                    channel=channel,
                    category=values["category"],
                    sale_date=sale_date,
                    defaults={
                        "sales_amount": values["sales_amount"],
                        "original_amount": values["sales_amount"],
                        "original_currency": "COP",
                        "exchange_rate": 1,
                        "quantity": values["quantity"],
                        "source_type": DailyProductCategorySale.SourceType.IMPORTED,
                        "source_file": path.name,
                        "notes": "Productos: " + ", ".join(sorted(values["products"])),
                    },
                )
                if was_created:
                    category_created += 1
                else:
                    category_updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Importacion completada. Ventas diarias creadas: {created}. Actualizadas: {updated}. "
                f"Categorias creadas: {category_created}. Actualizadas: {category_updated}. Eliminadas: {category_deleted}."
            )
        )
