from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from openpyxl import load_workbook

from reports.models import DailyAdSpend
from reports.services.sales_dashboard import ensure_ad_platform_catalogs, ensure_uva_catalogs


def parse_decimal(value):
    raw = str(value or "").strip().replace(",", "")
    if not raw:
        return Decimal("0")
    try:
        return Decimal(raw)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


class Command(BaseCommand):
    help = "Importa inversion diaria de Meta Ads y Google Ads desde un Excel consolidado por dia."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path")
        parser.add_argument("--country", default="CO")
        parser.add_argument("--end-date")

    def handle(self, *args, **options):
        path = Path(options["xlsx_path"]).expanduser()
        if not path.exists():
            raise CommandError(f"No existe el archivo: {path}")

        end_date = None
        if options.get("end_date"):
            end_date = datetime.strptime(options["end_date"], "%Y-%m-%d").date()

        catalogs = ensure_uva_catalogs()
        platforms = ensure_ad_platform_catalogs()
        country = catalogs["countries"].get(options["country"])
        if not country:
            raise CommandError("No fue posible resolver el pais para la importacion.")

        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        created = 0
        updated = 0
        try:
            sheet = workbook.worksheets[0]
            rows = sheet.iter_rows(min_row=2, values_only=True)
            for row in rows:
                if not row or not row[0]:
                    continue
                spend_date = row[0].date() if hasattr(row[0], "date") else None
                if not spend_date:
                    continue
                if end_date and spend_date > end_date:
                    continue

                for slug, column_index in (("meta-ads", 1), ("google-ads", 2)):
                    amount = parse_decimal(row[column_index] if len(row) > column_index else 0)
                    _, was_created = DailyAdSpend.objects.update_or_create(
                        business_unit=catalogs["business_unit"],
                        country=country,
                        ad_platform=platforms[slug],
                        spend_date=spend_date,
                        defaults={
                            "spend_amount": amount,
                            "source_type": DailyAdSpend.SourceType.IMPORTED,
                            "source_file": path.name,
                            "notes": "Importado desde consolidado diario de pauta.",
                        },
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
        finally:
            workbook.close()

        self.stdout.write(self.style.SUCCESS(f"Importacion completada. Creados: {created}. Actualizados: {updated}."))
