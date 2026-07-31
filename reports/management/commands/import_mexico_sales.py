import re
from collections import defaultdict
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from reports.models import Channel, Country, DailyAdSpend, DailyChannelSale, DailyProductCategoryMetric, DailyProductCategorySale, ProductCategory
from reports.services.sales_dashboard import ensure_ad_platform_catalogs, ensure_uva_catalogs, parse_excel_date, uva_category_slug_from_product_name, uva_exchange_rate_for_country
from reports.utils.numbers import normalize_header, parse_decimal


PRODUCT_QTY_RE = re.compile(r"^\s*(\d+)\s*[xX×]\s*(.+?)\s*$")




def category_for_product(product_name, file_name):
    slug = uva_category_slug_from_product_name(product_name)
    if not slug:
        return None
    category = ProductCategory.objects.filter(slug=slug).first()
    if category:
        return category
    return ProductCategory.objects.create(name=str(product_name).strip(), slug=slug, description=f"Categoria importada desde {file_name}.")


def split_products(raw_value):
    parts = [item.strip() for item in str(raw_value or "").split(",") if item and str(item).strip()]
    parsed = []
    for part in parts:
        match = PRODUCT_QTY_RE.match(part)
        if match:
            quantity = int(match.group(1))
            name = match.group(2).strip()
        else:
            quantity = 1
            name = part
        parsed.append((name, quantity))
    return parsed or [("Copa Menstrual", 1)]


class Command(BaseCommand):
    help = "Importa ventas y pauta diaria de Uva Mexico desde un mismo Excel."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)
        parser.add_argument("--sales-sheet", default="Hoja1")
        parser.add_argument("--ads-sheet", default="Hoja2")
        parser.add_argument("--end-date", default="")

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        if not file_path.exists():
            raise CommandError(f"No existe el archivo: {file_path}")

        end_date = parse_excel_date(options["end_date"]) if options["end_date"] else None
        catalogs = ensure_uva_catalogs()
        platforms = ensure_ad_platform_catalogs()
        business_unit = catalogs["business_unit"]
        country = Country.objects.filter(code="MX").first()
        web_channel = Channel.objects.filter(slug="ecommerce-uva", business_unit=business_unit).first()
        if not country or not web_channel:
            raise CommandError("No fue posible resolver el catalogo base para Mexico.")
        mxn_to_cop = uva_exchange_rate_for_country("MX", "MXN")

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        sales_by_day = defaultdict(lambda: {"sales_cop": parse_decimal(0), "orders": 0})
        sales_by_category = defaultdict(lambda: {"sales_cop": parse_decimal(0), "sales_mxn": parse_decimal(0), "quantity": 0, "products": set()})
        skipped_sales = 0

        try:
            if options["sales_sheet"] not in workbook.sheetnames:
                raise CommandError(f"No existe la hoja '{options['sales_sheet']}'. Hojas disponibles: {', '.join(workbook.sheetnames)}")

            sheet = workbook[options["sales_sheet"]]
            headers = [normalize_header(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            required = {"fecha", "producto(s)", "articulos vendidos", "ventas netas", "ventas cop"}
            if not required.issubset(set(headers)):
                raise CommandError("La hoja de ventas Mexico debe tener Fecha, Producto(s), Articulos vendidos, Ventas netas y Ventas COP.")
            column_map = {name: headers.index(name) for name in required}

            for row in sheet.iter_rows(min_row=2, values_only=True):
                sale_dt = parse_excel_date(row[column_map["fecha"]])
                if not sale_dt:
                    skipped_sales += 1
                    continue
                if end_date and sale_dt > end_date:
                    skipped_sales += 1
                    continue

                sales_mxn = parse_decimal(row[column_map["ventas netas"]])
                sales_cop = parse_decimal(row[column_map["ventas cop"]])
                if sales_mxn:
                    sales_cop = sales_mxn * mxn_to_cop
                if not sales_cop and not sales_mxn:
                    skipped_sales += 1
                    continue

                items = split_products(row[column_map["producto(s)"]])
                total_items = sum(quantity for _, quantity in items) or 1
                valid_sales_cop = parse_decimal(0)
                for product_name, quantity in items:
                    category = category_for_product(product_name, file_path.name)
                    if not category:
                        continue
                    key = (category.id, sale_dt)
                    allocated_cop = sales_cop * quantity / total_items
                    allocated_mxn = sales_mxn * quantity / total_items
                    sales_by_category[key]["category"] = category
                    sales_by_category[key]["sales_cop"] += allocated_cop
                    sales_by_category[key]["sales_mxn"] += allocated_mxn
                    sales_by_category[key]["quantity"] += quantity
                    sales_by_category[key]["products"].add(product_name)
                    valid_sales_cop += allocated_cop

                if valid_sales_cop:
                    sales_by_day[sale_dt]["sales_cop"] += valid_sales_cop
                    sales_by_day[sale_dt]["orders"] += 1
        finally:
            workbook.close()

        day_created = 0
        day_updated = 0
        for sale_date, values in sorted(sales_by_day.items(), key=lambda item: item[0]):
            _, was_created = DailyChannelSale.objects.update_or_create(
                business_unit=business_unit,
                country=country,
                channel=web_channel,
                sale_date=sale_date,
                defaults={
                    "sales_amount": values["sales_cop"],
                    "order_count": values["orders"],
                    "source_type": DailyChannelSale.SourceType.IMPORTED,
                    "source_file": file_path.name,
                    "notes": "Importado desde WooCommerce Mexico.",
                },
            )
            if was_created:
                day_created += 1
            else:
                day_updated += 1

        category_created = 0
        category_updated = 0
        for (_, sale_date), values in sorted(sales_by_category.items(), key=lambda item: (item[0][1], item[1]["category"].name)):
            sale, was_created = DailyProductCategorySale.objects.update_or_create(
                business_unit=business_unit,
                country=country,
                channel=web_channel,
                category=values["category"],
                sale_date=sale_date,
                defaults={
                    "original_amount": values["sales_mxn"],
                    "original_currency": "MXN",
                    "exchange_rate": mxn_to_cop,
                    "quantity": values["quantity"],
                    "source_type": DailyProductCategorySale.SourceType.IMPORTED,
                    "source_file": file_path.name,
                    "notes": "Productos: " + ", ".join(sorted(values["products"])),
                },
            )
            sale.save()
            if was_created:
                category_created += 1
            else:
                category_updated += 1

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        spend_created = 0
        spend_updated = 0
        metric_created = 0
        metric_updated = 0
        skipped_ads = 0
        daily_spend = defaultdict(lambda: {"meta": parse_decimal(0), "google": parse_decimal(0)})
        try:
            if options["ads_sheet"] not in workbook.sheetnames:
                raise CommandError(f"No existe la hoja '{options['ads_sheet']}'. Hojas disponibles: {', '.join(workbook.sheetnames)}")
            sheet = workbook[options["ads_sheet"]]
            headers = [normalize_header(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            required = {
                "fecha",
                "producto",
                "cpa meta ads",
                "cpa google ads",
                "inversion meta ads",
                "inversion google ads",
                "inversion total",
                "inversion total cop",
            }
            if not required.issubset(set(headers)):
                raise CommandError("La hoja de pauta Mexico debe tener Fecha, Producto, CPA Meta Ads, CPA Google Ads, Inversion Meta Ads, Inversion Google Ads, Inversion Total e Inversion total COP.")
            column_map = {name: headers.index(name) for name in required}

            for row in sheet.iter_rows(min_row=2, values_only=True):
                metric_date = parse_excel_date(row[column_map["fecha"]])
                product_name = str(row[column_map["producto"]] or "").strip()
                if not metric_date or not product_name:
                    skipped_ads += 1
                    continue
                if end_date and metric_date > end_date:
                    skipped_ads += 1
                    continue

                category = category_for_product(product_name, file_path.name)
                if not category:
                    skipped_ads += 1
                    continue
                spend_meta_mxn = parse_decimal(row[column_map["inversion meta ads"]])
                spend_google_mxn = parse_decimal(row[column_map["inversion google ads"]])
                total_spend_mxn = parse_decimal(row[column_map["inversion total"]]) or (spend_meta_mxn + spend_google_mxn)
                total_spend_cop = total_spend_mxn * mxn_to_cop
                fx = mxn_to_cop
                spend_meta_cop = spend_meta_mxn * fx
                spend_google_cop = spend_google_mxn * fx
                daily_spend[metric_date]["meta"] += spend_meta_cop
                daily_spend[metric_date]["google"] += spend_google_cop

                sales_amount = sales_by_category.get((category.id, metric_date), {}).get("sales_cop", parse_decimal(0))
                cpa_meta = parse_decimal(row[column_map["cpa meta ads"]])
                cpa_google = parse_decimal(row[column_map["cpa google ads"]])
                _, was_created = DailyProductCategoryMetric.objects.update_or_create(
                    business_unit=business_unit,
                    country=country,
                    category=category,
                    metric_date=metric_date,
                    defaults={
                        "cpa_meta": cpa_meta * fx if cpa_meta else parse_decimal(0),
                        "cpa_google": cpa_google * fx if cpa_google else parse_decimal(0),
                        "spend_meta": spend_meta_cop,
                        "spend_google": spend_google_cop,
                        "total_spend": total_spend_cop,
                        "sales_amount": sales_amount,
                        "notes": "Valores CPA e inversion convertidos desde MXN a COP.",
                        "source_type": DailyProductCategoryMetric.SourceType.IMPORTED,
                        "source_file": file_path.name,
                    },
                )
                if was_created:
                    metric_created += 1
                else:
                    metric_updated += 1
        finally:
            workbook.close()

        for spend_date, amounts in sorted(daily_spend.items(), key=lambda item: item[0]):
            for platform, amount in ((platforms["meta-ads"], amounts["meta"]), (platforms["google-ads"], amounts["google"])):
                _, was_created = DailyAdSpend.objects.update_or_create(
                    business_unit=business_unit,
                    country=country,
                    ad_platform=platform,
                    spend_date=spend_date,
                    defaults={
                        "spend_amount": amount,
                        "source_type": DailyAdSpend.SourceType.IMPORTED,
                        "source_file": file_path.name,
                        "notes": "Importado desde pauta Mexico en MXN convertida a COP.",
                    },
                )
                if was_created:
                    spend_created += 1
                else:
                    spend_updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                "Importacion Mexico completada. "
                f"Ventas diarias creadas: {day_created}. Ventas diarias actualizadas: {day_updated}. "
                f"Ventas categoria creadas: {category_created}. Ventas categoria actualizadas: {category_updated}. "
                f"Pauta creada: {spend_created}. Pauta actualizada: {spend_updated}. "
                f"Metricas creadas: {metric_created}. Metricas actualizadas: {metric_updated}. "
                f"Omitidos ventas: {skipped_sales}. Omitidos pauta: {skipped_ads}."
            )
        )
