from datetime import date, datetime
from decimal import Decimal
import re

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from openpyxl import load_workbook

from reports.models import UserTask


DAY_RE = re.compile(r"^(lunes|martes|miercoles|miércoles|jueves|viernes|sabado|sábado|domingo)\s+(\d{1,2})$", re.IGNORECASE)
MONTHS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


class Command(BaseCommand):
    help = "Importa la agenda de EdiTrafficker desde el Excel historico y reparte 8 horas por dia."

    def add_arguments(self, parser):
        parser.add_argument("--file", default="data/agenda_editrafficker.xlsx")
        parser.add_argument("--username", default="EdiTrafficker")
        parser.add_argument("--year", type=int, default=2026)
        parser.add_argument("--sheet", default="Agenda")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        user = User.objects.filter(username__iexact=options["username"]).first()
        if not user:
            raise CommandError(f"No existe el usuario {options['username']}.")

        workbook = load_workbook(options["file"], data_only=True)
        if options["sheet"] not in workbook.sheetnames:
            raise CommandError(f"No existe la hoja {options['sheet']}.")
        sheet = workbook[options["sheet"]]

        day_columns = self._extract_day_columns(sheet, options["year"])
        rows = self._extract_tasks(sheet, day_columns)
        today = timezone.localdate()
        created = 0
        updated = 0

        for row in rows:
            status = UserTask.Status.COMPLETED if row["due_date"] < today else UserTask.Status.PENDING
            completed_at = timezone.now() if status == UserTask.Status.COMPLETED else None
            defaults = {
                "description": "Importado desde Agenda Editrafficker.xlsx",
                "registered_hours": row["hours"],
                "status": status,
                "completed_at": completed_at,
            }
            if options["dry_run"]:
                self.stdout.write(f"{row['due_date']} | {row['hours']}h | {row['title']}")
                continue
            _, was_created = UserTask.objects.update_or_create(
                created_by=user,
                assigned_to=user,
                due_date=row["due_date"],
                title=row["title"],
                defaults=defaults,
            )
            created += int(was_created)
            updated += int(not was_created)

        self.stdout.write(self.style.SUCCESS(f"Agenda importada. Creadas: {created}, actualizadas: {updated}, filas: {len(rows)}."))

    def _extract_day_columns(self, sheet, year):
        day_columns = []
        current_month = None
        previous_date = None

        for row_index in range(1, sheet.max_row + 1):
            row_values = [sheet.cell(row_index, column).value for column in range(1, sheet.max_column + 1)]
            month_names = [str(value).strip().lower() for value in row_values if isinstance(value, str) and str(value).strip().lower() in MONTHS]
            header_cells = []
            for column, value in enumerate(row_values, start=1):
                if isinstance(value, str) and DAY_RE.match(value.strip()):
                    header_cells.append((column, value.strip()))
            if not header_cells:
                if month_names:
                    current_month = MONTHS[month_names[0]]
                continue

            for column, label in header_cells:
                match = DAY_RE.match(label)
                day = int(match.group(2))
                month = current_month or 1
                candidate = date(year, month, day)
                if previous_date and candidate < previous_date:
                    month += 1
                    if month > 12:
                        month = 1
                        year += 1
                    candidate = date(year, month, day)
                day_columns.append({"row": row_index, "column": column, "date": candidate, "label": label})
                previous_date = candidate

            if month_names:
                current_month = MONTHS[month_names[0]]

        return day_columns

    def _extract_tasks(self, sheet, day_columns):
        rows = []
        for index, day_info in enumerate(day_columns):
            next_header_row = sheet.max_row + 1
            for next_day in day_columns[index + 1 :]:
                if next_day["row"] > day_info["row"]:
                    next_header_row = next_day["row"]
                    break
            titles = []
            for row_index in range(day_info["row"] + 1, next_header_row):
                value = sheet.cell(row_index, day_info["column"]).value
                if not value or not str(value).strip():
                    continue
                text = str(value).strip()
                if text.lower() in MONTHS:
                    continue
                titles.append(text)
            if not titles:
                continue
            base_hours = (Decimal("8.00") / Decimal(len(titles))).quantize(Decimal("0.01"))
            total = Decimal("0.00")
            for title in titles[:-1]:
                rows.append({"due_date": day_info["date"], "title": title, "hours": base_hours})
                total += base_hours
            rows.append({"due_date": day_info["date"], "title": titles[-1], "hours": Decimal("8.00") - total})
        return rows
