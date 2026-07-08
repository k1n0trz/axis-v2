from collections import defaultdict
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from reports.models import Channel, Country, DailyProductCategorySale, ProductCategory
from reports.services.sales_dashboard import ensure_uva_catalogs, parse_decimal, parse_excel_date, parse_quantity, uva_category_slug_from_product_name


class Command(BaseCommand):
    help = "Importa ventas WhatsApp Detal por categoria de producto desde Excel."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)
        parser.add_argument("--sheet", default="Hoja1")
        parser.add_argument("--country", default="CO")
        parser.add_argument("--channel-slug", default="whatsapp-uva-co")
        parser.add_argument("--end-date", default="")
        parser.add_argument("--create-categories", action="store_true")
        parser.add_argument("--replace-existing", action="store_true", help="Reemplaza las ventas existentes del canal en las fechas incluidas en el archivo.")

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        if not file_path.exists():
            raise CommandError(f"No existe el archivo: {file_path}")

        catalogs = ensure_uva_catalogs()
        business_unit = catalogs["business_unit"]
        country = Country.objects.filter(code=options["country"]).first()
        channel = Channel.objects.filter(slug=options["channel_slug"], business_unit=business_unit).first()
        if not country or not channel:
            raise CommandError("No fue posible resolver pais o canal para la importacion.")

        end_date = parse_excel_date(options["end_date"]) if options["end_date"] else None
        workbook = load_workbook(file_path, data_only=True, read_only=True)
        aggregated = defaultdict(lambda: {"sales_amount": parse_decimal(0), "quantity": 0, "products": set()})
        skipped = 0
        skipped_unknown = 0

        try:
            if options["sheet"] not in workbook.sheetnames:
                raise CommandError(f"No existe la hoja '{options['sheet']}'. Hojas disponibles: {', '.join(workbook.sheetnames)}")

            sheet = workbook[options["sheet"]]
            headers = [str(value or "").strip().lower() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            required = {"producto", "fecha", "cantidad", "valor"}
            if not required.issubset(set(headers)):
                raise CommandError("La hoja debe tener columnas PRODUCTO, FECHA, CANTIDAD y VALOR.")

            column_map = {name: headers.index(name) for name in required}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                product_name = str(row[column_map["producto"]] or "").strip()
                sale_date = parse_excel_date(row[column_map["fecha"]])
                if not product_name or not sale_date:
                    skipped += 1
                    continue
                if end_date and sale_date > end_date:
                    skipped += 1
                    continue

                category_slug = uva_category_slug_from_product_name(product_name)
                if not category_slug:
                    skipped_unknown += 1
                    continue
                category = ProductCategory.objects.filter(slug=category_slug).first()
                if not category and options["create_categories"]:
                    category = ProductCategory.objects.create(
                        name=product_name,
                        slug=category_slug,
                        description=f"Categoria importada desde {file_path.name}.",
                    )
                if not category:
                    skipped_unknown += 1
                    continue

                key = (category.id, sale_date)
                aggregated[key]["category"] = category
                aggregated[key]["sales_amount"] += parse_decimal(row[column_map["valor"]])
                aggregated[key]["quantity"] += parse_quantity(row[column_map["cantidad"]])
                aggregated[key]["products"].add(product_name)
        finally:
            workbook.close()

        created = 0
        updated = 0
        deleted = 0
        with transaction.atomic():
            if options["replace_existing"] and aggregated:
                imported_dates = sorted({sale_date for _, sale_date in aggregated})
                deleted, _ = DailyProductCategorySale.objects.filter(
                    business_unit=business_unit,
                    country=country,
                    channel=channel,
                    sale_date__in=imported_dates,
                ).delete()

            for (_, sale_date), values in sorted(aggregated.items(), key=lambda item: (item[0][1], item[1]["category"].name)):
                _, was_created = DailyProductCategorySale.objects.update_or_create(
                    business_unit=business_unit,
                    country=country,
                    channel=channel,
                    category=values["category"],
                    sale_date=sale_date,
                    defaults={
                        "sales_amount": values["sales_amount"],
                        "quantity": values["quantity"],
                        "source_type": DailyProductCategorySale.SourceType.IMPORTED,
                        "source_file": file_path.name,
                        "notes": "Productos: " + ", ".join(sorted(values["products"])),
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Importacion WhatsApp Detal completada. Creados: {created}. Actualizados: {updated}. Eliminados: {deleted}. Omitidos: {skipped}. Productos sin categoria: {skipped_unknown}."
            )
        )
