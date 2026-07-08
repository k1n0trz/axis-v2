import json

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from reports.utils import onedrive


class Command(BaseCommand):
    help = "Inspecciona una hoja de OneDrive y muestra encabezados y primeras filas."

    def add_arguments(self, parser):
        parser.add_argument("--sheet", default="")
        parser.add_argument("--drive-path", default="")
        parser.add_argument("--rows", type=int, default=10)

    def handle(self, *args, **options):
        drive_path = options["drive_path"] or getattr(settings, "ONEDRIVE_SHARED_SALES_FILE_PATH", "")
        if not drive_path:
            raise CommandError("Falta --drive-path o ONEDRIVE_SHARED_SALES_FILE_PATH.")

        token_payload = onedrive.refresh_access_token()
        buffer = onedrive.download_file_content_by_path(token_payload["access_token"], drive_path)
        workbook = load_workbook(filename=buffer, read_only=True, data_only=True)
        try:
            sheet = workbook[options["sheet"]] if options["sheet"] else workbook.active
            rows = list(sheet.iter_rows(min_row=1, max_row=options["rows"], values_only=True))
            payload = {
                "sheet": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "rows": rows,
            }
        finally:
            workbook.close()

        self.stdout.write(json.dumps(payload, indent=2, default=str, ensure_ascii=False))
