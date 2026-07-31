from collections import defaultdict
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from reports.models import Channel, Country, DailyAdSpend, DailyProductCategoryMetric, DailyProductCategorySale, ProductCategory
from reports.services.sales_dashboard import ECUADOR_USD_TO_COP_RATE, ensure_ad_platform_catalogs, ensure_uva_catalogs, parse_excel_date, uva_category_slug_from_product_name
from reports.utils.numbers import normalize_header, parse_decimal, parse_quantity
from reports.utils.unit_price_audit import AuditorDePrecioUnitario


CHANNEL_BY_LABEL = {
    "pagina": "ecommerce-uva",
    "pagina web": "ecommerce-uva",
    "página": "ecommerce-uva",
    "página web": "ecommerce-uva",
    "whatsapp": "whatsapp-uva-ec",
    "web": "ecommerce-uva",
}




def category_for_product(product_name, file_name):
    slug = uva_category_slug_from_product_name(product_name)
    if not slug:
        return None
    category = ProductCategory.objects.filter(slug=slug).first()
    if category:
        return category
    display_name = {
        "lubricantes": "Lubricantes",
        "higiene-intima": "Higiene Intima",
    }.get(slug, str(product_name).strip())
    return ProductCategory.objects.create(name=display_name, slug=slug, description=f"Categoria importada desde {file_name}.")


class Command(BaseCommand):
    help = "Importa ventas, inversion diaria y metricas por categoria de Uva Ecuador desde un mismo Excel."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)
        parser.add_argument("--sheet", default="Hoja1")
        parser.add_argument("--ads-sheet", default="Hoja2")
        parser.add_argument("--end-date", default="")

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        if not file_path.exists():
            raise CommandError(f"No existe el archivo: {file_path}")

        end_date = parse_excel_date(options["end_date"]) if options["end_date"] else None
        catalogs = ensure_uva_catalogs()
        platforms = ensure_ad_platform_catalogs()
        business_unit = catalogs["business_unit"]
        country = Country.objects.filter(code="EC").first()
        if not country:
            raise CommandError("No existe el pais Ecuador con codigo EC.")

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        aggregated_sales = defaultdict(lambda: {"sales_cop": parse_decimal(0), "amount_usd": parse_decimal(0), "quantity": 0, "products": set()})
        auditor = AuditorDePrecioUnitario()
        sales_skipped = 0

        try:
            if options["sheet"] not in workbook.sheetnames:
                raise CommandError(f"No existe la hoja '{options['sheet']}'. Hojas disponibles: {', '.join(workbook.sheetnames)}")

            sheet = workbook[options["sheet"]]
            headers = [normalize_header(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            required = {"producto", "fecha", "canal", "cantidad"}
            has_detailed_totals = {"valor", "envio", "total cop"}.issubset(set(headers))
            has_simple_sales = "ventas" in headers
            if not required.issubset(set(headers)) or not (has_detailed_totals or has_simple_sales):
                raise CommandError("La hoja de ventas debe tener PRODUCTO, FECHA, Canal, CANTIDAD y VENTAS, o VALOR, ENVIO y Total COP.")
            column_map = {name: headers.index(name) for name in required}
            if has_detailed_totals:
                column_map.update({name: headers.index(name) for name in ("valor", "envio", "total cop")})
            if has_simple_sales:
                column_map["ventas"] = headers.index("ventas")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                product_name = str(row[column_map["producto"]] or "").strip()
                sale_date = parse_excel_date(row[column_map["fecha"]])
                channel_label = str(row[column_map["canal"]] or "").strip().lower()
                if not product_name or not sale_date or not channel_label:
                    sales_skipped += 1
                    continue
                if end_date and sale_date > end_date:
                    sales_skipped += 1
                    continue

                channel_slug = CHANNEL_BY_LABEL.get(channel_label)
                channel = Channel.objects.filter(slug=channel_slug, business_unit=business_unit).first() if channel_slug else None
                if not channel:
                    sales_skipped += 1
                    continue

                category = category_for_product(product_name, file_path.name)
                if not category:
                    sales_skipped += 1
                    continue
                row_quantity = parse_quantity(row[column_map["cantidad"]])
                if has_detailed_totals:
                    # VALOR es precio unitario: sin multiplicar por CANTIDAD,
                    # toda linea de 2 o mas unidades se contaba como una sola.
                    unit_value = parse_decimal(row[column_map["valor"]])
                    # La hoja tiene la convencion mezclada: algunas filas ya traen
                    # el total y multiplicarlas las duplica. El auditor las señala
                    # al final; no se corrigen aqui porque la correccion valida es
                    # editar el archivo fuente.
                    auditor.registrar(product_name, row_quantity, unit_value, referencia=sale_date.isoformat())
                    amount_usd = unit_value * (row_quantity if row_quantity > 0 else 1) + parse_decimal(row[column_map["envio"]])
                    sales_cop = amount_usd * ECUADOR_USD_TO_COP_RATE if amount_usd else parse_decimal(row[column_map["total cop"]])
                else:
                    amount_usd = parse_decimal(0)
                    sales_cop = parse_decimal(row[column_map["ventas"]])
                key = (category.id, channel.id, sale_date)
                aggregated_sales[key]["category"] = category
                aggregated_sales[key]["channel"] = channel
                aggregated_sales[key]["sales_cop"] += sales_cop
                aggregated_sales[key]["amount_usd"] += amount_usd
                aggregated_sales[key]["quantity"] += row_quantity
                aggregated_sales[key]["products"].add(product_name)
        finally:
            workbook.close()

        sales_created = 0
        sales_updated = 0
        for (_, _, sale_date), values in sorted(aggregated_sales.items(), key=lambda item: (item[0][2], item[1]["category"].name)):
            exchange_rate = values["sales_cop"] / values["amount_usd"] if values["amount_usd"] else parse_decimal(0)
            original_currency = "USD" if values["amount_usd"] else "COP"
            original_amount = values["amount_usd"] if values["amount_usd"] else values["sales_cop"]
            sale, was_created = DailyProductCategorySale.objects.update_or_create(
                business_unit=business_unit,
                country=country,
                channel=values["channel"],
                category=values["category"],
                sale_date=sale_date,
                defaults={
                    "sales_amount": values["sales_cop"],
                    "original_amount": original_amount,
                    "original_currency": original_currency,
                    "exchange_rate": exchange_rate or 1,
                    "quantity": values["quantity"],
                    "source_type": DailyProductCategorySale.SourceType.IMPORTED,
                    "source_file": file_path.name,
                    "notes": "Productos: " + ", ".join(sorted(values["products"])),
                },
            )
            sale.save()
            if was_created:
                sales_created += 1
            else:
                sales_updated += 1

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        ads_created = 0
        ads_updated = 0
        metric_created = 0
        metric_updated = 0
        metric_skipped = 0
        try:
            if options["ads_sheet"] not in workbook.sheetnames:
                self.stdout.write(f"No existe la hoja '{options['ads_sheet']}'. Se omite pauta/metrica y solo se importan ventas por categoria.")
                workbook.close()
                sospechosas = self._avisar_valores_sospechosos(auditor)
                self.stdout.write(
                    self.style.SUCCESS(
                        "Importacion Ecuador completada. "
                        f"Ventas creadas: {sales_created}. Ventas actualizadas: {sales_updated}. "
                        "Pauta creada: 0. Pauta actualizada: 0. Metricas creadas: 0. Metricas actualizadas: 0. "
                        f"Omitidos ventas: {sales_skipped}. Omitidos pauta: 0. "
                        f"Filas con VALOR sospechoso: {len(sospechosas)}."
                    )
                )
                return

            sheet = workbook[options["ads_sheet"]]
            headers = [normalize_header(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            required = {
                "fecha",
                "producto",
                "mensajes wp",
                "cpl meta ads",
                "cpa google ads",
                "inversion meta ads",
                "inversion google ads",
                "total gasto",
            }
            if not required.issubset(set(headers)):
                raise CommandError("La hoja de pauta Ecuador debe tener FECHA, PRODUCTO, MENSAJES WP, CPL Meta Ads, CPA Google Ads, Inversion Meta Ads, Inversion Google Ads y Total gasto.")
            column_map = {name: headers.index(name) for name in required}

            meta_platform = platforms["meta-ads"]
            google_platform = platforms["google-ads"]
            sales_by_date_category = defaultdict(lambda: parse_decimal(0))
            daily_spend_totals = defaultdict(lambda: {"meta": parse_decimal(0), "google": parse_decimal(0)})
            for (_, _, sale_date), values in aggregated_sales.items():
                sales_by_date_category[(values["category"].id, sale_date)] += values["sales_cop"]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                metric_date = parse_excel_date(row[column_map["fecha"]])
                product_name = str(row[column_map["producto"]] or "").strip()
                if not metric_date or not product_name:
                    metric_skipped += 1
                    continue
                if end_date and metric_date > end_date:
                    metric_skipped += 1
                    continue

                category = category_for_product(product_name, file_path.name)
                if not category:
                    metric_skipped += 1
                    continue
                spend_meta = parse_decimal(row[column_map["inversion meta ads"]])
                spend_google = parse_decimal(row[column_map["inversion google ads"]])
                total_spend = parse_decimal(row[column_map["total gasto"]]) or (spend_meta + spend_google)
                daily_spend_totals[metric_date]["meta"] += spend_meta
                daily_spend_totals[metric_date]["google"] += spend_google

                _, was_created = DailyProductCategoryMetric.objects.update_or_create(
                    business_unit=business_unit,
                    country=country,
                    category=category,
                    metric_date=metric_date,
                    defaults={
                        "cpa_meta": parse_decimal(row[column_map["cpl meta ads"]]),
                        "cpa_google": parse_decimal(row[column_map["cpa google ads"]]),
                        "spend_meta": spend_meta,
                        "spend_google": spend_google,
                        "total_spend": total_spend,
                        "sales_amount": sales_by_date_category.get((category.id, metric_date), parse_decimal(0)),
                        "notes": f"Mensajes WP: {parse_quantity(row[column_map['mensajes wp']])}. CPL Meta almacenado en campo Meta por compatibilidad visual.",
                        "source_type": DailyProductCategoryMetric.SourceType.IMPORTED,
                        "source_file": file_path.name,
                    },
                )
                if was_created:
                    metric_created += 1
                else:
                    metric_updated += 1

            for spend_date, spend_values in sorted(daily_spend_totals.items(), key=lambda item: item[0]):
                for platform, amount in ((meta_platform, spend_values["meta"]), (google_platform, spend_values["google"])):
                    _, was_created = DailyAdSpend.objects.update_or_create(
                        business_unit=business_unit,
                        country=country,
                        ad_platform=platform,
                        spend_date=spend_date,
                        defaults={
                            "spend_amount": amount,
                            "source_type": DailyAdSpend.SourceType.IMPORTED,
                            "source_file": file_path.name,
                            "notes": "Importado desde hoja de pauta Ecuador.",
                        },
                    )
                    if was_created:
                        ads_created += 1
                    else:
                        ads_updated += 1
        finally:
            workbook.close()

        sospechosas = self._avisar_valores_sospechosos(auditor)

        self.stdout.write(
            self.style.SUCCESS(
                "Importacion Ecuador completada. "
                f"Ventas creadas: {sales_created}. Ventas actualizadas: {sales_updated}. "
                f"Pauta creada: {ads_created}. Pauta actualizada: {ads_updated}. "
                f"Metricas creadas: {metric_created}. Metricas actualizadas: {metric_updated}. "
                f"Omitidos ventas: {sales_skipped}. Omitidos pauta: {metric_skipped}. "
                f"Filas con VALOR sospechoso: {len(sospechosas)}."
            )
        )

    def _avisar_valores_sospechosos(self, auditor):
        """Imprime las filas donde VALOR parece ser el total de la linea.

        Solo avisa. Corregir aqui seria peor: el archivo fuente y Axis dirian
        cosas distintas y nadie sabria cual creer.
        """
        sospechosas = auditor.sospechosas()
        if sospechosas:
            self.stdout.write(
                self.style.WARNING(
                    f"\n{len(sospechosas)} filas parecen traer el TOTAL de la linea en VALOR, no el precio unitario. "
                    "Si es asi, se estan contando de mas. Hay que corregirlas en el archivo fuente, "
                    "dejando el precio por unidad:"
                )
            )
            for aviso in sospechosas:
                self.stdout.write(self.style.WARNING(f"  {aviso['mensaje']}"))
        return sospechosas
