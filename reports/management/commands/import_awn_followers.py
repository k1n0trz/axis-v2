from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from reports.models import AwnInternationalFollowerMetric, Country
from reports.services.sales_dashboard import ensure_uva_catalogs, parse_excel_date
from reports.utils.numbers import normalize_header, parse_decimal




COUNTRY_ALIASES = {
    "ecuador": "EC",
    "mexico": "MX",
}


def find_column(headers, aliases, required=True):
    for alias in aliases:
        normalized = normalize_header(alias)
        if normalized in headers:
            return headers.index(normalized)
    if required:
        raise CommandError(f"La hoja debe incluir una columna compatible con: {', '.join(aliases)}.")
    return None


def merge_source_files(*sources):
    parts = []
    for source in sources:
        parts.extend(part.strip() for part in str(source or "").split(";") if part.strip())
    return "; ".join(dict.fromkeys(parts))


def import_awn_followers_workbook(workbook_source, source_name, sheet_name="Hoja1", end_date=None, preserve_spend=False):
    ensure_uva_catalogs()
    workbook = load_workbook(workbook_source, data_only=True, read_only=True)
    created = 0
    updated = 0
    skipped = 0
    try:
        if sheet_name not in workbook.sheetnames:
            if sheet_name == "Hoja1" and len(workbook.sheetnames) == 1:
                sheet_name = workbook.sheetnames[0]
            else:
                raise CommandError(f"No existe la hoja '{sheet_name}'. Hojas disponibles: {', '.join(workbook.sheetnames)}")

        sheet = workbook[sheet_name]
        headers = [normalize_header(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        column_map = {
            "country": find_column(headers, ["Pais", "PaÃ­s"]),
            "date": find_column(headers, ["Fecha"]),
            "visits": find_column(headers, ["Visitas al perfil de instagram", "Visitas Al Perfil De Instagram"]),
            "followers": find_column(headers, ["Nuevos seguidores", "Seguidores Nuevos"]),
            "spend": find_column(headers, ["Inversion", "InversiÃ³n"], required=False),
            "cpr": find_column(headers, ["CPR (COP)", "Cpr"], required=False),
            "cps": find_column(headers, ["CPS (Costo por seguidor) COP", "Cps"], required=False),
        }

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            country_label = normalize_header(row[column_map["country"]])
            metric_date = parse_excel_date(row[column_map["date"]])
            country_code = COUNTRY_ALIASES.get(country_label)
            if not country_code or not metric_date:
                skipped += 1
                continue
            if end_date and metric_date > end_date:
                skipped += 1
                continue

            country = Country.objects.filter(code=country_code).first()
            if not country:
                skipped += 1
                continue

            visits = int(parse_decimal(row[column_map["visits"]]) or 0)
            followers = int(parse_decimal(row[column_map["followers"]]) or 0)
            workbook_spend = parse_decimal(row[column_map["spend"]]) if column_map["spend"] is not None else parse_decimal(0)
            workbook_cpr = parse_decimal(row[column_map["cpr"]]) if column_map["cpr"] is not None else parse_decimal(0)
            workbook_cps = parse_decimal(row[column_map["cps"]]) if column_map["cps"] is not None else parse_decimal(0)
            if not workbook_spend:
                if workbook_cps and followers:
                    workbook_spend = workbook_cps * followers
                elif workbook_cpr and visits:
                    workbook_spend = workbook_cpr * visits

            existing = AwnInternationalFollowerMetric.objects.filter(country=country, metric_date=metric_date).first()
            spend = existing.spend_amount if preserve_spend and existing and existing.spend_amount else workbook_spend
            cpr = (spend / visits) if visits else parse_decimal(0)
            cps = (spend / followers) if followers else parse_decimal(0)
            if not preserve_spend:
                cpr = workbook_cpr or cpr
                cps = workbook_cps or cps

            _, was_created = AwnInternationalFollowerMetric.objects.update_or_create(
                country=country,
                metric_date=metric_date,
                defaults={
                    "instagram_profile_visits": visits,
                    "new_followers": followers,
                    "spend_amount": spend,
                    "cpr": cpr,
                    "cps": cps,
                    "source_type": AwnInternationalFollowerMetric.SourceType.IMPORTED,
                    "source_file": merge_source_files(existing.source_file if existing else "", source_name),
                    "source_row": row_index,
                    "notes": (
                        "Visitas y seguidores importados desde reporte Awareness; inversion preservada desde Meta cuando existe."
                        if preserve_spend
                        else "Importado desde reporte de seguidores Instagram."
                    ),
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1
    finally:
        workbook.close()
    return {"created": created, "updated": updated, "skipped": skipped, "sheet": sheet_name}


class Command(BaseCommand):
    help = "Importa los datos diarios de seguidores de Instagram para Awn Internacional."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)
        parser.add_argument("--sheet", default="Hoja1")
        parser.add_argument("--end-date", default="")

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        if not file_path.exists():
            raise CommandError(f"No existe el archivo: {file_path}")

        end_date = parse_excel_date(options["end_date"]) if options["end_date"] else None
        result = import_awn_followers_workbook(
            file_path,
            file_path.name,
            sheet_name=options["sheet"],
            end_date=end_date,
            preserve_spend=False,
        )

        self.stdout.write(
            self.style.SUCCESS(
                "Importacion Awn Internacional completada. "
                f"Creados: {result['created']}. Actualizados: {result['updated']}. Omitidos: {result['skipped']}."
            )
        )
