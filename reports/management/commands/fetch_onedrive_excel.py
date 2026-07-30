import json
from collections import defaultdict
from datetime import date, timedelta
from decimal import Decimal

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from reports.integrations.axis_sync import AxisSyncService
from reports.integrations.clients import ExchangeRateClient, MicrosoftGraphClient, load_json_mapping
from reports.integrations.schema import CategorySaleRecord, ChannelSaleRecord
from reports.services.sales_dashboard import category_slug_from_product_name, parse_excel_date, uva_category_slug_from_product_name, uva_exchange_rate_for_country
from reports.utils import onedrive
from reports.utils.numbers import normalize_header, parse_decimal




def display_name_for_category(slug):
    labels = {
        "copa-menstrual": "Copa Menstrual",
        "disco-menstrual": "Disco Menstrual",
        "dilatadores-vaginales": "Dilatadores Vaginales",
        "higiene-intima": "Higiene Intima",
        "kits": "Kits",
        "lubricantes": "Lubricantes",
        "panties-menstruales": "Panties Menstruales",
        "cubrepezones": "Cubrepezones",
        "copa-con-esterilizador-electrico": "Copa con Esterilizador Eléctrico",
    }
    return labels.get(slug, slug.replace("-", " ").title())


def channel_slug_for_row(country_code, row_value, fallback_slug):
    normalized = normalize_header(row_value)
    if not normalized:
        return fallback_slug
    if "whatsapp" in normalized:
        return f"whatsapp-uva-{country_code.lower()}"
    if "pagina" in normalized or "web" in normalized:
        return "ecommerce-uva"
    return fallback_slug


class Command(BaseCommand):
    help = "Descarga un Excel desde OneDrive via Microsoft Graph y consolida ventas."

    def _resolve_target_dates(self, options):
        if options.get("date"):
            return [date.fromisoformat(options["date"])]
        if not options.get("date_from") or not options.get("date_to"):
            raise CommandError("Debes enviar --date o el rango --date-from/--date-to.")
        start_date = date.fromisoformat(options["date_from"])
        end_date = date.fromisoformat(options["date_to"])
        if start_date > end_date:
            raise CommandError("--date-from no puede ser mayor que --date-to.")
        dates = []
        current = start_date
        while current <= end_date:
            dates.append(current)
            current += timedelta(days=1)
        return dates

    def add_arguments(self, parser):
        parser.add_argument("--date", default="")
        parser.add_argument("--date-from", default="")
        parser.add_argument("--date-to", default="")
        parser.add_argument("--country", required=True)
        parser.add_argument("--business-unit", default="uva")
        parser.add_argument("--channel-slug", default="whatsapp-uva-co")
        parser.add_argument("--sheet", default="")
        parser.add_argument("--drive-user-id", default="")
        parser.add_argument("--drive-path", default="")
        parser.add_argument("--tenant-id", default="")
        parser.add_argument("--client-id", default="")
        parser.add_argument("--client-secret", default="")
        parser.add_argument("--column-map", default="")
        parser.add_argument("--default-currency", default="COP")
        parser.add_argument("--exchange-rate", default="1")
        parser.add_argument("--auth-mode", choices=("delegated", "application"), default=getattr(settings, "ONEDRIVE_AUTH_MODE", "delegated"))
        parser.add_argument("--sync-axis", action="store_true")

    def handle(self, *args, **options):
        tenant_id = options["tenant_id"] or getattr(settings, "ONEDRIVE_TENANT_ID", "")
        client_id = options["client_id"] or getattr(settings, "ONEDRIVE_CLIENT_ID", "")
        client_secret = options["client_secret"] or getattr(settings, "ONEDRIVE_CLIENT_SECRET", "")
        drive_user_id = options["drive_user_id"] or getattr(settings, "ONEDRIVE_USER_ID", "")
        default_path = getattr(settings, "ONEDRIVE_WHATSAPP_FILE_PATH", "") or getattr(settings, "ONEDRIVE_SHARED_SALES_FILE_PATH", "")
        if options["country"].upper() == "EC":
            default_path = getattr(settings, "ONEDRIVE_ECUADOR_FILE_PATH", "") or getattr(settings, "ONEDRIVE_SHARED_SALES_FILE_PATH", "") or default_path
            if not options["sheet"]:
                options["sheet"] = getattr(settings, "ONEDRIVE_ECUADOR_SHEET", "Ecuador")
        elif options["country"].upper() == "CO" and not options["sheet"]:
            options["sheet"] = getattr(settings, "ONEDRIVE_COLOMBIA_SHEET", "Colombia")
        drive_path = options["drive_path"] or default_path
        if not all([tenant_id, client_id, client_secret, drive_path]):
            raise CommandError("Faltan credenciales o ubicacion del archivo en OneDrive.")

        mapping = load_json_mapping(options["column_map"]) if options["column_map"] else {
            "date": ["fecha"],
            "product": ["producto"],
            # `amount` son columnas que ya traen el total de la linea.
            # `unit_amount` son precios por unidad: hay que multiplicarlos por
            # CANTIDAD. En las hojas de despachos, VALOR es precio unitario, y
            # tratarlo como total hacia que toda linea de 2 o mas unidades se
            # contara como una sola.
            "amount": ["ventas", "total cop"],
            "unit_amount": ["valor"],
            "shipping": ["envio"],
            "quantity": ["cantidad"],
            "currency": ["moneda"],
            "channel": ["centro de costos", "origen", "canal"],
        }
        target_dates = self._resolve_target_dates(options)
        target_date_set = set(target_dates)
        fallback_currency = options["default_currency"].upper()
        if fallback_currency == "COP" and options["country"].upper() == "EC":
            fallback_currency = "USD"
        rate = Decimal(str(options["exchange_rate"]))
        aggregated = defaultdict(lambda: {"sales": Decimal("0"), "original": Decimal("0"), "qty": 0, "products": set()})
        channel_totals = defaultdict(lambda: {"sales": Decimal("0"), "original": Decimal("0"), "qty": 0, "rows": 0})

        def fallback_amount(normalized_row):
            keys = list(normalized_row.keys())
            quantity_keys = [normalize_header(alias) for alias in mapping.get("quantity", [])]
            shipping_keys = [normalize_header(alias) for alias in mapping.get("shipping", [])]
            quantity_index = next((keys.index(key) for key in quantity_keys if key in normalized_row), -1)
            shipping_index = next((keys.index(key) for key in shipping_keys if key in normalized_row), len(keys))
            for key in keys[quantity_index + 1 : shipping_index]:
                value = parse_decimal(normalized_row.get(key))
                if value:
                    return value
            return Decimal("0")

        workbook = None
        if options["auth_mode"] == "delegated":
            token_payload = onedrive.refresh_access_token()
            buffer = onedrive.download_file_content_by_path(
                token_payload["access_token"],
                drive_path,
                user_id=drive_user_id or None,
            )
            workbook = load_workbook(filename=buffer, read_only=True, data_only=True)
            sheet = workbook[options["sheet"]] if options["sheet"] else workbook.active
            headers = [str(cell or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            rows = (dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True))
        else:
            if not drive_user_id:
                raise CommandError("En modo application necesitas ONEDRIVE_USER_ID.")
            client = MicrosoftGraphClient(tenant_id, client_id, client_secret)
            rows = client.workbook_rows(drive_user_id, drive_path, sheet_name=options["sheet"] or None)

        fx_client = None
        fx_cache = {}

        try:
            for row in rows:
                normalized_row = {normalize_header(key): value for key, value in row.items()}

                def pick(alias_list):
                    if isinstance(alias_list, str):
                        alias_list = [alias_list]
                    for alias in alias_list or []:
                        normalized_alias = normalize_header(alias)
                        if normalized_alias in normalized_row:
                            return normalized_row.get(normalized_alias)
                    return None

                raw_date = pick(mapping["date"])
                row_date = parse_excel_date(raw_date)
                if row_date not in target_date_set:
                    continue
                product_name = str(pick(mapping["product"]) or "").strip()
                if not product_name:
                    continue
                qty = int(parse_decimal(pick(mapping["quantity"])) or 0)
                # Una linea sin cantidad legible se cuenta como una unidad, que
                # es el comportamiento anterior.
                qty_factor = qty if qty > 0 else 1

                line_amount = parse_decimal(pick(mapping.get("amount", [])))
                if not line_amount:
                    # VALOR (y el barrido de respaldo, que apunta a la misma
                    # columna) es precio unitario: se multiplica por la cantidad.
                    unit_amount = parse_decimal(pick(mapping.get("unit_amount", [])))
                    if not unit_amount:
                        unit_amount = fallback_amount(normalized_row)
                    line_amount = unit_amount * qty_factor

                shipping_amount = parse_decimal(pick(mapping.get("shipping", [])))
                original_amount = line_amount + shipping_amount
                currency_value = pick(mapping.get("currency", []))
                currency = str(currency_value or fallback_currency).upper()
                row_channel_slug = channel_slug_for_row(options["country"], pick(mapping.get("channel", [])), options["channel_slug"])
                effective_rate = uva_exchange_rate_for_country(options["country"].upper(), currency, rate)
                if currency != "COP" and effective_rate == Decimal("1"):
                    cache_key = (currency, row_date.isoformat())
                    if cache_key not in fx_cache:
                        if fx_client is None:
                            fx_client = ExchangeRateClient(
                                getattr(settings, "EXCHANGE_RATE_API_URL", "https://api.exchangerate.host"),
                                api_key=getattr(settings, "EXCHANGE_RATE_API_KEY", ""),
                            )
                        fx_cache[cache_key] = fx_client.convert(currency, "COP", Decimal("1"), target_date=row_date)
                    effective_rate = fx_cache[cache_key]
                sales_cop = original_amount if currency == "COP" else original_amount * effective_rate
                total_key = (row_date, row_channel_slug, currency, effective_rate)
                channel_totals[total_key]["sales"] += sales_cop
                channel_totals[total_key]["original"] += original_amount
                channel_totals[total_key]["qty"] += qty
                channel_totals[total_key]["rows"] += 1
                slug = (
                    uva_category_slug_from_product_name(product_name)
                    if str(options["business_unit"]).strip().lower() == "uva"
                    else category_slug_from_product_name(product_name)
                )
                if not slug:
                    continue
                key = (slug, row_date, row_channel_slug, currency, effective_rate)
                aggregated[key]["sales"] += sales_cop
                aggregated[key]["original"] += original_amount
                aggregated[key]["qty"] += qty
                aggregated[key]["products"].add(product_name)
        finally:
            if workbook is not None:
                workbook.close()

        for sale_date, channel_slug, currency, effective_rate in sorted(channel_totals):
            categorized_sales = sum(
                values["sales"]
                for (_, row_sale_date, row_channel_slug, row_currency, row_rate), values in aggregated.items()
                if row_sale_date == sale_date and row_channel_slug == channel_slug and row_currency == currency and row_rate == effective_rate
            )
            residual_sales = channel_totals[(sale_date, channel_slug, currency, effective_rate)]["sales"] - categorized_sales
            if residual_sales:
                residual_original = residual_sales if currency == "COP" else residual_sales / effective_rate
                key = ("otros-uva", sale_date, channel_slug, currency, effective_rate)
                aggregated[key]["sales"] += residual_sales
                aggregated[key]["original"] += residual_original
                aggregated[key]["products"].add("Ajuste ventas sin categoria")

        channel_records = [
            ChannelSaleRecord(
                business_unit_slug=options["business_unit"],
                country_code=options["country"].upper(),
                channel_slug=channel_slug,
                sale_date=sale_date,
                sales_amount=values["sales"],
                order_count=values["rows"],
                units=values["qty"],
                source_file=drive_path,
                notes="Importado desde OneDrive.",
            )
            for (sale_date, channel_slug, _, _), values in sorted(channel_totals.items())
        ]
        records = [
            CategorySaleRecord(
                business_unit_slug=options["business_unit"],
                country_code=options["country"].upper(),
                channel_slug=channel_slug,
                category_slug=slug,
                category_name=display_name_for_category(slug),
                sale_date=sale_date,
                sales_amount=values["sales"],
                original_amount=values["original"],
                original_currency=currency,
                exchange_rate=effective_rate,
                quantity=values["qty"],
                source_file=drive_path,
                notes="Productos fuente: " + ", ".join(sorted(values["products"])),
            )
            for (slug, sale_date, channel_slug, currency, effective_rate), values in sorted(aggregated.items())
        ]

        if options["sync_axis"]:
            sync = AxisSyncService()
            sync.sync_channel_sales(channel_records)
            sync.sync_category_sales(records)

        self.stdout.write(
            json.dumps(
                {
                    "channel_sales": [item.to_dict() for item in channel_records],
                    "category_sales": [item.to_dict() for item in records],
                },
                indent=2,
                default=str,
            )
        )
