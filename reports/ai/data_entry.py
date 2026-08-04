"""Registrar un dato hablando, sin archivo.

Esto es lo que faltaba. Las Etapas G y H asumieron que actualizar datos era subir un
Excel, y para la mayoria de la gente no lo es: es decir "ayer el gasto de Meta en Bali
fue de 320.000" y que quede. El archivo es un camino, no el camino.

Como funciona, igual que todo lo que escribe en Axis:

**Lista blanca por tipo de dato.** Cada entrada declara en que tabla escribe, cual es su
clave unica real, que campos pide y que permiso hace falta. Lo que no este aqui no se
puede registrar por este camino.

**Acotado a las marcas de la persona.** Karen ve Marketplace: puede registrar gasto de
Marketplace y nada mas. La misma regla que las lecturas y que la carga de archivos.

**Si falta un dato, se pregunta.** `plan_entry` no rellena huecos con supuestos: devuelve
que falta y el modelo lo pide. Un asistente que adivina la fecha o el pais escribe en el
dia equivocado y nadie se entera hasta el cierre de mes.

**Si ya habia un valor, se muestra.** `update_or_create` sobre la clave unica es lo
correcto --volver a registrar el mismo dia corrige, no duplica-- pero pisar 400.000 con
320.000 sin que nadie lo vea seria peor que duplicar. El plan trae el antes.
"""
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from django.utils import timezone

from ..models import AdPlatform, BusinessUnit, Channel, Country, DailyAdSpend, DailyChannelSale
from ..services.common import normalize_text


class EntryError(ValueError):
    """No se puede registrar. El mensaje se le muestra a la persona."""


class MissingData(EntryError):
    """Falta un dato para poder registrar. El modelo tiene que preguntarlo."""


def _amount(valor, campo):
    try:
        numero = Decimal(str(valor).strip().replace(".", "").replace(",", "."))
    except (TypeError, InvalidOperation):
        raise EntryError(f"'{valor}' no es un monto que pueda leer para {campo}.")
    if numero < 0:
        raise EntryError(f"{campo} no puede ser negativo.")
    return numero


def _count(valor, campo):
    try:
        numero = int(Decimal(str(valor).strip()))
    except (TypeError, InvalidOperation, ValueError):
        raise EntryError(f"'{valor}' no es un numero entero para {campo}.")
    if numero < 0:
        raise EntryError(f"{campo} no puede ser negativo.")
    return numero


def _day(valor):
    """La fecha, sin adivinar. 'ayer' y 'hoy' se aceptan porque son inequivocos."""
    texto = normalize_text(valor)
    hoy = timezone.localdate()
    if texto in {"hoy", "today"}:
        return hoy
    if texto in {"ayer", "yesterday"}:
        return hoy - timedelta(days=1)
    if texto in {"antier", "anteayer"}:
        return hoy - timedelta(days=2)
    try:
        dia = date.fromisoformat(str(valor).strip())
    except (TypeError, ValueError):
        raise MissingData(
            f"No entiendo la fecha '{valor}'. Preguntale el dia exacto (AAAA-MM-DD)."
        )
    if dia > hoy:
        raise EntryError(f"{dia.isoformat()} es en el futuro: no se puede registrar todavia.")
    return dia


def _resolve(modelo, nombre, etiqueta, activos=True):
    consulta = modelo.objects.filter(is_active=True) if activos else modelo.objects.all()
    pedido = normalize_text(nombre)
    if not pedido:
        raise MissingData(f"Falta el {etiqueta}. Preguntalo antes de registrar.")
    for objeto in consulta:
        if pedido in (normalize_text(objeto.name), normalize_text(getattr(objeto, "slug", ""))):
            return objeto
    for objeto in consulta:
        if pedido in normalize_text(objeto.name) or normalize_text(objeto.name) in pedido:
            return objeto
    disponibles = ", ".join(o.name for o in consulta[:15])
    raise EntryError(f"No encuentro el {etiqueta} '{nombre}'. Hay: {disponibles}.")


# Lo que se puede registrar hablando. Corto y explicito: cada entrada tiene detras una
# tabla con clave unica real, y por eso volver a registrar corrige en vez de duplicar.
ENTRY_TYPES = {
    "gasto_publicitario": {
        "label": "Inversion publicitaria de un dia",
        "permission": "reports.change_dailyadspend",
        "campos": "marca, pais, plataforma, fecha y monto",
        "ejemplo": "El gasto de Meta en Bali Colombia del 3 de agosto fue 320.000",
    },
    "ventas_de_canal": {
        "label": "Ventas de un canal en un dia",
        "permission": "reports.change_dailychannelsale",
        "campos": "marca, pais, canal, fecha y monto (pedidos y unidades opcionales)",
        "ejemplo": "Ayer WhatsApp de Bali vendio 1.250.000 en 8 pedidos",
    },
}


def describe_entry_types(user):
    """Que puede registrar esta persona hablando, y con que datos."""
    from .tools import allowed_business_units

    return {
        "marcas_que_puede_actualizar": [b.name for b in allowed_business_units(user)],
        "tipos": [
            {
                "tipo": clave,
                "que_registra": spec["label"],
                "datos_que_necesito": spec["campos"],
                "ejemplo": spec["ejemplo"],
                "puede": user.has_perm(spec["permission"]),
            }
            for clave, spec in ENTRY_TYPES.items()
        ],
        "nota": (
            "Registrar el mismo dia dos veces corrige el valor, no lo suma: la tabla tiene "
            "clave unica por marca, pais, canal/plataforma y fecha."
        ),
    }


def _resolve_country(marca, nombre):
    """El pais del dato. Si la marca solo vende en uno, no se pregunta.

    Bali vende solo en Colombia: preguntarle a Estefy "¿de que pais?" todos los dias es
    ruido, y el ruido entrena a la gente a contestar sin leer. Se pregunta cuando hay mas
    de una respuesta posible, no por costumbre.
    """
    paises = list(marca.countries.filter(is_active=True))
    if not nombre and len(paises) == 1:
        return paises[0]
    if not nombre and paises:
        raise MissingData(
            f"{marca.name} vende en {', '.join(p.name for p in paises)}. Preguntale de cual es."
        )
    pais = _resolve(Country, nombre, "pais")
    if paises and pais.pk not in {p.pk for p in paises}:
        raise EntryError(
            f"{marca.name} no tiene {pais.name} entre sus paises "
            f"({', '.join(p.name for p in paises)}). Confirmalo antes de registrar."
        )
    return pais


def _check_brand(user, marca):
    from .tools import allowed_business_units

    permitidas = allowed_business_units(user)
    if marca.pk not in {b.pk for b in permitidas}:
        raise EntryError(
            f"'{marca.name}' no esta entre las marcas que ves en Axis "
            f"({', '.join(b.name for b in permitidas)}), asi que no puedes registrar sus datos."
        )


def plan_entry(user, kind, **datos):
    """Valida el dato y devuelve antes/despues. No escribe nada."""
    tipo = str(kind or "").strip().lower()
    if tipo not in ENTRY_TYPES:
        raise EntryError(f"'{kind}' no se puede registrar. Se puede: {', '.join(ENTRY_TYPES)}.")

    marca = _resolve(BusinessUnit, datos.get("marca"), "marca")
    _check_brand(user, marca)
    pais = _resolve_country(marca, datos.get("pais"))
    dia = _day(datos.get("fecha"))
    monto = _amount(datos.get("monto"), "el monto")

    if tipo == "gasto_publicitario":
        plataforma = _resolve(AdPlatform, datos.get("plataforma"), "plataforma de pauta")
        existente = DailyAdSpend.objects.filter(
            business_unit=marca, country=pais, ad_platform=plataforma, spend_date=dia
        ).first()
        return {
            "tipo": tipo,
            "que": f"Inversion de {plataforma.name} en {marca.name} {pais.name}",
            "fecha": dia.isoformat(),
            "antes": str(existente.spend_amount) if existente else "sin dato",
            "despues": str(monto),
            "claves": {
                "marca": marca.slug, "pais": pais.code,
                "plataforma": plataforma.slug, "fecha": dia.isoformat(),
            },
            "monto": str(monto),
            "permiso": ENTRY_TYPES[tipo]["permission"],
            "aviso": (
                f"Ya habia un dato para ese dia ({existente.spend_amount}): esto lo reemplaza."
                if existente
                else ""
            ),
        }

    canal = _resolve(Channel, datos.get("canal"), "canal")
    pedidos = _count(datos.get("pedidos") or 0, "los pedidos")
    unidades = _count(datos.get("unidades") or 0, "las unidades")
    existente = DailyChannelSale.objects.filter(
        business_unit=marca, country=pais, channel=canal, sale_date=dia
    ).first()
    return {
        "tipo": tipo,
        "que": f"Ventas de {canal.name} en {marca.name} {pais.name}",
        "fecha": dia.isoformat(),
        "antes": str(existente.sales_amount) if existente else "sin dato",
        "despues": str(monto),
        "claves": {
            "marca": marca.slug, "pais": pais.code,
            "canal": canal.slug, "fecha": dia.isoformat(),
        },
        "monto": str(monto),
        "pedidos": pedidos,
        "unidades": unidades,
        "permiso": ENTRY_TYPES[tipo]["permission"],
        "aviso": (
            f"Ya habia ventas para ese dia ({existente.sales_amount}): esto las reemplaza."
            if existente
            else ""
        ),
    }


def apply_entry(user, kind, **datos):
    """Registra el dato. Solo se llama tras confirmacion humana."""
    from ..integrations.run_log import track_run
    from ..models import AiConfigChange

    plan = plan_entry(user, kind, **datos)
    if not user.has_perm(plan["permiso"]):
        raise EntryError(f"Te falta el permiso {plan['permiso']}.")

    claves = plan["claves"]
    marca = BusinessUnit.objects.get(slug=claves["marca"])
    pais = Country.objects.get(code=claves["pais"])
    dia = date.fromisoformat(claves["fecha"])
    monto = Decimal(plan["monto"])

    if plan["tipo"] == "gasto_publicitario":
        DailyAdSpend.objects.update_or_create(
            business_unit=marca, country=pais,
            ad_platform=AdPlatform.objects.get(slug=claves["plataforma"]), spend_date=dia,
            defaults={
                "spend_amount": monto,
                "source_type": DailyAdSpend.SourceType.MANUAL,
                "source_file": f"asistente ({user.get_username()})",
            },
        )
    else:
        DailyChannelSale.objects.update_or_create(
            business_unit=marca, country=pais,
            channel=Channel.objects.get(slug=claves["canal"]), sale_date=dia,
            defaults={
                "sales_amount": monto,
                "order_count": plan["pedidos"],
                "units": plan["unidades"],
                "source_type": DailyChannelSale.SourceType.MANUAL,
                "source_file": f"asistente ({user.get_username()})",
            },
        )

    with track_run(
        "IA dato manual", command=f"{plan['tipo']} (via asistente, {user.get_username()})",
        target_date=dia,
    ) as run:
        run.summary = f"{plan['que']} {plan['fecha']}: {plan['antes']} -> {plan['despues']}"

    AiConfigChange.objects.create(
        user=user, target=plan["tipo"], object_label=f"{plan['que']} {plan['fecha']}",
        field="monto", old_value=plan["antes"], new_value=plan["despues"],
    )
    return {**plan, "aplicado": True, "nota": "Este dato SI quedo registrado en Axis."}


# La plantilla que el admin ya exporta ("DESCARGAR PLANTILLA" en Ventas Marketplace).
# Cabecera en la fila 1 y una fila por dia/pais/canal.
PLANTILLA_COLUMNAS = ("fecha", "pais", "canal", "ventas", "inversion", "pedidos", "unidades")
PLANTILLA_MINIMAS = {"fecha", "pais", "canal"}
MAX_FILAS_PLANTILLA = 400


def plan_workbook(user, ruta, marca, hoja_nombre=""):
    """Cada fila de la plantilla, validada como si la hubieran dictado.

    No hay importador nuevo: la plantilla es N entradas, y pasan por el mismo
    `plan_entry` que el chat. Asi un archivo no puede escribir lo que un dictado no
    podria --otra marca, otro pais, una fecha futura-- ni saltarse el aviso de que ya
    habia un valor.
    """
    from openpyxl import load_workbook

    libro = load_workbook(filename=ruta, read_only=True, data_only=True)
    try:
        hoja = libro[hoja_nombre] if hoja_nombre else libro.active
        filas = list(hoja.iter_rows(min_row=1, values_only=True))
    finally:
        libro.close()

    if not filas:
        raise EntryError("El archivo esta vacio.")
    cabecera = [normalize_text(c) for c in filas[0]]
    if not PLANTILLA_MINIMAS.issubset(set(cabecera)):
        raise EntryError(
            "Esta hoja no tiene la forma de la plantilla. Se esperan las columnas "
            f"{', '.join(PLANTILLA_COLUMNAS)} en la primera fila."
        )

    indice = {nombre: cabecera.index(nombre) for nombre in cabecera if nombre}
    planes = []
    problemas = []
    for numero, fila in enumerate(filas[1:], start=2):
        if not any(c is not None and str(c).strip() for c in (fila or [])):
            continue
        if len(planes) + len(problemas) >= MAX_FILAS_PLANTILLA:
            problemas.append({"fila": numero, "error": "Se alcanzo el tope de filas por archivo."})
            break

        def celda(nombre):
            posicion = indice.get(nombre)
            if posicion is None or posicion >= len(fila):
                return None
            return fila[posicion]

        fecha = celda("fecha")
        # openpyxl devuelve datetime en las celdas con formato de fecha.
        fecha = fecha.date().isoformat() if hasattr(fecha, "date") else fecha
        comun = {"marca": marca, "pais": celda("pais"), "fecha": fecha}

        for tipo, columna, extra in (
            ("ventas_de_canal", "ventas", {"canal": celda("canal"), "pedidos": celda("pedidos") or 0,
                                           "unidades": celda("unidades") or 0}),
            ("gasto_publicitario", "inversion", {"plataforma": celda("canal")}),
        ):
            valor = celda(columna)
            if valor is None or str(valor).strip() in ("", "0"):
                # Una celda vacia o en cero no es un dato: la plantilla trae las dos
                # columnas siempre, y escribir ceros borraria lo que ya hubiera.
                continue
            try:
                planes.append({"fila": numero, **plan_entry(user, tipo, monto=valor, **comun, **extra)})
            except EntryError as exc:
                problemas.append({"fila": numero, "columna": columna, "error": str(exc)})

    return {"planes": planes, "problemas": problemas}


def apply_workbook(user, ruta, marca, hoja_nombre=""):
    """Registra todas las filas validadas. Las que fallaron se reportan, no se inventan."""
    from ..integrations.run_log import track_run

    revision = plan_workbook(user, ruta, marca, hoja_nombre)
    aplicados = 0
    with track_run(
        "IA plantilla", command=f"{marca} (via asistente, {user.get_username()})"
    ) as run:
        for plan in revision["planes"]:
            datos = {"marca": marca, "pais": None, "fecha": plan["claves"]["fecha"],
                     "monto": plan["monto"]}
            claves = plan["claves"]
            datos["pais"] = claves["pais"]
            if plan["tipo"] == "gasto_publicitario":
                datos["plataforma"] = claves["plataforma"]
            else:
                datos["canal"] = claves["canal"]
                datos["pedidos"] = plan.get("pedidos") or 0
                datos["unidades"] = plan.get("unidades") or 0
            apply_entry(user, plan["tipo"], **datos)
            aplicados += 1
        run.summary = (
            f"{aplicados} filas registradas, {len(revision['problemas'])} con problema"
        )
    return {**revision, "aplicados": aplicados,
            "nota": f"{aplicados} datos quedaron registrados en Axis."}
