"""Diagnostico y previsualizacion de un Excel que alguien le subio al asistente.

Dos decisiones cargan todo el peso de este modulo:

**La IA no escribe codigo de parseo.** Ni aqui ni en ningun lado. Las columnas se
normalizan con `normalize_header`, los numeros con `parse_decimal` y las fechas con
`parse_excel_date` —las mismas funciones que usan los importadores. Cuando habia cinco
copias de "convertir a numero", el mismo bug de la coma decimal hubo que arreglarlo dos
veces. Una sexta copia dentro de la IA seria la peor de todas: nadie la revisaria.

**El diff sale de correr el importador de verdad y deshacer la transaccion.** No de
predecir lo que haria. Un diff calculado aparte es una segunda implementacion que se
desincroniza con la primera, y en el peor momento: cuando alguien confia en la vista
previa para aprobar una escritura. Aqui el importador corre completo dentro de un
`atomic()` que se revierte, se miden los totales antes y despues, y se descarta.
"""
import tempfile
from contextlib import contextmanager
from pathlib import Path

from django.core.management import call_command
from django.db import transaction
from django.db.models import Sum
from openpyxl import load_workbook

from ..models import DailyAdSpend, DailyChannelSale, DailyProductCategorySale
from ..services.common import parse_excel_date
from ..utils.numbers import normalize_header, parse_decimal
from .attachments import list_attachments

# Filas que se revisan buscando la cabecera. Las hojas de despachos la traen en la 14,
# asi que asumir la 1 descartaba el archivo entero.
HEADER_SEARCH_ROWS = 30
SAMPLE_ROWS = 5
MIN_HEADER_CELLS = 3

# Que forma tiene cada archivo conocido y con que importador se procesa. La lista es
# corta a proposito: solo entra lo que ya tiene un importador probado detras.
KNOWN_SHAPES = (
    {
        "key": "despachos_ecuador",
        "label": "Despachos / ventas de Uva Ecuador",
        "required": {"fecha", "producto", "cantidad"},
        "optional": {"valor", "moneda", "centro de costos", "envio"},
        "command": "fetch_onedrive_excel",
        # El canal por defecto del importador es el de Colombia. Sin fijarlo aqui, un
        # archivo de Ecuador cuyas filas no traigan CENTRO DE COSTOS aterrizaba en
        # whatsapp-uva-co: ventas de Ecuador contadas como de Colombia.
        "extra_args": [
            "--country", "EC", "--business-unit", "uva",
            "--channel-slug", "whatsapp-uva-ec", "--sync-axis",
        ],
        "note": (
            "En esta hoja VALOR es precio unitario: hay que multiplicarlo por CANTIDAD. "
            "El importador ya lo hace."
        ),
    },
    {
        "key": "ventas_whatsapp_colombia",
        "label": "Ventas por WhatsApp de Uva Colombia",
        "required": {"fecha", "producto", "cantidad"},
        "optional": {"ventas", "total cop", "valor", "canal", "origen"},
        "command": "fetch_onedrive_excel",
        "extra_args": [
            "--country", "CO", "--business-unit", "uva",
            "--channel-slug", "whatsapp-uva-co", "--sync-axis",
        ],
        "note": "",
    },
)


def _sheet_headers(hoja):
    """La fila que parece cabecera y sus columnas normalizadas."""
    mejor = (0, [], 0)
    for indice, fila in enumerate(
        hoja.iter_rows(min_row=1, max_row=HEADER_SEARCH_ROWS, values_only=True), start=1
    ):
        textos = [str(c).strip() for c in (fila or []) if c is not None and str(c).strip()]
        # Una cabecera es texto, no numeros: una fila de datos tiene importes.
        candidatas = [t for t in textos if parse_decimal(t, default=None) is None]
        if len(candidatas) >= MIN_HEADER_CELLS and len(candidatas) > mejor[2]:
            mejor = (indice, [normalize_header(t) for t in candidatas], len(candidatas))
    return mejor[0], mejor[1]


def _date_range(hoja, columnas, fila_cabecera):
    """Desde y hasta que dia trae la hoja, con el parser de fechas de siempre.

    El importador exige rango, y el archivo ya lo dice: pedirselo a la persona seria
    hacerle leer el Excel que acaba de subir para poder subirlo.
    """
    if "fecha" not in columnas:
        return None, None
    indice = columnas.index("fecha")
    fechas = []
    for fila in hoja.iter_rows(min_row=fila_cabecera + 1, values_only=True):
        if indice >= len(fila or []):
            continue
        dia = parse_excel_date(fila[indice])
        if dia:
            fechas.append(dia)
    if not fechas:
        return None, None
    return min(fechas), max(fechas)


def match_shape(columnas):
    """La forma conocida que mejor encaja, o None."""
    presentes = set(columnas)
    mejor = None
    for forma in KNOWN_SHAPES:
        if not forma["required"].issubset(presentes):
            continue
        puntaje = len(forma["optional"] & presentes)
        if mejor is None or puntaje > mejor[0]:
            mejor = (puntaje, forma)
    return mejor[1] if mejor else None


def describe_attachment(attachment):
    """Que hay dentro del archivo. Solo lectura, sin tocar la base."""
    if not attachment.original_name.lower().endswith((".xlsx", ".xls")):
        return {"error": "Por ahora solo puedo mirar dentro de archivos .xlsx o .xls."}

    try:
        hojas = _sheets_of(attachment)
    except AttachmentGone as exc:
        return {"error": str(exc)}

    return {
        "archivo": attachment.original_name,
        "hojas": hojas,
        "nota": (
            "Si ninguna hoja tiene forma reconocida, no hay importador para este archivo "
            "y no se puede cargar. Decirlo es la respuesta correcta."
        ),
    }


def _sheets_of(attachment):
    """Lo que trae cada hoja. Separado para que el error de storage se atrape arriba."""
    with materialized(attachment) as ruta:
        libro = load_workbook(filename=ruta, read_only=True, data_only=True)
        try:
            hojas = []
            for nombre in libro.sheetnames:
                hoja = libro[nombre]
                fila_cabecera, columnas = _sheet_headers(hoja)
                muestra = []
                if fila_cabecera:
                    for fila in hoja.iter_rows(
                        min_row=fila_cabecera + 1,
                        max_row=fila_cabecera + SAMPLE_ROWS,
                        values_only=True,
                    ):
                        if any(c is not None for c in (fila or [])):
                            muestra.append([str(c)[:40] if c is not None else "" for c in fila[:10]])
                forma = match_shape(columnas)
                desde, hasta = _date_range(hoja, columnas, fila_cabecera) if fila_cabecera else (None, None)
                hojas.append({
                    "hoja": nombre,
                    "filas": hoja.max_row,
                    "fila_de_cabecera": fila_cabecera or None,
                    "columnas": columnas,
                    "desde": desde.isoformat() if desde else None,
                    "hasta": hasta.isoformat() if hasta else None,
                    "forma_reconocida": forma["key"] if forma else None,
                    "forma_label": forma["label"] if forma else "",
                    "nota_de_la_forma": forma["note"] if forma else "",
                    "muestra": muestra,
                })
        finally:
            libro.close()
    return hojas


class AttachmentGone(Exception):
    """La fila existe pero el archivo ya no esta en el storage."""


@contextmanager
def materialized(attachment):
    """Baja el archivo del bucket a un temporal, porque los importadores leen rutas."""
    sufijo = Path(attachment.original_name).suffix or ".xlsx"
    temporal = tempfile.NamedTemporaryFile(suffix=sufijo, delete=False)
    try:
        try:
            attachment.file.open("rb")
        except (FileNotFoundError, OSError) as exc:
            raise AttachmentGone(
                f"'{attachment.original_name}' esta registrado pero su contenido ya no "
                "esta guardado. Habria que volver a subirlo."
            ) from exc
        try:
            for trozo in attachment.file.chunks():
                temporal.write(trozo)
        finally:
            attachment.file.close()
        temporal.close()
        yield temporal.name
    finally:
        # En Windows no se puede borrar un archivo que siga abierto. Si algo fallo antes
        # de cerrarlo, sin esto el PermissionError de la limpieza tapaba el error real.
        if not temporal.closed:
            temporal.close()
        try:
            Path(temporal.name).unlink(missing_ok=True)
        except OSError:
            # Dejar un temporal huerfano no vale tumbar una respuesta del chat.
            pass


def _totals():
    """Los totales que el importador podria mover. Se miden antes y despues."""
    return {
        "ventas_por_canal": {
            "filas": DailyChannelSale.objects.count(),
            "monto": float(
                DailyChannelSale.objects.aggregate(t=Sum("sales_amount"))["t"] or 0
            ),
        },
        "ventas_por_categoria": {
            "filas": DailyProductCategorySale.objects.count(),
            "monto": float(
                DailyProductCategorySale.objects.aggregate(t=Sum("sales_amount"))["t"] or 0
            ),
        },
        "inversion": {
            "filas": DailyAdSpend.objects.count(),
            "monto": float(DailyAdSpend.objects.aggregate(t=Sum("spend_amount"))["t"] or 0),
        },
    }


def _diff(antes, despues):
    salida = {}
    for tabla, valores in despues.items():
        salida[tabla] = {
            "filas_nuevas": valores["filas"] - antes[tabla]["filas"],
            "cambio_de_monto": round(valores["monto"] - antes[tabla]["monto"], 2),
        }
    return salida


class ImportNotPossible(Exception):
    """No hay importador para este archivo, o falta un dato para correrlo."""


def _plan(attachment, sheet_name="", shape_key=""):
    """El comando exacto que se correria, sin correrlo."""
    descripcion = describe_attachment(attachment)
    if descripcion.get("error"):
        raise ImportNotPossible(descripcion["error"])


    candidatas = [h for h in descripcion["hojas"] if h["forma_reconocida"]]
    if sheet_name:
        candidatas = [h for h in candidatas if h["hoja"] == sheet_name]
    if shape_key:
        candidatas = [h for h in candidatas if h["forma_reconocida"] == shape_key]
    if not candidatas:
        raise ImportNotPossible(
            "Ninguna hoja de este archivo tiene una forma que Axis sepa importar. "
            f"Hojas revisadas: {', '.join(h['hoja'] for h in descripcion['hojas'])}."
        )
    if len(candidatas) > 1:
        raise ImportNotPossible(
            "Hay varias hojas importables: "
            f"{', '.join(h['hoja'] for h in candidatas)}. Pide que elijan una."
        )

    hoja = candidatas[0]
    if not hoja["desde"] or not hoja["hasta"]:
        raise ImportNotPossible(
            f"La hoja '{hoja['hoja']}' no tiene ninguna fecha que se pueda leer en la "
            "columna FECHA, y sin rango de fechas no hay nada que cargar."
        )
    forma = next(f for f in KNOWN_SHAPES if f["key"] == hoja["forma_reconocida"])
    return hoja, forma


def preview_import(attachment, sheet_name="", shape_key=""):
    """Corre el importador de verdad y deshace la transaccion. Nada queda escrito."""
    hoja, forma = _plan(attachment, sheet_name, shape_key)

    with materialized(attachment) as ruta:
        argumentos = [
            *forma["extra_args"],
            "--file", ruta,
            "--sheet", hoja["hoja"],
            "--header-row", str(hoja["fila_de_cabecera"] or 1),
            # El rango sale del archivo, no de lo que alguien recuerde.
            "--date-from", hoja["desde"],
            "--date-to", hoja["hasta"],
        ]
        antes = _totals()
        # El rollback es el que garantiza que esto sea una vista previa: si algo del
        # importador escribe, se revierte igual.
        try:
            with transaction.atomic():
                call_command(forma["command"], *argumentos)
                despues = _totals()
                diferencia = _diff(antes, despues)
                raise _Rollback()
        except _Rollback:
            pass

    return {
        "archivo": attachment.original_name,
        "hoja": hoja["hoja"],
        "forma": forma["label"],
        "periodo_del_archivo": f"{hoja['desde']} a {hoja['hasta']}",
        "comando": f"{forma['command']} {' '.join(forma['extra_args'])}",
        "cambios": diferencia,
        "nota": (
            "Esto es una simulacion: se corrio el importador real y se deshizo la "
            "transaccion. Nada quedo escrito." + (f" {forma['note']}" if forma["note"] else "")
        ),
    }


class _Rollback(Exception):
    """Aborta el atomic() a proposito para deshacer la simulacion."""


def apply_import(attachment, user, sheet_name="", shape_key=""):
    """Corre el importador de verdad y **si** deja los cambios escritos.

    Es el mismo camino de `preview_import` sin el rollback: el mismo comando, los mismos
    argumentos, el mismo rango sacado del archivo. Que sean el mismo camino es el punto:
    lo que la persona aprobo en la vista previa es lo que se ejecuta.
    """
    from ..integrations.run_log import track_run

    hoja, forma = _plan(attachment, sheet_name, shape_key)

    with materialized(attachment) as ruta:
        argumentos = [
            *forma["extra_args"],
            "--file", ruta,
            "--sheet", hoja["hoja"],
            "--header-row", str(hoja["fila_de_cabecera"] or 1),
            "--date-from", hoja["desde"],
            "--date-to", hoja["hasta"],
        ]
        antes = _totals()
        # Queda en la bitacora igual que cualquier importacion: quien la lanzo tiene que
        # poder verse despues, no solo las que corren en el job diario.
        with track_run(
            f"IA carga {forma['key']}",
            command=f"{forma['command']} (via asistente, {user.get_username()})",
        ) as run:
            call_command(forma["command"], *argumentos)
            despues = _totals()
            cambios = _diff(antes, despues)
            run.summary = (
                f"{attachment.original_name} hoja {hoja['hoja']} "
                f"({hoja['desde']}..{hoja['hasta']}): "
                f"{cambios['ventas_por_canal']['filas_nuevas']} filas de venta, "
                f"{cambios['ventas_por_canal']['cambio_de_monto']:.0f} COP"
            )

    return {
        "archivo": attachment.original_name,
        "hoja": hoja["hoja"],
        "periodo_del_archivo": f"{hoja['desde']} a {hoja['hasta']}",
        "cambios": cambios,
        "nota": "Esto SI quedo escrito en Axis.",
    }


def attachment_for(user, attachment_id):
    return list_attachments(user).filter(pk=attachment_id).first()
